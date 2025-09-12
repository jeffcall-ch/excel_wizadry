import pandas as pd
import numpy as np
from datetime import datetime
import os
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
except Exception:
    # openpyxl is optional; if missing, XLSX output will be skipped
    load_workbook = None
    PatternFill = None
    get_column_letter = None
    Alignment = None

def compare_po_vs_bom(po_file_path, bom_file_path, output_dir=None):
    """
    Compare PO and BOM files and generate a comprehensive comparison report.
    
    Args:
        po_file_path (str): Path to the PO CSV file
        bom_file_path (str): Path to the BOM CSV file
        output_dir (str): Directory to save the output file (optional)
    
    Returns:
        str: Path to the generated output file
    """
    
    # Read the CSV files
    print("Reading PO file...")
    po_df = pd.read_csv(po_file_path)
    
    print("Reading BOM file...")
    bom_df = pd.read_csv(bom_file_path)
    
    # Display basic info about the files
    print(f"PO file contains {len(po_df)} rows")
    print(f"BOM file contains {len(bom_df)} rows")
    
    # Create a copy of PO dataframe to work with
    result_df = po_df.copy()
    
    # Step 1: Add new columns to the PO layout
    print("Step 1: Adding new BOM columns to PO layout...")
    result_df['BOM QTY [pcs/m]'] = np.nan
    result_df['BOM Total weight [kg]'] = np.nan
    
    # Step 3: Add difference columns
    result_df['QTY Difference [pcs/m]'] = np.nan
    result_df['Weight Difference [kg]'] = pd.Series(dtype='object')  # Allow mixed data types (numbers and strings)
    result_df['Status'] = ''
    
    # Step 2: Match PO lines with BOM lines and populate BOM data
    print("Step 2: Matching PO lines with BOM data...")
    
    # Create a mapping dictionary for faster lookup
    bom_mapping = {}
    duplicate_count = 0
    for idx, row in bom_df.iterrows():
        pipe_component = str(row['Pipe Component']).strip()
        if pipe_component in bom_mapping:
            duplicate_count += 1
            # For duplicates, keep the first occurrence but update the tracking
            continue
        bom_mapping[pipe_component] = {
            'Total weight [kg]': row['Total weight [kg]'],
            'QTY [pcs/m]': row['QTY [pcs/m]'],
            'matched': False  # Track which BOM items have been matched
        }
    
    if duplicate_count > 0:
        print(f"Note: Found {duplicate_count} duplicate pipe components in BOM, keeping first occurrence of each")
    
    matched_bom_items = set()
    
    for idx, row in result_df.iterrows():
        po_description = str(row['DESCRIPTION/ Pipe Component']).strip()
        
        if po_description in bom_mapping:
            # Found a match
            bom_data = bom_mapping[po_description]
            
            # Populate BOM data
            result_df.at[idx, 'BOM QTY [pcs/m]'] = bom_data['QTY [pcs/m]']
            result_df.at[idx, 'BOM Total weight [kg]'] = bom_data['Total weight [kg]']
            
            # Mark this BOM item as matched
            matched_bom_items.add(po_description)
            bom_mapping[po_description]['matched'] = True
            
            # Step 3: Calculate differences
            # Handle PO quantities - they might be in different columns
            po_qty = pd.to_numeric(row.get('PO QUANTITY\n (pcs./m)', 0), errors='coerce')
            if pd.isna(po_qty):
                po_qty = 0
                
            po_weight = pd.to_numeric(row.get('PO WEIGHT\n [kg]', 0), errors='coerce')
            if pd.isna(po_weight):
                po_weight = 0
            
            bom_qty = pd.to_numeric(bom_data['QTY [pcs/m]'], errors='coerce')
            bom_weight_raw = bom_data['Total weight [kg]']
            bom_weight = pd.to_numeric(bom_weight_raw, errors='coerce')
            
            if pd.isna(bom_qty):
                bom_qty = 0
            
            # Calculate quantity difference (BOM - PO, positive means BOM has more)
            qty_diff = bom_qty - po_qty
            
            # Only show quantity differences if they are not zero
            if qty_diff != 0:
                result_df.at[idx, 'QTY Difference [pcs/m]'] = qty_diff
            
            # Handle weight difference - only calculate if BOM has weight data
            if pd.isna(bom_weight) or bom_weight_raw == '' or str(bom_weight_raw).strip() == '':
                # BOM weight is missing - show N/A instead of calculating difference
                result_df.at[idx, 'Weight Difference [kg]'] = 'N/A (BOM weight missing)'
            else:
                # BOM has weight data - calculate difference
                weight_diff = bom_weight - po_weight
                if weight_diff != 0:
                    result_df.at[idx, 'Weight Difference [kg]'] = weight_diff
                
            result_df.at[idx, 'Status'] = 'Matched'
            
        else:
            # Step 4: Item is in PO but not in BOM (removed item)
            po_qty = pd.to_numeric(row.get('PO QUANTITY\n (pcs./m)', 0), errors='coerce')
            po_weight = pd.to_numeric(row.get('PO WEIGHT\n [kg]', 0), errors='coerce')
            
            if pd.isna(po_qty):
                po_qty = 0
            if pd.isna(po_weight):
                po_weight = 0
            
            # Show negative values to indicate removal
            if po_qty > 0:
                result_df.at[idx, 'QTY Difference [pcs/m]'] = -po_qty
            if po_weight > 0:
                result_df.at[idx, 'Weight Difference [kg]'] = -po_weight
                
            result_df.at[idx, 'Status'] = 'Removed from BOM'
    
    # Step 5: Add new BOM items that are not in PO
    print("Step 5: Adding new BOM items not present in PO...")
    
    # We'll insert new BOM items into the result while preserving the PO row order.
    # Insertion strategy (hard-coded per request):
    # 1) Insert after last PO row where both Category and TYPE match (case-insensitive, trimmed).
    # 2) If no TYPE match, insert after last PO row with same Category.
    # 3) If no Category match, append at the end.
    # Multiple new items that map to the same insertion point are inserted as a contiguous block
    # in the order they appear in the BOM.

    # Prepare a map insertion_index -> list of new rows
    insertion_map = {}

    # Use the current result_df (which reflects the PO order and includes any matched/removed marks)
    po_snapshot = result_df.copy()
    po_len = len(po_snapshot)

    for pipe_component, bom_data in bom_mapping.items():
        if not bom_data['matched']:
            # Find the original BOM row for this item
            bom_rows = bom_df[bom_df['Pipe Component'] == pipe_component]
            if bom_rows.empty:
                print(f"Warning: Could not find BOM row for component: {pipe_component}")
                continue

            bom_row = bom_rows.iloc[0]  # Take the first match if there are duplicates

            # Create a new row based on PO structure but with BOM data
            new_row = pd.Series(index=result_df.columns, dtype=object)

            # Fill in available data from BOM
            new_row['Category'] = bom_row.get('Category', '')
            new_row['MATERIAL'] = bom_row.get('Material', '')
            new_row['TYPE'] = bom_row.get('Type', '')
            new_row['DESCRIPTION/ Pipe Component'] = pipe_component

            # BOM quantities
            new_row['BOM QTY [pcs/m]'] = bom_data['QTY [pcs/m]']
            new_row['BOM Total weight [kg]'] = bom_data['Total weight [kg]']

            # Since this is new, differences equal BOM values
            bom_qty = pd.to_numeric(bom_data['QTY [pcs/m]'], errors='coerce')
            bom_weight_raw = bom_data['Total weight [kg]']
            bom_weight = pd.to_numeric(bom_weight_raw, errors='coerce')

            if not pd.isna(bom_qty) and bom_qty > 0:
                new_row['QTY Difference [pcs/m]'] = bom_qty

            # Handle weight difference - only show if BOM has weight data
            if pd.isna(bom_weight) or bom_weight_raw == '' or str(bom_weight_raw).strip() == '':
                new_row['Weight Difference [kg]'] = 'N/A (BOM weight missing)'
            elif bom_weight > 0:
                new_row['Weight Difference [kg]'] = bom_weight

            new_row['Status'] = 'New in BOM'

            # Determine insertion index based on Category and TYPE (case-insensitive)
            bom_cat = str(bom_row.get('Category', '')).strip().lower()
            bom_type = str(bom_row.get('Type', '')).strip().lower()

            insertion_idx = None

            if bom_cat != '' and bom_type != '':
                matches = po_snapshot[
                    (po_snapshot['Category'].astype(str).str.strip().str.lower() == bom_cat) &
                    (po_snapshot['TYPE'].astype(str).str.strip().str.lower() == bom_type)
                ]
                if not matches.empty:
                    insertion_idx = int(matches.index.max())

            if insertion_idx is None and bom_cat != '':
                cat_matches = po_snapshot[
                    (po_snapshot['Category'].astype(str).str.strip().str.lower() == bom_cat)
                ]
                if not cat_matches.empty:
                    insertion_idx = int(cat_matches.index.max())

            # If no match found, append to end
            if insertion_idx is None:
                insertion_idx = po_len

            insertion_map.setdefault(insertion_idx, []).append(new_row)

    # Rebuild result dataframe by iterating PO rows and inserting blocks at the right points
    if insertion_map:
        rebuilt_rows = []
        for i in range(po_len):
            # Append the existing PO-derived row (preserves any earlier modifications)
            rebuilt_rows.append(po_snapshot.iloc[i].to_dict())

            # After this row, insert any new rows mapped to this index
            if i in insertion_map:
                for nr in insertion_map[i]:
                    rebuilt_rows.append(nr.to_dict())

        # Handle items that should be appended at the end
        if po_len in insertion_map:
            for nr in insertion_map[po_len]:
                rebuilt_rows.append(nr.to_dict())

        # Create a new DataFrame preserving original columns order
        result_df = pd.DataFrame(rebuilt_rows, columns=result_df.columns)
    
    # Step 6: Generate output file with smart naming
    print("Step 6: Generating output file...")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"PO_vs_BOM_Comparison_Report_{timestamp}.csv"
    
    if output_dir:
        output_path = os.path.join(output_dir, output_filename)
    else:
        # Save in the same directory as the script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_path = os.path.join(script_dir, output_filename)
    
    # Save the result
    result_df.to_csv(output_path, index=False)
    # Also save an Excel version with formatting if openpyxl is available
    try:
        xlsx_path = os.path.splitext(output_path)[0] + '.xlsx'
        if load_workbook is not None:
            # Use pandas ExcelWriter which will use openpyxl engine by default when available
            with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Comparison')

            # Load workbook to apply formatting
            wb = load_workbook(xlsx_path)
            ws = wb.active

            # Freeze panes at F2 (freeze rows above row 2 and columns left of F)
            try:
                ws.freeze_panes = 'F2'
            except Exception:
                pass

            # Add autofilter for header row over full data range
            try:
                last_col = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
            except Exception:
                pass

            # Auto width columns based on max length in each column
            try:
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    for cell in ws[col_letter]:
                        try:
                            val = cell.value
                            if val is None:
                                length = 0
                            else:
                                length = len(str(val))
                        except Exception:
                            length = 0
                        if length > max_length:
                            max_length = length
                    # Set a sensible cap and add padding
                    adjusted_width = (max_length + 2) if max_length > 0 else 8
                    if adjusted_width > 60:
                        adjusted_width = 60
                    ws.column_dimensions[col_letter].width = adjusted_width
            except Exception:
                pass

            # Center text horizontally and vertically for all cells
            try:
                if Alignment is not None:
                    center_align = Alignment(horizontal='center', vertical='center')
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            # preserve wrap_text if already set for given cell later
                            cell.alignment = center_align
            except Exception:
                pass

            # Wrap text in column E (5th column)
            try:
                if Alignment is not None and ws.max_column >= 5:
                    col_e_idx = 5
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_e_idx, max_col=col_e_idx):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except Exception:
                pass

            # Auto height rows based on approximate wrapped lines for column widths
            try:
                # Build column width map (fallback to 8)
                col_widths = {}
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    w = ws.column_dimensions[col_letter].width
                    col_widths[col_idx] = float(w) if w is not None else 8.0

                for row_idx in range(1, ws.max_row + 1):
                    max_lines = 1
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        val = '' if cell.value is None else str(cell.value)
                        # Only consider wrapping for column E; other columns treated as single line
                        if col_idx == 5:
                            width = max(1.0, col_widths.get(col_idx, 8.0))
                            # approximate characters per line ~= width
                            lines = max(1, int((len(val) + width - 1) // width))
                        else:
                            lines = 1
                        if lines > max_lines:
                            max_lines = lines
                    # approximate row height: 15 points per line
                    row_height = max(15, int(max_lines * 15))
                    ws.row_dimensions[row_idx].height = row_height
            except Exception:
                pass

            # Color rows by Status: New in BOM -> blue, Removed from BOM -> red
            if PatternFill is not None:
                blue_fill = PatternFill(fill_type='solid', start_color='ADD8E6')
                red_fill = PatternFill(fill_type='solid', start_color='FFC7CE')
                # find Status column index
                status_col = None
                for cell in ws[1]:
                    if str(cell.value).strip() == 'Status':
                        status_col = cell.column
                        break

                if status_col is not None:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=status_col)
                        try:
                            val = str(cell.value).strip()
                        except Exception:
                            val = ''
                        if val == 'New in BOM':
                            for c in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=c).fill = blue_fill
                        elif val == 'Removed from BOM':
                            for c in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=c).fill = red_fill

            # Save workbook
            try:
                wb.save(xlsx_path)
            except Exception:
                pass
        else:
            print('openpyxl not available: skipping XLSX output')
    except Exception:
        print('Warning: failed to write XLSX output')
    
    # Generate summary statistics
    print("\n" + "="*60)
    print("COMPARISON SUMMARY")
    print("="*60)
    
    matched_items = len(result_df[result_df['Status'] == 'Matched'])
    removed_items = len(result_df[result_df['Status'] == 'Removed from BOM'])
    new_items = len(result_df[result_df['Status'] == 'New in BOM'])
    
    print(f"Total items in PO: {len(po_df)}")
    print(f"Total items in BOM: {len(bom_df)}")
    print(f"Matched items: {matched_items}")
    print(f"Items removed from BOM: {removed_items}")
    print(f"New items in BOM: {new_items}")
    print(f"Total items in comparison report: {len(result_df)}")
    
    print(f"\nOutput file saved as: {output_path}")
    print("="*60)
    
    return output_path

def main():
    """
    Main function to run the PO vs BOM comparison.
    """
    # Get the script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Define file paths
    po_file = os.path.join(script_dir, "PO.csv")
    bom_file = os.path.join(script_dir, "BOM.csv")
    
    # Check if files exist
    if not os.path.exists(po_file):
        print(f"Error: PO.csv not found at {po_file}")
        return
    
    if not os.path.exists(bom_file):
        print(f"Error: BOM.csv not found at {bom_file}")
        return
    
    print("Starting PO vs BOM Comparison...")
    print(f"PO file: {po_file}")
    print(f"BOM file: {bom_file}")
    
    try:
        output_path = compare_po_vs_bom(po_file, bom_file, script_dir)
        print(f"\nComparison completed successfully!")
        print(f"Report saved to: {output_path}")
        
    except Exception as e:
        print(f"Error during comparison: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
