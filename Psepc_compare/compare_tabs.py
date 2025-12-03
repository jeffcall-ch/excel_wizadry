"""
Excel Tab Comparison Script
Compares two tabs in an Excel file and creates a detailed comparison report.
Baseline: First tab (ACDE)
Comparison: Second tab (ACDM)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

# File path
FILE_PATH = r'C:\Users\szil\Repos\excel_wizadry\Psepc_compare\ACDE_ACDM_ADI.xlsx'

def normalize_columns(df):
    """Strip whitespace from column names and handle unnamed columns."""
    df.columns = df.columns.str.strip()
    # Remove completely unnamed/empty columns
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    return df

def create_row_key(row, key_columns):
    """Create a unique key for a row based on specified columns."""
    return tuple(str(row[col]).strip() if pd.notna(row[col]) else '' for col in key_columns)

def compare_tabs(file_path):
    """
    Compare two tabs in an Excel file.
    Returns comparison results and statistics.
    """
    print(f"Reading Excel file: {file_path}")
    xl = pd.ExcelFile(file_path)
    
    # Read both sheets
    sheet1_name = xl.sheet_names[0]
    sheet2_name = xl.sheet_names[1]
    
    print(f"Baseline sheet: {sheet1_name}")
    print(f"Comparison sheet: {sheet2_name}")
    
    df1 = pd.read_excel(xl, sheet1_name)
    df2 = pd.read_excel(xl, sheet2_name)
    
    # Normalize columns
    df1 = normalize_columns(df1)
    df2 = normalize_columns(df2)
    
    print(f"\n{sheet1_name} shape: {df1.shape}")
    print(f"{sheet2_name} shape: {df2.shape}")
    print(f"\n{sheet1_name} columns: {list(df1.columns)}")
    print(f"{sheet2_name} columns: {list(df2.columns)}")
    
    # Use all columns for comparison (or specify key columns)
    # Assuming all columns are relevant for identifying unique rows
    key_columns = list(df1.columns)
    
    # Create sets of row keys for comparison
    df1_keys = set()
    df1_dict = {}
    for idx, row in df1.iterrows():
        key = create_row_key(row, key_columns)
        df1_keys.add(key)
        df1_dict[key] = idx
    
    df2_keys = set()
    df2_dict = {}
    # Only use columns that exist in df1 for fair comparison
    comparison_columns = [col for col in key_columns if col in df2.columns]
    for idx, row in df2.iterrows():
        key = create_row_key(row, comparison_columns)
        df2_keys.add(key)
        df2_dict[key] = idx
    
    # Find differences
    identical_keys = df1_keys & df2_keys  # In both
    missing_in_sheet2 = df1_keys - df2_keys  # In sheet1 but not in sheet2
    extra_in_sheet2 = df2_keys - df1_keys  # In sheet2 but not in sheet1
    
    # Create comparison dataframe
    comparison_data = []
    
    # Add identical rows
    for key in identical_keys:
        row_data = df1.iloc[df1_dict[key]].to_dict()
        row_data['Status'] = 'IDENTICAL'
        row_data['Sheet1_Row'] = df1_dict[key] + 2  # +2 for Excel row (1-indexed + header)
        row_data['Sheet2_Row'] = df2_dict[key] + 2
        comparison_data.append(row_data)
    
    # Add missing rows (in baseline but not in comparison)
    for key in missing_in_sheet2:
        row_data = df1.iloc[df1_dict[key]].to_dict()
        row_data['Status'] = 'MISSING_IN_' + sheet2_name
        row_data['Sheet1_Row'] = df1_dict[key] + 2
        row_data['Sheet2_Row'] = ''
        comparison_data.append(row_data)
    
    # Add extra rows (in comparison but not in baseline)
    for key in extra_in_sheet2:
        row_data = df2.iloc[df2_dict[key]].to_dict()
        row_data['Status'] = 'EXTRA_IN_' + sheet2_name
        row_data['Sheet1_Row'] = ''
        row_data['Sheet2_Row'] = df2_dict[key] + 2
        comparison_data.append(row_data)
    
    # Create comparison dataframe
    df_comparison = pd.DataFrame(comparison_data)
    
    # Reorder columns to put Status and row references first
    status_cols = ['Status', 'Sheet1_Row', 'Sheet2_Row']
    other_cols = [col for col in df_comparison.columns if col not in status_cols]
    df_comparison = df_comparison[status_cols + other_cols]
    
    # Statistics
    stats = {
        'baseline_sheet': sheet1_name,
        'comparison_sheet': sheet2_name,
        'baseline_total': len(df1),
        'comparison_total': len(df2),
        'identical_count': len(identical_keys),
        'missing_count': len(missing_in_sheet2),
        'extra_count': len(extra_in_sheet2),
        'identical_pct': len(identical_keys) / len(df1) * 100 if len(df1) > 0 else 0,
        'missing_pct': len(missing_in_sheet2) / len(df1) * 100 if len(df1) > 0 else 0,
        'extra_pct': len(extra_in_sheet2) / len(df2) * 100 if len(df2) > 0 else 0,
    }
    
    return df_comparison, stats, sheet1_name, sheet2_name

def create_summary_sheet(stats):
    """Create a summary statistics dataframe."""
    summary_data = [
        ['Comparison Summary', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['Baseline Sheet', stats['baseline_sheet']],
        ['Comparison Sheet', stats['comparison_sheet']],
        ['', ''],
        ['Metric', 'Value'],
        ['Baseline Total Rows', stats['baseline_total']],
        ['Comparison Total Rows', stats['comparison_total']],
        ['', ''],
        ['Identical Rows', f"{stats['identical_count']} ({stats['identical_pct']:.1f}%)"],
        [f"Missing in {stats['comparison_sheet']}", f"{stats['missing_count']} ({stats['missing_pct']:.1f}%)"],
        [f"Extra in {stats['comparison_sheet']}", f"{stats['extra_count']} ({stats['extra_pct']:.1f}%)"],
        ['', ''],
        ['Status Legend', ''],
        ['IDENTICAL', 'Row exists in both sheets with same values'],
        [f"MISSING_IN_{stats['comparison_sheet']}", f"Row exists in {stats['baseline_sheet']} but not in {stats['comparison_sheet']}"],
        [f"EXTRA_IN_{stats['comparison_sheet']}", f"Row exists in {stats['comparison_sheet']} but not in {stats['baseline_sheet']}"],
    ]
    
    return pd.DataFrame(summary_data, columns=['Item', 'Description'])

def apply_formatting(file_path, sheet_name, stats):
    """Apply color coding and formatting to the comparison sheet."""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Define color fills
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue
    
    # Font styles
    header_font = Font(bold=True, color='FFFFFF')
    bold_font = Font(bold=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Apply conditional formatting based on Status column
    for row in range(2, ws.max_row + 1):
        status_cell = ws[f'A{row}']
        status_value = status_cell.value
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            if status_value == 'IDENTICAL':
                cell.fill = green_fill
            elif status_value and 'MISSING' in str(status_value):
                cell.fill = red_fill
            elif status_value and 'EXTRA' in str(status_value):
                cell.fill = yellow_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(file_path)
    print(f"Formatting applied to sheet: {sheet_name}")

def format_summary_sheet(file_path, sheet_name):
    """Apply formatting to the summary sheet."""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Define styles
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_font = Font(bold=True, size=14, color='FFFFFF')
    header_font = Font(bold=True, size=11)
    
    # Format title
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws.merge_cells('A1:B1')
    
    # Bold specific rows
    bold_rows = [4, 5, 7, 15]
    for row_num in bold_rows:
        ws[f'A{row_num}'].font = header_font
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    print(f"Formatting applied to summary sheet: {sheet_name}")

def main():
    """Main execution function."""
    print("="*80)
    print("Excel Tab Comparison Tool")
    print("="*80)
    
    # Perform comparison
    df_comparison, stats, sheet1_name, sheet2_name = compare_tabs(FILE_PATH)
    
    # Print statistics
    print("\n" + "="*80)
    print("COMPARISON RESULTS")
    print("="*80)
    print(f"Baseline: {sheet1_name} ({stats['baseline_total']} rows)")
    print(f"Comparison: {sheet2_name} ({stats['comparison_total']} rows)")
    print(f"\nIdentical rows: {stats['identical_count']} ({stats['identical_pct']:.1f}%)")
    print(f"Missing in {sheet2_name}: {stats['missing_count']} ({stats['missing_pct']:.1f}%)")
    print(f"Extra in {sheet2_name}: {stats['extra_count']} ({stats['extra_pct']:.1f}%)")
    print("="*80)
    
    # Create summary sheet
    df_summary = create_summary_sheet(stats)
    
    # Write results to Excel
    print("\nWriting comparison results to Excel...")
    with pd.ExcelWriter(FILE_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_summary.to_excel(writer, sheet_name='Summary', index=False, header=False)
        df_comparison.to_excel(writer, sheet_name='Comparison', index=False)
    
    print("Results written successfully!")
    
    # Apply formatting
    print("\nApplying formatting...")
    apply_formatting(FILE_PATH, 'Comparison', stats)
    format_summary_sheet(FILE_PATH, 'Summary')
    
    print("\n" + "="*80)
    print("COMPLETE!")
    print("="*80)
    print(f"Output file: {FILE_PATH}")
    print("New sheets created:")
    print("  - 'Summary': Overview and statistics")
    print("  - 'Comparison': Detailed row-by-row comparison")
    print("\nColor Legend:")
    print("  GREEN: Identical rows")
    print("  RED: Missing rows (in baseline but not in comparison)")
    print("  YELLOW: Extra rows (in comparison but not in baseline)")
    print("="*80)

if __name__ == "__main__":
    main()
