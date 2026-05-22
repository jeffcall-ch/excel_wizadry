"""
Sikla BoM Processor – Primary & Secondary Support Weight Analysis
Source: CA100-KVI-50296373_2.0 - BoM General Hangers and support material.xlsm

Rules:
  - SB = Nominal Bore <= 40,  LB = Nominal Bore > 40
  - Primary support weight  : sum of Total Weight [kg] per KKS code (Per_Pipe_Support)
  - Secondary support weight: Support_List Total Weight minus primary weight for that structure
  - SB/LB classification for secondary: max NB from Support_List column
  - Total Qty sheet used for validation (both coating sections combined)
  - Structures/KKS codes with unknown NB are excluded from the SB/LB split
"""

import os
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

FILE = r"CA100-KVI-50296373_2.0 - BoM General Hangers and support material.xlsm"
SB_THRESHOLD = 40  # NB <= 40 → Small Bore, NB > 40 → Large Bore


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def load_sheet(sheet_name: str) -> pd.DataFrame:
    """Load a sheet whose header row is Excel row 7 (pandas header=6).
    Drops fully-blank rows that appear after blank separator rows."""
    df = pd.read_excel(FILE, sheet_name=sheet_name, header=6, engine="openpyxl")
    return df.dropna(how="all").reset_index(drop=True)


def parse_max_nb(value) -> float:
    """Return the maximum NB from a value that may be a number or a
    comma-delimited string like '25, 50, 65, 80'."""
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float)):
        return float(value)
    parts = [p.strip() for p in str(value).split(",")]
    nums = []
    for p in parts:
        try:
            nums.append(float(p))
        except ValueError:
            pass
    return float(max(nums)) if nums else np.nan


def classify(nb: float) -> str:
    if pd.isna(nb):
        return "Unknown"
    return "SB" if nb <= SB_THRESHOLD else "LB"


# ──────────────────────────────────────────────────────────────────────────────
# Step 1 – Per_Pipe_Support  (primary supports only)
# ──────────────────────────────────────────────────────────────────────────────
print("Loading Per_Pipe_Support …")
pps = load_sheet("Per_Pipe_Support")

pps = pps.rename(columns={
    "Total Weight [kg]": "row_weight",
    "Nominal Bore":      "NB",
})
pps["NB"]         = pd.to_numeric(pps["NB"],         errors="coerce")
pps["row_weight"] = pd.to_numeric(pps["row_weight"], errors="coerce")
pps = pps.dropna(subset=["Pipe Support"])  # rows without a KKS are not primary supports

# Check if any KKS code spans multiple NB values (data quality)
nb_per_kks = pps.groupby("Pipe Support")["NB"].nunique()
multi_nb_kks = nb_per_kks[nb_per_kks > 1]
if not multi_nb_kks.empty:
    print(f"  ⚠  {len(multi_nb_kks)} KKS code(s) have multiple NB values – using max.")

# Aggregate per KKS code
kks = pps.groupby("Pipe Support", as_index=False).agg(
    structure = ("Support Structure", "first"),
    max_nb    = ("NB",                "max"),
    total_wt  = ("row_weight",        "sum"),
)
kks["class"] = kks["max_nb"].apply(classify)

# Per-structure totals from Per_Pipe_Support (needed to derive secondary weight)
struct_primary = pps.groupby("Support Structure", as_index=False).agg(
    primary_wt = ("row_weight", "sum"),
)


# ──────────────────────────────────────────────────────────────────────────────
# Step 2 – Support_List  (total weight per structure → derive secondary weight)
# ──────────────────────────────────────────────────────────────────────────────
print("Loading Support_List …")
sl = load_sheet("Support_List")

sl = sl.rename(columns={
    "Total Weight [kg]": "total_wt_sl",
    "Nominal Bore":      "NB_raw",
})
sl = sl.dropna(subset=["Support Structure"])
sl["max_nb_sl"]   = sl["NB_raw"].apply(parse_max_nb)
sl["total_wt_sl"] = pd.to_numeric(sl["total_wt_sl"], errors="coerce")

# Aggregate in case a structure somehow appears more than once
sl_agg = sl.groupby("Support Structure", as_index=False).agg(
    max_nb_sl   = ("max_nb_sl",   "max"),
    total_wt_sl = ("total_wt_sl", "sum"),
)

# Attach primary weight; structures not in Per_Pipe_Support get primary_wt = 0
sl_agg = sl_agg.merge(struct_primary, on="Support Structure", how="left")
sl_agg["primary_wt"]   = sl_agg["primary_wt"].fillna(0)
sl_agg["secondary_wt"] = sl_agg["total_wt_sl"] - sl_agg["primary_wt"]
sl_agg["class"]        = sl_agg["max_nb_sl"].apply(classify)

# Structures in Per_Pipe_Support but NOT in Support_List (anomaly check)
missing_in_sl = set(pps["Support Structure"].dropna()) - set(sl_agg["Support Structure"])
if missing_in_sl:
    print(f"  ⚠  {len(missing_in_sl)} structure(s) found in Per_Pipe_Support "
          f"but missing from Support_List – their secondary weight is unaccounted for.")

# Structures in Support_List with zero primary weight (pure secondary structures)
pure_secondary = sl_agg[sl_agg["primary_wt"] == 0]


# ──────────────────────────────────────────────────────────────────────────────
# Step 3 – Total Qty  (validation – combine both coating sections)
# ──────────────────────────────────────────────────────────────────────────────
print("Loading Total Qty …")
tq = load_sheet("Total Qty")
tq["total_wt_tq"] = pd.to_numeric(tq["Total Weight [kg]"], errors="coerce")
grand_total_tq    = tq["total_wt_tq"].sum()


# ──────────────────────────────────────────────────────────────────────────────
# Results  (Unknown NB excluded from SB/LB split)
# ──────────────────────────────────────────────────────────────────────────────
sb_kks = kks[kks["class"] == "SB"]
lb_kks = kks[kks["class"] == "LB"]
uk_kks = kks[kks["class"] == "Unknown"]   # excluded from output

sb_sec = sl_agg[sl_agg["class"] == "SB"]
lb_sec = sl_agg[sl_agg["class"] == "LB"]
uk_sec = sl_agg[sl_agg["class"] == "Unknown"]  # excluded from output

total_primary_sb   = sb_kks["total_wt"].sum()
total_primary_lb   = lb_kks["total_wt"].sum()
total_primary      = total_primary_sb + total_primary_lb

total_secondary_sb = sb_sec["secondary_wt"].sum()
total_secondary_lb = lb_sec["secondary_wt"].sum()
total_secondary    = total_secondary_sb + total_secondary_lb

grand_total_calc   = total_primary + total_secondary
excl_wt            = uk_kks["total_wt"].sum() + uk_sec["secondary_wt"].sum()
grand_total_check  = grand_total_calc + excl_wt   # add back excluded for validation
diff               = grand_total_check - grand_total_tq
status             = "OK" if abs(diff) < 1.0 else "CHECK"

# ── console echo ──────────────────────────────────────────────────────────────
print(f"  SB primary  : {len(sb_kks):>5} pcs   {total_primary_sb:>10.2f} kg")
print(f"  LB primary  : {len(lb_kks):>5} pcs   {total_primary_lb:>10.2f} kg")
print(f"  SB secondary:          {total_secondary_sb:>10.2f} kg")
print(f"  LB secondary:          {total_secondary_lb:>10.2f} kg")
print(f"  Grand total : {grand_total_calc:>10.2f} kg  |  "
      f"Total Qty: {grand_total_tq:.2f} kg  |  diff (incl. excl.): {diff:+.2f} kg  [{status}]")
if excl_wt > 0:
    print(f"  (Excluded unknown-NB weight: {excl_wt:.2f} kg – not in SB/LB split)")

# ──────────────────────────────────────────────────────────────────────────────
# Excel output
# ──────────────────────────────────────────────────────────────────────────────
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

timestamp  = datetime.now().strftime("%Y-%m-%d_%H%M%S")
input_stem = os.path.splitext(os.path.basename(FILE))[0]
out_name   = f"Support_Weight_Analysis_{input_stem}_{timestamp}.xlsx"
out_path   = os.path.join(os.path.dirname(os.path.abspath(FILE)), out_name)

# ── palette & style helpers ───────────────────────────────────────────────────
C_HEADER = "1F4E79"
C_SB     = "D6E4F0"
C_LB     = "FCE4D6"
C_TOTAL  = "E2EFDA"
C_META   = "F2F2F2"
C_OK     = "C6EFCE"
C_WARN   = "FFEB9C"

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

thin = Side(style="thin",   color="AAAAAA")
med  = Side(style="medium", color="555555")

def border(top_med=False, bot_med=False):
    t = med if top_med else thin
    b = med if bot_med else thin
    return Border(left=thin, right=thin, top=t, bottom=b)

L = Alignment(horizontal="left",   vertical="center")
R = Alignment(horizontal="right",  vertical="center")
C = Alignment(horizontal="center", vertical="center")

FMT_INT = "#,##0"
FMT_KG  = "#,##0.00"
FMT_SGN = '+#,##0.00;-#,##0.00;"-"'

def section_header(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(C_HEADER); c.font = font(bold=True, color="FFFFFF", size=11)
    c.alignment = C
    c.border = Border(left=med, right=med, top=med, bottom=med)

def col_hdr_row(ws, row, labels):
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row=row, column=ci, value=lbl)
        c.fill = fill(C_HEADER); c.font = font(bold=True, color="FFFFFF")
        c.alignment = C if ci > 1 else L
        c.border = border(top_med=True, bot_med=True)

def data_row(ws, row, values, bg, bold=False, top_med=False):
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = fill(bg); c.font = font(bold=bold)
        c.border = border(top_med=top_med)
        c.alignment = L if ci == 1 else R
        if ci == 2 and isinstance(val, (int, float)):
            c.number_format = FMT_INT
        elif ci > 2 and isinstance(val, (int, float)):
            c.number_format = FMT_KG

# ── build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Summary"
ws.freeze_panes = "A2"

for col, w in zip("ABCDE", [34, 15, 15, 15, 15]):
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 24

# ── title & meta ──────────────────────────────────────────────────────────────
row = 1
ws.merge_cells(f"A{row}:E{row}")
c = ws.cell(row=row, column=1, value="Sikla BoM – Support Weight Analysis")
c.font = Font(bold=True, color="1F4E79", size=14, name="Calibri")
c.fill = fill(C_META); c.alignment = C
row += 1

for label, value in [
    ("Input file",     os.path.basename(FILE)),
    ("Generated",      datetime.now().strftime("%Y-%m-%d  %H:%M:%S")),
    ("SB / LB split",  f"Nominal Bore ≤ {SB_THRESHOLD} → Small Bore  |  > {SB_THRESHOLD} → Large Bore"),
]:
    ws.cell(row=row, column=1, value=label).font = font(bold=True, color="444444")
    ws.cell(row=row, column=1).fill = fill(C_META)
    ws.cell(row=row, column=1).alignment = L
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = ws.cell(row=row, column=2, value=value)
    c.fill = fill(C_META); c.font = font(color="000000"); c.alignment = L
    row += 1

row += 1  # spacer

# ── primary + secondary combined table ───────────────────────────────────────
section_header(ws, row, "SUPPORT WEIGHT SUMMARY"); row += 1
col_hdr_row(ws, row, ["Category", "KKS count (pcs)", "Primary wt (kg)",
                      "Secondary wt (kg)", "Total wt (kg)"]); row += 1

data_row(ws, row,
         ["Small Bore  (NB ≤ 40)", len(sb_kks),
          total_primary_sb, total_secondary_sb,
          total_primary_sb + total_secondary_sb],
         C_SB); row += 1

data_row(ws, row,
         ["Large Bore  (NB > 40)", len(lb_kks),
          total_primary_lb, total_secondary_lb,
          total_primary_lb + total_secondary_lb],
         C_LB); row += 1

data_row(ws, row,
         ["TOTAL  (SB + LB)", len(sb_kks) + len(lb_kks),
          total_primary, total_secondary,
          total_primary + total_secondary],
         C_TOTAL, bold=True, top_med=True); row += 1

row += 1  # spacer

# ── structure counts note ─────────────────────────────────────────────────────
section_header(ws, row, "DATA STATISTICS"); row += 1
for label, value in [
    ("Total structures in Support_List",        len(sl_agg)),
    ("Structures with primary clamps",          len(sl_agg) - len(pure_secondary)),
    ("Pure secondary structures (no clamps)",   len(pure_secondary)),
    ("Excluded – unknown Nominal Bore",         len(uk_sec) + len(uk_kks)),
]:
    ws.cell(row=row, column=1, value=label).font = font()
    ws.cell(row=row, column=1).fill = fill(C_META); ws.cell(row=row, column=1).alignment = L
    ws.cell(row=row, column=1).border = border()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = ws.cell(row=row, column=2, value=value)
    c.fill = fill(C_META); c.font = font(bold=True); c.alignment = R
    c.number_format = FMT_INT; c.border = border()
    row += 1

row += 1  # spacer

# ── validation ────────────────────────────────────────────────────────────────
section_header(ws, row, "VALIDATION  (Total Qty sheet – both coating sections)"); row += 1
val_bg    = C_OK   if status == "OK" else C_WARN
val_color = "375623" if status == "OK" else "9C5700"

for label, value, fmt in [
    ("Total Qty sheet grand total (kg)",        grand_total_tq,    FMT_KG),
    ("Calculated SB+LB total (kg)",             grand_total_calc,  FMT_KG),
    ("Excluded unknown-NB weight (kg)",         excl_wt,           FMT_KG),
    ("Calculated total incl. excluded (kg)",    grand_total_check, FMT_KG),
    ("Difference (kg)",                         diff,              FMT_SGN),
    ("Status",                                  status,            "@"),
]:
    ws.cell(row=row, column=1, value=label).font = font(bold=True)
    ws.cell(row=row, column=1).fill = fill(C_META); ws.cell(row=row, column=1).alignment = L
    ws.cell(row=row, column=1).border = border()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = ws.cell(row=row, column=2, value=value)
    c.fill = fill(val_bg); c.font = Font(bold=True, color=val_color, name="Calibri", size=11)
    c.alignment = R; c.number_format = fmt; c.border = border()
    row += 1

if excl_wt > 0:
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1,
                value=f"Note: {excl_wt:.2f} kg with unknown Nominal Bore is excluded "
                      f"from the SB/LB split but IS counted in the validation check above.")
    c.fill = fill("FFEB9C"); c.font = Font(color="9C5700", size=10, name="Calibri")
    c.alignment = L

wb.save(out_path)
print(f"\nOutput saved → {out_path}")
