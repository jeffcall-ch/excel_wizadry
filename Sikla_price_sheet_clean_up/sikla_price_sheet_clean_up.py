
import pandas as pd
import sys
import os

# Get input file path from command line argument
if len(sys.argv) < 2:
	print("Usage: python sikla_price_sheet_clean_up.py <input_excel_file>")
	sys.exit(1)

excel_path = sys.argv[1]
if not os.path.isfile(excel_path):
	print(f"Input file not found: {excel_path}")
	sys.exit(1)

# 1. Read all sheets
sheets = pd.read_excel(excel_path, sheet_name=None, header=None)

# 2. Identify the sheet with the most rows
max_rows = 0
max_sheet = None
for name, df in sheets.items():
	if len(df) > max_rows:
		max_rows = len(df)
		max_sheet = name

df = sheets[max_sheet]

# 3. Find the header row
header_keywords = ["description", "weight", "price"]
header_row_idx = None

for idx, row in df.iterrows():
	row_strs = [str(cell).lower() for cell in row]
	if all(any(keyword in cell for cell in row_strs) for keyword in header_keywords):
		header_row_idx = idx
		break


# 4. Drop all rows above the header row
df_clean = df.iloc[header_row_idx:].reset_index(drop=True)

# 5. Find the description column index
header_row = df_clean.iloc[0].astype(str).str.lower().tolist()
desc_col_idx = None
for i, col in enumerate(header_row):
	if "description" in col:
		desc_col_idx = i
		break

if desc_col_idx is None:
	raise ValueError("Description column not found in header row.")


# 6. Drop rows below header where description column is empty (including whitespace and NaN), always keep header row
desc_series = df_clean.iloc[:, desc_col_idx]
mask_not_empty = desc_series.notna() & (desc_series.astype(str).str.strip() != "")
mask_not_empty.iloc[0] = True  # Always keep header row
df_clean = df_clean[mask_not_empty]

print(f"Sheet with most rows: {max_sheet}")
print(f"Header row index: {header_row_idx}")
print(f"Header row values: {df.iloc[header_row_idx].tolist()}")


# 7. Drop rows where description column contains 'description' (substring or full match), but always keep the header row
desc_col = df_clean.iloc[:, desc_col_idx].astype(str).str.lower()
mask = ~desc_col.str.contains("description")
# Always keep the header row (index 0)
mask.iloc[0] = True
df_clean = df_clean[mask]

# 8. Export to Excel with timestamped filename

from datetime import datetime
now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
import os
base, ext = os.path.splitext(os.path.basename(excel_path))
output_name = f"{base}_cleaned_up_{now_str}.xlsx"
output_path = os.path.join(os.path.dirname(excel_path), output_name)

# 9. Export to Excel with formatting: auto-width columns and filters on top row
from openpyxl import load_workbook
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
	df_clean.to_excel(writer, index=False, header=False)
	worksheet = writer.sheets[df_clean.columns[0] if hasattr(df_clean.columns[0], 'sheet_title') else writer.book.sheetnames[0]]


	# Set column A width to 18
	worksheet.column_dimensions['A'].width = 18
	# Set column C width to 60 and wrap text
	worksheet.column_dimensions['C'].width = 60
	for cell in worksheet['C']:
		cell.alignment = cell.alignment.copy(wrap_text=True)

	# Format all decimal numbers to two decimal places
	from openpyxl.styles import numbers
	for row in worksheet.iter_rows():
		for cell in row:
			if isinstance(cell.value, float):
				cell.value = round(cell.value, 2)
				cell.number_format = '0.00'

	# Auto-width for other columns
	for col in worksheet.columns:
		column = col[0].column_letter
		if column not in ['A', 'C']:
			max_length = 0
			for cell in col:
				try:
					if cell.value:
						max_length = max(max_length, len(str(cell.value)))
				except:
					pass
			adjusted_width = (max_length + 2)
			worksheet.column_dimensions[column].width = adjusted_width

	# Auto height for all rows
	for row in worksheet.iter_rows():
		max_height = 15  # default row height
		for cell in row:
			if cell.value:
				lines = str(cell.value).count('\n') + 1
				max_height = max(max_height, lines * 15)
		worksheet.row_dimensions[row[0].row].height = max_height


	# Add filter to top row
	worksheet.auto_filter.ref = worksheet.dimensions

	# Split and freeze at D2
	worksheet.freeze_panes = 'D2'

print(f"Exported cleaned file to: {output_path}")
