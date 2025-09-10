import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def clean_and_convert_baseline(df):
    """Clean and convert baseline tab data according to specifications"""
    # Strip whitespace from all text columns
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace('nan', np.nan)
    
    # Convert data types
    df['Item Number'] = pd.to_numeric(df['Item Number'], errors='coerce').astype('Int64')
    df['Qty BOM_6.0'] = pd.to_numeric(df['Qty BOM_6.0'], errors='coerce').round(2)
    df['Total Weight [kg] BOM_6.0'] = pd.to_numeric(df['Total Weight [kg] BOM_6.0'], errors='coerce').round(2)
    df['Qty Spare'] = pd.to_numeric(df['Qty Spare'], errors='coerce').round(2)
    df['Qty PO'] = pd.to_numeric(df['Qty PO'], errors='coerce').round(2)
    df['Total Weight [kg] PO'] = pd.to_numeric(df['Total Weight [kg] PO'], errors='coerce').round(2)
    
    return df

def clean_and_convert_update(df):
    """Clean and convert update tab data according to specifications"""
    # Strip whitespace from all text columns
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace('nan', np.nan)
    
    # Convert data types
    df['Item Number'] = pd.to_numeric(df['Item Number'], errors='coerce').astype('Int64')
    df['Qty'] = pd.to_numeric(df['Qty'], errors='coerce').round(2)
    df['Total Weight [kg]'] = pd.to_numeric(df['Total Weight [kg]'], errors='coerce').round(2)
    
    return df

def infer_item_category(description, baseline_df):
    """Infer item category from description by looking for similar descriptions in baseline"""
    if pd.isna(description):
        return 'Unknown'
    
    # Look for exact matches first
    exact_matches = baseline_df[baseline_df['Description'] == description]
    if not exact_matches.empty:
        return exact_matches['Item Category'].iloc[0]
    
    # Look for partial matches (descriptions containing similar words)
    description_words = str(description).lower().split()
    if description_words:
        # Try to find descriptions with common words
        for _, row in baseline_df.iterrows():
            baseline_desc = str(row['Description']).lower()
            common_words = sum(1 for word in description_words if word in baseline_desc)
            if common_words >= 2:  # At least 2 common words
                return row['Item Category']
    
    # Default category if no match found
    return 'Unknown'

def check_baseline_duplicates(baseline_df):
    """Check for duplicate item numbers within same corrosion category in baseline"""
    print("Checking for duplicates in baseline data...")
    
    duplicates = baseline_df.groupby(['Item Number', 'Corrosion category']).size()
    duplicate_pairs = duplicates[duplicates > 1]
    
    if len(duplicate_pairs) > 0:
        print(f"WARNING: Found {len(duplicate_pairs)} duplicate item-category combinations in baseline:")
        for (item_num, corr_cat), count in duplicate_pairs.items():
            print(f"  Item {item_num} in {corr_cat}: appears {count} times")
        
        print("\nDetailed view of baseline duplicates:")
        for (item_num, corr_cat), count in duplicate_pairs.head(5).items():
            subset = baseline_df[(baseline_df['Item Number'] == item_num) & 
                               (baseline_df['Corrosion category'] == corr_cat)]
            print(f"\nItem {item_num} - {corr_cat}:")
            print(subset[['Item Number', 'Description', 'Qty PO', 'Remarks']])
        
        return len(duplicate_pairs)
    else:
        print("✓ No duplicate item-category combinations found in baseline.")
        return 0

def check_update_duplicates(update_df):
    """Check for duplicate item numbers within same corrosion category in update"""
    print("Checking for duplicates in update data...")
    
    duplicates = update_df.groupby(['Item Number', 'Corrosion Category']).size()
    duplicate_pairs = duplicates[duplicates > 1]
    
    if len(duplicate_pairs) > 0:
        print(f"Found {len(duplicate_pairs)} duplicate item-category combinations in update:")
        special_count = 0
        for (item_num, corr_cat), count in duplicate_pairs.items():
            subset = update_df[(update_df['Item Number'] == item_num) & 
                             (update_df['Corrosion Category'] == corr_cat)]
            has_special = subset['Remarks'].str.contains('special', case=False, na=False).any()
            special_indicator = " (contains 'special' remarks)" if has_special else ""
            print(f"  Item {item_num} in {corr_cat}: appears {count} times{special_indicator}")
            if has_special:
                special_count += 1
        
        print(f"✓ {special_count} of these duplicates have 'special' remarks (expected)")
        return len(duplicate_pairs)
    else:
        print("✓ No duplicate item-category combinations found in update.")
        return 0

def process_po_vs_bom(baseline_df, update_df):
    """Process the comparison between baseline and update data"""
    results = []
    new_items = []  # Collect new items separately for proper insertion
    
    # Process each baseline item
    for _, baseline_row in baseline_df.iterrows():
        result_row = baseline_row.copy()
        
        # Initialize new columns
        result_row['Qty BOM update'] = np.nan
        result_row['Total Weight [kg] BOM update'] = np.nan
        result_row['Remarks BOM update'] = np.nan
        result_row['Special Remarks'] = np.nan
        result_row['Qty Difference'] = np.nan
        result_row['New'] = False
        result_row['Deleted'] = False
        
        # Find matching items in update (same item number and corrosion category)
        matches = update_df[
            (update_df['Item Number'] == baseline_row['Item Number']) & 
            (update_df['Corrosion Category'] == baseline_row['Corrosion category'])
        ]
        
        if matches.empty:
            # Item not found in update - mark as deleted
            result_row['Qty Difference'] = -baseline_row['Qty PO'] if pd.notna(baseline_row['Qty PO']) else 0
            result_row['Deleted'] = True
        else:
            # Handle multiple matches (special items)
            if len(matches) > 1:
                # Multiple matches - handle special items
                for _, match in matches.iterrows():
                    special_row = result_row.copy()
                    special_row['Qty BOM update'] = match['Qty']
                    special_row['Total Weight [kg] BOM update'] = match['Total Weight [kg]']
                    special_row['Remarks BOM update'] = match['Remarks']
                    
                    # Handle special remarks
                    if pd.notna(match['Remarks']) and 'special' in str(match['Remarks']).lower():
                        special_row['Special Remarks'] = match['Remarks']
                    
                    # Calculate quantity difference
                    if pd.notna(baseline_row['Qty PO']) and pd.notna(match['Qty']):
                        special_row['Qty Difference'] = match['Qty'] - baseline_row['Qty PO']
                    
                    results.append(special_row)
            else:
                # Single match
                match = matches.iloc[0]
                result_row['Qty BOM update'] = match['Qty']
                result_row['Total Weight [kg] BOM update'] = match['Total Weight [kg]']
                result_row['Remarks BOM update'] = match['Remarks']
                
                # Handle special remarks
                if pd.notna(match['Remarks']) and 'special' in str(match['Remarks']).lower():
                    result_row['Special Remarks'] = match['Remarks']
                
                # Calculate quantity difference
                if pd.notna(baseline_row['Qty PO']) and pd.notna(match['Qty']):
                    result_row['Qty Difference'] = match['Qty'] - baseline_row['Qty PO']
                
                results.append(result_row)
        
        # Add single match or deleted item if not already added
        if matches.empty or len(matches) == 1:
            if len(matches) == 0:
                # Only add deleted items once
                results.append(result_row)
    
    # Find new items in update that are not in baseline
    for _, update_row in update_df.iterrows():
        baseline_matches = baseline_df[
            (baseline_df['Item Number'] == update_row['Item Number']) & 
            (baseline_df['Corrosion category'] == update_row['Corrosion Category'])
        ]
        
        if baseline_matches.empty:
            # New item - create new row
            new_row = pd.Series(dtype=object)
            
            # Infer item category
            new_row['Item Category'] = infer_item_category(update_row['Description'], baseline_df)
            new_row['Item Number'] = update_row['Item Number']
            new_row['Description'] = update_row['Description']
            new_row['Qty BOM_6.0'] = np.nan
            new_row['Total Weight [kg] BOM_6.0'] = np.nan
            new_row['Qty Spare'] = np.nan
            new_row['Qty PO'] = np.nan
            new_row['Total Weight [kg] PO'] = np.nan
            new_row['Remarks'] = np.nan
            new_row['Corrosion category'] = update_row['Corrosion Category']
            
            # Fill update columns
            new_row['Qty BOM update'] = update_row['Qty']
            new_row['Total Weight [kg] BOM update'] = update_row['Total Weight [kg]']
            new_row['Remarks BOM update'] = update_row['Remarks']
            
            # Handle special remarks
            if pd.notna(update_row['Remarks']) and 'special' in str(update_row['Remarks']).lower():
                new_row['Special Remarks'] = update_row['Remarks']
            else:
                new_row['Special Remarks'] = np.nan
            
            new_row['Qty Difference'] = update_row['Qty'] if pd.notna(update_row['Qty']) else 0
            new_row['New'] = True
            new_row['Deleted'] = False
            
            results.append(new_row)
    
    # Convert results to DataFrame and sort appropriately
    results_df = pd.DataFrame(results)
    
    # Sort by corrosion category, item category, then item number for proper grouping
    results_df = results_df.sort_values(['Corrosion category', 'Item Category', 'Item Number']).reset_index(drop=True)
    
    return results_df

def insert_new_items_by_category(results_df, new_items):
    """Insert new items at the bottom of their respective Item Category and Corrosion Category groups"""
    # Convert new_items to DataFrame
    new_items_df = pd.DataFrame(new_items)
    
    # Sort results by Item Category and Corrosion Category to group them
    results_df = results_df.sort_values(['Item Category', 'Corrosion category', 'Item Number']).reset_index(drop=True)
    
    # Group new items by their category and corrosion category
    new_items_grouped = new_items_df.groupby(['Item Category', 'Corrosion category'])
    
    final_results = []
    current_category = None
    current_corrosion = None
    
    # Process each row in the sorted results
    for idx, row in results_df.iterrows():
        row_category = row['Item Category']
        row_corrosion = row['Corrosion category']
        
        # Check if we're moving to a new category/corrosion combination
        if (current_category != row_category or current_corrosion != row_corrosion):
            # Insert any new items that belong to the previous category before moving on
            if current_category is not None and current_corrosion is not None:
                if (current_category, current_corrosion) in new_items_grouped.groups:
                    new_group = new_items_grouped.get_group((current_category, current_corrosion))
                    for _, new_row in new_group.iterrows():
                        final_results.append(new_row)
            
            current_category = row_category
            current_corrosion = row_corrosion
        
        # Add the current baseline/existing row
        final_results.append(row)
    
    # Add any remaining new items for the last category
    if current_category is not None and current_corrosion is not None:
        if (current_category, current_corrosion) in new_items_grouped.groups:
            new_group = new_items_grouped.get_group((current_category, current_corrosion))
            for _, new_row in new_group.iterrows():
                final_results.append(new_row)
    
    # Handle new items with categories that don't exist in baseline
    for (category, corrosion), group in new_items_grouped:
        # Check if this combination was already processed
        existing_combo = results_df[
            (results_df['Item Category'] == category) & 
            (results_df['Corrosion category'] == corrosion)
        ]
        
        if existing_combo.empty:
            # This is a completely new category/corrosion combination
            for _, new_row in group.iterrows():
                final_results.append(new_row)
    
    return pd.DataFrame(final_results).reset_index(drop=True)

def save_to_excel_with_formatting(df, filename):
    """Save DataFrame to Excel with conditional formatting, filters, and freeze panes"""
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "PO vs BOM Comparison"
    
    # Define fill colors
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
    blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Light blue
    
    # Write headers
    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data with formatting
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        # Determine row color based on New/Deleted status
        is_new = row.get('New', False)
        is_deleted = row.get('Deleted', False)
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply formatting
            if is_deleted:
                cell.fill = red_fill
            elif is_new:
                cell.fill = blue_fill
    
    # Add autofilter to the data range
    from openpyxl.utils import get_column_letter
    max_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"
    
    # Split freeze panes at D2 (column D, row 2)
    ws.freeze_panes = "D2"
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)

def main():
    """Main function to process PO vs BOM comparison"""
    input_file = 'POvsBOM_10.09.xlsx'
    
    try:
        # Read the input Excel file
        print(f"Reading {input_file}...")
        baseline_df = pd.read_excel(input_file, sheet_name='baseline')
        update_df = pd.read_excel(input_file, sheet_name='update')
        
        print(f"Baseline data: {baseline_df.shape[0]} rows")
        print(f"Update data: {update_df.shape[0]} rows")
        
        # Clean and convert data
        print("Cleaning and converting data...")
        baseline_df = clean_and_convert_baseline(baseline_df)
        update_df = clean_and_convert_update(update_df)
        
        # Check for duplicates in both datasets
        baseline_duplicates = check_baseline_duplicates(baseline_df)
        update_duplicates = check_update_duplicates(update_df)
        
        # Process comparison
        print("Processing PO vs BOM comparison...")
        result_df = process_po_vs_bom(baseline_df, update_df)
        
        print(f"Results: {result_df.shape[0]} rows")
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        base_name = os.path.splitext(input_file)[0]
        output_file = f"{base_name}_processed_{timestamp}.xlsx"
        
        # Save to Excel with formatting
        print(f"Saving results to {output_file}...")
        save_to_excel_with_formatting(result_df, output_file)
        
        # Display summary
        print("\n--- Summary ---")
        print(f"Baseline duplicates found: {baseline_duplicates}")
        print(f"Update duplicates found: {update_duplicates}")
        print(f"Total processed items: {len(result_df)}")
        print(f"Deleted items (red): {result_df['Deleted'].sum()}")
        print(f"New items (blue): {result_df['New'].sum()}")
        print(f"Items with special remarks: {result_df['Special Remarks'].notna().sum()}")
        print(f"Items with quantity differences: {result_df['Qty Difference'].notna().sum()}")
        
        print(f"\nSuccessfully saved to: {output_file}")
        
    except FileNotFoundError:
        print(f"Error: File {input_file} not found in current directory")
    except Exception as e:
        print(f"Error processing files: {e}")

if __name__ == "__main__":
    main()
