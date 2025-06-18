import os
import math
import pandas as pd
from openpyxl import load_workbook
import openai
from dotenv import load_dotenv
import json
import time
import re
from tqdm import tqdm

# --- LOAD ENVIRONMENT VARIABLES ---
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY not set in .env file!")
openai.api_key = OPENAI_API_KEY

# --- CONFIGURATION ---
FOLDER = "pipe_class_summary"
SUMMARY_SHEET_FILENAME = "Summary.xlsx"
SUMMARY_SHEET_NAME = "Pipe Class Summary"
BATCH_SIZE = 5
MAX_ROWS = 200  # Increased to ensure we capture row 83 and beyond
MAX_COLS = 25
BATCH_DELAY = 2  # seconds between batches
CONTEXT_WINDOW = 50  # Increased context window to capture more data around pipe class matches
SAMPLE_SIZE = 20  # rows to include if no specific pipe class found
DEBUG_FOLDER = "debug_logs"  # folder to save prompts and responses

summary_sheet_path = os.path.join(FOLDER, SUMMARY_SHEET_FILENAME)

source_sheet_paths = [
    os.path.join(FOLDER, f)
    for f in os.listdir(FOLDER)
    if (f.endswith(".xlsx") or f.endswith(".xlsm")) and f != SUMMARY_SHEET_FILENAME
]

if not os.path.exists(summary_sheet_path):
    raise FileNotFoundError(f"Summary sheet not found at {summary_sheet_path}")
if not source_sheet_paths:
    raise FileNotFoundError(f"No source Excel files found in {FOLDER}")

# Create debug folder for saving prompts and responses
if not os.path.exists(DEBUG_FOLDER):
    os.makedirs(DEBUG_FOLDER)

# Global counter for prompt/response file naming
prompt_counter = 1
response_counter = 1

# --- 1. Load Summary Sheet ---
summary_df = pd.read_excel(summary_sheet_path, sheet_name=SUMMARY_SHEET_NAME)
pipe_classes = summary_df['Pipe Class Name'].tolist()
summary_columns = summary_df.columns.tolist()

# --- 2. Load and collect all source data once, excluding empty cells ---
all_source_data = []
pipe_class_index = {}  # Index to quickly find relevant data for each pipe class

print("Loading and indexing source data...")
for src_path in source_sheet_paths:
    wb = load_workbook(src_path, data_only=True)
    for tab in wb.sheetnames:
        print(f"Loading source file: {os.path.basename(src_path)}, tab: {tab}")
        ws = wb[tab]
        tab_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=MAX_ROWS, max_col=MAX_COLS)):
            if i >= MAX_ROWS:
                break
            filtered_row = [cell for cell in list(row)[:MAX_COLS] if cell not in (None, "", " ")]
            if filtered_row:
                tab_data.append(filtered_row)
                # Index rows that might contain pipe class names
                for cell in filtered_row:
                    if isinstance(cell, str) and any(pc.lower() in cell.lower() for pc in pipe_classes):
                        for pc in pipe_classes:
                            if pc.lower() in cell.lower():
                                if pc not in pipe_class_index:
                                    pipe_class_index[pc] = []
                                pipe_class_index[pc].append({
                                    "file": os.path.basename(src_path),
                                    "sheet": tab,
                                    "row_index": len(tab_data) - 1,
                                    "data": filtered_row
                                })
        if tab_data:
            all_source_data.append({
                "file": os.path.basename(src_path),
                "sheet": tab,
                "data": tab_data
            })

print(f"Indexed {len(pipe_class_index)} pipe classes across {len(all_source_data)} sheets")

def save_prompt_to_file(prompt, batch_info):
    """Save prompt to a markdown file"""
    global prompt_counter
    filename = os.path.join(DEBUG_FOLDER, f"prompt_{prompt_counter:03d}.md")
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"# Prompt {prompt_counter}\n\n")
        f.write(f"**Batch Info:** {batch_info}\n\n")
        f.write(f"**Timestamp:** {time.strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("## Prompt Content\n\n")
        f.write("```\n")
        f.write(prompt)
        f.write("\n```\n")
    
    print(f"Prompt saved to: {filename}")
    prompt_counter += 1
    return filename

def save_response_to_file(response, prompt_file):
    """Save LLM response to a markdown file"""
    global response_counter
    filename = os.path.join(DEBUG_FOLDER, f"response_{response_counter:03d}.md")
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"# Response {response_counter}\n\n")
        f.write(f"**Corresponding Prompt:** {os.path.basename(prompt_file)}\n\n")
        f.write(f"**Timestamp:** {time.strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("## Response Content\n\n")
        f.write("```json\n")
        f.write(response)
        f.write("\n```\n")
    
    print(f"Response saved to: {filename}")
    response_counter += 1
    return filename

def get_relevant_source_data(batch_pipe_classes):
    """Get only source data relevant to the current batch of pipe classes"""
    relevant_data = []
    
    # First, try to get indexed data for specific pipe classes
    for pipe_class in batch_pipe_classes:
        # Check if we have indexed data for this pipe class
        if pipe_class in pipe_class_index:
            for entry in pipe_class_index[pipe_class]:
                # Find the corresponding sheet data
                for sheet_data in all_source_data:
                    if sheet_data["file"] == entry["file"] and sheet_data["sheet"] == entry["sheet"]:
                        # Include a larger context window around the found row
                        start_row = max(0, entry["row_index"] - CONTEXT_WINDOW)
                        end_row = min(len(sheet_data["data"]), entry["row_index"] + CONTEXT_WINDOW + 1)
                        context_data = sheet_data["data"][start_row:end_row]
                        
                        relevant_data.append({
                            "file": entry["file"],
                            "sheet": entry["sheet"],
                            "data": context_data,
                            "pipe_class": pipe_class
                        })
                        break
        
        # Additionally, if the sheet name matches the pipe class, include the entire sheet
        # This handles cases where the sheet itself is dedicated to the pipe class (like EFDX sheet)
        for sheet_data in all_source_data:
            if sheet_data["sheet"].upper() == pipe_class.upper():
                # Include the entire sheet data for sheets named after pipe classes
                relevant_data.append({
                    "file": sheet_data["file"],
                    "sheet": sheet_data["sheet"],
                    "data": sheet_data["data"],  # Include all data from the sheet
                    "pipe_class": pipe_class
                })
                print(f"Including entire sheet '{sheet_data['sheet']}' for pipe class '{pipe_class}'")
    
    # If no indexed data found, include a sample from each sheet for context
    if not relevant_data:
        print(f"No indexed data found for pipe classes: {batch_pipe_classes}")
        print("Including sample data from all sheets for context...")
        for sheet_data in all_source_data:
            # Include first SAMPLE_SIZE rows as sample
            sample_data = sheet_data["data"][:SAMPLE_SIZE]
            relevant_data.append({
                "file": sheet_data["file"],
                "sheet": sheet_data["sheet"],
                "data": sample_data
            })
    
    return relevant_data

def extract_key_properties(data_rows):
    """Extract key properties from source data rows to highlight them"""
    import re  # Import re at the beginning of the function
    extracted = {}
    
    for row in data_rows:
        if not row:
            continue
            
        row_str = ' '.join(str(cell) for cell in row if cell is not None)
        row_lower = row_str.lower()
        
        # Extract corrosion allowance
        if 'corrosion allowance' in row_lower:
            for cell in row:
                if cell and isinstance(cell, str) and ('mm' in cell.lower() or cell.isdigit()):
                    # Extract numeric value from corrosion allowance
                    match = re.search(r'(\d+(?:\.\d+)?)', str(cell))
                    if match:
                        extracted['Corrosion allowance'] = float(match.group(1))
                        break
        
        # Extract material codes (numbers like 1.4401)
        if any(isinstance(cell, (int, float)) and 1.0 <= cell <= 2.0 for cell in row):
            for cell in row:
                if isinstance(cell, (int, float)) and 1.0 <= cell <= 2.0:
                    extracted['Material Code'] = cell
                    break
        
        # Extract pressure class
        if any('pn' in str(cell).lower() for cell in row):
            for cell in row:
                if cell and 'pn' in str(cell).lower():
                    match = re.search(r'pn\s*(\d+)', str(cell).lower())
                    if match:
                        extracted['Pressure class [PN, CL]'] = int(match.group(1))
                        break
        
        # Extract temperature ranges
        if any('°c' in str(cell).lower() for cell in row):
            temps = []
            for cell in row:
                if cell and '°c' in str(cell).lower():
                    match = re.search(r'(-?\d+)°?c', str(cell).lower())
                    if match:
                        temps.append(int(match.group(1)))
            if temps:
                if len(temps) >= 2:
                    extracted['Min temperature (°C)'] = min(temps)
                    extracted['Max temperature (°C)'] = max(temps)
    
    return extracted

def create_prompt(batch_pipe_classes, batch_summary_df):
    # Get only relevant source data for this batch
    relevant_source_data = get_relevant_source_data(batch_pipe_classes)
    
    # Extract key properties for easier matching
    extracted_properties = {}
    for data_entry in relevant_source_data:
        pipe_class = data_entry.get('pipe_class', 'unknown')
        if pipe_class != 'unknown':
            extracted_properties[pipe_class] = extract_key_properties(data_entry['data'])
    
    prompt = f"""
You are a piping engineering assistant.

You are given:
- A summary table of pipe classes, where each row represents a pipe class and columns are its properties.
- Relevant source Excel data that may contain information about these specific pipe classes.
- Extracted key properties from the source data for easier matching.

The source data only includes non-empty cell values and is filtered to show only relevant sections.

IMPORTANT MATCHING RULES:
1. For "Corrosion allowance": Look for values like "0 mm", "1 mm", "4.0 mm", "1.0mm", etc. Maybe there is extra text after it.  The summary shows just the number (0, 1), while source may show "0 mm", "1 mm".
2. For "Material Code": Look for decimal numbers like 1.4401, 1.4404, etc.
3. For "Pressure class": Look for "PN 16", "PN 25", etc. Match the number only.
4. For temperatures: Look for values with °C, may appear in ranges like "-15°C to 100°C".
5. Property names may vary slightly (e.g., "Material" vs "Material Code").
6. For "Diameter from" and "Diameter to" look for a row with diameters mentioned. Look for "DN" or "OD" then series of numbers. We need the smallest and largest here. 
7. For "Min termperature" and "Max temperature" look for a row which mentions "temperature" or "temp" and then series of numbers. We need the smallest and largest here. If a cell contains "RT" this is room temperature, so we can use 20°C as a default value equivalent to 20.
8. For "min pressure" and "max pressure" look for a row which mentions "pressure" or "press" and then series of numbers. We need the smallest and largest here. If a cell contains "atm" this is atmospheric pressure, so we can use 1 atm as a default value equivalent to 1.

Your tasks:
1. For each pipe class in the summary batch, find and extract the corresponding properties from the source data, even if the layouts differ.
2. Compare each property in the summary with the value found in the sources.
3. Return, for each pipe class, a dictionary mapping each property to "match" (if the summary and source agree), "mismatch" (if they differ), or "missing" (if not found in the sources).

IMPORTANT: For "Pipe Class Name" field, always use the ACTUAL pipe class name (e.g., "EFDX", "EFDJ"), NOT "match"/"mismatch"/"missing".

Summary pipe classes:
{batch_pipe_classes}

Summary columns:
{summary_columns}

Summary table for current batch:
{batch_summary_df.to_dict(orient='records')}

Extracted key properties from source data:
{extracted_properties}

Full relevant source data (filtered for these pipe classes):
{relevant_source_data}

Return only a valid JSON array as your entire response, with no introduction, explanation, or trailing text.
Each element of the array should be a dictionary containing ALL columns from the summary table.

REQUIRED OUTPUT FORMAT - You MUST include ALL {len(summary_columns)} columns:
For "Pipe Class Name": Use the actual pipe class name (e.g., "EFDX", "EFDJ")
For all other columns: Use "match"/"mismatch"/"missing"

Example for pipe class EFDX:
{{
    "Pipe Class Name": "EFDX",
{chr(10).join([f'    "{col}": "match"/"mismatch"/"missing"' + ("," if col != summary_columns[-1] else "") for col in summary_columns if col != "Pipe Class Name"])}
}}

CRITICAL: Your response must include exactly these {len(summary_columns)} columns: {summary_columns}
"""
    return prompt

def extract_first_json_array(text):
    match = re.search(r'\[\s*{.*?}\s*\]', text, re.DOTALL)
    if match:
        return match.group(0)
    else:
        raise ValueError("No JSON array found in AI response.")

def estimate_tokens(text):
    # Roughly estimate: 1 token ≈ 4 characters (for English)
    return len(text) // 4

def call_openai_with_retry(prompt, batch_info, max_retries=5):
    # Save prompt to file
    prompt_file = save_prompt_to_file(prompt, batch_info)
    
    for attempt in range(max_retries):
        try:
            prompt_len = len(prompt)
            prompt_tokens = estimate_tokens(prompt)
            print(f"\n--- Sending data to LLM API ---")
            print(f"Prompt size: {prompt_len} characters (~{prompt_tokens} tokens)")
            start_time = time.perf_counter()
            response = openai.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.0,
                max_tokens=4096
            )
            elapsed = time.perf_counter() - start_time
            print(f"--- Received response from LLM API (elapsed: {elapsed:.2f}s) ---")
            content = response.choices[0].message.content.strip()
              # Save response to file
            save_response_to_file(content, prompt_file)
            json_str = extract_first_json_array(content)
            result = json.loads(json_str)
            return result
        except openai.RateLimitError:
            wait = 2 ** attempt
            print(f"Rate limited. Waiting {wait} seconds before retrying...")
            time.sleep(wait)
        except Exception as e:
            print(f"Error parsing AI response or other error: {e}")
            if attempt == max_retries - 1:
                print("Max retries reached. Returning empty result for this batch.")
                # Save error response to file
                error_content = f"Error after {max_retries} attempts: {str(e)}"
                save_response_to_file(error_content, prompt_file)
                return []
            else:
                time.sleep(2 ** attempt)
    return []

# --- 3. Process pipe classes in batches with progress bar ---
results = []
num_batches = math.ceil(len(pipe_classes) / BATCH_SIZE)

print(f"Processing {len(pipe_classes)} pipe classes in {num_batches} batches of size {BATCH_SIZE}...")

with tqdm(total=num_batches, desc="Processing Pipe Classes") as pbar:
    for batch_idx in range(num_batches):
        batch_pipe_classes = pipe_classes[batch_idx*BATCH_SIZE:(batch_idx+1)*BATCH_SIZE]
        batch_summary_df = summary_df[summary_df['Pipe Class Name'].isin(batch_pipe_classes)]
        
        # Show which pipe class is being processed
        pbar.set_description(f"Pipe Class: {batch_pipe_classes[0]}")
        
        # Get relevant data for this batch only
        relevant_data = get_relevant_source_data(batch_pipe_classes)
        print(f"\nProcessing batch {batch_idx + 1}/{num_batches}")
        print(f"Pipe classes: {batch_pipe_classes}")
        print(f"Found {len(relevant_data)} relevant data sections")
        
        prompt = create_prompt(batch_pipe_classes, batch_summary_df)
        batch_info = f"Batch {batch_idx + 1}/{num_batches}: {batch_pipe_classes}"
        batch_results = call_openai_with_retry(prompt, batch_info)
        
        if batch_results:
            # Ensure pipe class names are properly set
            for i, result in enumerate(batch_results):
                if i < len(batch_pipe_classes):
                    result["Pipe Class Name"] = batch_pipe_classes[i]
            results.extend(batch_results)
        else:
            for pc in batch_pipe_classes:
                empty_result = {"Pipe Class Name": pc}
                for col in summary_columns:
                    if col != "Pipe Class Name":
                        empty_result[col] = "missing"
                results.append(empty_result)
        print(f"Waiting {BATCH_DELAY} seconds before next batch...\n")
        time.sleep(BATCH_DELAY)
        pbar.update(1)

# --- 4. Save combined results with conditional formatting ---
results_df = pd.DataFrame(results)

output_path = os.path.join(FOLDER, "pipe_class_validation_report.xlsx")

# Save with conditional formatting
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# First save the dataframe to Excel
results_df.to_excel(output_path, index=False)

# Then load the workbook to apply conditional formatting
wb = load_workbook(output_path)
ws = wb.active

# Define color fills
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light grey

# Get the data range (excluding header row)
max_row = ws.max_row
max_col = ws.max_column

# Apply conditional formatting to each cell in the data range
for row in range(2, max_row + 1):  # Start from row 2 (skip header)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell_value = str(cell.value).lower() if cell.value else ""
        
        if cell_value == "match":
            cell.fill = green_fill
        elif cell_value == "mismatch":
            cell.fill = red_fill
        elif cell_value == "missing":
            cell.fill = grey_fill

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # Add padding and cap at 50
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook with formatting
wb.save(output_path)
print(f"Validation report with conditional formatting saved as '{output_path}'.")
