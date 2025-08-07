import os
import tempfile
import pandas as pd
import openpyxl
import csv
from excel_file_search import search_excel_for_text, search_directory_for_excels

def create_test_xlsx(path, sheet_name, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in data:
        ws.append(row)
    wb.save(path)

def create_test_csv(path, data):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(data)

def test_search_excel_for_text():
    with tempfile.TemporaryDirectory() as tmpdir:
        # Create XLSX file
        xlsx_path = os.path.join(tmpdir, 'test.xlsx')
        create_test_xlsx(xlsx_path, 'Sheet1', [['foo', 'bar'], ['baz', 'searchPhrase']])
        # Create CSV file
        csv_path = os.path.join(tmpdir, 'test.csv')
        create_test_csv(csv_path, [['hello', 'searchPhrase'], ['world', 'test']])
        # Test XLSX
        matches_xlsx = search_excel_for_text(xlsx_path, 'searchPhrase')
        assert any('searchPhrase' in m['match_context'] for m in matches_xlsx)
        # Test CSV
        matches_csv = search_excel_for_text(csv_path, 'searchPhrase')
        assert any('searchPhrase' in m['match_context'] for m in matches_csv)
        # Test directory search
        matches_dir = search_directory_for_excels(tmpdir, 'searchPhrase')
        assert len(matches_dir) == 2

        # Test Excel report output
        output_path = os.path.join(tmpdir, 'excel_search_results.xlsx')
        from excel_file_search import create_excel_report
        create_excel_report(matches_dir, output_path, 'searchPhrase')
        # Check the hyperlink for all files in the report
        import openpyxl
        wb = openpyxl.load_workbook(output_path, data_only=False)
        ws = wb['Search Results']
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        file_name_idx = header.index('file_name')
        file_path_idx = header.index('file_path')
        found_hyperlink = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            file_name_cell = row[file_name_idx]
            file_path_cell = row[file_path_idx]
            if file_name_cell.hyperlink and str(file_path_cell.value) in str(file_name_cell.hyperlink.target):
                found_hyperlink = True
        wb.close()
        assert found_hyperlink, "File hyperlink not found in report."
        print('All tests passed.')

if __name__ == "__main__":
    test_search_excel_for_text()
