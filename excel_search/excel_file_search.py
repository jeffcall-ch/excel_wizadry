import os
import argparse
import pandas as pd
import openpyxl
import xlrd
import csv
from datetime import datetime
from pathlib import Path

def search_excel_for_text(file_path, search_text):
    """
    Search an Excel file (xlsx, xls, xlsm, xlsb) for the specified text (case-insensitive, partial matches).
    Returns a list of dicts: file_path, sheet_name, cell, match_context
    """
    matches = []
    ext = file_path.lower().split('.')[-1]
    search_pattern = search_text.lower()
    try:
        if ext in ['xlsx', 'xlsm', 'xlsb']:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and search_pattern in str(cell.value).lower():
                                context = str(cell.value)
                                matches.append({
                                    'file_path': file_path,
                                    'sheet_name': sheet_name,
                                    'cell': cell.coordinate,
                                    'match_context': context
                                })
            finally:
                wb.close()
        elif ext == 'xls':
            wb = xlrd.open_workbook(file_path)
            for sheet in wb.sheets():
                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        val = sheet.cell_value(row_idx, col_idx)
                        if val and search_pattern in str(val).lower():
                            context = str(val)
                            cell = f"R{row_idx+1}C{col_idx+1}"
                            matches.append({
                                'file_path': file_path,
                                'sheet_name': sheet.name,
                                'cell': cell,
                                'match_context': context
                            })
        elif ext == 'csv':
            with open(file_path, encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader):
                    for col_idx, val in enumerate(row):
                        if val and search_pattern in str(val).lower():
                            context = str(val)
                            cell = f"Row {row_idx+1}, Col {col_idx+1}"
                            matches.append({
                                'file_path': file_path,
                                'sheet_name': 'CSV',
                                'cell': cell,
                                'match_context': context
                            })
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
    return matches

def search_directory_for_excels(directory, search_text):
    """
    Search all Excel/CSV files in the directory and subdirectories.
    """
    all_matches = []
    excel_exts = ['.xlsx', '.xls', '.xlsm', '.xlsb', '.csv']
    file_count = 0
    match_count = 0
    for root, _, files in os.walk(directory):
        for file in files:
            if any(file.lower().endswith(ext) for ext in excel_exts):
                file_path = os.path.join(root, file)
                file_count += 1
                print(f"Searching {file_path}...")
                matches = search_excel_for_text(file_path, search_text)
                if matches:
                    match_count += len(matches)
                    all_matches.extend(matches)
    print(f"\nSearch completed. Found {match_count} matches across {file_count} Excel/CSV files.")
    return all_matches

def create_excel_report(matches, output_file, search_text):
    """
    Create an Excel report with match results and hyperlinks to the original files.
    """
    if not matches:
        print("No matches found. No Excel report generated.")
        return
    df = pd.DataFrame(matches)
    df['file_name'] = df['file_path'].apply(lambda x: os.path.basename(x))
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Search Results')
        workbook = writer.book
        worksheet = writer.sheets['Search Results']
        for row_idx, file_path in enumerate(df['file_path'], start=2):
            cell = worksheet.cell(row=row_idx, column=df.columns.get_loc('file_name') + 1)
            cell.hyperlink = file_path
            cell.style = "Hyperlink"
        for idx, column in enumerate(df.columns):
            column_width = max(len(str(column)), df[column].astype(str).str.len().max())
            column_letter = worksheet.cell(1, idx+1).column_letter
            worksheet.column_dimensions[column_letter].width = min(column_width + 2, 50)
        summary_data = {
            'Date of Search': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Search Term': search_text,
            'Total Excel/CSV Files Searched': len(set(df['file_path'])),
            'Total Matches Found': len(df),
            'Report Generated By': 'Excel Search Tool'
        }
        summary_df = pd.DataFrame(list(summary_data.items()), columns=['Item', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

def main():
    parser = argparse.ArgumentParser(description='Search Excel/CSV files for specific text and generate Excel report')
    parser.add_argument('search_text', help='Text to search for in Excel/CSV files')
    parser.add_argument('directory', help='Directory containing Excel/CSV files to search')
    parser.add_argument('--output', '-o', help='Output file path for Excel report', 
                        default=f'excel_search_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    args = parser.parse_args()
    print(f"Searching for '{args.search_text}' in {args.directory}...")
    matches = search_directory_for_excels(args.directory, args.search_text)
    create_excel_report(matches, args.output, args.search_text)
    if matches:
        print(f"Excel report created: {args.output}")

if __name__ == "__main__":
    main()
