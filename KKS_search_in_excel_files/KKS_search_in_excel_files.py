#!/usr/bin/env python3
"""
KKS Code Search Tool - CSV Output Version
Searches for KKS codes in Excel files and exports results to CSV.
"""

import os
import re
import argparse
import csv
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries


def find_excel_files(directory):
    """Find all Excel files in the given directory."""
    excel_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                excel_files.append(os.path.join(root, file))
    return excel_files


def has_autofilter(sheet):
    """Check if the sheet has autofilter applied."""
    return sheet.auto_filter is not None


def search_sheet(sheet, kks_code):
    """Search for KKS code in a sheet and return header row and matching rows."""
    kks_code_lower = kks_code.lower()
    hit_rows = []
    header_row = None
    header_row_idx = None
    
    # Find autofilter range and header row index
    filter_ref = getattr(sheet.auto_filter, 'ref', None)
    if filter_ref:
        min_col, min_row, max_col, max_row = range_boundaries(filter_ref)
        header_row_idx = min_row
    
    # Search through all rows
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_strs = [str(cell).lower() if cell is not None else '' for cell in row]
        
        # Capture header row if we know its position
        if header_row is None and header_row_idx is not None and idx == header_row_idx:
            header_row = row
        
        # Check if any cell contains the KKS code
        if any(kks_code_lower in cell for cell in row_strs):
            hit_rows.append(row)
    
    return header_row, hit_rows


def process_xlsx_file(filepath, kks_code, csv_writer):
    """Process XLSX/XLSM files and write results to CSV."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
        return 0
    
    matches_found = 0
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        if not has_autofilter(sheet):
            continue
            
        header_row, hit_rows = search_sheet(sheet, kks_code)
        if header_row and hit_rows:
            # Add source information row
            source_info = [""] * max(len(header_row), 3)
            source_info[0] = f"=== FILE: {os.path.basename(filepath)} | SHEET: {sheetname} ==="
            csv_writer.writerow(source_info)
            
            # Write header row
            csv_writer.writerow([str(cell) if cell is not None else "" for cell in header_row])
            
            # Write matching rows
            for row in hit_rows:
                csv_writer.writerow([str(cell) if cell is not None else "" for cell in row])
                matches_found += 1
            
            # Add separator row
            csv_writer.writerow([""])
    
    return matches_found


def process_xls_file(filepath, kks_code, csv_writer):
    """Process XLS files and write results to CSV."""
    try:
        df_dict = pd.read_excel(filepath, sheet_name=None, engine='xlrd')
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
        return 0
    
    matches_found = 0
    for sheetname, df in df_dict.items():
        # Check each row for KKS code matches
        hit_mask = df.apply(
            lambda row: row.astype(str).str.lower().str.contains(kks_code.lower()).any(), 
            axis=1
        )
        
        if hit_mask.any():
            # Add source information row
            source_info = [""] * max(len(df.columns), 3)
            source_info[0] = f"=== FILE: {os.path.basename(filepath)} | SHEET: {sheetname} ==="
            csv_writer.writerow(source_info)
            
            # Write header row (column names)
            csv_writer.writerow(list(df.columns))
            
            # Write matching rows
            df_hits = df[hit_mask]
            for _, row in df_hits.iterrows():
                csv_writer.writerow([str(cell) if pd.notna(cell) else "" for cell in row])
                matches_found += 1
            
            # Add separator row
            csv_writer.writerow([""])
    
    return matches_found


def main():
    """Main function to run the KKS search."""
    parser = argparse.ArgumentParser(description="KKS code search in Excel files - CSV output")
    parser.add_argument("search_term", help="KKS code to search for (partial match, case-insensitive)")
    parser.add_argument("directory", help="Directory to search")
    args = parser.parse_args()
    
    # Validate directory
    if not os.path.isdir(args.directory):
        print(f"Error: Directory '{args.directory}' does not exist.")
        return
    
    # Find Excel files
    excel_files = find_excel_files(args.directory)
    if not excel_files:
        print(f"No Excel files found in '{args.directory}'")
        return
    
    print(f"Found {len(excel_files)} Excel files to process...")
    
    # Generate output filename (always in search directory with search term and timestamp)
    safe_search = re.sub(r'[^A-Za-z0-9]+', '_', args.search_term)[:40]
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f"RESULT_{safe_search}_{date_str}.csv"
    output_path = os.path.join(args.directory, output_filename)
    
    # Process files and write to CSV
    total_matches = 0
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile)
            
            # Write header information
            csv_writer.writerow([f"KKS Search Results for: {args.search_term}"])
            csv_writer.writerow([f"Search performed on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            csv_writer.writerow([f"Directory: {args.directory}"])
            csv_writer.writerow([""])
            
            # Process each file
            for fpath in excel_files:
                print(f"Processing: {os.path.basename(fpath)}")
                ext = os.path.splitext(fpath)[1].lower()
                
                if ext in ['.xlsx', '.xlsm']:
                    matches = process_xlsx_file(fpath, args.search_term, csv_writer)
                    total_matches += matches
                elif ext == '.xls':
                    matches = process_xls_file(fpath, args.search_term, csv_writer)
                    total_matches += matches
        
        if total_matches > 0:
            print(f"Results saved to: {output_path}")
            print(f"Found {total_matches} matching row(s) across all files")
        else:
            # Close the current file and create a NO_RESULT file instead
            pass
        
        # Handle no results case - create NO_RESULT file
        if total_matches == 0:
            # Remove the original file
            os.remove(output_path)
            
            # Create NO_RESULT filename
            no_result_filename = f"NO_RESULT_{safe_search}_{date_str}.csv"
            no_result_path = os.path.join(args.directory, no_result_filename)
            
            # Create empty results file with search info
            with open(no_result_path, 'w', newline='', encoding='utf-8') as csvfile:
                csv_writer = csv.writer(csvfile)
                csv_writer.writerow([f"KKS Search Results for: {args.search_term}"])
                csv_writer.writerow([f"Search performed on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                csv_writer.writerow([f"Directory: {args.directory}"])
                csv_writer.writerow([f"Files processed: {len(excel_files)}"])
                csv_writer.writerow([""])
                csv_writer.writerow(["NO MATCHES FOUND"])
            
            print(f"No matches found for '{args.search_term}'")
            print(f"Empty result file saved to: {no_result_path}")
            
    except Exception as e:
        print(f"Error writing CSV file: {e}")


if __name__ == "__main__":
    main()