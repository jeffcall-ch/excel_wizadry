from __future__ import annotations

import argparse
import collections
import decimal
import difflib
import logging
import re
import sqlite3
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from logging_utils import DEFAULT_LOG_DIR, setup_csv_logging
from settings_markdown import load_markdown_settings, sections_hint

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

try:
    from Spare_calcs_for_price_sheet.run_spare_calcs_on_aggregated import run_spare_calcs as integrated_run_spare_calcs
except Exception:
    integrated_run_spare_calcs = None

NA_VALUE = "N/A"

TARGET_MATERIAL_TYPES = ("ss", "cs", "mapress")
REQUIRED_FILE_CATEGORIES = {
    ("ss", False),
    ("ss", True),
    ("cs", False),
    ("cs", True),
    ("mapress", False),
    ("mapress", True),
}

NORMAL_REQUIRED_KEYS = (
    "system",
    "pipe",
    "type",
    "pipe_component",
    "dn",
    "material",
    "total_weight",
    "qty_pcs",
    "qty_m",
)

ERECTION_REQUIRED_KEYS = (
    "pipe_kks",
    "type",
    "pipe_component",
    "dn",
    "material",
    "qty_pcs_per_m",
)

PAINT_REQUIRED_KEYS = (
    "name_of_system",
    "name_of_pipe",
    "type",
    "description",
    "dn_1",
    "material",
    "quantity_pcs",
    "quantity_m",
    "external_surface_m2",
    "corrosion_class",
    "painting_colour",
    "insulated",
)

ORIFICE_REQUIRED_KEYS = (
    "kks",
    "nominal_pipe_size",
)

SQLITE_NUMERIC_COLUMNS = {
    "UID",
    "Size",
    "Qty pcs",
    "Qty m",
    "Weight kg",
    "Surface m2",
    "Order qty",
    "Weight w spare",
    "Unit price",
    "Total price",
    "Delta qty to previous revision",
}

AGGREGATE_SUM_COLUMNS = ("Qty pcs", "Qty m", "Weight kg", "Surface m2")
AGGREGATE_MATERIAL_TYPE_SORT_PREFIXES = ("mapress", "ss", "cs", "paint")

SHOP_SUFFIX = "shop"
ERECTION_SUFFIX = "erection"
PAINT_MATERIAL_TYPE_LABEL = "05 Paint shop"
AGGREGATE_TABLE_SUFFIX = "aggreated"

PAINT_FILE_KEYWORDS = ("paint", "painting")
ORIFICE_FILE_KEYWORDS = ("orifice", "orifices")

MATERIAL_TYPE_BASE_LABELS = {"ss": "SS", "cs": "CS", "mapress": "Mapress"}

BASE_MATERIAL_TYPE_VALUES = {
    ("mapress", False): "01 Mapress shop",
    ("ss", False): "02 SS shop",
    ("cs", False): "03 CS shop",
    ("mapress", True): "04a Mapress erection",
    ("ss", True): "04b SS erection",
    ("cs", True): "05b CS erection",
}

NORMAL_SOURCE_TO_SQLITE = {
    "system": "System",
    "pipe": "KKS",
    "type": "Type",
    "pipe_component": "Description",
    "dn": "Size",
    "material": "Material",
    "total_weight": "Weight kg",
    "qty_pcs": "Qty pcs",
    "qty_m": "Qty m",
}

ERECTION_SOURCE_TO_SQLITE = {
    "pipe_kks": "KKS",
    "type": "Type",
    "pipe_component": "Description",
    "dn": "Size",
    "material": "Material",
    "qty_pcs_per_m": "Qty pcs",
}

PAINT_SOURCE_TO_SQLITE = {
    "name_of_system": "System",
    "name_of_pipe": "KKS",
    "type": "Type",
    "description": "Description",
    "dn_1": "Size",
    "material": "Material",
    "quantity_pcs": "Qty pcs",
    "quantity_m": "Qty m",
    "external_surface_m2": "Surface m2",
    "corrosion_class": "Corrosion category",
    "painting_colour": "Paint colour",
    "insulated": "Insulated",
}

ORIFICE_SOURCE_TO_SQLITE = {
    "kks": "KKS",
    "nominal_pipe_size": "Size",
}

ORIFICE_MATERIAL_TYPE_LABEL = "08 Orifice Plates"
ORIFICE_TYPE_VALUE = "ORIFICE PLATE"
ORIFICE_DESCRIPTION_TEXT = "Orifice Plate between flanges EN-1092-1/11/B1/PN16"
ORIFICE_ADD_INFO_TEMPLATE = "Orifice plate thickness S1={thickness} mm (DN {dn})."
ORIFICE_ADD_INFO_MISSING_TEMPLATE = "Orifice plate thickness S1=N/A (DN {dn})."
ORIFICE_QTY_PER_ROW = 1
ORIFICE_MATERIAL_VALUE = "1.4571 or 1.4404"
ORIFICE_THICKNESS_BY_DN: Dict[str, str] = {}

MAPRESS_SEAL_NOTE = 'Note: The required seal rings are listed separately with "SEAL" type this list.'
MAPRESS_SEAL_ORDER_SIZE: Dict[str, str] = {}

ALIAS_MAP = {
    "system": ["system"],
    "pipe": ["pipe"],
    "pipe_kks": ["pipe kks", "pipe"],
    "type": ["type"],
    "pipe_component": ["pipe component", "pipecomponent"],
    "material": ["material"],
    "total_weight": ["total weight", "total weight kg", "total weight kg"],
    "qty_pcs": ["qty pcs", "qty pcs ", "qty pcs"],
    "qty_m": ["qty m"],
    "qty_pcs_per_m": ["qty pcs m", "qty pcs m", "qty pcs m"],
    "name_of_system": ["name of system"],
    "name_of_pipe": ["name of pipe"],
    "description": ["description"],
    "dn_1": ["dn 1", "dn1"],
    "quantity_pcs": ["quantity pcs", "quantity pcs"],
    "quantity_m": ["quantity m"],
    "external_surface_m2": ["external surface m2", "external surface"],
    "corrosion_class": ["corrosion class"],
    "insulated": ["insulated"],
    "kks": ["kks"],
    "nominal_pipe_size": ["nominal pipe size", "dn", "dn size", "nominal size"],
}

AGGREGATION_MATERIAL_TYPE_RULES = [
    ("endswith", "erection", "04 Erection"),
    ("contains", "paint shop", "05 Paint shop"),
]


logger = logging.getLogger(__name__)


def settings_sheet_hint(*sheet_names: str) -> str:
    return sections_hint(*sheet_names)


@dataclass
class ClassifiedFile:
    path: Path
    material_type: str
    is_erection: bool


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def tokenize_filename(filename: str) -> List[str]:
    return [token for token in re.split(r"[^a-z0-9]+", filename.lower()) if token]


def classify_file(path: Path) -> Optional[ClassifiedFile]:
    tokens = tokenize_filename(path.name)
    if "material" not in tokens:
        return None

    material_type = None
    for candidate in TARGET_MATERIAL_TYPES:
        if candidate in tokens:
            material_type = candidate
            break

    if material_type is None:
        return None

    return ClassifiedFile(path=path, material_type=material_type, is_erection=("erection" in tokens))


def is_paint_file(path: Path) -> bool:
    tokens = tokenize_filename(path.name)
    return any(any(token.startswith(keyword) for keyword in PAINT_FILE_KEYWORDS) for token in tokens)


def is_orifice_file(path: Path) -> bool:
    tokens = tokenize_filename(path.name)
    return any(any(token.startswith(keyword) for keyword in ORIFICE_FILE_KEYWORDS) for token in tokens)


def locate_required_files(
    input_dir: Path,
) -> Tuple[Dict[Tuple[str, bool], ClassifiedFile], List[str], List[Path], List[Path], List[Path]]:
    found: Dict[Tuple[str, bool], ClassifiedFile] = {}
    warnings: List[str] = []
    paint_files: List[Path] = []
    orifice_files: List[Path] = []
    ignored_files: List[Path] = []

    for path in sorted(input_dir.glob("*.xls*")):
        if path.name.startswith("~$"):
            continue

        classified = classify_file(path)
        if classified is None:
            if is_paint_file(path):
                paint_files.append(path)
            elif is_orifice_file(path):
                orifice_files.append(path)
            else:
                ignored_files.append(path)
            continue

        key = (classified.material_type, classified.is_erection)
        if key in found:
            warnings.append(
                f"Duplicate file category for {format_category(key)}: '{found[key].path.name}' and '{path.name}'. "
                f"{settings_sheet_hint('RequiredFileCategories', 'MaterialTypes')}"
            )
            continue
        found[key] = classified

    missing = sorted(REQUIRED_FILE_CATEGORIES - set(found.keys()))
    for category in missing:
        warnings.append(
            f"Missing required file for {format_category(category)}. "
            f"{settings_sheet_hint('RequiredFileCategories', 'MaterialTypes')}"
        )

    return found, warnings, paint_files, orifice_files, ignored_files


def format_category(category: Tuple[str, bool]) -> str:
    material_type, is_erection = category
    label = material_type.upper() if material_type != "mapress" else "MAPRESS"
    return f"{label} {'erection material' if is_erection else 'material'}"


def material_type_value(material_type: str, is_erection: bool) -> str:
    key = (material_type, is_erection)
    if key in BASE_MATERIAL_TYPE_VALUES:
        return BASE_MATERIAL_TYPE_VALUES[key]
    base = MATERIAL_TYPE_BASE_LABELS[material_type]
    return f"{base} {ERECTION_SUFFIX}" if is_erection else f"{base} {SHOP_SUFFIX}"


def choose_sheet_with_max_rows(workbook) -> str:
    return max(workbook.sheetnames, key=lambda name: workbook[name].max_row)


def load_sqlite_headers(headers_workbook: Path) -> List[str]:
    wb = load_workbook(headers_workbook, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1] if cell.value is not None and str(cell.value).strip()]
    return [str(h).strip() for h in headers]


def is_dn_header(normalized: str) -> bool:
    return normalized == "dn"


def aliases_for_key(key: str) -> List[str]:
    return ALIAS_MAP.get(key, [key])


def parse_bool_like(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _load_mapping_sheet(records: List[Dict[str, object]]) -> Tuple[Dict[str, str], Tuple[str, ...]]:
    if not records:
        return {}, tuple()

    mapping: Dict[str, str] = {}
    required: List[str] = []
    for record in records:
        source_key_raw = record.get("SourceKey")
        sqlite_col_raw = record.get("SQLiteColumn")
        if source_key_raw is None or sqlite_col_raw is None:
            continue
        source_key = str(source_key_raw).strip()
        sqlite_col = str(sqlite_col_raw).strip()
        if not source_key or not sqlite_col:
            continue
        mapping[source_key] = sqlite_col
        if parse_bool_like(record.get("Required")):
            required.append(source_key)

    return mapping, tuple(required)


def load_import_settings(settings_workbook: Path) -> None:
    global TARGET_MATERIAL_TYPES
    global REQUIRED_FILE_CATEGORIES
    global NORMAL_REQUIRED_KEYS
    global ERECTION_REQUIRED_KEYS
    global PAINT_REQUIRED_KEYS
    global ORIFICE_REQUIRED_KEYS
    global SQLITE_NUMERIC_COLUMNS
    global AGGREGATE_SUM_COLUMNS
    global SHOP_SUFFIX
    global ERECTION_SUFFIX
    global PAINT_MATERIAL_TYPE_LABEL
    global AGGREGATE_TABLE_SUFFIX
    global PAINT_FILE_KEYWORDS
    global ORIFICE_FILE_KEYWORDS
    global MATERIAL_TYPE_BASE_LABELS
    global NORMAL_SOURCE_TO_SQLITE
    global ERECTION_SOURCE_TO_SQLITE
    global PAINT_SOURCE_TO_SQLITE
    global ORIFICE_SOURCE_TO_SQLITE
    global ORIFICE_MATERIAL_TYPE_LABEL
    global ORIFICE_TYPE_VALUE
    global ORIFICE_DESCRIPTION_TEXT
    global ORIFICE_ADD_INFO_TEMPLATE
    global ORIFICE_ADD_INFO_MISSING_TEMPLATE
    global ORIFICE_QTY_PER_ROW
    global ORIFICE_MATERIAL_VALUE
    global ORIFICE_THICKNESS_BY_DN
    global MAPRESS_SEAL_ORDER_SIZE
    global ALIAS_MAP
    global AGGREGATION_MATERIAL_TYPE_RULES

    settings_doc = load_markdown_settings(settings_workbook)

    general_records = settings_doc.get_records("General")
    general = {}
    for rec in general_records:
        key = str(rec.get("Key", "")).strip()
        val = rec.get("Value")
        if key:
            general[key] = "" if val is None else str(val).strip()

    SHOP_SUFFIX = general.get("shop_suffix", SHOP_SUFFIX)
    ERECTION_SUFFIX = general.get("erection_suffix", ERECTION_SUFFIX)
    PAINT_MATERIAL_TYPE_LABEL = general.get("paint_material_type_label", PAINT_MATERIAL_TYPE_LABEL)
    AGGREGATE_TABLE_SUFFIX = general.get("aggregate_table_suffix", AGGREGATE_TABLE_SUFFIX)

    material_type_records = settings_doc.get_records("MaterialTypes")
    loaded_types = [str(r.get("MaterialType", "")).strip().lower() for r in material_type_records]
    loaded_types = [x for x in loaded_types if x]
    if loaded_types:
        TARGET_MATERIAL_TYPES = tuple(loaded_types)

    required_file_records = settings_doc.get_records("RequiredFileCategories")
    loaded_required_categories = set()
    for rec in required_file_records:
        material_type = str(rec.get("MaterialType", "")).strip().lower()
        if not material_type:
            continue
        is_erection = parse_bool_like(rec.get("IsErection"))
        loaded_required_categories.add((material_type, is_erection))
    if loaded_required_categories:
        REQUIRED_FILE_CATEGORIES = loaded_required_categories

    material_type_label_records = settings_doc.get_records("MaterialTypeLabels")
    loaded_labels = {}
    for rec in material_type_label_records:
        material_type = str(rec.get("MaterialType", "")).strip().lower()
        label = str(rec.get("Label", "")).strip()
        if material_type and label:
            loaded_labels[material_type] = label
    if loaded_labels:
        MATERIAL_TYPE_BASE_LABELS = loaded_labels

    mapping, required = _load_mapping_sheet(settings_doc.get_records("ColumnMappings_Normal"))
    if mapping:
        NORMAL_SOURCE_TO_SQLITE = mapping
    if required:
        NORMAL_REQUIRED_KEYS = required

    mapping, required = _load_mapping_sheet(settings_doc.get_records("ColumnMappings_Erection"))
    if mapping:
        ERECTION_SOURCE_TO_SQLITE = mapping
    if required:
        ERECTION_REQUIRED_KEYS = required

    mapping, required = _load_mapping_sheet(settings_doc.get_records("ColumnMappings_Paint"))
    if mapping:
        PAINT_SOURCE_TO_SQLITE = mapping
    if required:
        PAINT_REQUIRED_KEYS = required

    mapping, required = _load_mapping_sheet(settings_doc.get_records("ColumnMappings_Orifice"))
    if mapping:
        ORIFICE_SOURCE_TO_SQLITE = mapping
    if required:
        ORIFICE_REQUIRED_KEYS = required

    alias_records = settings_doc.get_records("Aliases")
    if alias_records:
        grouped_aliases: Dict[str, List[str]] = collections.defaultdict(list)
        for rec in alias_records:
            key = str(rec.get("Key", "")).strip()
            alias = str(rec.get("Alias", "")).strip()
            if key and alias:
                grouped_aliases[key].append(alias)
        if grouped_aliases:
            ALIAS_MAP = dict(grouped_aliases)

    keyword_records = settings_doc.get_records("FileKeywords")
    paint_keywords = [
        str(rec.get("Keyword", "")).strip().lower()
        for rec in keyword_records
        if str(rec.get("Kind", "")).strip().lower() == "paint"
    ]
    paint_keywords = [k for k in paint_keywords if k]
    if paint_keywords:
        PAINT_FILE_KEYWORDS = tuple(paint_keywords)

    orifice_keywords = [
        str(rec.get("Keyword", "")).strip().lower()
        for rec in keyword_records
        if str(rec.get("Kind", "")).strip().lower() == "orifice"
    ]
    orifice_keywords = [k for k in orifice_keywords if k]
    if orifice_keywords:
        ORIFICE_FILE_KEYWORDS = tuple(orifice_keywords)

    numeric_col_records = settings_doc.get_records("SqliteNumericColumns")
    loaded_numeric_cols = [str(rec.get("Column", "")).strip() for rec in numeric_col_records]
    loaded_numeric_cols = [col for col in loaded_numeric_cols if col]
    if loaded_numeric_cols:
        SQLITE_NUMERIC_COLUMNS = set(loaded_numeric_cols)

    agg_sum_records = settings_doc.get_records("AggregateSumColumns")
    loaded_agg_sum_cols = [str(rec.get("Column", "")).strip() for rec in agg_sum_records]
    loaded_agg_sum_cols = [col for col in loaded_agg_sum_cols if col]
    if loaded_agg_sum_cols:
        AGGREGATE_SUM_COLUMNS = tuple(loaded_agg_sum_cols)

    agg_rule_records = settings_doc.get_records("AggregationMaterialTypeRules")
    loaded_rules = []
    for rec in agg_rule_records:
        match_type = str(rec.get("MatchType", "")).strip().lower()
        pattern = str(rec.get("Pattern", "")).strip().lower()
        output = str(rec.get("Output", "")).strip()
        if match_type and pattern and output:
            loaded_rules.append((match_type, pattern, output))
    if loaded_rules:
        AGGREGATION_MATERIAL_TYPE_RULES = loaded_rules

    orifice_defaults_records = settings_doc.get_records("OrificeDefaults")
    orifice_defaults = {str(r.get("Key", "")).strip(): r.get("Value") for r in orifice_defaults_records}
    ORIFICE_MATERIAL_TYPE_LABEL = str(orifice_defaults.get("material_type_label", ORIFICE_MATERIAL_TYPE_LABEL) or ORIFICE_MATERIAL_TYPE_LABEL)
    ORIFICE_TYPE_VALUE = str(orifice_defaults.get("type_value", ORIFICE_TYPE_VALUE) or ORIFICE_TYPE_VALUE)
    ORIFICE_DESCRIPTION_TEXT = str(orifice_defaults.get("description", ORIFICE_DESCRIPTION_TEXT) or ORIFICE_DESCRIPTION_TEXT)
    ORIFICE_ADD_INFO_TEMPLATE = str(orifice_defaults.get("add_info_template", ORIFICE_ADD_INFO_TEMPLATE) or ORIFICE_ADD_INFO_TEMPLATE)
    ORIFICE_ADD_INFO_MISSING_TEMPLATE = str(
        orifice_defaults.get("add_info_missing_template", ORIFICE_ADD_INFO_MISSING_TEMPLATE)
        or ORIFICE_ADD_INFO_MISSING_TEMPLATE
    )
    ORIFICE_QTY_PER_ROW = int(orifice_defaults.get("qty_per_row", ORIFICE_QTY_PER_ROW) or ORIFICE_QTY_PER_ROW)
    ORIFICE_MATERIAL_VALUE = str(orifice_defaults.get("material_value", ORIFICE_MATERIAL_VALUE) or ORIFICE_MATERIAL_VALUE)

    thickness_records = settings_doc.get_records("BlindDiskThickness")
    loaded_thickness: Dict[str, str] = {}
    for rec in thickness_records:
        dn = str(rec.get("DN", "")).strip()
        s1 = str(rec.get("S1 mm", "")).strip()
        if dn and s1:
            loaded_thickness[dn] = s1
    ORIFICE_THICKNESS_BY_DN = loaded_thickness

    mapress_seal_records = settings_doc.get_records("MapressSealOrderSize")
    loaded_seal_sizes: Dict[str, str] = {}
    for rec in mapress_seal_records:
        order_no = str(rec.get("OrderNo", "")).strip()
        size = str(rec.get("Size", "")).strip()
        if order_no and size:
            loaded_seal_sizes[order_no] = size
    MAPRESS_SEAL_ORDER_SIZE = loaded_seal_sizes


def smart_find_paint_colour_column(
    normalized_map: Dict[str, int],
    used_indices: set[int],
) -> Optional[int]:
    strong_matches: List[int] = []
    weak_matches: List[int] = []

    for normalized, idx in normalized_map.items():
        if idx in used_indices:
            continue
        has_paint = "paint" in normalized
        has_color = ("colour" in normalized) or ("color" in normalized)
        if has_paint and has_color:
            strong_matches.append(idx)
        elif has_color:
            weak_matches.append(idx)

    if len(strong_matches) == 1:
        return strong_matches[0]
    if len(strong_matches) > 1:
        return None
    if len(weak_matches) == 1:
        return weak_matches[0]
    return None


def best_fuzzy_match(target: str, available: Dict[str, int], used_indices: set[int]) -> Optional[int]:
    candidates: List[Tuple[float, int]] = []
    for normalized, idx in available.items():
        if idx in used_indices:
            continue
        ratio = difflib.SequenceMatcher(None, target, normalized).ratio()
        if ratio >= 0.88:
            candidates.append((ratio, idx))

    if not candidates:
        return None

    candidates.sort(reverse=True)
    if len(candidates) > 1 and abs(candidates[0][0] - candidates[1][0]) < 0.05:
        return None

    return candidates[0][1]


def detect_header_row(ws, max_scan_rows: int = 40) -> int:
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [normalize_text(v) for v in values if normalize_text(v)]
        if not normalized:
            continue
        # Prefer rows that look like headers but also contain many non-empty cells.
        score = len(set(normalized)) * 10 + len(normalized)
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def resolve_required_column_indices(
    ws,
    header_row_idx: int,
    is_erection: bool,
) -> Tuple[Dict[str, int], List[str]]:
    required_keys = ERECTION_REQUIRED_KEYS if is_erection else NORMAL_REQUIRED_KEYS
    warnings: List[str] = []

    headers = [ws.cell(row=header_row_idx, column=col).value for col in range(1, ws.max_column + 1)]
    normalized_map: Dict[str, int] = {}

    for idx, value in enumerate(headers, start=1):
        normalized = normalize_text(value)
        if normalized:
            normalized_map[normalized] = idx

    matched: Dict[str, int] = {}
    used_indices: set[int] = set()

    for key in required_keys:
        if key == "dn":
            dn_idx = None
            for normalized, idx in normalized_map.items():
                if is_dn_header(normalized):
                    dn_idx = idx
                    break
            if dn_idx is None:
                raise ValueError(
                    "Required column 'DN' not found with exact header 'DN'. "
                    f"{settings_sheet_hint('ColumnMappings_Normal', 'ColumnMappings_Erection', 'Aliases')}"
                )
            matched[key] = dn_idx
            used_indices.add(dn_idx)
            continue

        aliases = aliases_for_key(key)
        exact_idx = None
        for alias in aliases:
            alias_norm = normalize_text(alias)
            if alias_norm in normalized_map and normalized_map[alias_norm] not in used_indices:
                exact_idx = normalized_map[alias_norm]
                break

        if exact_idx is not None:
            matched[key] = exact_idx
            used_indices.add(exact_idx)
            continue

        fuzzy_idx = None
        for alias in aliases:
            alias_norm = normalize_text(alias)
            fuzzy_idx = best_fuzzy_match(alias_norm, normalized_map, used_indices)
            if fuzzy_idx is not None:
                warnings.append(
                    f"Fuzzy matched source column for '{key}' to header '{headers[fuzzy_idx - 1]}'."
                )
                break

        if fuzzy_idx is None:
            target_sheet = "ColumnMappings_Erection" if is_erection else "ColumnMappings_Normal"
            raise ValueError(
                f"Required column '{key}' was not found. "
                f"{settings_sheet_hint(target_sheet, 'Aliases')}"
            )

        matched[key] = fuzzy_idx
        used_indices.add(fuzzy_idx)

    return matched, warnings


def resolve_paint_required_column_indices(
    ws,
    header_row_idx: int,
) -> Tuple[Dict[str, int], List[str]]:
    warnings: List[str] = []
    headers = [ws.cell(row=header_row_idx, column=col).value for col in range(1, ws.max_column + 1)]
    normalized_map: Dict[str, int] = {}

    for idx, value in enumerate(headers, start=1):
        normalized = normalize_text(value)
        if normalized:
            normalized_map[normalized] = idx

    matched: Dict[str, int] = {}
    used_indices: set[int] = set()

    for key in PAINT_REQUIRED_KEYS:
        if key == "dn_1":
            dn_idx = None
            for normalized, idx in normalized_map.items():
                if normalized == "dn 1":
                    dn_idx = idx
                    break
            if dn_idx is None:
                raise ValueError(
                    "Required paint column 'DN 1' not found. "
                    f"{settings_sheet_hint('ColumnMappings_Paint', 'Aliases')}"
                )
            matched[key] = dn_idx
            used_indices.add(dn_idx)
            continue

        if key == "painting_colour":
            color_idx = smart_find_paint_colour_column(normalized_map, used_indices)
            if color_idx is None:
                raise ValueError(
                    "Required paint column for 'Painting colour' not found (color/colour). "
                    f"{settings_sheet_hint('ColumnMappings_Paint', 'Aliases')}"
                )
            matched[key] = color_idx
            used_indices.add(color_idx)
            if normalize_text(headers[color_idx - 1]) not in {"painting colour", "painting color"}:
                warnings.append(
                    f"Smart matched paint color column to header '{headers[color_idx - 1]}'."
                )
            continue

        aliases = aliases_for_key(key)
        exact_idx = None
        for alias in aliases:
            alias_norm = normalize_text(alias)
            if alias_norm in normalized_map and normalized_map[alias_norm] not in used_indices:
                exact_idx = normalized_map[alias_norm]
                break

        if exact_idx is not None:
            matched[key] = exact_idx
            used_indices.add(exact_idx)
            continue

        fuzzy_idx = None
        for alias in aliases:
            alias_norm = normalize_text(alias)
            fuzzy_idx = best_fuzzy_match(alias_norm, normalized_map, used_indices)
            if fuzzy_idx is not None:
                warnings.append(
                    f"Fuzzy matched paint source column for '{key}' to header '{headers[fuzzy_idx - 1]}'."
                )
                break

        if fuzzy_idx is None:
            raise ValueError(
                f"Required paint column '{key}' was not found. "
                f"{settings_sheet_hint('ColumnMappings_Paint', 'Aliases')}"
            )

        matched[key] = fuzzy_idx
        used_indices.add(fuzzy_idx)

    return matched, warnings


def resolve_orifice_required_column_indices(
    ws,
    header_row_idx: int,
) -> Tuple[Dict[str, int], List[str]]:
    warnings: List[str] = []
    headers = [ws.cell(row=header_row_idx, column=col).value for col in range(1, ws.max_column + 1)]
    normalized_map: Dict[str, int] = {}

    for idx, value in enumerate(headers, start=1):
        normalized = normalize_text(value)
        if normalized:
            normalized_map[normalized] = idx

    matched: Dict[str, int] = {}
    used_indices: set[int] = set()

    for key in ORIFICE_REQUIRED_KEYS:
        aliases = aliases_for_key(key)
        exact_idx = None
        for alias in aliases:
            alias_norm = normalize_text(alias)
            if alias_norm in normalized_map and normalized_map[alias_norm] not in used_indices:
                exact_idx = normalized_map[alias_norm]
                break

        if exact_idx is not None:
            matched[key] = exact_idx
            used_indices.add(exact_idx)
            continue

        fuzzy_idx = None
        for alias in aliases:
            alias_norm = normalize_text(alias)
            fuzzy_idx = best_fuzzy_match(alias_norm, normalized_map, used_indices)
            if fuzzy_idx is not None:
                warnings.append(
                    f"Fuzzy matched orifice source column for '{key}' to header '{headers[fuzzy_idx - 1]}'."
                )
                break

        if fuzzy_idx is None:
            raise ValueError(
                f"Required orifice column '{key}' was not found. "
                f"{settings_sheet_hint('ColumnMappings_Orifice', 'Aliases')}"
            )

        matched[key] = fuzzy_idx
        used_indices.add(fuzzy_idx)

    return matched, warnings


def get_cell_text(value: object) -> str:
    if value is None:
        return NA_VALUE
    text = str(value).strip()
    return text if text else NA_VALUE


def dn_numeric_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if not text or text.upper() == NA_VALUE:
        return NA_VALUE
    match = re.search(r"(\d+(?:[\.,]\d+)?)", text)
    if not match:
        return text
    raw = match.group(1).replace(",", ".")
    try:
        num = float(raw)
        if abs(num - round(num)) < 1e-12:
            return str(int(round(num)))
        return str(num)
    except ValueError:
        return text


def _find_mapress_seal_marker_index(description: str) -> int:
    lowered = description.lower()
    markers = ["seal ring fkm blue order no.", "seal ring fmk blue order no."]
    hits = [lowered.find(m) for m in markers if lowered.find(m) >= 0]
    return min(hits) if hits else -1


def rewrite_mapress_seal_description(description: str) -> str:
    idx = _find_mapress_seal_marker_index(description)
    if idx < 0:
        return description
    prefix = description[:idx].rstrip()
    if not prefix:
        return MAPRESS_SEAL_NOTE
    return f"{prefix} {MAPRESS_SEAL_NOTE}"


def extract_mapress_seal_orders(description: str) -> List[Tuple[str, int]]:
    idx = _find_mapress_seal_marker_index(description)
    if idx < 0:
        return []
    tail = description[idx:]
    out: List[Tuple[str, int]] = []
    for m in re.finditer(r"\b(\d{5})\b", tail):
        order_no = m.group(1)
        before = tail[max(0, m.start() - 16):m.start()]
        after = tail[m.end():m.end() + 16]
        mult = 1
        mb = re.search(r"(\d+)\s*x(?:\s+and)?\s*$", before, flags=re.IGNORECASE)
        if mb:
            mult = int(mb.group(1))
        else:
            ma = re.search(r"^\s*(\d+)\s*x\b", after, flags=re.IGNORECASE)
            if ma:
                mult = int(ma.group(1))
            else:
                ma2 = re.search(r"^\s*x\s*(\d+)\b", after, flags=re.IGNORECASE)
                if ma2:
                    mult = int(ma2.group(1))
        out.append((order_no, max(1, mult)))
    return out


def _decimal_places_from_excel_number_format(number_format: str) -> Optional[int]:
    if not number_format:
        return None

    # Use only the positive section in formats such as "0.00;[Red]-0.00".
    fmt = number_format.split(";", 1)[0]
    fmt = re.sub(r'"[^"]*"', "", fmt)
    fmt = re.sub(r"\\.", "", fmt)
    fmt = re.sub(r"\[[^\]]*\]", "", fmt)
    fmt = fmt.replace("*", "").replace("_", "")

    if "." not in fmt:
        # Numeric format without decimal separator means integer-like display.
        if re.search(r"[0#?]", fmt):
            return 0
        return None

    fractional = fmt.split(".", 1)[1]
    placeholders = re.findall(r"[0#?]", fractional)
    return len(placeholders)


def cell_to_sql_value(cell: Any) -> Any:
    value = cell.value
    if value is None:
        return NA_VALUE

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        dec_value = decimal.Decimal(str(value))
        decimals = _decimal_places_from_excel_number_format(str(getattr(cell, "number_format", "")))

        if decimals is not None:
            quantizer = decimal.Decimal("1").scaleb(-decimals)
            dec_value = dec_value.quantize(quantizer, rounding=decimal.ROUND_HALF_UP)

        if decimals == 0:
            return int(dec_value)

        # Float keeps SQLite numeric storage class while honoring displayed precision.
        return float(dec_value)

    text = str(value).strip()
    return text if text else NA_VALUE


def parse_bom_revision(filename: str) -> str:
    match = re.search(r"_(\d+(?:\.\d+)?)\b", filename)
    if not match:
        return NA_VALUE
    return match.group(1)


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    cur = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ? LIMIT 1;",
        (table_name,),
    )
    return cur.fetchone() is not None


def db_has_any_rows(conn: sqlite3.Connection) -> bool:
    table_names = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table';")]
    for table_name in table_names:
        cur = conn.execute(f'SELECT 1 FROM "{table_name}" LIMIT 1;')
        if cur.fetchone() is not None:
            return True
    return False


def sanitize_table_name(raw_name: str) -> str:
    candidate = re.sub(r"[^a-zA-Z0-9_]", "_", raw_name)
    if not re.match(r"^[a-zA-Z_]", candidate):
        candidate = f"rev_{candidate}"
    return candidate


def create_table(conn: sqlite3.Connection, table_name: str, headers: Iterable[str]) -> None:
    col_defs = []
    for header in headers:
        sqlite_type = "NUMERIC" if header in SQLITE_NUMERIC_COLUMNS else "TEXT"
        col_defs.append(f'"{header}" {sqlite_type}')
    columns_sql = ", ".join(col_defs)
    conn.execute(f'CREATE TABLE "{table_name}" ({columns_sql});')


def build_rows_for_file(
    classified: ClassifiedFile,
    sqlite_headers: List[str],
    uid_start: int,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    wb = load_workbook(classified.path, data_only=True, read_only=True)
    sheet_name = choose_sheet_with_max_rows(wb)
    ws = wb[sheet_name]

    header_row_idx = detect_header_row(ws)
    source_columns, warnings = resolve_required_column_indices(ws, header_row_idx, classified.is_erection)

    source_to_sqlite = ERECTION_SOURCE_TO_SQLITE if classified.is_erection else NORMAL_SOURCE_TO_SQLITE

    rows: List[Dict[str, Any]] = []
    uid = uid_start

    for row_cells in ws.iter_rows(
        min_row=header_row_idx + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
        values_only=False,
    ):
        source_values: Dict[str, Any] = {}
        for source_key, col_idx in source_columns.items():
            source_values[source_key] = cell_to_sql_value(row_cells[col_idx - 1])

        if all(value == NA_VALUE for value in source_values.values()):
            continue

        record: Dict[str, Any] = {header: NA_VALUE for header in sqlite_headers}

        record["UID"] = uid
        uid += 1

        record["Material type"] = material_type_value(classified.material_type, classified.is_erection)
        record["File name - complete"] = classified.path.name
        record["BoM rev from file name"] = parse_bom_revision(classified.path.name)

        for source_key, sqlite_col in source_to_sqlite.items():
            if sqlite_col in record:
                record[sqlite_col] = source_values.get(source_key, NA_VALUE)

        rows.append(record)

    if not rows:
        raise ValueError(
            f"No data rows found in file '{classified.path.name}' after applying required column mapping."
        )

    return rows, warnings


def build_rows_for_paint_file(
    paint_file: Path,
    sqlite_headers: List[str],
    uid_start: int,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    wb = load_workbook(paint_file, data_only=True, read_only=True)
    sheet_name = choose_sheet_with_max_rows(wb)
    ws = wb[sheet_name]

    header_row_idx = detect_header_row(ws)
    source_columns, warnings = resolve_paint_required_column_indices(ws, header_row_idx)

    source_to_sqlite = PAINT_SOURCE_TO_SQLITE

    rows: List[Dict[str, Any]] = []
    uid = uid_start

    for row_cells in ws.iter_rows(
        min_row=header_row_idx + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
        values_only=False,
    ):
        source_values: Dict[str, Any] = {}
        for source_key, col_idx in source_columns.items():
            source_values[source_key] = cell_to_sql_value(row_cells[col_idx - 1])

        if all(value == NA_VALUE for value in source_values.values()):
            continue

        record: Dict[str, Any] = {header: NA_VALUE for header in sqlite_headers}
        record["UID"] = uid
        uid += 1

        record["Material type"] = PAINT_MATERIAL_TYPE_LABEL
        record["File name - complete"] = paint_file.name
        record["BoM rev from file name"] = parse_bom_revision(paint_file.name)

        for source_key, sqlite_col in source_to_sqlite.items():
            if sqlite_col in record:
                record[sqlite_col] = source_values.get(source_key, NA_VALUE)

        rows.append(record)

    if not rows:
        raise ValueError(
            f"No paint data rows found in file '{paint_file.name}' after applying required column mapping."
        )

    return rows, warnings


def build_rows_for_orifice_file(
    orifice_file: Path,
    sqlite_headers: List[str],
    uid_start: int,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    wb = load_workbook(orifice_file, data_only=True, read_only=True)
    sheet_name = choose_sheet_with_max_rows(wb)
    ws = wb[sheet_name]

    header_row_idx = detect_header_row(ws)
    source_columns, warnings = resolve_orifice_required_column_indices(ws, header_row_idx)

    rows: List[Dict[str, Any]] = []
    uid = uid_start

    for row_cells in ws.iter_rows(
        min_row=header_row_idx + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
        values_only=False,
    ):
        source_values: Dict[str, Any] = {}
        for source_key, col_idx in source_columns.items():
            source_values[source_key] = cell_to_sql_value(row_cells[col_idx - 1])

        if all(value == NA_VALUE for value in source_values.values()):
            continue

        record: Dict[str, Any] = {header: NA_VALUE for header in sqlite_headers}
        record["UID"] = uid
        uid += 1

        record["Material type"] = ORIFICE_MATERIAL_TYPE_LABEL
        record["File name - complete"] = orifice_file.name
        record["BoM rev from file name"] = parse_bom_revision(orifice_file.name)

        kks_col = ORIFICE_SOURCE_TO_SQLITE.get("kks", "KKS")
        size_col = ORIFICE_SOURCE_TO_SQLITE.get("nominal_pipe_size", "Size")
        mat_col = "Material"

        if kks_col in record:
            record[kks_col] = source_values.get("kks", NA_VALUE)
        size_dn = dn_numeric_text(source_values.get("nominal_pipe_size", NA_VALUE))
        if size_col in record:
            record[size_col] = size_dn
        if mat_col in record:
            record[mat_col] = ORIFICE_MATERIAL_VALUE

        if "Description" in record:
            record["Description"] = ORIFICE_DESCRIPTION_TEXT
        if "Type" in record:
            record["Type"] = ORIFICE_TYPE_VALUE

        thickness = ORIFICE_THICKNESS_BY_DN.get(size_dn, NA_VALUE)
        if "Add. Info" in record:
            if thickness != NA_VALUE:
                record["Add. Info"] = ORIFICE_ADD_INFO_TEMPLATE.format(thickness=thickness, dn=size_dn)
            else:
                record["Add. Info"] = ORIFICE_ADD_INFO_MISSING_TEMPLATE.format(dn=size_dn)

        if "Qty pcs" in record:
            record["Qty pcs"] = ORIFICE_QTY_PER_ROW

        rows.append(record)

    if not rows:
        raise ValueError(
            f"No orifice data rows found in file '{orifice_file.name}' after applying required column mapping."
        )

    return rows, warnings


def insert_rows(conn: sqlite3.Connection, table_name: str, sqlite_headers: List[str], rows: List[Dict[str, Any]]) -> None:
    placeholders = ", ".join(["?" for _ in sqlite_headers])
    columns = ", ".join([f'"{h}"' for h in sqlite_headers])
    sql = f'INSERT INTO "{table_name}" ({columns}) VALUES ({placeholders});'

    values = [[row.get(header, NA_VALUE) for header in sqlite_headers] for row in rows]
    conn.executemany(sql, values)


def normalize_material_type_for_aggregation(material_type: str) -> str:
    text = material_type.strip().lower()

    if "paint shop" in text:
        return "05 Paint shop"
    if text.endswith("erection") and "paint" not in text:
        return "04 Erection"

    for match_type, pattern, output in AGGREGATION_MATERIAL_TYPE_RULES:
        if match_type == "equals" and text == pattern:
            return output
        if match_type == "startswith" and text.startswith(pattern):
            return output
        if match_type == "endswith" and text.endswith(pattern):
            return output
        if match_type == "contains" and pattern in text:
            return output
    return material_type


def to_float_or_zero(value: Any) -> float:
    if value in (None, "", NA_VALUE):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        raw = value.strip()
        if not raw or raw.upper() == NA_VALUE:
            return 0.0
        normalized = raw.replace(" ", "")
        if "," in normalized and "." not in normalized:
            normalized = normalized.replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return 0.0
    return 0.0


def int_if_whole(value: float) -> Any:
    if abs(value - round(value)) < 1e-12:
        return int(round(value))
    return value


def material_type_sort_rank(material_type: str) -> int:
    mt = material_type.strip().lower()
    if "mapress shop" in mt:
        return 0
    if "ss shop" in mt:
        return 1
    if "cs shop" in mt:
        return 2
    if "paint shop" in mt:
        return 3
    if "04 erection" in mt or ("erection" in mt and "paint" not in mt):
        return 4
    if "paint erection" in mt:
        return 5
    for idx, prefix in enumerate(AGGREGATE_MATERIAL_TYPE_SORT_PREFIXES):
        if mt.startswith(prefix):
            return idx + 6
    return len(AGGREGATE_MATERIAL_TYPE_SORT_PREFIXES) + 6


def size_sort_key(size_value: Any) -> Tuple[int, float, str]:
    text = "" if size_value is None else str(size_value).strip()
    if not text or text.upper() == NA_VALUE:
        return (1, float("inf"), text)

    normalized = text.replace(" ", "")
    if "," in normalized and "." not in normalized:
        normalized = normalized.replace(",", ".")

    try:
        return (0, float(normalized), text)
    except ValueError:
        return (1, float("inf"), text)


def create_aggregated_table(
    conn: sqlite3.Connection,
    source_table_name: str,
    sqlite_headers: List[str],
) -> str:
    aggregate_table_name = f"{source_table_name}{AGGREGATE_TABLE_SUFFIX}"

    if table_exists(conn, aggregate_table_name):
        raise ValueError(
            f"Table '{aggregate_table_name}' already exists. Import stopped to protect existing aggregation data. "
            f"{settings_sheet_hint('General')}"
        )

    cols_sql = ", ".join([f'"{h}"' for h in sqlite_headers])
    rows = conn.execute(f'SELECT {cols_sql} FROM "{source_table_name}";').fetchall()

    grouped: Dict[Tuple[str, str, str, str, str, str, str], Dict[str, Any]] = {}
    orifice_kks_by_group: Dict[Tuple[str, str, str, str, str, str, str], set[str]] = {}
    mapress_seal_qty_by_order: Dict[str, float] = {}
    for db_row in rows:
        record = dict(zip(sqlite_headers, db_row))
        type_text_raw = str(record.get("Type", NA_VALUE))
        description_raw = str(record.get("Description", NA_VALUE))

        if type_text_raw.strip().upper() == "OLET" and "branch" in description_raw.lower():
            continue

        mt_raw_l = str(record.get("Material type", NA_VALUE)).strip().lower()
        is_mapress_fkm_install = (
            "mapress" in description_raw.lower()
            and "fkm blue" in description_raw.lower()
            and "must be installed" in description_raw.lower()
        )
        if mt_raw_l == "01 mapress shop" and is_mapress_fkm_install:
            qty_base = to_float_or_zero(record.get("Qty pcs", 0.0))
            if qty_base <= 0:
                qty_base = to_float_or_zero(record.get("Qty m", 0.0))
            for order_no, mult in extract_mapress_seal_orders(description_raw):
                mapress_seal_qty_by_order[order_no] = mapress_seal_qty_by_order.get(order_no, 0.0) + (qty_base * float(mult))
            record["Description"] = rewrite_mapress_seal_description(description_raw)
        material_type_norm = normalize_material_type_for_aggregation(str(record.get("Material type", NA_VALUE)))
        mt_norm_l = material_type_norm.strip().lower()
        is_paint_shop = "paint shop" in mt_norm_l
        corrosion_category = str(record.get("Corrosion category", NA_VALUE)) if is_paint_shop else NA_VALUE
        paint_colour = str(record.get("Paint colour", NA_VALUE)) if is_paint_shop else NA_VALUE

        key = (
            str(record.get("Description", NA_VALUE)),
            str(record.get("Material", NA_VALUE)),
            str(record.get("Type", NA_VALUE)),
            str(record.get("Size", NA_VALUE)),
            material_type_norm,
            corrosion_category,
            paint_colour,
        )

        if key not in grouped:
            base = {header: NA_VALUE for header in sqlite_headers}
            base["Description"] = key[0]
            base["Material"] = key[1]
            base["Type"] = key[2]
            base["Size"] = key[3]
            base["Material type"] = key[4]
            if "Corrosion category" in base:
                base["Corrosion category"] = key[5]
            if "Paint colour" in base:
                base["Paint colour"] = key[6]
            for col in AGGREGATE_SUM_COLUMNS:
                base[col] = 0.0
            grouped[key] = base

        if material_type_norm.strip().lower() == "08 orifice plates":
            kks_text = str(record.get("KKS", NA_VALUE)).strip()
            if kks_text and kks_text.upper() != NA_VALUE:
                orifice_kks_by_group.setdefault(key, set()).add(kks_text)

        group_record = grouped[key]
        for col in AGGREGATE_SUM_COLUMNS:
            group_record[col] = to_float_or_zero(group_record.get(col, 0.0)) + to_float_or_zero(record.get(col, 0.0))

    sorted_groups = sorted(
        grouped.values(),
        key=lambda r: (
            material_type_sort_rank(str(r.get("Material type", ""))),
            str(r.get("Material", "")),
            str(r.get("Type", "")),
            size_sort_key(r.get("Size", "")),
            str(r.get("Material type", "")),
        ),
    )

    pre_uid_rows: List[Dict[str, Any]] = []
    for row in sorted_groups:
        for col in AGGREGATE_SUM_COLUMNS:
            row[col] = int_if_whole(to_float_or_zero(row[col]))

        if str(row.get("Material type", "")).strip().lower() == "08 orifice plates":
            key = (
                str(row.get("Description", NA_VALUE)),
                str(row.get("Material", NA_VALUE)),
                str(row.get("Type", NA_VALUE)),
                str(row.get("Size", NA_VALUE)),
                str(row.get("Material type", NA_VALUE)),
                str(row.get("Corrosion category", NA_VALUE)),
                str(row.get("Paint colour", NA_VALUE)),
            )
            kks_values = sorted(orifice_kks_by_group.get(key, set()))
            if "Add. Info" in row:
                row["Add. Info"] = f"KKS: {', '.join(kks_values)}" if kks_values else NA_VALUE

            # Keep non-used orifice columns as N/A in aggregated output.
            for col in ("Qty m", "Weight kg", "Surface m2", "Corrosion category", "Paint colour", "System", "Insulated"):
                if col in row:
                    row[col] = NA_VALUE

        pre_uid_rows.append(row)

    generated_seal_rows: List[Dict[str, Any]] = []
    mapress_shop_label = material_type_value("mapress", False)
    for order_no, qty_sum in sorted(mapress_seal_qty_by_order.items(), key=lambda kv: int(kv[0])):
        row = {header: NA_VALUE for header in sqlite_headers}
        row["Material type"] = mapress_shop_label
        row["Material"] = "FKM"
        row["Type"] = "SEAL"
        row["Description"] = f"Seal ring FKM blue Order no. {order_no}"
        row["Size"] = MAPRESS_SEAL_ORDER_SIZE.get(order_no, NA_VALUE)
        row["Qty pcs"] = int_if_whole(qty_sum)
        generated_seal_rows.append(row)

    if generated_seal_rows:
        insert_pos = 0
        for idx, row in enumerate(pre_uid_rows):
            if str(row.get("Material type", "")).strip().lower() == mapress_shop_label.lower():
                insert_pos = idx + 1
        pre_uid_rows[insert_pos:insert_pos] = generated_seal_rows

    uid = 1
    final_rows: List[Dict[str, Any]] = []
    for row in pre_uid_rows:
        row["UID"] = uid
        uid += 1
        final_rows.append(row)

    create_table(conn, aggregate_table_name, sqlite_headers)
    insert_rows(conn, aggregate_table_name, sqlite_headers, final_rows)
    return aggregate_table_name


def run_spare_calculation_after_import(
    db_path: Path,
    revision_table: str,
    settings_workbook: Path,
) -> Tuple[bool, str]:
    if integrated_run_spare_calcs is None:
        return (
            False,
            "Integrated spare function is unavailable. "
            + settings_sheet_hint("SpareGeneral", "SpareMainRules", "SpareErectionRules", "SpareKeywords"),
        )

    result_code = integrated_run_spare_calcs(
        db_path=db_path,
        revision_table=revision_table,
        settings_workbook=settings_workbook,
    )
    if int(result_code) != 0:
        return (
            False,
            "Spare calculation returned non-zero result. "
            + settings_sheet_hint("SpareGeneral", "SpareMainRules", "SpareErectionRules", "SpareKeywords"),
        )
    return True, "Spare calculation completed."


def run_import(
    revision_folder: Path,
    db_path: Path,
    headers_workbook: Path,
    input_subdir: str,
    settings_workbook: Path,
) -> int:
    input_dir = revision_folder / input_subdir
    if not input_dir.exists():
        logger.error("Input directory not found: %s", input_dir)
        return 1

    sqlite_headers = load_sqlite_headers(headers_workbook)
    table_name = sanitize_table_name(revision_folder.name)

    found_files, file_warnings, paint_files, orifice_files, ignored_files = locate_required_files(input_dir)

    if ignored_files:
        for path in ignored_files:
            logger.info("Ignoring non-target file: %s", path.name)

    if file_warnings:
        for warning in file_warnings:
            logger.warning("%s", warning)
        logger.error(
            "ERROR: Required input file validation failed. Import stopped. "
            f"{settings_sheet_hint('RequiredFileCategories', 'MaterialTypes', 'FileKeywords')}"
        )
        return 1

    db_path.parent.mkdir(parents=True, exist_ok=True)
    db_exists_before = db_path.exists()

    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA journal_mode=WAL;")

        if db_exists_before:
            has_data = db_has_any_rows(conn)
            logger.info("Shared database exists: %s", db_path)
            logger.info("Existing database has data: %s", "yes" if has_data else "no")
        else:
            logger.info("Shared database does not exist. It will be created: %s", db_path)

        if table_exists(conn, table_name):
            logger.warning("Table '%s' already exists. Import stopped to protect existing revision data.", table_name)
            return 1

        all_rows: List[Dict[str, Any]] = []
        uid_counter = 1

        for category in sorted(REQUIRED_FILE_CATEGORIES):
            classified = found_files[category]
            logger.info(
                "Processing source file category=%s erection=%s path=%s",
                classified.material_type,
                classified.is_erection,
                classified.path,
            )
            try:
                rows, mapping_warnings = build_rows_for_file(classified, sqlite_headers, uid_counter)
            except ValueError as exc:
                logger.error("%s", exc)
                logger.error("Import stopped due to missing required column/data. file=%s", classified.path)
                return 1
            except Exception:
                logger.exception("Unexpected failure while processing file=%s", classified.path)
                return 1

            uid_counter += len(rows)
            all_rows.extend(rows)
            for warning in mapping_warnings:
                logger.warning("%s: %s", classified.path.name, warning)

        for paint_file in sorted(paint_files):
            logger.info("Processing paint file path=%s", paint_file)
            try:
                rows, mapping_warnings = build_rows_for_paint_file(paint_file, sqlite_headers, uid_counter)
            except ValueError as exc:
                logger.error("%s", exc)
                logger.error("Import stopped due to missing required column/data. file=%s", paint_file)
                return 1
            except Exception:
                logger.exception("Unexpected failure while processing paint file=%s", paint_file)
                return 1

            uid_counter += len(rows)
            all_rows.extend(rows)
            for warning in mapping_warnings:
                logger.warning("%s: %s", paint_file.name, warning)

        for orifice_file in sorted(orifice_files):
            logger.info("Processing orifice file path=%s", orifice_file)
            try:
                rows, mapping_warnings = build_rows_for_orifice_file(orifice_file, sqlite_headers, uid_counter)
            except ValueError as exc:
                logger.error("%s", exc)
                logger.error("Import stopped due to missing required column/data. file=%s", orifice_file)
                return 1
            except Exception:
                logger.exception("Unexpected failure while processing orifice file=%s", orifice_file)
                return 1

            uid_counter += len(rows)
            all_rows.extend(rows)
            for warning in mapping_warnings:
                logger.warning("%s: %s", orifice_file.name, warning)

        create_table(conn, table_name, sqlite_headers)
        insert_rows(conn, table_name, sqlite_headers, all_rows)

        try:
            aggregate_table_name = create_aggregated_table(conn, table_name, sqlite_headers)
        except ValueError as exc:
            logger.warning("%s", exc)
            return 1

        conn.commit()

        logger.info("Imported rows=%s into table='%s' db=%s", len(all_rows), table_name, db_path)
        agg_count = conn.execute(f'SELECT COUNT(*) FROM "{aggregate_table_name}";').fetchone()[0]
        logger.info("Created aggregated table='%s' rows=%s", aggregate_table_name, agg_count)

    spare_ok, spare_msg = run_spare_calculation_after_import(
        db_path=db_path,
        revision_table=table_name,
        settings_workbook=settings_workbook,
    )
    if not spare_ok:
        logger.error("Spare calculation failed after import. %s", spare_msg)
        return 1

    logger.info("%s", spare_msg)

    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import Price Sheet revision input material files to SQLite with strict validation.",
    )
    parser.add_argument(
        "--revision-folder",
        type=Path,
        default=Path("Price_Sheet_New_Format/rev0"),
        help="Path to the revision folder (contains input subfolder).",
    )
    parser.add_argument(
        "--input-subdir",
        type=str,
        default="input",
        help="Input subfolder name under revision folder.",
    )
    parser.add_argument(
        "--db-path",
        type=Path,
        default=Path("Price_Sheet_New_Format/price_sheet.db"),
        help="Path to the shared SQLite database file.",
    )
    parser.add_argument(
        "--headers-workbook",
        type=Path,
        default=Path("Price_Sheet_New_Format/sqlite_table_headers.xlsx"),
        help="Workbook containing the mandatory SQLite table column headers in row 1.",
    )
    parser.add_argument(
        "--settings-workbook",
        type=Path,
        default=Path("Price_Sheet_New_Format/import_settings.md"),
        help="Markdown settings file containing configurable import mappings and rules.",
    )
    parser.add_argument(
        "--log-dir",
        type=Path,
        default=DEFAULT_LOG_DIR,
        help="Directory where timestamped CSV run logs are written.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    log_file = setup_csv_logging(run_name="price_sheet_import", log_dir=args.log_dir)
    logger.info("CSV logging initialized. log_file=%s", log_file)
    try:
        load_import_settings(args.settings_workbook)
    except Exception as exc:
        logger.exception(
            f"ERROR: Failed to load settings file: {exc} "
            f"{settings_sheet_hint('General', 'MaterialTypes', 'RequiredFileCategories', 'ColumnMappings_Normal', 'ColumnMappings_Erection', 'ColumnMappings_Paint', 'ColumnMappings_Orifice', 'OrificeDefaults', 'Aliases')}"
        )
        return 1

    return run_import(
        revision_folder=args.revision_folder,
        db_path=args.db_path,
        headers_workbook=args.headers_workbook,
        input_subdir=args.input_subdir,
        settings_workbook=args.settings_workbook,
    )


if __name__ == "__main__":
    raise SystemExit(main())
