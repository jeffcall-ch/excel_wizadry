from __future__ import annotations

import argparse
import logging
import re
import sqlite3
from pathlib import Path
from typing import Dict, List, Set

from logging_utils import DEFAULT_LOG_DIR, setup_csv_logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


INVALID_SHEET_CHARS = r'[\[\]\*\?/\\:]'
MAX_EXCEL_COLUMN_WIDTH = 255
NUMERIC_TYPE_TOKENS = ("INT", "REAL", "NUM", "FLOA", "DOUB", "DEC")


logger = logging.getLogger(__name__)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export all user tables from a SQLite database to a single XLSX workbook.",
    )
    parser.add_argument(
        "--db-path",
        type=Path,
        default=Path("Price_Sheet_New_Format/price_sheet.db"),
        help="Path to the source SQLite database.",
    )
    parser.add_argument(
        "--output-path",
        type=Path,
        default=Path("Price_Sheet_New_Format/price_sheet_export.xlsx"),
        help="Path to the output XLSX workbook.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite output workbook if it already exists.",
    )
    parser.add_argument(
        "--log-dir",
        type=Path,
        default=DEFAULT_LOG_DIR,
        help="Directory where timestamped CSV run logs are written.",
    )
    return parser.parse_args()


def get_user_table_names(conn: sqlite3.Connection) -> List[str]:
    rows = conn.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table'
          AND name NOT LIKE 'sqlite_%'
        ORDER BY name;
        """
    ).fetchall()
    return [row[0] for row in rows]


def make_valid_sheet_name(name: str, used: Dict[str, int]) -> str:
    sanitized = re.sub(INVALID_SHEET_CHARS, "_", name).strip()
    sanitized = sanitized or "Sheet"

    # Excel sheet names max length is 31.
    base = sanitized[:31]

    if base not in used:
        used[base] = 1
        return base

    counter = used[base]
    while True:
        suffix = f"_{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in used:
            used[base] += 1
            used[candidate] = 1
            return candidate
        counter += 1


def update_column_width(max_widths: Dict[int, int], col_idx: int, value: object) -> None:
    text = "" if value is None else str(value)
    width = len(text)
    if width > max_widths.get(col_idx, 0):
        max_widths[col_idx] = width


def apply_sheet_formatting(ws, max_widths: Dict[int, int]) -> None:
    if ws.max_column > 0 and ws.max_row > 0:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    ws.freeze_panes = "A2"

    for col_idx in range(1, ws.max_column + 1):
        max_len = max_widths.get(col_idx, 0)
        # Add a small padding so content is not clipped.
        width = min(max_len + 2, MAX_EXCEL_COLUMN_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(8, width)


def parse_numeric_text(value: str):
    text = value.strip()
    if not text:
        return value

    # Preserve explicit placeholders.
    if text.upper() == "N/A":
        return value

    normalized = text.replace(" ", "")
    # Support decimal comma if dot is not present.
    if "," in normalized and "." not in normalized:
        normalized = normalized.replace(",", ".")

    if re.fullmatch(r"[+-]?\d+", normalized):
        try:
            return int(normalized)
        except ValueError:
            return value

    if re.fullmatch(r"[+-]?(?:\d+\.\d*|\d*\.\d+|\d+)(?:[eE][+-]?\d+)?", normalized):
        try:
            return float(normalized)
        except ValueError:
            return value

    return value


def get_numeric_headers_for_table(conn: sqlite3.Connection, table_name: str) -> Set[str]:
    rows = conn.execute(f'PRAGMA table_info("{table_name}");').fetchall()
    numeric_headers: Set[str] = set()
    for row in rows:
        # PRAGMA table_info columns: cid, name, type, notnull, dflt_value, pk
        column_name = str(row[1]) if len(row) > 1 else ""
        declared_type = str(row[2]).upper() if len(row) > 2 and row[2] is not None else ""
        if column_name and any(token in declared_type for token in NUMERIC_TYPE_TOKENS):
            numeric_headers.add(column_name)
    return numeric_headers


def coerce_cell_for_excel(header: str, value, numeric_headers: Set[str]):
    if value is None:
        return None

    if header in numeric_headers:
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            return parse_numeric_text(value)

    return value


def export_sqlite_to_xlsm(db_path: Path, output_path: Path, overwrite: bool) -> int:
    if not db_path.exists():
        logger.error("Database file not found: %s", db_path)
        return 1

    if output_path.suffix.lower() != ".xlsx":
        logger.error("Output file must use .xlsx extension.")
        return 1

    if output_path.exists() and not overwrite:
        logger.error("Output already exists: %s", output_path)
        logger.error("Use --overwrite to replace it.")
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(db_path) as conn:
        table_names = get_user_table_names(conn)
        if not table_names:
            logger.error("No user tables found in the database.")
            return 1

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        used_sheet_names: Dict[str, int] = {}

        for table_name in table_names:
            sheet_name = make_valid_sheet_name(table_name, used_sheet_names)
            ws = wb.create_sheet(title=sheet_name)
            max_widths: Dict[int, int] = {}
            numeric_headers = get_numeric_headers_for_table(conn, table_name)

            header_cur = conn.execute(f'SELECT * FROM "{table_name}" LIMIT 0;')
            headers = [col[0] for col in header_cur.description] if header_cur.description else []
            if "UID" in headers:
                cur = conn.execute(f'SELECT * FROM "{table_name}" ORDER BY CAST("UID" AS NUMERIC), rowid;')
            else:
                cur = conn.execute(f'SELECT * FROM "{table_name}";')

            if headers:
                ws.append(headers)
                for col_idx, value in enumerate(headers, start=1):
                    update_column_width(max_widths, col_idx, value)

            row_count = 0
            for row in cur:
                converted_row = [
                    coerce_cell_for_excel(headers[col_idx], value, numeric_headers)
                    for col_idx, value in enumerate(row)
                ]
                ws.append(converted_row)
                for col_idx, value in enumerate(converted_row, start=1):
                    update_column_width(max_widths, col_idx, value)
                row_count += 1

            apply_sheet_formatting(ws, max_widths)

            logger.info("Exported table '%s' to sheet '%s' (%s rows).", table_name, sheet_name, row_count)

        wb.save(output_path)

    logger.info("Export completed to %s", output_path)
    return 0


def main() -> int:
    args = parse_args()
    log_file = setup_csv_logging(run_name="sqlite_export", log_dir=args.log_dir)
    logger.info("CSV logging initialized. log_file=%s", log_file)
    return export_sqlite_to_xlsm(
        db_path=args.db_path,
        output_path=args.output_path,
        overwrite=args.overwrite,
    )


if __name__ == "__main__":
    raise SystemExit(main())
