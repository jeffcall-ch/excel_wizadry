#!/usr/bin/env python3
"""
Price Sheet Template Generator

Creates an Excel workbook template (.xlsx) that transforms BoM input data
into a formatted price sheet using M365 dynamic array formulas.

Workflow:
  1. Run this script to generate the template
  2. Open the .xlsx in Excel 365
  3. Paste input data into the IN_* sheets
  4. The output sheets auto-calculate via formulas

No VBA, no Power Query — pure Excel formulas (requires Microsoft 365).
"""

import math
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_DATA_ROWS = 500          # pre-fill formulas for up to this many data rows
GRAND_TOTAL_ROW = 510        # fixed row for Grand Total on output sheets

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CALC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SUPPLIER_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

BOLD = Font(bold=True)
BOLD_12 = Font(bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helper — write a row of headers with formatting
# ---------------------------------------------------------------------------
def write_headers(ws, row, headers, fill=HEADER_FILL, start_col=1):
    for j, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=j, value=h)
        c.font = BOLD
        c.fill = fill
        c.alignment = WRAP
        c.border = THIN_BORDER


def set_col_widths(ws, widths: dict):
    """widths = {1: 20, 2: 10, ...}  (col_number → width)."""
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ========================================================================
# INPUT SHEETS
# ========================================================================
def create_input_sheet_material(wb, name, title, is_erection=False):
    """Create an input sheet for a material BoM 'Total Piping Material' tab."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = "FFC000"  # amber tab

    ws.cell(row=1, column=1, value=f"PASTE '{title}' Total Piping Material DATA BELOW (starting row 3)")
    ws.cell(row=1, column=1).font = BOLD_12

    if is_erection:
        headers = ["Type", "Pipe Component", "DN", "Material",
                   "QTY [pcs./m]", "QTY prev. rev.", "Difference"]
        widths = {1: 10, 2: 50, 3: 10, 4: 30, 5: 14, 6: 16, 7: 12}
    else:
        headers = ["Type", "Pipe Component", "DN", "DN2", "Material",
                   "Weight [kg]", "Total weight [kg]", "QTY [pcs.]",
                   "QTY [m]", "QTY prev. rev.", "Difference"]
        widths = {1: 10, 2: 50, 3: 8, 4: 8, 5: 30, 6: 12, 7: 14,
                  8: 12, 9: 12, 10: 16, 11: 12}

    write_headers(ws, 2, headers)
    set_col_widths(ws, widths)
    # Light input fill for the data area hint
    for r in range(3, 8):
        for j in range(1, len(headers) + 1):
            ws.cell(row=r, column=j).fill = INPUT_FILL
    return ws


def create_input_sheet_painting(wb):
    """Create the painting BoM input sheet."""
    ws = wb.create_sheet("IN_Paint")
    ws.sheet_properties.tabColor = "FFC000"

    ws.cell(row=1, column=1,
            value="PASTE 'PIPING PAINTING MATERIAL' DATA BELOW (starting row 3)")
    ws.cell(row=1, column=1).font = BOLD_12

    headers = [
        "Name of System", "Name of Pipe", "Type",
        "Description", "Medium", "DN 1", "DN 2", "Material",
        "Pipe Class", "Building Section", "AIC",
        "Quantity [pcs]", "Quantity [m]", "External Surface [m2]",
        "Corrosion class", "Painting colour (acc. BS4800)", "Insulated",
    ]
    write_headers(ws, 2, headers)
    set_col_widths(ws, {1: 28, 2: 18, 3: 8, 4: 50,
                        5: 10, 6: 8, 7: 8, 8: 28, 9: 10,
                        10: 16, 11: 10, 12: 12, 13: 12, 14: 16,
                        15: 12, 16: 20, 17: 10})
    return ws


def create_input_sheet_kks(wb, name, title):
    """Create a KKS detail input sheet (for flange guard extraction)."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = "FFC000"

    ws.cell(row=1, column=1,
            value=f"PASTE '{title}' KKS Piping Material DATA BELOW (starting row 3)")
    ws.cell(row=1, column=1).font = BOLD_12

    headers = [
        "System", "Pipe", "Type", "Pipe Component", "DN", "DN2",
        "Material", "Pipe Class", "Building section", "AIC",
        "Weight [kg]", "Total weight [kg]", "QTY [pcs.]", "QTY [m]",
    ]
    write_headers(ws, 2, headers)
    set_col_widths(ws, {1: 28, 2: 18, 3: 8, 4: 50, 5: 8, 6: 8,
                        7: 28, 8: 10, 9: 16, 10: 10,
                        11: 12, 12: 14, 13: 12, 14: 12})
    return ws


# ========================================================================
# CONFIG SHEET
# ========================================================================
def create_config_sheet(wb):
    ws = wb.create_sheet("CFG")
    ws.sheet_properties.tabColor = "92D050"

    # --- Blind Disk DN → Thickness ---
    ws.cell(row=1, column=1, value="Blind Disk DN → S1 Thickness").font = BOLD_12
    write_headers(ws, 2, ["DN", "S1 Thickness (mm)"])
    blind_disk_data = [
        (15, 6), (20, 6), (25, 6), (32, 8), (40, 8),
        (50, 10), (65, 10), (80, 12), (100, 12), (125, 12),
        (150, 12), (200, 12), (250, 16), (300, 16),
    ]
    for i, (dn, s1) in enumerate(blind_disk_data, start=3):
        ws.cell(row=i, column=1, value=dn)
        ws.cell(row=i, column=2, value=s1)

    # --- Flange Guard Systems ---
    ws.cell(row=1, column=4, value="Flange Guard Systems (edit as needed)").font = BOLD_12
    write_headers(ws, 2, ["System Name keyword"], start_col=4)
    fg_systems = ["Fuel Oil", "Ammonia", "Urea"]
    for i, s in enumerate(fg_systems, start=3):
        ws.cell(row=i, column=4, value=s)

    # --- Spare Rules Reference ---
    ws.cell(row=1, column=7, value="Spare Rules — Main Material").font = BOLD_12
    rules_text = [
        "Priority: SEAL > CRITICAL > TUBI > STANDARD",
        "",
        "SEAL (desc has ALL of: seal,ring,fkm; NOT: elbow,flange,tee,reducer)",
        "  QTY ≤ 10 → Spare = 5",
        "  QTY > 10 → Spare = MAX(5, CEILING(QTY×10%))",
        "",
        "CRITICAL (desc has ANY of: nipple,union,connection,welding,adaptor,coupling,threaded)",
        "  QTY ≤ 10 → Spare = MAX(3, CEILING(QTY×50%))",
        "  QTY > 10 → Spare = MAX(3, CEILING(QTY×10%))",
        "",
        "TUBI (Type = TUBI)",
        "  DN < 50: base = CEILING(QTY_m × 15%); total = CEILING_TO_6(QTY_m + base)",
        "  DN ≥ 50: base = CEILING(QTY_m × 10%); total = CEILING_TO_6(QTY_m + base)",
        "  Spare = total − QTY_m",
        "",
        "STANDARD (pcs)",
        "  QTY ≤ 5 → 0",
        "  6–20 → 1",
        "  >20 & DN<50 → CEILING(QTY×7.5%)",
        "  >20 & DN≥50 → CEILING(QTY×5%)",
        "",
        "STANDARD (m, non-TUBI)",
        "  DN<50 → CEILING(QTY_m×15%)",
        "  DN≥50 → CEILING(QTY_m×10%)",
    ]
    for i, t in enumerate(rules_text, start=2):
        ws.cell(row=i, column=7, value=t)

    # --- Erection Spare Rules ---
    r = 30
    ws.cell(row=r, column=7, value="Spare Rules — Erection Material").font = BOLD_12
    erec_rules = [
        "Priority: NUT/WASHER > BOLT > GASKET > STANDARD",
        "",
        "NUT/WASHER (desc has: nut or washer)",
        "  QTY ≤ 30 → base=100%; total rounded (≤100→next10, 101-500→next50, >500→next100)",
        "  31–100 → base=50%; same rounding",
        "  101–400 → base=35%; same rounding",
        "  >400 → base=20%; same rounding",
        "",
        "BOLT (desc has: bolt) — same % tiers, but always round to next 10",
        "",
        "GASKET (desc has: gasket)",
        "  QTY ≤ 30 → CEILING(QTY×100%)",
        "  31–500 → base=50%, round total to next 10",
        "  >500 → base=30%, round total to next 10",
        "",
        "STANDARD (all others)",
        "  QTY < 400 → CEILING(QTY×35%)",
        "  QTY ≥ 400 → CEILING(QTY×20%)",
    ]
    for i, t in enumerate(erec_rules, start=r + 1):
        ws.cell(row=i, column=7, value=t)

    set_col_widths(ws, {1: 10, 2: 18, 3: 4, 4: 28, 5: 4, 6: 4, 7: 80})
    return ws


# ========================================================================
# CALCULATION SHEET — Painting Pivot
# ========================================================================
def create_calc_painting(wb):
    """Hidden calculation sheet for painting pivot + CS TUBI feedback."""
    ws = wb.create_sheet("CALC_Paint")
    ws.sheet_properties.tabColor = "A9D18E"

    # --- Section A: Painting TUBI pivot (by Description + Color) ---
    ws.cell(row=1, column=1, value="Painting TUBI Pivot (auto-calculated)").font = BOLD_12
    write_headers(ws, 2, [
        "Description", "Painting Colour", "Base Length [m]",
        "Spare %", "Length + Spare [m]", "Rounded to 6m",
        "Surface per m [m2/m]", "Surface [m2]",
    ])

    # A3: UNIQUE pairs of (Description, Colour) from TUBI rows in painting input
    # Using SORT(UNIQUE(FILTER(...)))
    paint_ref = "IN_Paint!A$3:R$5000"
    # Filter TUBI rows, pick cols 5 (Description) and 17 (Colour)
    desc_col = f"INDEX({paint_ref},,3)"   # Type column (D)
    filt_cond = f'{desc_col}="TUBI"'

    # Dynamic array: unique (Description, Colour) pairs for TUBI
    unique_formula = (
        '=IFERROR(_xlfn.SORT(_xlfn.UNIQUE(_xlfn._xlpm.FILTER('
        'CHOOSE({1,2},INDEX(IN_Paint!$A$3:$R$5000,,3),INDEX(IN_Paint!$A$3:$R$5000,,16)),'
        'INDEX(IN_Paint!$A$3:$R$5000,,3)="TUBI"'
        ')),{1,2},{1,1}),"")'
    )
    ws.cell(row=3, column=1).value = unique_formula

    # C3: SUMIFS for base length per (Description, Colour)
    for r in range(3, 3 + MAX_DATA_ROWS):
        row_s = str(r)
        # Base length = SUMIFS(qty_m, type="TUBI", description=A, colour=B)
        ws.cell(row=r, column=3).value = (
            f'=IF(A{row_s}="","",SUMPRODUCT('
            f'(INDEX(IN_Paint!$A$3:$R$5000,,3)="TUBI")*'
            f'(INDEX(IN_Paint!$A$3:$R$5000,,3)=A{row_s})*'
            f'(INDEX(IN_Paint!$A$3:$R$5000,,16)=B{row_s})*'
            f'VALUE(INDEX(IN_Paint!$A$3:$R$5000,,13))))'
        )
        # D: Spare %  — extract DN from description for threshold
        # Pipe description like "Tube 33.7x2.6 EN10216-2" → OD is second token
        # DN < 50mm OD → 15%, else 10%.  OD 33.7 → DN25 → <50.  OD 60.3 → DN50 → ≥50
        # Simpler: if OD<60 then 15% else 10%  (OD<60 covers DN15-DN40)
        ws.cell(row=r, column=4).value = (
            f'=IF(A{row_s}="","",IF(IFERROR(VALUE(MID(A{row_s},'
            f'FIND(" ",A{row_s})+1,FIND("x",A{row_s})-FIND(" ",A{row_s})-1)),999)<60,0.15,0.1))'
        )
        # E: Length with spare = C * (1 + D)
        ws.cell(row=r, column=5).value = (
            f'=IF(A{row_s}="","",C{row_s}*(1+D{row_s}))'
        )
        # F: Rounded to 6m = CEILING(E, 6)
        ws.cell(row=r, column=6).value = (
            f'=IF(A{row_s}="","",_xlfn.CEILING.MATH(E{row_s},6))'
        )
        # G: Surface per meter = SUMPRODUCT of surface / SUMPRODUCT of length for matching rows
        ws.cell(row=r, column=7).value = (
            f'=IF(A{row_s}="","",IFERROR(SUMPRODUCT('
            f'(INDEX(IN_Paint!$A$3:$R$5000,,3)="TUBI")*'
            f'(INDEX(IN_Paint!$A$3:$R$5000,,3)=A{row_s})*'
            f'(INDEX(IN_Paint!$A$3:$R$5000,,16)=B{row_s})*'
            f'VALUE(INDEX(IN_Paint!$A$3:$R$5000,,13)))/C{row_s},0))'
        )
        # H: Surface = F * G (order length × surface rate)
        ws.cell(row=r, column=8).value = (
            f'=IF(A{row_s}="","",F{row_s}*G{row_s})'
        )

    set_col_widths(ws, {1: 35, 2: 18, 3: 14, 4: 10, 5: 16, 6: 14, 7: 14, 8: 14})

    # --- Section B: CS TUBI order qty (sum of painting rounded by description) ---
    # Column K onwards
    ws.cell(row=1, column=11, value="CS TUBI For-Order (from painting)").font = BOLD_12
    write_headers(ws, 2, ["Tube Description", "For Order [m]"], start_col=11)

    # K3: Unique tube descriptions
    ws.cell(row=3, column=11).value = (
        '=IFERROR(_xlfn.SORT(_xlfn.UNIQUE(_xlfn._xlpm.FILTER('
        'INDEX(IN_Paint!$A$3:$R$5000,,3),'
        'INDEX(IN_Paint!$A$3:$R$5000,,3)="TUBI"'
        '))),"")'
    )
    # L3: Sum of rounded-to-6 by description
    for r in range(3, 3 + MAX_DATA_ROWS):
        row_s = str(r)
        ws.cell(row=r, column=12).value = (
            f'=IF(K{row_s}="","",SUMIFS(F$3:F${3+MAX_DATA_ROWS},A$3:A${3+MAX_DATA_ROWS},K{row_s}))'
        )

    set_col_widths(ws, {11: 35, 12: 14})

    # --- Section C: Painting FITTINGS pivot (non-TUBI by Description + Color) ---
    ws.cell(row=1, column=15, value="Painting FITTINGS surface (non-TUBI)").font = BOLD_12
    write_headers(ws, 2, ["Painting Colour", "Total Surface [m2]"], start_col=15)
    # O3: Unique colours for non-TUBI
    ws.cell(row=3, column=15).value = (
        '=IFERROR(_xlfn.SORT(_xlfn.UNIQUE(_xlfn._xlpm.FILTER('
        'INDEX(IN_Paint!$A$3:$R$5000,,16),'
        'INDEX(IN_Paint!$A$3:$R$5000,,3)<>"TUBI"'
        '))),"")'
    )
    # P3: Sum surface per colour
    for r in range(3, 3 + 50):
        row_s = str(r)
        ws.cell(row=r, column=16).value = (
            f'=IF(O{row_s}="","",SUMPRODUCT('
            f'(INDEX(IN_Paint!$A$3:$R$5000,,3)<>"TUBI")*'
            f'(INDEX(IN_Paint!$A$3:$R$5000,,16)=O{row_s})*'
            f'VALUE(INDEX(IN_Paint!$A$3:$R$5000,,13))))'
        )

    # --- Section D: Painting TUBI surface by colour only (for cans calc) ---
    ws.cell(row=1, column=19, value="TUBI surface by colour").font = BOLD_12
    write_headers(ws, 2, ["Painting Colour", "Total TUBI Surface [m2]"], start_col=19)
    ws.cell(row=3, column=19).value = (
        '=IFERROR(_xlfn.SORT(_xlfn.UNIQUE(_xlfn._xlpm.FILTER('
        'INDEX(IN_Paint!$A$3:$R$5000,,16),'
        'INDEX(IN_Paint!$A$3:$R$5000,,3)="TUBI"'
        '))),"")'
    )
    for r in range(3, 3 + 50):
        row_s = str(r)
        ws.cell(row=r, column=20).value = (
            f'=IF(S{row_s}="","",SUMPRODUCT('
            f'(INDEX(IN_Paint!$A$3:$R$5000,,3)="TUBI")*'
            f'(INDEX(IN_Paint!$A$3:$R$5000,,16)=S{row_s})*'
            f'VALUE(INDEX(IN_Paint!$A$3:$R$5000,,13))))'
        )

    set_col_widths(ws, {15: 18, 16: 16, 19: 18, 20: 20})
    return ws


# ========================================================================
# SPARE FORMULAS
# ========================================================================
def main_material_spare_formula(row, is_cs=False):
    """
    Return the spare formula for main material sheets (1. Mapress, 2. SS, 3. CS).
    Layout: A=MATERIAL, B=TYPE, C=DN, D=DESCRIPTION, E=QTY_pcs, F=QTY_m, G=WEIGHT
    Spare goes in H.
    For CS TUBI rows, spare is overridden by painting calculation.
    """
    r = str(row)

    # Component classification checks (on DESCRIPTION = column D)
    is_seal = (
        f'AND(ISNUMBER(SEARCH("seal",D{r})),ISNUMBER(SEARCH("ring",D{r})),'
        f'ISNUMBER(SEARCH("fkm",D{r})),'
        f'ISERROR(SEARCH("elbow",D{r})),ISERROR(SEARCH("flange",D{r})),'
        f'ISERROR(SEARCH("tee",D{r})),ISERROR(SEARCH("reducer",D{r})))'
    )
    is_critical = (
        f'OR(ISNUMBER(SEARCH("nipple",D{r})),ISNUMBER(SEARCH("union",D{r})),'
        f'ISNUMBER(SEARCH("connection",D{r})),ISNUMBER(SEARCH("welding",D{r})),'
        f'ISNUMBER(SEARCH("adaptor",D{r})),ISNUMBER(SEARCH("coupling",D{r})),'
        f'ISNUMBER(SEARCH("threaded",D{r})))'
    )
    is_tubi = f'B{r}="TUBI"'

    # Spare sub-formulas
    seal_spare = (
        f'IF(E{r}<=10,5,MAX(5,_xlfn.CEILING.MATH(E{r}*0.1,1)))'
    )
    critical_spare = (
        f'IF(E{r}<=10,MAX(3,_xlfn.CEILING.MATH(E{r}*0.5,1)),'
        f'MAX(3,_xlfn.CEILING.MATH(E{r}*0.1,1)))'
    )

    if is_cs:
        # CS TUBI: spare = painting for-order − base qty
        tubi_spare = (
            f'MAX(0,IFERROR(INDEX(CALC_Paint!$L$3:$L${3+MAX_DATA_ROWS},'
            f'MATCH(D{r},CALC_Paint!$K$3:$K${3+MAX_DATA_ROWS},0)),0)-F{r})'
        )
    else:
        # Non-CS TUBI: normal spare + round to 6
        tubi_spare = (
            f'MAX(0,_xlfn.CEILING.MATH(F{r}+_xlfn.CEILING.MATH('
            f'F{r}*IF(C{r}<50,0.15,0.1),1),6)-F{r})'
        )

    std_pcs_spare = (
        f'_xlfn.IFS(E{r}<=5,0,E{r}<=20,1,'
        f'C{r}<50,_xlfn.CEILING.MATH(E{r}*0.075,1),'
        f'TRUE,_xlfn.CEILING.MATH(E{r}*0.05,1))'
    )
    std_m_spare = (
        f'IF(C{r}<50,_xlfn.CEILING.MATH(F{r}*0.15,1),'
        f'_xlfn.CEILING.MATH(F{r}*0.1,1))'
    )

    # Main decision tree
    formula = (
        f'=IF(A{r}="","",'
        f'IF({is_seal},{seal_spare},'
        f'IF({is_critical},{critical_spare},'
        f'IF({is_tubi},{tubi_spare},'
        f'IF(E{r}>0,{std_pcs_spare},'
        f'IF(F{r}>0,{std_m_spare},0))))))'
    )
    return formula


def erection_spare_formula(row):
    """
    Erection material spare formula.
    Layout: A=TYPE, B=DN, C=DESCRIPTION, D=MATERIAL, E=QTY, Spare in F.
    """
    r = str(row)

    is_nut_washer = f'OR(ISNUMBER(SEARCH("nut",C{r})),ISNUMBER(SEARCH("washer",C{r})))'
    is_bolt = f'ISNUMBER(SEARCH("bolt",C{r}))'
    is_gasket = f'ISNUMBER(SEARCH("gasket",C{r}))'

    # NutWasher rounding: ≤100→ceil10, 101-500→ceil50, >500→ceil100
    def nw_round(total_expr):
        return (
            f'IF({total_expr}<=100,_xlfn.CEILING.MATH({total_expr},10),'
            f'IF({total_expr}<=500,_xlfn.CEILING.MATH({total_expr},50),'
            f'_xlfn.CEILING.MATH({total_expr},100)))'
        )

    def nw_spare(pct):
        base = f'_xlfn.CEILING.MATH(E{r}*{pct},1)'
        total = f'E{r}+{base}'
        return f'{nw_round(total)}-E{r}'

    nw_formula = (
        f'_xlfn.IFS(E{r}<=30,{nw_spare(1)},'
        f'E{r}<=100,{nw_spare(0.5)},'
        f'E{r}<=400,{nw_spare(0.35)},'
        f'TRUE,{nw_spare(0.2)})'
    )

    def bolt_spare(pct):
        base = f'_xlfn.CEILING.MATH(E{r}*{pct},1)'
        return f'_xlfn.CEILING.MATH(E{r}+{base},10)-E{r}'

    bolt_formula = (
        f'_xlfn.IFS(E{r}<=30,{bolt_spare(1)},'
        f'E{r}<=100,{bolt_spare(0.5)},'
        f'E{r}<=400,{bolt_spare(0.35)},'
        f'TRUE,{bolt_spare(0.2)})'
    )

    gasket_formula = (
        f'_xlfn.IFS(E{r}<=30,_xlfn.CEILING.MATH(E{r}*1,1),'
        f'E{r}<=500,_xlfn.CEILING.MATH(E{r}+_xlfn.CEILING.MATH(E{r}*0.5,1),10)-E{r},'
        f'TRUE,_xlfn.CEILING.MATH(E{r}+_xlfn.CEILING.MATH(E{r}*0.3,1),10)-E{r})'
    )

    std_formula = (
        f'IF(E{r}<400,_xlfn.CEILING.MATH(E{r}*0.35,1),'
        f'_xlfn.CEILING.MATH(E{r}*0.2,1))'
    )

    return (
        f'=IF(A{r}="","",'
        f'IF({is_nut_washer},{nw_formula},'
        f'IF({is_bolt},{bolt_formula},'
        f'IF({is_gasket},{gasket_formula},'
        f'{std_formula}))))'
    )


# ========================================================================
# OUTPUT SHEETS — Material (1. Mapress, 2. SS, 3. CS)
# ========================================================================
def create_output_material(wb, sheet_name, input_sheet, section_num, is_cs=False):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "4472C4"

    # --- Rows 1–5: Spare rules header ---
    ws.cell(row=1, column=2, value="Supplier to fill in").font = BOLD
    ws.cell(row=1, column=2).fill = SUPPLIER_FILL
    ws.cell(row=1, column=5, value="Spare rules: ").font = BOLD
    ws.cell(row=1, column=7, value="QTY <=5 add 0%")
    ws.cell(row=1, column=9, value="Pipe <DN50 add 15%")
    ws.cell(row=2, column=3, value="Supplier:")
    ws.cell(row=2, column=3).fill = SUPPLIER_FILL
    ws.cell(row=2, column=7, value="QTY >5 to 20 add 1")
    ws.cell(row=2, column=9, value="Pipe >=DN50 add 10%")
    ws.cell(row=2, column=11, value="Currency")
    ws.cell(row=2, column=11).fill = SUPPLIER_FILL
    ws.cell(row=3, column=7, value=" <DN50 and QTY >20 add 7.5%")
    ws.cell(row=4, column=7, value=" >=DN50 and QTY >20 add 5% ")

    # Row 6: Section title
    ws.cell(row=6, column=1, value=f"{section_num}. {sheet_name.split('. ',1)[-1]}").font = BOLD
    ws.cell(row=6, column=5, value="As per BOM Rev.")
    ws.cell(row=6, column=8, value="Spare").font = BOLD
    ws.cell(row=6, column=9, value="For Order").font = BOLD
    ws.cell(row=6, column=11, value="Material Supply").font = BOLD

    # Row 7: Column headers
    headers = [
        "MATERIAL", "TYPE", "DN\n(mm)", "DESCRIPTION/ Pipe Component",
        "QUANTITY\n (pcs.)", "QUANTITY\n (m)", "WEIGHT\n [kg]",
        "Spare", "QUANTITY\n (pcs./m)", "WEIGHT\n [kg]",
        "Unit  Price  (pcs./m)", "Total Price",
    ]
    write_headers(ws, 7, headers)

    # Row 8: Dynamic array formula — sorted + reordered from input
    # Input layout: A=Type, B=PipeComp, C=DN, D=DN2, E=Material,
    #               F=Weight, G=TotalWeight, H=QTY_pcs, I=QTY_m
    inp = input_sheet
    data_range = f"{inp}!$A$3:$K$5000"
    spill_formula = (
        f'=IFERROR(_xlfn.SORT(_xlfn._xlpm.FILTER('
        f'CHOOSE({{1,2,3,4,5,6,7}},'
        f'INDEX({data_range},,3),'   # Material → A
        f'INDEX({data_range},,1),'   # Type → B
        f'INDEX({data_range},,3),'   # DN → C
        f'INDEX({data_range},,2),'   # Pipe Component → D
        f'INDEX({data_range},,8),'   # QTY pcs → E
        f'INDEX({data_range},,9),'   # QTY m → F
        f'INDEX({data_range},,7)'    # Total weight → G
        f'),INDEX({data_range},,1)<>""),'
        f'{{1,2,3,4}},{{1,1,1,1}}),"")'
    )
    ws.cell(row=8, column=1).value = spill_formula

    # Columns H–L: Row formulas (pre-filled for MAX_DATA_ROWS rows)
    for r in range(8, 8 + MAX_DATA_ROWS):
        rs = str(r)
        # H: Spare
        ws.cell(row=r, column=8).value = main_material_spare_formula(r, is_cs=is_cs)
        ws.cell(row=r, column=8).fill = CALC_FILL

        if is_cs:
            # I: For Order — TUBI uses painting total, others = QTY + spare
            ws.cell(row=r, column=9).value = (
                f'=IF(A{rs}="","",IF(B{rs}="TUBI",'
                f'IFERROR(INDEX(CALC_Paint!$L$3:$L${3+MAX_DATA_ROWS},'
                f'MATCH(D{rs},CALC_Paint!$K$3:$K${3+MAX_DATA_ROWS},0)),F{rs}+H{rs}),'
                f'IF(E{rs}>0,E{rs}+H{rs},IF(F{rs}>0,F{rs}+H{rs},0))))'
            )
        else:
            # I: For Order = QTY + Spare (pcs or m)
            ws.cell(row=r, column=9).value = (
                f'=IF(A{rs}="","",IF(E{rs}>0,E{rs}+H{rs},IF(F{rs}>0,F{rs}+H{rs},0)))'
            )

        # J: For Order Weight = weight × (for_order / max(qty_pcs, qty_m))
        ws.cell(row=r, column=10).value = (
            f'=IF(A{rs}="","",IFERROR(G{rs}*I{rs}/IF(E{rs}>0,E{rs},F{rs}),0))'
        )

        # K: Unit Price — blank for supplier
        ws.cell(row=r, column=11).fill = SUPPLIER_FILL

        # L: Total Price = For Order × Unit Price
        ws.cell(row=r, column=12).value = (
            f'=IF(A{rs}="","",IFERROR(I{rs}*K{rs},0))'
        )

    # Grand Total row
    gt = GRAND_TOTAL_ROW
    ws.cell(row=gt, column=1, value=f"GRAND TOTAL PRICE {sheet_name.upper()}").font = BOLD
    for col in [5, 6, 7, 8, 10]:
        ws.cell(row=gt, column=col).value = f"=SUM({get_column_letter(col)}8:{get_column_letter(col)}{8+MAX_DATA_ROWS-1})"
        ws.cell(row=gt, column=col).font = BOLD
    ws.cell(row=gt, column=12).value = f"=SUM(L8:L{8+MAX_DATA_ROWS-1})"
    ws.cell(row=gt, column=12).font = BOLD

    # Column widths
    set_col_widths(ws, {1: 28, 2: 8, 3: 8, 4: 52, 5: 12, 6: 12, 7: 12,
                        8: 10, 9: 14, 10: 12, 11: 14, 12: 14})

    # Number formats
    for r in range(8, 8 + MAX_DATA_ROWS + 1):
        for c in [5, 6, 7, 8, 9, 10]:
            ws.cell(row=r, column=c).number_format = '#,##0.00'
        for c in [11, 12]:
            ws.cell(row=r, column=c).number_format = '#,##0.00'

    return ws


# ========================================================================
# OUTPUT SHEET — 4. Erection Material
# ========================================================================
def create_output_erection(wb):
    ws = wb.create_sheet("4. Erection Material")
    ws.sheet_properties.tabColor = "4472C4"

    # Header area
    ws.cell(row=1, column=2, value="Supplier to fill in").font = BOLD
    ws.cell(row=1, column=2).fill = SUPPLIER_FILL
    ws.cell(row=1, column=4, value="Spare rules:").font = BOLD
    ws.cell(row=1, column=5, value="QTY <400 add 30-40%")
    ws.cell(row=2, column=2, value="Supplier:").fill = SUPPLIER_FILL
    ws.cell(row=2, column=5, value="QTY >=400 add more than 20%")
    ws.cell(row=2, column=8, value="Currency").fill = SUPPLIER_FILL
    ws.cell(row=3, column=5, value="Gaskets add 40%-60% for Hydrotest issues")

    ws.cell(row=5, column=1, value="4. Erection Material").font = BOLD
    ws.cell(row=5, column=5, value="As per\n BOM rev")
    ws.cell(row=5, column=6, value="Spare").font = BOLD
    ws.cell(row=5, column=7, value="For Order").font = BOLD
    ws.cell(row=5, column=8, value="Material Supply").font = BOLD

    headers = [
        "TYPE", "DN\n(mm)", "DESCRIPTION/ Pipe Component", "MATERIAL",
        "QUANTITY\n (pcs./m)", "Spare", "QUANTITY\n (pcs./m)",
        "Unit  Price  (pcs./m)", "Total \nPrice",
    ]
    write_headers(ws, 6, headers)

    # Row 7: Dynamic array — VSTACK + SORT the 3 erection inputs
    # Erection input: A=Type, B=PipeComp, C=DN, D=Material, E=QTY
    # Output: A=Type, B=DN, C=PipeComp, D=Material, E=QTY
    def erec_block(inp):
        dr = f"{inp}!$A$3:$G$5000"
        return (
            f'CHOOSE({{1,2,3,4,5}},'
            f'INDEX({dr},,1),'   # Type
            f'INDEX({dr},,3),'   # DN
            f'INDEX({dr},,2),'   # Pipe Component
            f'INDEX({dr},,3),'   # Material
            f'INDEX({dr},,3))'   # QTY
        )

    spill_formula = (
        f'=IFERROR(_xlfn.SORT(_xlfn._xlpm.FILTER('
        f'_xlfn.VSTACK({erec_block("IN_Erect_SS")},{erec_block("IN_Erect_CS")},{erec_block("IN_Erect_MAP")}),'
        f'_xlfn.VSTACK('
        f'INDEX(IN_Erect_SS!$A$3:$A$5000,,1),'
        f'INDEX(IN_Erect_CS!$A$3:$A$5000,,1),'
        f'INDEX(IN_Erect_MAP!$A$3:$A$5000,,1))<>""),'
        f'{{1,2,3,4}},{{1,1,1,1}}),"")'
    )
    ws.cell(row=7, column=1).value = spill_formula

    # F–I: Row formulas
    for r in range(7, 7 + MAX_DATA_ROWS):
        rs = str(r)
        # F: Spare
        ws.cell(row=r, column=6).value = erection_spare_formula(r)
        ws.cell(row=r, column=6).fill = CALC_FILL
        # G: For Order = QTY + Spare
        ws.cell(row=r, column=7).value = f'=IF(A{rs}="","",E{rs}+F{rs})'
        # H: Unit Price — blank
        ws.cell(row=r, column=8).fill = SUPPLIER_FILL
        # I: Total Price
        ws.cell(row=r, column=9).value = f'=IF(A{rs}="","",IFERROR(G{rs}*H{rs},0))'

    gt = GRAND_TOTAL_ROW
    ws.cell(row=gt, column=1, value="GRAND TOTAL PRICE ERECTION MATERIAL").font = BOLD
    for col in [5, 6, 7]:
        ws.cell(row=gt, column=col).value = f"=SUM({get_column_letter(col)}7:{get_column_letter(col)}{7+MAX_DATA_ROWS-1})"
        ws.cell(row=gt, column=col).font = BOLD
    ws.cell(row=gt, column=9).value = f"=SUM(I7:I{7+MAX_DATA_ROWS-1})"
    ws.cell(row=gt, column=9).font = BOLD

    set_col_widths(ws, {1: 10, 2: 10, 3: 52, 4: 30, 5: 14, 6: 10,
                        7: 14, 8: 14, 9: 14})
    return ws


# ========================================================================
# OUTPUT SHEET — 5. Painting + 6. Additional Parts
# ========================================================================
def create_output_painting_addparts(wb):
    ws = wb.create_sheet("5. Painting +6. Add Part")
    ws.sheet_properties.tabColor = "4472C4"

    # --- Section 5: Painting ---
    ws.cell(row=1, column=2, value="Supplier to fill in").font = BOLD
    ws.cell(row=1, column=2).fill = SUPPLIER_FILL
    ws.cell(row=2, column=2, value="Supplier:").fill = SUPPLIER_FILL
    ws.cell(row=2, column=6, value="Currency").fill = SUPPLIER_FILL

    ws.cell(row=4, column=1, value="5. Painting Material for Carbon Steel piping").font = BOLD
    ws.cell(row=4, column=6, value="For Order").font = BOLD
    ws.cell(row=4, column=7, value="Material Supply").font = BOLD

    headers_paint = [
        "Painting Pipe Component", "PIPE \nOD (mm)",
        "Color acc. to\nBS 1710", "Corrosivity\n category",
        "PIPE length\n (m)", "Surface in (m2)",
        "Unit   (m2)", "Total Price",
    ]
    write_headers(ws, 5, headers_paint)

    # Helper column headers (columns J onwards)
    ws.cell(row=5, column=10, value="Description").font = Font(bold=True, color="808080")
    ws.cell(row=5, column=11, value="color").font = Font(bold=True, color="808080")
    ws.cell(row=5, column=12, value="len [m]").font = Font(bold=True, color="808080")
    ws.cell(row=5, column=13, value="surface [m2]").font = Font(bold=True, color="808080")
    ws.cell(row=5, column=14, value="[m2/m]").font = Font(bold=True, color="808080")
    ws.cell(row=5, column=15, value="len with spare [m]").font = Font(bold=True, color="808080")
    ws.cell(row=5, column=16, value="len rounded to 6 [m]").font = Font(bold=True, color="808080")

    # Row 6: Dynamic array pulling from CALC_Paint (cols A-B = Description+Colour, F=rounded, H=surface)
    # Main visible columns A-H pull from CALC_Paint
    # A=Description, B=OD (extract from desc), C=Colour, D=Corrosion class (C4), E=Rounded length, F=Surface
    ws.cell(row=6, column=1).value = (
        '=IFERROR(_xlfn._xlpm.FILTER(CALC_Paint!A$3:A$502,CALC_Paint!A$3:A$502<>""),"")'
    )
    # B: OD extracted from description
    for r in range(6, 6 + MAX_DATA_ROWS):
        rs = str(r)
        ws.cell(row=r, column=2).value = (
            f'=IF(A{rs}="","",IFERROR(VALUE(MID(A{rs},FIND(" ",A{rs})+1,FIND("x",A{rs})-FIND(" ",A{rs})-1)),0))'
        )
        # C: Colour from CALC_Paint col B
        ws.cell(row=r, column=3).value = (
            f'=IF(A{rs}="","",INDEX(CALC_Paint!B$3:B$502,ROW()-5))'
        )
        # D: Corrosivity category (static C4 for CS)
        ws.cell(row=r, column=4).value = f'=IF(A{rs}="","","C4")'
        # E: Rounded pipe length from CALC_Paint col F
        ws.cell(row=r, column=5).value = (
            f'=IF(A{rs}="","",INDEX(CALC_Paint!F$3:F$502,ROW()-5))'
        )
        # F: Surface from CALC_Paint col H
        ws.cell(row=r, column=6).value = (
            f'=IF(A{rs}="","",INDEX(CALC_Paint!H$3:H$502,ROW()-5))'
        )
        # G: Unit Price — blank
        ws.cell(row=r, column=7).fill = SUPPLIER_FILL
        # H: Total Price = Surface × Unit Price
        ws.cell(row=r, column=8).value = (
            f'=IF(A{rs}="","",IFERROR(F{rs}*G{rs},0))'
        )

        # Helper columns J-P (for transparency)
        ws.cell(row=r, column=10).value = f'=A{rs}'
        ws.cell(row=r, column=11).value = f'=C{rs}'
        ws.cell(row=r, column=12).value = f'=IF(A{rs}="","",INDEX(CALC_Paint!C$3:C$502,ROW()-5))'
        ws.cell(row=r, column=13).value = (
            f'=IF(A{rs}="","",INDEX(CALC_Paint!C$3:C$502,ROW()-5)*INDEX(CALC_Paint!G$3:G$502,ROW()-5))'
        )
        ws.cell(row=r, column=14).value = f'=IF(A{rs}="","",INDEX(CALC_Paint!G$3:G$502,ROW()-5))'
        ws.cell(row=r, column=15).value = f'=IF(A{rs}="","",INDEX(CALC_Paint!E$3:E$502,ROW()-5))'
        ws.cell(row=r, column=16).value = f'=IF(A{rs}="","",INDEX(CALC_Paint!F$3:F$502,ROW()-5))'

    # SHOP COATING Grand Total
    shop_gt_row = 200
    ws.cell(row=shop_gt_row, column=1, value="SHOP COATING FOR CARBON STEEL PIPES").font = BOLD
    ws.cell(row=shop_gt_row, column=5).value = f'=SUM(E6:E{6+MAX_DATA_ROWS-1})'
    ws.cell(row=shop_gt_row, column=5).font = BOLD
    ws.cell(row=shop_gt_row, column=6).value = f'=SUM(F6:F{6+MAX_DATA_ROWS-1})'
    ws.cell(row=shop_gt_row, column=6).font = BOLD

    # --- Paint Cans Section ---
    cans_row = shop_gt_row + 3
    ws.cell(row=cans_row, column=1, value="Deliver primer and final paint in Cans for site painting (touch up and fittings)").font = BOLD
    ws.cell(row=cans_row, column=5, value="Estimated (litre)")
    ws.cell(row=cans_row, column=6, value="Surface in (m2)")
    ws.cell(row=cans_row, column=7, value="Unit   (litre)").font = BOLD
    ws.cell(row=cans_row, column=10, value="Touch up - 5% of TUBI pipes m2").font = Font(bold=True, color="808080")
    ws.cell(row=cans_row, column=11, value="Fittings m2").font = Font(bold=True, color="808080")

    # Primer row
    pr = cans_row + 1
    ws.cell(row=pr, column=1, value=" Cans for site painting (touch up and fittings) ")
    ws.cell(row=pr, column=3, value="Primer coat")
    # Surface = for colored: sum(color_touch + color_fit) + for primer-only: 2×(primer_touch + primer_fit)
    # Touch up (TUBI) = 5% of TUBI surface for ALL colours (because colored pipes also need primer)
    # + primer-only TUBI doubled
    # Fittings = 100% of fitting surface for ALL colours + primer-only doubled
    # Simplified: Each color contributes its touch+fit to primer. Primer-only gets doubled.
    ws.cell(row=pr, column=6).value = (
        f'=J{pr}+K{pr}'
    )
    # Touch (col J): 5% of all TUBI surface (primer-only doubled)
    ws.cell(row=pr, column=10).value = (
        f'=SUMPRODUCT((CALC_Paint!S$3:S$52<>"")*'
        f'IF(LOWER(CALC_Paint!S$3:S$52)="basic only",2,1)*'
        f'CALC_Paint!T$3:T$52)*0.05'
    )
    # Fittings (col K): 100% of all fitting surface (primer-only doubled)
    ws.cell(row=pr, column=11).value = (
        f'=SUMPRODUCT((CALC_Paint!O$3:O$52<>"")*'
        f'IF(LOWER(CALC_Paint!O$3:O$52)="basic only",2,1)*'
        f'CALC_Paint!P$3:P$52)'
    )
    # Litre = surface / 9, round to next 20; if >15 add 20
    ws.cell(row=pr, column=5).value = (
        f'=IF(F{pr}=0,0,IF(F{pr}/9>15,'
        f'_xlfn.CEILING.MATH(F{pr}/9,20)+20,'
        f'_xlfn.CEILING.MATH(F{pr}/9,20)))'
    )
    ws.cell(row=pr, column=7).fill = SUPPLIER_FILL
    ws.cell(row=pr, column=8).value = f'=IFERROR(E{pr}*G{pr},0)'

    # Dynamic colour rows for cans (non-primer colours)
    # List unique non-"basic only" colours from painting input
    color_start = pr + 1
    ws.cell(row=color_start, column=1, value=" Cans for site painting (touch up and fittings) ")
    ws.cell(row=color_start, column=3).value = (
        '=IFERROR(_xlfn._xlpm.FILTER('
        '_xlfn.UNIQUE(INDEX(IN_Paint!$A$3:$R$5000,,16)),'
        'LOWER(_xlfn.UNIQUE(INDEX(IN_Paint!$A$3:$R$5000,,16)))<>"basic only"'
        '),"")'
    )

    for r in range(color_start, color_start + 20):
        rs = str(r)
        if r > color_start:
            ws.cell(row=r, column=1, value=" Cans for site painting (touch up and fittings) ")
        # Surface = touch + fittings
        ws.cell(row=r, column=6).value = f'=IF(C{rs}="","",J{rs}+K{rs})'
        # Touch = 5% of TUBI surface for this colour
        ws.cell(row=r, column=10).value = (
            f'=IF(C{rs}="","",IFERROR(SUMPRODUCT('
            f'(INDEX(IN_Paint!$A$3:$R$5000,,3)="TUBI")*'
            f'(INDEX(IN_Paint!$A$3:$R$5000,,16)=C{rs})*'
            f'VALUE(INDEX(IN_Paint!$A$3:$R$5000,,13))),0)*0.05)'
        )
        # Fittings = 100% fitting surface for this colour
        ws.cell(row=r, column=11).value = (
            f'=IF(C{rs}="","",IFERROR(SUMPRODUCT('
            f'(INDEX(IN_Paint!$A$3:$R$5000,,3)<>"TUBI")*'
            f'(INDEX(IN_Paint!$A$3:$R$5000,,16)=C{rs})*'
            f'VALUE(INDEX(IN_Paint!$A$3:$R$5000,,13))),0))'
        )
        # Litre
        ws.cell(row=r, column=5).value = (
            f'=IF(C{rs}="","",IF(F{rs}/9>15,'
            f'_xlfn.CEILING.MATH(F{rs}/9,20)+20,'
            f'_xlfn.CEILING.MATH(F{rs}/9,20)))'
        )
        ws.cell(row=r, column=7).fill = SUPPLIER_FILL
        ws.cell(row=r, column=8).value = f'=IF(C{rs}="","",IFERROR(E{rs}*G{rs},0))'

    cans_gt = color_start + 21
    ws.cell(row=cans_gt, column=1, value="GRAND TOTAL PRICE SHOP COATING AND CANS FOR CARBON STEEL").font = BOLD
    ws.cell(row=cans_gt, column=8).value = f'=SUM(H6:H{6+MAX_DATA_ROWS-1})+SUM(H{pr}:H{color_start+19})'
    ws.cell(row=cans_gt, column=8).font = BOLD

    ws.cell(row=cans_gt + 1, column=1, value="Note: All paint in point 5. shall come from the same OEM.")

    # --- Section 6A: Blind Disks ---
    bd_start = cans_gt + 4
    ws.cell(row=bd_start, column=1, value="6. Additional Parts.").font = BOLD_12
    ws.cell(row=bd_start + 1, column=1,
            value="6A. Blind disk for Hydrotest material according TSD General Piping Material").font = BOLD
    ws.cell(row=bd_start + 2, column=1,
            value='PN16 between Flanges EN 1092-1 Form AA (unfinished)')
    ws.cell(row=bd_start + 2, column=6, value="For Order").font = BOLD
    ws.cell(row=bd_start + 2, column=7, value="Material Supply").font = BOLD

    bd_headers = ["MATERIAL", "TYPE", "DN\n(mm)", "S1\nThickness",
                  "Total Flanges", "QUANTITY\n (pcs)", "Unit  Price  (pcs.)", "Total Price"]
    write_headers(ws, bd_start + 3, bd_headers)

    hdr = bd_start + 4
    ws.cell(row=hdr, column=1, value="S235JRG2/1.0038\nor equal")
    ws.cell(row=hdr, column=2, value="Blind\n disk ")

    for i in range(14):  # 14 standard DN sizes in CFG
        r = hdr + i
        rs = str(r)
        # C: DN from CFG
        ws.cell(row=r, column=3).value = f'=CFG!A{3+i}'
        # D: S1 from CFG
        ws.cell(row=r, column=4).value = f'=CFG!B{3+i}'
        # E: Total flanges = SUMIFS across all 3 material input FLAN + FBLI rows for this DN
        ws.cell(row=r, column=5).value = (
            f'=SUMPRODUCT((IN_SS!$A$3:$A$5000="FLAN")*(IN_SS!$C$3:$C$5000=C{rs})*IN_SS!$H$3:$H$5000)'
            f'+SUMPRODUCT((IN_SS!$A$3:$A$5000="FBLI")*(IN_SS!$C$3:$C$5000=C{rs})*IN_SS!$H$3:$H$5000)'
            f'+SUMPRODUCT((IN_CS!$A$3:$A$5000="FLAN")*(IN_CS!$C$3:$C$5000=C{rs})*IN_CS!$H$3:$H$5000)'
            f'+SUMPRODUCT((IN_MAPRESS!$A$3:$A$5000="FLAN")*(IN_MAPRESS!$C$3:$C$5000=C{rs})*IN_MAPRESS!$H$3:$H$5000)'
        )
        # F: For Order based on rules: >50→10, 30-50→4, ≤30→2
        ws.cell(row=r, column=6).value = (
            f'=IF(E{rs}=0,0,_xlfn.IFS(E{rs}>50,10,E{rs}>30,4,TRUE,2))'
        )
        ws.cell(row=r, column=7).fill = SUPPLIER_FILL
        ws.cell(row=r, column=8).value = f'=IFERROR(F{rs}*G{rs},0)'

    # --- Section 6B: Flange Guards ---
    fg_start = hdr + 16
    ws.cell(row=fg_start, column=1,
            value="6B. Steel Flange guards — for fuel oil, ammonia, urea flange connections").font = BOLD
    ws.cell(row=fg_start + 1, column=6, value="For Order").font = BOLD
    ws.cell(row=fg_start + 1, column=7, value="Material Supply").font = BOLD

    fg_headers = ["TYPE", "Flange material", "DN\n(mm)", "DESCRIPTION/ Pipe Component",
                  "From BOQ QUANTITY\n (pcs)", "QUANTITY\n (pcs)", "Unit  Price  (pcs.)", "Total Price"]
    write_headers(ws, fg_start + 2, fg_headers)

    # Dynamic array: filter flanges from KKS inputs for flange-guard systems
    fg_data_row = fg_start + 3
    ws.cell(row=fg_data_row, column=1, value="Flange guards ")
    # Pull flange guard data from KKS inputs (CS + SS) for matching systems
    # We sort unique (Material, DN) from filtered flanges and sum quantities
    # This uses SORT(UNIQUE(FILTER(...)))
    ws.cell(row=fg_data_row, column=2).value = (
        '=IFERROR(_xlfn.SORT(_xlfn.UNIQUE(_xlfn._xlpm.FILTER('
        'CHOOSE({1,2},'
        'INDEX(IN_CS_KKS!$A$3:$N$5000,,7),'  # Material
        'INDEX(IN_CS_KKS!$A$3:$N$5000,,3)),' # DN
        'MMULT(--(LOWER(INDEX(IN_CS_KKS!$A$3:$N$5000,,1))=TRANSPOSE(LOWER(CFG!$D$3:$D$5))),'
        'ROW(CFG!$D$3:$D$5)^0)>0'
        '*(INDEX(IN_CS_KKS!$A$3:$N$5000,,3)="FLAN")'
        ')),{1,2},{1,1}),"")'
    )

    for r in range(fg_data_row, fg_data_row + 50):
        rs = str(r)
        ws.cell(row=r, column=4, value="Flange Guards")
        # E: From BOQ qty = SUMIFS
        ws.cell(row=r, column=5).value = (
            f'=IF(B{rs}="","",SUMPRODUCT('
            f'MMULT(--(LOWER(IN_CS_KKS!$A$3:$A$5000)=TRANSPOSE(LOWER(CFG!$D$3:$D$5))),'
            f'ROW(CFG!$D$3:$D$5)^0)>0'
            f'*(IN_CS_KKS!$C$3:$C$5000="FLAN")'
            f'*(IN_CS_KKS!$G$3:$G$5000=B{rs})'
            f'*(IN_CS_KKS!$E$3:$E$5000=C{rs})'
            f'*IN_CS_KKS!$M$3:$M$5000))'
        )
        # F: For Order = BOQ qty + 1
        ws.cell(row=r, column=6).value = f'=IF(B{rs}="","",E{rs}+1)'
        ws.cell(row=r, column=7).fill = SUPPLIER_FILL
        ws.cell(row=r, column=8).value = f'=IF(B{rs}="","",IFERROR(F{rs}*G{rs},0))'

    # --- Section 6C: Welding Nozzles ---
    wn_start = fg_data_row + 52
    ws.cell(row=wn_start, column=1,
            value="6C. Welding Nozzle for Temperature Measurement Weldolet").font = BOLD
    wn_headers = ["MATERIAL", "", "DESCRIPTION/ Pipe Component", "",
                  "", "QUANTITY\n (pcs)", "Unit  Price  (pcs.)", "Total Price"]
    write_headers(ws, wn_start + 1, wn_headers)
    ws.cell(row=wn_start + 2, column=1, value="P235GH or equal")
    ws.cell(row=wn_start + 2, column=3,
            value="Welding Nozzle for Temperature Measurement Weldolet for Thermowell OD40/ID25.5 L=50 EN 10222-1")
    ws.cell(row=wn_start + 2, column=6, value=0)
    ws.cell(row=wn_start + 2, column=7).fill = SUPPLIER_FILL
    ws.cell(row=wn_start + 2, column=8).value = f'=F{wn_start+2}*G{wn_start+2}'

    # Additional Parts Grand Total
    ap_gt = wn_start + 5
    ws.cell(row=ap_gt, column=1,
            value="GRAND TOTAL PRICE Additional Parts + Blind Disk + Olet").font = BOLD
    ws.cell(row=ap_gt, column=8).value = (
        f'=SUM(H{hdr}:H{hdr+13})'
        f'+SUM(H{fg_data_row}:H{fg_data_row+49})'
        f'+H{wn_start+2}'
    )
    ws.cell(row=ap_gt, column=8).font = BOLD

    set_col_widths(ws, {1: 40, 2: 12, 3: 14, 4: 12, 5: 14, 6: 14,
                        7: 14, 8: 14, 9: 4, 10: 30, 11: 14,
                        12: 12, 13: 12, 14: 10, 15: 16, 16: 18})
    return ws


# ========================================================================
# PRICE SUMMARY SHEET
# ========================================================================
def create_price_summary(wb):
    ws = wb.create_sheet("Price summary sheet")
    ws.sheet_properties.tabColor = "4472C4"

    ws.cell(row=4, column=2, value="Supplier to fill in").fill = SUPPLIER_FILL
    ws.cell(row=4, column=5, value="Currency")
    ws.cell(row=4, column=6, value="EURO").fill = SUPPLIER_FILL
    ws.cell(row=6, column=1, value="Supplier:").fill = SUPPLIER_FILL

    ws.cell(row=8, column=4, value="RFQ ;")
    ws.cell(row=9, column=4, value="Material Supply")

    headers = ["ITEM", "DESCRIPTION", "", "Length [m]", "WEIGHT\n [kg] ", "Price"]
    write_headers(ws, 10, headers)

    ws.cell(row=12, column=1, value="I: MATERIAL SUPPLY").font = BOLD_12

    gt = GRAND_TOTAL_ROW  # the row we used in output sheets
    items = [
        (14, "1. Mapress Geberit",
         f"='1. Mapress'!F{gt}", f"='1. Mapress'!J{gt}", f"='1. Mapress'!L{gt}"),
        (16, "2. Stainless Steel Material",
         f"='2. Stainless Steel'!F{gt}", f"='2. Stainless Steel'!J{gt}", f"='2. Stainless Steel'!L{gt}"),
        (18, "3. Carbon Steel Material",
         f"='3. Carbon Steel'!F{gt}", f"='3. Carbon Steel'!J{gt}", f"='3. Carbon Steel'!L{gt}"),
        (20, "4. Erection Material", "", "", f"='4. Erection Material'!I{gt}"),
        (22, "5. Painting Material for Carbon Steel piping in m2:",
         f"='5. Painting +6. Add Part'!F200", "", f"='5. Painting +6. Add Part'!H224"),
        (24, "6. Additional Parts + Blind Disk + Welding Nozzle",
         "", "", "='5. Painting +6. Add Part'!H" + str(200 + 3 + 21 + 4 + 52 + 5)),
    ]
    for row, desc, length_f, weight_f, price_f in items:
        ws.cell(row=row, column=1, value=desc).font = BOLD
        if length_f:
            ws.cell(row=row, column=4).value = length_f
        if weight_f:
            ws.cell(row=row, column=5).value = weight_f
        ws.cell(row=row, column=6).value = price_f

    ws.cell(row=26, column=1, value="GRAND TOTAL PRICE FOR MATERIAL SUPPLY").font = BOLD_12
    ws.cell(row=26, column=4).value = "=SUM(D14,D16,D18)"
    ws.cell(row=26, column=5).value = "=SUM(E14,E16,E18)"
    ws.cell(row=26, column=6).value = "=SUM(F14,F16,F18,F20,F22,F24)"
    ws.cell(row=26, column=6).font = BOLD_12

    ws.cell(row=28, column=1, value="Option").font = BOLD
    ws.cell(row=28, column=5, value="Service Supply")
    ws.cell(row=29, column=5, value="Estimated quantity of 20' standard containers")
    ws.cell(row=29, column=6, value="Total transportation \nPrice")
    ws.cell(row=30, column=1, value="TRANSPORTATION SERVICES TO CONSTRUCTION SITE").font = BOLD

    set_col_widths(ws, {1: 50, 2: 14, 3: 10, 4: 14, 5: 14, 6: 16})
    return ws


# ========================================================================
# MAIN
# ========================================================================
def main():
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # --- Input sheets ---
    create_input_sheet_material(wb, "IN_SS", "SS Material")
    create_input_sheet_material(wb, "IN_CS", "CS Material")
    create_input_sheet_material(wb, "IN_MAPRESS", "MAPRESS Material")
    create_input_sheet_material(wb, "IN_Erect_SS", "SS Erection", is_erection=True)
    create_input_sheet_material(wb, "IN_Erect_CS", "CS Erection", is_erection=True)
    create_input_sheet_material(wb, "IN_Erect_MAP", "MAPRESS Erection", is_erection=True)
    create_input_sheet_painting(wb)
    create_input_sheet_kks(wb, "IN_CS_KKS", "CS")
    create_input_sheet_kks(wb, "IN_SS_KKS", "SS")

    # --- Config ---
    create_config_sheet(wb)

    # --- Calculation (helper) ---
    create_calc_painting(wb)

    # --- Output sheets ---
    create_output_material(wb, "1. Mapress", "IN_MAPRESS", 1)
    create_output_material(wb, "2. Stainless Steel", "IN_SS", 2)
    create_output_material(wb, "3. Carbon Steel", "IN_CS", 3, is_cs=True)
    create_output_erection(wb)
    create_output_painting_addparts(wb)
    create_price_summary(wb)

    # --- Save ---
    out_path = Path(__file__).parent / "Price_Sheet_Template.xlsx"
    wb.save(str(out_path))
    print(f"Template created: {out_path}")
    print(f"  Input sheets:  {[s.title for s in wb.worksheets if s.title.startswith('IN_')]}")
    print(f"  Config:        CFG")
    print(f"  Calculation:   CALC_Paint")
    print(f"  Output sheets: {[s.title for s in wb.worksheets if not s.title.startswith(('IN_','CFG','CALC'))]}")
    print()
    print("WORKFLOW:")
    print("  1. Open Price_Sheet_Template.xlsx in Excel 365")
    print("  2. Paste 'Total Piping Material' data into each IN_* sheet (starting row 3)")
    print("  3. Paste painting BoM into IN_Paint, KKS data into IN_CS_KKS / IN_SS_KKS")
    print("  4. Output sheets auto-calculate. Supplier fills Unit Price columns.")


if __name__ == "__main__":
    main()
