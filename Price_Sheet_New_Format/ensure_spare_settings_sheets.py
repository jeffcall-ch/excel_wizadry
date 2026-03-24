from pathlib import Path

from openpyxl import load_workbook


def ensure_sheet(wb, name: str, headers: list[str], rows: list[list[object]]) -> None:
    if name in wb.sheetnames:
        return
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def upsert_key_value_sheet(wb, name: str, rows: list[list[object]]) -> None:
    headers = ["Key", "Value", "Description", "Example"]
    existing_values = {}

    if name in wb.sheetnames:
        ws_old = wb[name]
        hdr = [ws_old.cell(1, c).value for c in range(1, ws_old.max_column + 1)]
        hdr_map = {str(v).strip(): idx for idx, v in enumerate(hdr, start=1) if v is not None}
        if "Key" in hdr_map and "Value" in hdr_map:
            for r in range(2, ws_old.max_row + 1):
                key_val = ws_old.cell(r, hdr_map["Key"]).value
                if key_val is None:
                    continue
                existing_values[str(key_val).strip()] = ws_old.cell(r, hdr_map["Value"]).value
        del wb[name]

    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for key, value, desc, ex in rows:
        key_s = str(key)
        current_value = existing_values.get(key_s, value)
        ws.append([key_s, current_value, desc, ex])


def upsert_keyword_sheet(wb, name: str, rows: list[list[object]]) -> None:
    headers = ["Name", "Keywords", "Description", "Example"]
    existing_values = {}

    if name in wb.sheetnames:
        ws_old = wb[name]
        hdr = [ws_old.cell(1, c).value for c in range(1, ws_old.max_column + 1)]
        hdr_map = {str(v).strip(): idx for idx, v in enumerate(hdr, start=1) if v is not None}
        if "Name" in hdr_map and "Keywords" in hdr_map:
            for r in range(2, ws_old.max_row + 1):
                name_val = ws_old.cell(r, hdr_map["Name"]).value
                if name_val is None:
                    continue
                existing_values[str(name_val).strip()] = ws_old.cell(r, hdr_map["Keywords"]).value
        del wb[name]

    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for name_val, keywords, desc, ex in rows:
        name_s = str(name_val)
        current_keywords = existing_values.get(name_s, keywords)
        ws.append([name_s, current_keywords, desc, ex])


def replace_guide_sheet(wb) -> None:
    name = "SpareRuleGuide"
    if name in wb.sheetnames:
        del wb[name]
    ensure_sheet(
        wb,
        name,
        ["Area", "Rule", "Condition", "Formula", "Rounding", "Example"],
        [
            ["Main/Shop", "MAIN#1", "Seal and qty<=seal_qty_threshold", "spare=seal_min_spare_low", "integer", "qty=7 -> spare=5"],
            ["Main/Shop", "MAIN#2", "Seal and qty>seal_qty_threshold", "spare=max(seal_min_spare_high, ceil(qty*seal_percent))", "ceil", "qty=40 -> max(5,4)=5"],
            ["Main/Shop", "MAIN#3", "Critical and qty<=critical_qty_threshold", "spare=max(critical_min_spare, ceil(qty*critical_percent_low))", "ceil", "qty=8 -> max(3,4)=4"],
            ["Main/Shop", "MAIN#4", "Critical and qty>critical_qty_threshold", "spare=max(critical_min_spare, ceil(qty*critical_percent_high))", "ceil", "qty=80 -> max(3,8)=8"],
            ["Main/Shop", "MAIN#T1", "TUBI row", "base=ceil(qty*pct); total=round_up_to_multiple(qty+base, tubi_round_multiple); spare=total-qty", "ceil + multiple", "qty=35,base=4,total=42,spare=7"],
            ["Main/Shop", "MAIN#5", "Standard pcs and qty<=threshold_1", "spare=0", "none", "qty=4 -> 0"],
            ["Main/Shop", "MAIN#6", "Standard pcs and threshold_1<qty<=threshold_2", "spare=1", "none", "qty=12 -> 1"],
            ["Main/Shop", "MAIN#7a", "Standard pcs and qty>threshold_2 and size<size_threshold", "spare=ceil(qty*percent_pcs_small)", "ceil", "qty=100,size=25 -> 8"],
            ["Main/Shop", "MAIN#7b", "Standard pcs and qty>threshold_2 and size>=size_threshold", "spare=ceil(qty*percent_pcs_large)", "ceil", "qty=100,size=80 -> 5"],
            ["Main/Shop", "MAIN#8a", "Standard meter and size<size_threshold", "spare=ceil(qty_m*percent_m_small)", "ceil", "qty_m=30,size=25 -> 5"],
            ["Main/Shop", "MAIN#8b", "Standard meter and size>=size_threshold", "spare=ceil(qty_m*percent_m_large)", "ceil", "qty_m=30,size=80 -> 3"],
            ["Erection", "ER#1/2/3/4", "NWB by thresholds", "base=ceil(qty*nwb_percent_*); total=round_up_nut_washer(qty+base); spare=total-qty", "ceil + step rounding", "qty=156 -> spare=94"],
            ["Erection", "ER#1B/2B/3B/4B", "Bolt by thresholds", "base=ceil(qty*nwb_percent_*); total=round_up_to_next_10(qty+base); spare=total-qty", "ceil + next10", "qty=108 -> spare=42"],
            ["Erection", "ER#5/6/7", "Gasket by thresholds", "low:ceil(qty*gasket_low); mid/high: base+round_to_next10", "ceil + next10", "qty=200 -> spare=100"],
            ["Erection", "ER#8/9", "Standard erection", "qty<standard_threshold -> ceil(qty*standard_percent_low), else ceil(qty*standard_percent_high)", "ceil", "qty=500 -> 100"],
            ["Scope", "Skip", "Paint and unknown material types", "no spare update", "n/a", "Paint shop rows unchanged"],
        ],
    )


def main() -> None:
    settings = Path("Price_Sheet_New_Format/import_settings.xlsx")
    wb = load_workbook(settings)

    upsert_key_value_sheet(
        wb,
        "SpareGeneral",
        [
            ["aggregate_table_suffix", "aggreated", "Suffix appended to revision table to get aggregated table.", "rev0 + aggreated -> rev0aggreated"],
            ["spare_column", "Spare", "Target column to store calculated spare quantity.", "Spare"],
            ["spare_info_column", "Spare info", "Dedicated column to store applied spare rule explanation text.", "Rule #3: Critical components qty <= 10 -> ..."],
            ["material_type_column", "Material type", "Column used to classify shop/erection/paint rows.", "SS shop, Erection, Paint shop"],
            ["description_column", "Description", "Description source for keyword-based classification.", "Hexagon Nut DIN-EN-ISO-4032 M12"],
            ["type_label_column", "Type label", "Optional type label source used by TUBI detection; leave as default if not present.", "TUBI"],
            ["type_column", "Type", "Type source used by some rules and TUBI detection.", "PIPE, TUBI"],
            ["size_column", "Size", "DN/size source used for threshold split logic.", "25, 50, 80"],
            ["qty_pcs_column", "Qty pcs", "Piece quantity source column.", "120"],
            ["qty_m_column", "Qty m", "Meter quantity source column (used when Qty pcs is 0).", "35.5"],
            ["weight_column", "Weight kg", "Weight source used to derive unit weight (Weight / base qty).", "120.5"],
            ["total_qty_column", "Order qty", "Output column for total quantity = base qty + spare.", "137"],
            ["weight_w_spare_column", "Weight w spare", "Output column for total weight with spare = total qty * unit weight.", "138.22"],
            ["paint_prefix", "paint", "Rows with material type starting with this are skipped in spare calc.", "Paint shop"],
            ["erection_exact", "Erection", "Material type value that is treated as erection logic.", "Erection"],
            ["shop_suffix", "shop", "Material type suffix that marks main/shop rows.", "SS shop"],
        ],
    )

    upsert_key_value_sheet(
        wb,
        "SpareMainRules",
        [
            ["seal_qty_threshold", 10, "Seal low/high split quantity threshold.", "qty<=10 -> fixed minimum"],
            ["seal_min_spare_low", 5, "Fixed spare for low-qty seal rows.", "qty=7 -> spare=5"],
            ["seal_min_spare_high", 5, "Minimum spare for high-qty seal rows.", "max(5, ceil(qty*seal_percent))"],
            ["seal_percent", 0.10, "Seal percentage for qty above threshold.", "qty=40 -> max(5,4)=5"],
            ["critical_qty_threshold", 10, "Critical low/high split threshold.", "qty<=10 uses critical_percent_low"],
            ["critical_min_spare", 3, "Minimum critical spare regardless of percentage result.", "qty=2 with 50% -> max(3,1)=3"],
            ["critical_percent_low", 0.50, "Critical percentage for low quantity.", "qty=8 -> ceil(4)=4"],
            ["critical_percent_high", 0.10, "Critical percentage for high quantity.", "qty=80 -> ceil(8)=8"],
            ["threshold_1", 5, "Standard rule first threshold.", "qty<=5 -> spare=0"],
            ["threshold_2", 20, "Standard rule second threshold.", "5<qty<=20 -> spare=1"],
            ["size_threshold", 50, "Size split between small and large percentages.", "size<50 uses small percentages"],
            ["percent_pcs_small", 0.075, "Standard pcs percentage when size is small.", "qty=100,size=25 -> 8"],
            ["percent_pcs_large", 0.05, "Standard pcs percentage when size is large.", "qty=100,size=80 -> 5"],
            ["percent_m_small", 0.15, "Standard meter percentage when size is small.", "qty_m=30,size=25 -> 5"],
            ["percent_m_large", 0.10, "Standard meter percentage when size is large.", "qty_m=30,size=80 -> 3"],
            ["tubi_round_multiple", 6, "Round TUBI total quantity to this multiple.", "qty=35,base=4,total=42,spare=7"],
        ],
    )

    upsert_key_value_sheet(
        wb,
        "SpareErectionRules",
        [
            ["nwb_threshold_1", 30, "NWB first threshold.", "qty<=30 uses nwb_percent_low"],
            ["nwb_threshold_2", 100, "NWB second threshold.", "30<qty<=100 uses nwb_percent_mid"],
            ["nwb_threshold_3", 400, "NWB third threshold.", "100<qty<=400 uses nwb_percent_high"],
            ["nwb_percent_low", 1.0, "NWB percent for lowest range.", "100% base spare"],
            ["nwb_percent_mid", 0.5, "NWB percent for mid range.", "50% base spare"],
            ["nwb_percent_high", 0.35, "NWB percent for high range.", "35% base spare"],
            ["nwb_percent_very_high", 0.2, "NWB percent for very high range.", "20% base spare"],
            ["gasket_threshold_1", 30, "Gasket first threshold.", "qty<=30 uses gasket_percent_low"],
            ["gasket_threshold_2", 500, "Gasket second threshold.", "30<qty<=500 uses gasket_percent_mid"],
            ["gasket_percent_low", 1.0, "Gasket percent for low range.", "100%"],
            ["gasket_percent_mid", 0.5, "Gasket percent for mid range.", "50%"],
            ["gasket_percent_high", 0.3, "Gasket percent for high range.", "30%"],
            ["standard_threshold", 400, "Standard erection threshold.", "qty<400 low%, else high%"],
            ["standard_percent_low", 0.35, "Standard erection percent below threshold.", "35%"],
            ["standard_percent_high", 0.2, "Standard erection percent at/above threshold.", "20%"],
        ],
    )

    upsert_keyword_sheet(
        wb,
        "SpareKeywords",
        [
            ["seal_required", "seal,ring,fkm", "All keywords required for seal classification.", "'FKM seal ring'"],
            ["seal_excluded", "elbow,flange,tee,reducer", "If any appears, row is not seal class.", "'Seal ring elbow' excluded"],
            ["critical", "nipple,nipples,union,connection,connections,welding,adaptor,coupling,threaded", "Any keyword marks row as critical (after seal).", "'Threaded union'"],
            ["tubi", "tubi", "Any keyword marks row as TUBI class (after seal/critical).", "Type='TUBI'"],
            ["nwb", "nut,washer", "Any keyword marks erection row as NWB class.", "'Hexagon Nut'"],
            ["bolt", "bolt", "Any keyword marks erection row as bolt class.", "'Hexagon Bolt'"],
            ["gasket", "gasket", "Any keyword marks erection row as gasket class.", "'Flat gasket'"],
        ],
    )

    replace_guide_sheet(wb)

    wb.save(settings)
    print(f"Ensured spare settings sheets in: {settings}")


if __name__ == "__main__":
    main()
