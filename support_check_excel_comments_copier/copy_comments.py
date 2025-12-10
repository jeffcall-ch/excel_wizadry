"""
Script to copy Comments, Comment entered for BQ Zone, and Good Examples from old Excel file to new Excel file.
Matches rows based on SU/STEEL ref column (column D).
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil


def copy_comments_from_old_to_new(old_file, new_file, output_file):
    """
    Copy Comments, Comment entered for BQ Zone, and Good Examples from old file to new file based on SU/STEEL ref matching.
    
    Args:
        old_file: Path to the old Excel file (source for comments)
        new_file: Path to the new Excel file (destination)
        output_file: Path to save the output Excel file
    """
    # Load both Excel files
    print(f"Loading old file: {old_file}")
    df_old = pd.read_excel(old_file, sheet_name='main')
    
    print(f"Loading new file: {new_file}")
    df_new = pd.read_excel(new_file, sheet_name='main')
    
    # Get unique SU/STEEL ref values from old file along with their Comments, Comment entered for BQ Zone, and Good Examples
    # Filter out NaN values in SU/STEEL ref
    df_old_filtered = df_old[df_old['SU/STEEL ref'].notna()].copy()
    
    # Create a mapping dictionary: SU/STEEL ref -> (Comments, Comment entered for BQ Zone, Good Examples)
    # For duplicate SU/STEEL refs, take the first non-NaN value
    comments_map = {}
    
    for steel_ref in df_old_filtered['SU/STEEL ref'].unique():
        # Get all rows with this steel ref
        matching_rows = df_old_filtered[df_old_filtered['SU/STEEL ref'] == steel_ref]
        
        # Get the first non-NaN comment if available
        comments = matching_rows['Comments'].dropna()
        comment_value = comments.iloc[0] if len(comments) > 0 else None
        
        # Get the first non-NaN BQ Zone comment if available
        bq_zone_comments = matching_rows['Comment entered for BQ Zone'].dropna()
        bq_zone_comment_value = bq_zone_comments.iloc[0] if len(bq_zone_comments) > 0 else None
        
        # Get the first non-NaN good example if available
        good_examples = matching_rows['Good examples'].dropna()
        good_example_value = good_examples.iloc[0] if len(good_examples) > 0 else None
        
        comments_map[steel_ref] = (comment_value, bq_zone_comment_value, good_example_value)
    
    print(f"Found {len(comments_map)} unique SU/STEEL ref values in old file")
    
    # Ensure Comments, Comment entered for BQ Zone, and Good examples columns exist in new dataframe
    if 'Comments' not in df_new.columns:
        print("Adding 'Comments' column to new file")
        df_new['Comments'] = None
    
    if 'Comment entered for BQ Zone' not in df_new.columns:
        print("Adding 'Comment entered for BQ Zone' column to new file")
        df_new['Comment entered for BQ Zone'] = None
    
    if 'Good examples' not in df_new.columns:
        print("Adding 'Good examples' column to new file")
        df_new['Good examples'] = None
    
    # Copy comments, BQ Zone comments, and good examples to all matching rows in new file
    matches_found = 0
    for idx, row in df_new.iterrows():
        steel_ref = row['SU/STEEL ref']
        
        if pd.notna(steel_ref) and steel_ref in comments_map:
            comment_value, bq_zone_comment_value, good_example_value = comments_map[steel_ref]
            
            if comment_value is not None:
                df_new.at[idx, 'Comments'] = comment_value
                matches_found += 1
            
            if bq_zone_comment_value is not None:
                df_new.at[idx, 'Comment entered for BQ Zone'] = bq_zone_comment_value
            
            if good_example_value is not None:
                df_new.at[idx, 'Good examples'] = good_example_value
    
    print(f"Copied comments to {matches_found} matching rows")
    
    # Reorder columns to ensure Comments, Comment entered for BQ Zone, Good examples are in sequence
    # Get all column names
    cols = df_new.columns.tolist()
    
    # Remove the three comment columns from their current positions
    for col in ['Comments', 'Comment entered for BQ Zone', 'Good examples']:
        if col in cols:
            cols.remove(col)
    
    # Insert them in the desired order at the end (or you can specify a position)
    # Inserting at the end for now
    cols.extend(['Comments', 'Comment entered for BQ Zone', 'Good examples'])
    
    # Reorder the dataframe
    df_new = df_new[cols]
    
    # Create a copy of the new file
    print(f"Creating copy of new file: {output_file}")
    shutil.copy2(new_file, output_file)
    
    # Apply formatting using openpyxl
    print("Applying formatting...")
    
    # Load old workbook to get column widths by column name
    print("Reading column widths from old file...")
    wb_old = load_workbook(old_file)
    ws_old = wb_old['main']
    
    # Get old file column headers
    old_headers = [cell.value for cell in ws_old[1]]
    
    # Store column widths from old file by column name
    old_column_widths_by_name = {}
    for col_idx, header in enumerate(old_headers, 1):
        col_letter = get_column_letter(col_idx)
        old_column_widths_by_name[header] = ws_old.column_dimensions[col_letter].width
    
    # Load the copied workbook
    wb = load_workbook(output_file, keep_vba=True)
    
    # Create new sheet with processed data
    print("Creating new sheet 'main_processed'...")
    ws = wb.create_sheet('main_processed')
    
    # Write dataframe to the new sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_new, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Enable filters on header row
    ws.auto_filter.ref = ws.dimensions
    
    # Apply column widths from old file based on column names
    print("Applying column widths from old file...")
    new_headers = [cell.value for cell in ws[1]]
    for col_idx, header in enumerate(new_headers, 1):
        col_letter = get_column_letter(col_idx)
        if header in old_column_widths_by_name and old_column_widths_by_name[header] is not None:
            ws.column_dimensions[col_letter].width = old_column_widths_by_name[header]
    
    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Wrap text
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # Auto-fit row heights (set a reasonable default)
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = None  # Auto height
    
    # Freeze panes at E2 (column E, row 2)
    ws.freeze_panes = 'E2'
    
    # Save the formatted workbook
    wb.save(output_file)
    print("Done! File saved and formatted.")


if __name__ == "__main__":
    import os
    
    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Define file paths
    old_file = os.path.join(script_dir, "old.xlsm")
    new_file = os.path.join(script_dir, "new.xlsm")
    output_file = os.path.join(script_dir, "new_processed.xlsm")
    
    # Run the script
    copy_comments_from_old_to_new(old_file, new_file, output_file)
