#!/usr/bin/env python3
"""
Price Sheet Excel Comparator

Compares two versions of a piping component pricing Excel file (.xlsm).
Produces a timestamped copy of the new file with colour-coded changes:

  GREEN      — value present in NEW but absent in OLD, or entire row is new
  YELLOW     — value changed between OLD and NEW (non-numeric, or numeric NEW > OLD)
  LIGHT BLUE — numeric value in NEW is SMALLER than in OLD
  RED        — value present in OLD is absent in NEW (cell), or DELETED row at bottom

Row identification uses a compound key: Column A + B + C + D.
Deleted rows (present in OLD but absent in NEW) are appended below the last
populated row in each sheet.

Sheet processing rules
  - Skip sheet named 'Coversheet'
  - Skip sheets whose name contains 'summary' (case-insensitive)
  - Skip sheets whose name contains 'painting' or '+' in the name
  - Skip sheets with no MATERIAL or TYPE header found in column A

Header detection: first row where column A, stripped and lower-cased, equals
exactly 'material' or 'type'.
Last data row   : one row ABOVE the first row (after the header) whose column A
                  contains 'grand' (case-insensitive).

Usage:
    python price_sheet_compare.py                      # default paths
    python price_sheet_compare.py old.xlsm new.xlsm
"""

import shutil
import logging
import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------
class Colors:
    GREEN      = "00FF00"   # Added value / new row
    YELLOW     = "FFFF00"   # Changed value (or numeric NEW > OLD)
    LIGHT_BLUE = "C0C0FF"   # Numeric value decreased (NEW < OLD)
    RED        = "FFC0C0"   # Deleted value / deleted row


# ---------------------------------------------------------------------------
# Main comparator
# ---------------------------------------------------------------------------
class PriceSheetComparator:
    """
    Compares two piping price sheet Excel files (.xlsm) and writes a
    colour-coded output as a timestamped copy of the new file.
    """

    def __init__(self, old_file: str, new_file: str):
        self.old_file   = Path(old_file)
        self.new_file   = Path(new_file)
        suffix          = self.new_file.suffix          # keep .xlsm
        timestamp       = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_file = (
            self.new_file.parent
            / f"{self.new_file.stem}_{timestamp}_comparison{suffix}"
        )

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.FileHandler("price_comparison.log", mode="w"),
                logging.StreamHandler(),
            ],
        )
        self.logger = logging.getLogger(__name__)

    # ── Sheet filtering ──────────────────────────────────────────────────────

    def _should_skip_sheet(self, name: str) -> bool:
        """Return True when the sheet must be skipped entirely."""
        name_lower = name.lower()
        if name.strip() == "Coversheet":
            return True
        if "summary" in name_lower:
            return True
        if "painting" in name_lower or "+" in name:
            return True
        return False

    # ── Row / column detection ───────────────────────────────────────────────

    def _find_header_row_idx(self, rows: List[tuple]) -> Optional[int]:
        """
        Return the 0-based index of the header row.

        Criterion: column A, stripped and lower-cased, equals exactly
        'material' or 'type'.
        """
        for i, row in enumerate(rows):
            if row and row[0] is not None:
                val = str(row[0]).strip().lower()
                if val in ("material", "type"):
                    return i
        return None

    def _find_grand_row_idx(
        self, rows: List[tuple], after_idx: int
    ) -> Optional[int]:
        """
        Return the 0-based index of the GRAND TOTAL row — the first row
        after *after_idx* whose column A contains 'grand' (case-insensitive).
        """
        for i in range(after_idx + 1, len(rows)):
            if rows[i] and rows[i][0] is not None:
                if "grand" in str(rows[i][0]).lower():
                    return i
        return None

    def _find_last_header_col_idx(self, header_row: tuple) -> int:
        """Return the 0-based index of the rightmost non-None header column."""
        last = 0
        for j, v in enumerate(header_row):
            if v is not None:
                last = j
        return last

    def _find_last_populated_excel_row(self, ws) -> int:
        """
        Return the 1-based last row that has ANY non-None cell value.
        Scans from the bottom up so it stops as soon as content is found.
        """
        max_r = ws.max_row    or 1
        max_c = ws.max_column or 1
        for r in range(max_r, 0, -1):
            for c in range(1, max_c + 1):
                if ws.cell(row=r, column=c).value is not None:
                    return r
        return 0

    # ── Key and value helpers ────────────────────────────────────────────────

    def _build_key(self, row_values: tuple) -> str:
        """Compound key from first four columns (A|B|C|D)."""
        parts = []
        for i in range(4):
            v = row_values[i] if i < len(row_values) else None
            parts.append("" if v is None else str(v).strip())
        return "|".join(parts)

    def _build_partial_key(self, row_values: tuple) -> str:
        """Partial key from first three columns (A|B|C) — used as fallback."""
        parts = []
        for i in range(3):
            v = row_values[i] if i < len(row_values) else None
            parts.append("" if v is None else str(v).strip())
        return "|".join(parts)

    def _cell_text(self, value) -> str:
        """Normalise a cell value to a comparable string."""
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    # ── Core comparison ──────────────────────────────────────────────────────

    def _process_sheet(
        self,
        sheet_name: str,
        ws_output,
        rows_old: List[tuple],
        rows_new: List[tuple],
    ) -> None:
        """
        Compare old/new data for one sheet, apply colour fills to ws_output,
        append a 'Deleted Rows' section, and write a change-marker column.
        """

        # ── Locate header and data bounds ────────────────────────────────────
        header_idx_old = self._find_header_row_idx(rows_old)
        header_idx_new = self._find_header_row_idx(rows_new)

        if header_idx_old is None or header_idx_new is None:
            self.logger.info(
                f"  Skipping '{sheet_name}': no MATERIAL / TYPE header found in column A"
            )
            return

        grand_idx_old = self._find_grand_row_idx(rows_old, header_idx_old)
        grand_idx_new = self._find_grand_row_idx(rows_new, header_idx_new)

        if grand_idx_old is None or grand_idx_new is None:
            self.logger.warning(
                f"  Skipping '{sheet_name}': no GRAND TOTAL row found"
            )
            return

        # ── Column metadata from the NEW header row ──────────────────────────
        header_row     = rows_new[header_idx_new]
        last_col_idx   = self._find_last_header_col_idx(header_row)
        n_cols         = last_col_idx + 1
        col_names      = [
            (str(header_row[i]) if header_row[i] is not None else f"Col_{i+1}")
            for i in range(n_cols)
        ]
        header_excel_row = header_idx_new + 1   # convert to 1-based

        # ── Extract and pad data rows ────────────────────────────────────────
        def _padded(row: tuple) -> tuple:
            return tuple(row[i] if i < len(row) else None for i in range(n_cols))

        data_old = [
            _padded(rows_old[i])
            for i in range(header_idx_old + 1, grand_idx_old)
        ]
        data_new = [
            (i + 1, _padded(rows_new[i]))          # (excel_row, values)
            for i in range(header_idx_new + 1, grand_idx_new)
        ]

        self.logger.info(
            f"  '{sheet_name}': {len(data_old)} old rows, {len(data_new)} new rows, "
            f"{n_cols} columns  [{col_names[0]}  …  {col_names[-1]}]"
        )

        fill_green  = PatternFill(start_color=Colors.GREEN,      end_color=Colors.GREEN,      fill_type="solid")
        fill_yellow = PatternFill(start_color=Colors.YELLOW,     end_color=Colors.YELLOW,     fill_type="solid")
        fill_lblue  = PatternFill(start_color=Colors.LIGHT_BLUE, end_color=Colors.LIGHT_BLUE, fill_type="solid")
        fill_red    = PatternFill(start_color=Colors.RED,        end_color=Colors.RED,        fill_type="solid")

        # ── Pass 1: exact 4-key matching (A|B|C|D) ───────────────────────────
        old_lookup_4: Dict[str, List[tuple]] = defaultdict(list)
        for row_values in data_old:
            old_lookup_4[self._build_key(row_values)].append(row_values)
        old_cursor_4: Dict[str, int] = defaultdict(int)

        matched_pairs: List[tuple] = []     # (excel_row, new_vals, old_vals)
        unmatched_new: List[tuple] = []     # (excel_row, new_vals)

        for excel_row, new_vals in data_new:
            key4 = self._build_key(new_vals)
            if key4 in old_lookup_4 and old_cursor_4[key4] < len(old_lookup_4[key4]):
                old_vals = old_lookup_4[key4][old_cursor_4[key4]]
                old_cursor_4[key4] += 1
                matched_pairs.append((excel_row, new_vals, old_vals))
            else:
                unmatched_new.append((excel_row, new_vals))

        # OLD rows not consumed by pass 1
        unmatched_old: List[tuple] = []
        for key4, old_rows_list in old_lookup_4.items():
            unmatched_old.extend(old_rows_list[old_cursor_4.get(key4, 0):])

        # ── Pass 2: partial 3-key fallback (A|B|C) for unmatched rows ────────
        # Rows where only the description (D) changed should not become add+delete.
        old_lookup_3: Dict[str, List[tuple]] = defaultdict(list)
        for old_vals in unmatched_old:
            old_lookup_3[self._build_partial_key(old_vals)].append(old_vals)
        old_cursor_3: Dict[str, int] = defaultdict(int)

        truly_unmatched_new: List[tuple] = []   # no OLD counterpart at all

        for excel_row, new_vals in unmatched_new:
            key3 = self._build_partial_key(new_vals)
            if key3 in old_lookup_3 and old_cursor_3[key3] < len(old_lookup_3[key3]):
                old_vals = old_lookup_3[key3][old_cursor_3[key3]]
                old_cursor_3[key3] += 1
                matched_pairs.append((excel_row, new_vals, old_vals))
            else:
                truly_unmatched_new.append((excel_row, new_vals))

        # OLD rows not consumed by pass 2 → truly deleted
        deleted_rows: List[tuple] = []
        for key3, old_rows_list in old_lookup_3.items():
            deleted_rows.extend(old_rows_list[old_cursor_3.get(key3, 0):])

        # ── Apply colour fills ────────────────────────────────────────────────
        row_markers: Dict[int, str] = {}

        # Matched rows (pass 1 or pass 2): cell-by-cell comparison
        for excel_row, new_vals, old_vals in matched_pairs:
            change_types = set()
            for col in range(n_cols):
                n_txt = self._cell_text(new_vals[col])
                o_txt = self._cell_text(old_vals[col])
                if n_txt == o_txt:
                    continue

                if n_txt and not o_txt:
                    fill = fill_green
                    change_types.add("Added")
                elif not n_txt and o_txt:
                    fill = fill_red
                    change_types.add("Deleted")
                else:
                    # Both have values — check for numeric decrease
                    fill = fill_yellow
                    change_types.add("Changed")
                    try:
                        n_num = float(n_txt)
                        o_num = float(o_txt)
                        if n_num < o_num:
                            fill = fill_lblue
                            change_types.discard("Changed")
                            change_types.add("Smaller")
                    except ValueError:
                        pass  # non-numeric: keep yellow

                ws_output.cell(row=excel_row, column=col + 1).fill = fill

            if change_types:
                parts = [t for t in ("Deleted", "Smaller", "Changed", "Added") if t in change_types]
                row_markers[excel_row] = ", ".join(parts)

        # ── Write 'Deleted Rows' section below all existing data ─────────────
        if deleted_rows:
            last_row      = self._find_last_populated_excel_row(ws_output)
            section_start = last_row + 2

            # Section title
            title_cell       = ws_output.cell(row=section_start, column=1)
            title_cell.value = "Deleted Rows"
            title_cell.font  = Font(bold=True)

            # Column-name sub-header
            for c_idx, c_name in enumerate(col_names):
                sub_cell       = ws_output.cell(row=section_start + 1, column=c_idx + 1)
                sub_cell.value = c_name
                sub_cell.font  = Font(bold=True)

            # Deleted data rows (red fill)
            for i, del_vals in enumerate(deleted_rows):
                del_excel_row = section_start + 2 + i
                for c_idx in range(n_cols):
                    cell        = ws_output.cell(row=del_excel_row, column=c_idx + 1)
                    cell.value  = del_vals[c_idx]
                    cell.fill   = fill_red
                row_markers[del_excel_row] = "Deleted"

        # ── Write change-marker column ────────────────────────────────────────
        marker_col    = n_cols + 2                  # one empty column gap after data
        marker_letter = get_column_letter(marker_col)
        ws_output.column_dimensions[marker_letter].width = 16

        # Marker column header (aligned with data header row)
        hdr_marker        = ws_output.cell(row=header_excel_row, column=marker_col)
        hdr_marker.value  = "Changed"
        hdr_marker.font   = Font(bold=True)

        for row_num, marker_text in row_markers.items():
            ws_output.cell(row=row_num, column=marker_col).value = marker_text

        # ── Summary log ──────────────────────────────────────────────────────
        n_added   = sum(1 for m in row_markers.values() if m == "Added")
        n_deleted = len(deleted_rows)
        n_changed = sum(1 for m in row_markers.values() if "Changed" in m or "Smaller" in m)
        self.logger.info(
            f"  '{sheet_name}' done — "
            f"{n_added} added, {n_changed} changed, {n_deleted} deleted rows"
        )

    # ── Main workflow ─────────────────────────────────────────────────────────

    def run(self) -> None:
        start = datetime.now()

        self.logger.info("=" * 60)
        self.logger.info("Price Sheet Comparison — Starting")
        self.logger.info(f"  OLD : {self.old_file}")
        self.logger.info(f"  NEW : {self.new_file}")
        self.logger.info(f"  OUT : {self.output_file}")
        self.logger.info("=" * 60)

        # Step 1 — copy the new file as the output base
        shutil.copy2(self.new_file, self.output_file)
        self.logger.info(f"Output base created: {self.output_file}")

        # Step 2 — load source workbooks (read-only, formula values only)
        wb_old = load_workbook(
            self.old_file, read_only=True, data_only=True, keep_vba=False
        )
        wb_new = load_workbook(
            self.new_file, read_only=True, data_only=True, keep_vba=False
        )

        # Step 3 — load the output copy for editing (keep VBA so .xlsm is valid)
        wb_out = load_workbook(self.output_file, keep_vba=True, data_only=False)

        old_sheets = set(wb_old.sheetnames)

        for sheet_name in wb_new.sheetnames:
            # ── Apply skip rules ──────────────────────────────────────────────
            if self._should_skip_sheet(sheet_name):
                self.logger.info(f"Skipping (rule): '{sheet_name}'")
                continue

            if sheet_name not in old_sheets:
                self.logger.warning(
                    f"'{sheet_name}' not found in OLD file — skipping"
                )
                continue

            if sheet_name not in wb_out.sheetnames:
                self.logger.warning(
                    f"'{sheet_name}' not found in output file — skipping"
                )
                continue

            self.logger.info(f"Processing: '{sheet_name}'")

            rows_old = list(wb_old[sheet_name].iter_rows(values_only=True))
            rows_new = list(wb_new[sheet_name].iter_rows(values_only=True))
            ws_out   = wb_out[sheet_name]

            # Clear any existing auto-filter in this sheet
            ws_out.auto_filter.ref = None

            self._process_sheet(sheet_name, ws_out, rows_old, rows_new)

        # Step 4 — save
        wb_out.save(self.output_file)

        duration = (datetime.now() - start).total_seconds()
        self.logger.info("=" * 60)
        self.logger.info(f"Completed in {duration:.2f}s -> {self.output_file}")
        self.logger.info("=" * 60)

        print(f"\n{'='*60}")
        print("COMPARISON COMPLETE!")
        print(f"  Output  : {self.output_file}")
        print(f"  Duration: {duration:.2f}s")
        print(f"{'='*60}\n")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main() -> None:
    _DEFAULT_OLD = r"C:\Users\szil\Repos\excel_wizadry\Price_Sheet_Excel_Compare\Price_sheet_old.xlsm"
    _DEFAULT_NEW = r"C:\Users\szil\Repos\excel_wizadry\Price_Sheet_Excel_Compare\Price_sheet_new.xlsm"

    parser = argparse.ArgumentParser(
        description="Compare two piping price sheet Excel files (.xlsm)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python price_sheet_compare.py
  python price_sheet_compare.py Price_sheet_old.xlsm Price_sheet_new.xlsm
        """,
    )
    parser.add_argument(
        "old_file", nargs="?", default=_DEFAULT_OLD,
        help="Path to the OLD price sheet (.xlsm)"
    )
    parser.add_argument(
        "new_file", nargs="?", default=_DEFAULT_NEW,
        help="Path to the NEW price sheet (.xlsm)"
    )
    args = parser.parse_args()

    for path, label in [(args.old_file, "Old"), (args.new_file, "New")]:
        if not Path(path).exists():
            print(f"ERROR: {label} file not found: {path}")
            return

    PriceSheetComparator(args.old_file, args.new_file).run()


if __name__ == "__main__":
    main()
