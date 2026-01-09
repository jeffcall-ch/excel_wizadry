"""Convert numeric-looking text cells in Excel files into real numbers.

The script accepts a root folder, walks every subdirectory, and processes each
``.xlsx``/``.xlsm`` workbook it finds. Every sheet is opened via ``pandas`` to
honor the user's requirement, while ``openpyxl`` performs the in-place cell
updates so existing formatting is preserved. Only text cells that represent a
number (with ``.`` as the decimal separator) are touched; their numeric value is
stored and the number format is set to retain the original count of decimals.
"""

from __future__ import annotations

import argparse
import logging
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


NUMBER_PATTERN = re.compile(r"^[+-]?\d+(?:\.(\d+))?$")


@dataclass
class ConversionStats:
	"""Light-weight accumulator for per-file processing metrics."""

	files_seen: int = 0
	files_changed: int = 0
	sheets_seen: int = 0
	cells_converted: int = 0

	def log_file(self, changed: bool, sheet_count: int, converted_cells: int) -> None:
		self.files_seen += 1
		self.sheets_seen += sheet_count
		self.cells_converted += converted_cells
		if changed:
			self.files_changed += 1


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description=(
			"Recursively scan a folder and convert numeric-looking text cells in "
			"Excel workbooks into true numbers while preserving decimal places."
		)
	)
	parser.add_argument(
		"root",
		type=Path,
		help="Folder that will be scanned recursively for .xlsx/.xlsm files.",
	)
	parser.add_argument(
		"--dry-run",
		action="store_true",
		help="Discover candidates without writing changes back to disk.",
	)
	parser.add_argument(
		"--log-level",
		default="INFO",
		help="Python logging level (DEBUG, INFO, WARNING, ...).",
	)
	return parser.parse_args()


def iter_excel_files(root: Path) -> Iterable[Path]:
	"""Yield every .xlsx/.xlsm file under *root* (sub-folders included)."""

	suffixes = {".xlsx", ".xlsm"}
	for path in root.rglob("*"):
		if path.is_file() and path.suffix.lower() in suffixes:
			yield path


def open_with_pandas(file_path: Path) -> list[str]:
	"""Open every sheet via pandas to fulfill the requirement.

	The DataFrames are not retained because ``openpyxl`` provides better control
	over per-cell formatting, but parsing with pandas ensures each sheet is
	actually opened before the conversion phase.
	"""

	sheet_names: list[str] = []
	with pd.ExcelFile(file_path, engine="openpyxl") as xls:
		for sheet in xls.sheet_names:
			logging.debug("Reading sheet '%s' from %s via pandas", sheet, file_path)
			# Load the sheet to guarantee pandas touches the data, then discard.
			_ = xls.parse(sheet_name=sheet, header=None)
			sheet_names.append(sheet)
	return sheet_names


def parse_decimal_string(value: str) -> tuple[Decimal, int] | None:
	"""Return the numeric value and decimal places when *value* is valid."""

	stripped = value.strip()
	match = NUMBER_PATTERN.fullmatch(stripped)
	if not match:
		return None
	fractional_digits = match.group(1) or ""
	try:
		numeric_value = Decimal(stripped)
	except InvalidOperation:
		return None
	return numeric_value, len(fractional_digits)


def decimal_format(decimal_places: int) -> str:
	if decimal_places <= 0:
		return "0"
	return "0." + ("0" * decimal_places)


def convert_sheet(sheet: Worksheet) -> int:
	"""Convert eligible cells inside *sheet* and return how many changed."""

	updated = 0
	for row in sheet.iter_rows():
		for cell in row:
			if not isinstance(cell.value, str):
				continue
			parsed = parse_decimal_string(cell.value)
			if not parsed:
				continue
			numeric_value, decimals = parsed
			cell.value = numeric_value
			cell.number_format = decimal_format(decimals)
			updated += 1
	return updated


def handle_workbook(file_path: Path, dry_run: bool) -> tuple[int, int]:
	"""Process a single workbook and return (sheet_count, converted_cells)."""

	try:
		sheet_names = open_with_pandas(file_path)
	except Exception as exc:  # noqa: BLE001 - surface to operator
		logging.error("Failed to open %s via pandas: %s", file_path, exc)
		return 0, 0

	keep_vba = file_path.suffix.lower() == ".xlsm"
	try:
		workbook = load_workbook(file_path, keep_vba=keep_vba)
	except Exception as exc:  # noqa: BLE001
		logging.error("Failed to open %s via openpyxl: %s", file_path, exc)
		return 0, 0

	converted = 0
	for sheet_name in sheet_names:
		if sheet_name not in workbook.sheetnames:
			logging.warning(
				"Sheet '%s' present in pandas view but missing in workbook %s",
				sheet_name,
				file_path,
			)
			continue
		conversions = convert_sheet(workbook[sheet_name])
		logging.debug(
			"Converted %s cells on sheet '%s' in %s", conversions, sheet_name, file_path
		)
		converted += conversions

	if converted and not dry_run:
		workbook.save(file_path)
		logging.info("Saved %s (converted %s cells)", file_path, converted)
	elif converted:
		logging.info("Dry run: %s would change %s cells", file_path, converted)
	else:
		logging.info("No changes needed for %s", file_path)

	return len(sheet_names), converted


def run(root: Path, dry_run: bool) -> ConversionStats:
	stats = ConversionStats()
	for workbook_path in iter_excel_files(root):
		logging.info("Processing %s", workbook_path)
		sheets_seen, converted = handle_workbook(workbook_path, dry_run=dry_run)
		stats.log_file(changed=converted > 0, sheet_count=sheets_seen, converted_cells=converted)
	return stats


def main() -> None:
	args = parse_args()
	logging.basicConfig(
		level=getattr(logging, args.log_level.upper(), logging.INFO),
		format="%(levelname)s: %(message)s",
	)
	root = args.root.expanduser().resolve()
	if not root.exists():
		raise SystemExit(f"Folder does not exist: {root}")

	stats = run(root, dry_run=args.dry_run)
	logging.info(
		"Finished. Files: %s (changed %s) | Sheets: %s | Cells converted: %s",
		stats.files_seen,
		stats.files_changed,
		stats.sheets_seen,
		stats.cells_converted,
	)


if __name__ == "__main__":
	main()

