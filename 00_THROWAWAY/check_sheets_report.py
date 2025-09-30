"""Scan all .xlsx and .xlsm files in this folder and report sheets with 200+ rows

Produces a CSV file `sheets_report.csv` with columns:
filename,sheet_name,has_DN1,has_Material,has_Total_weight_kg,all_present

This script uses openpyxl (no macros executed) and does a case-insensitive header check.
"""
from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import List

try:
    import openpyxl
except Exception as e:
    raise SystemExit(
        "openpyxl is required. Install with: pip install openpyxl" "\n" + str(e)
    )


TARGET_COLS = ["dn1", "material", "total weight [kg]"]
OUT_CSV = "sheets_report.csv"


def check_sheet(ws) -> List[bool]:
    """Return booleans for presence of TARGET_COLS in header row (case-insensitive).

    The function looks at the first non-empty row as header.
    """
    header = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row and any(cell is not None and str(cell).strip() != "" for cell in row):
            header = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            break
    if header is None:
        return [False] * len(TARGET_COLS)

    result = []
    for target in TARGET_COLS:
        found = any(target == col for col in header)
        # also allow header that contains the target as substring (e.g., 'Total weight [kg] (per m)')
        if not found:
            found = any(target in col for col in header)
        result.append(bool(found))
    return result


def scan_folder(folder: Path) -> List[tuple]:
    rows = []
    for p in sorted(folder.iterdir()):
        if not p.is_file():
            continue
        if p.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception as e:
            rows.append((p.name, "<workbook-error>", False, False, False, False))
            continue

        found_boq = False
        for sheetname in wb.sheetnames:
            # Only consider sheets with 'BOQ' in the sheet name (case-insensitive)
            if "boq" not in sheetname.lower():
                continue
            found_boq = True
            ws = wb[sheetname]
            try:
                max_row = int(ws.max_row or 0)
            except Exception:
                max_row = 0

            has_dn1, has_material, has_total = check_sheet(ws)
            all_present = has_dn1 and has_material and has_total
            # Include BOQ sheets even if they have fewer than 200 rows; report their max_row
            rows.append((p.name, sheetname, max_row, has_dn1, has_material, has_total, all_present))

        if not found_boq:
            # Record that the workbook had no BOQ-named sheets
            rows.append((p.name, "<no-BOQ-sheet>", 0, False, False, False, False))

        wb.close()
    return rows


def write_csv(folder: Path, rows: List[tuple]):
    out = folder / OUT_CSV
    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "filename",
            "sheet_name",
            "max_row",
            "has_DN1",
            "has_Material",
            "has_Total_weight_kg",
            "all_present",
        ])
        for r in rows:
            writer.writerow(r)


def main():
    folder = Path(__file__).resolve().parent
    rows = scan_folder(folder)
    write_csv(folder, rows)
    print(f"Wrote report to: {folder / OUT_CSV} (rows: {len(rows)})")


if __name__ == "__main__":
    main()
