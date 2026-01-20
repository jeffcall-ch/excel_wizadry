"""
Script to flag rows containing empty cells in an Excel file.
Adds a new column before column A with header "Contains empty cells".
Marks rows with "X" if they contain any empty cells.
Colors empty cells with yellow background.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path


def flag_empty_cells(input_file):
    """
    Process Excel file and add a column to flag rows with empty cells.
    
    Args:
        input_file: Path to the input Excel file
    """
    # Load the data with pandas
    print(f"Loading workbook: {input_file}")
    df = pd.read_excel(input_file)
    
    print(f"Processing {len(df)} rows and {len(df.columns)} columns...")
    
    # Columns to exclude from empty cell checking
    exclude_columns = ["WT [mm]", "Pos. on PID"]
    
    # Check for empty cells in each row (NaN or empty string)
    # Create a mask for empty cells
    empty_mask = df.isna() | (df.astype(str).map(lambda x: x.strip() == ''))
    
    # Exclude specified columns from the mask
    for col in exclude_columns:
        if col in empty_mask.columns:
            empty_mask[col] = False
    
    # Handle "Insulation Type" conditionally based on "Insulation Thickness [mm]"
    if "Insulation Type" in df.columns and "Insulation Thickness [mm]" in df.columns:
        # Only flag "Insulation Type" as empty if "Insulation Thickness [mm]" > 0
        insulation_thickness = pd.to_numeric(df["Insulation Thickness [mm]"], errors='coerce')
        # Set to False where thickness is 0 or NaN (meaning we don't check Insulation Type)
        empty_mask.loc[insulation_thickness.isna() | (insulation_thickness <= 0), "Insulation Type"] = False
    
    # Add new column at the beginning to flag rows with empty cells
    df.insert(0, 'Contains empty cells', '')
    
    # Mark rows that have any empty cells with "X"
    has_empty_cells = empty_mask.any(axis=1)
    df.loc[has_empty_cells, 'Contains empty cells'] = 'X'
    
    # Generate output filename
    input_path = Path(input_file)
    output_file = input_path.parent / f"{input_path.stem}_empty_cells{input_path.suffix}"
    
    # Save the dataframe to Excel
    print(f"Saving to: {output_file}")
    df.to_excel(output_file, index=False)
    
    # Now apply yellow background to empty cells using openpyxl
    print("Applying yellow background to empty cells...")
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # Define yellow fill for empty cells
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Apply yellow fill to empty cells (starting from row 2, column 2)
    # Column 1 is the new flag column, so we start from column 2
    for row_idx in range(len(df)):
        for col_idx, col_name in enumerate(df.columns[1:], start=1):  # Skip first column
            if empty_mask.iloc[row_idx, col_idx - 1]:  # col_idx-1 because empty_mask doesn't have the new column
                cell = ws.cell(row=row_idx + 2, column=col_idx + 1)  # +2 for header row and 1-based, +1 for inserted column
                cell.fill = yellow_fill
    
    # Apply yellow fill to empty cells (starting from row 2, column 2)
    # Column 1 is the new flag column, so we start from column 2
    for row_idx in range(len(df)):
        for col_idx, col_name in enumerate(df.columns[1:], start=1):  # Skip first column
            if empty_mask.iloc[row_idx, col_idx - 1]:  # col_idx-1 because empty_mask doesn't have the new column
                cell = ws.cell(row=row_idx + 2, column=col_idx + 1)  # +2 for header row and 1-based, +1 for inserted column
                cell.fill = yellow_fill
    
    # Save the workbook with formatting
    wb.save(output_file)
    print("Done!")
    
    return output_file


if __name__ == "__main__":
    # Input file path
    input_file = r"C:\Users\szil\Repos\excel_wizadry\00_THROWAWAY\flag_empty_cells\TBY_Pipe_List_15.01.2026_.xlsx"
    
    # Process the file
    output_file = flag_empty_cells(input_file)
    print(f"\nOutput saved to: {output_file}")

