"""
Reformat accommodations_results_150.xlsx for easy reading on small Android screens.
- Human-friendly column headers
- 0/1 boolean columns -> ✓ / ✗
- Direct booking URL -> clickable "Book ↗" hyperlink
- Frozen header row + auto-filter
- Alternating row colors, bold header, sensible column widths
- Hide internal/redundant columns (validation_method, source_label, timestamp_checked)
- Summary sheet lightly styled too
"""

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

INPUT  = "00_THROWAWAY/reformat/accommodations_results_150.xlsx"
OUTPUT = "00_THROWAWAY/reformat/accommodations_results_150_REFORMATTED.xlsx"

# ── colours ─────────────────────────────────────────────────────────────────
HEADER_BG   = "1F4E79"   # dark-navy
HEADER_FG   = "FFFFFF"
ROW_ODD     = "FFFFFF"
ROW_EVEN    = "D6E4F0"   # light-blue
LINK_COLOR  = "1155CC"
SCORE_HIGH  = "C6EFCE"   # green
SCORE_MID   = "FFEB9C"   # yellow
SCORE_LOW   = "FFC7CE"   # red

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

# ── column spec for "Verified Results" ──────────────────────────────────────
# (original_header, new_header, width, horizontal_align, hidden)
COL_SPEC = [
    ("property_id",           "#",            4,   "center", True),
    ("location_name",         "Location",    12,   "left",   False),
    ("country",               "Country",      9,   "left",   False),
    ("property_name",         "Property",    24,   "left",   False),
    ("price_eur",             "Price (€)",    9,   "center", False),
    ("rating_aggregate",      "Rating",       7,   "center", False),
    ("detour_minutes",        "Detour\n(min)",9,   "center", False),
    ("time_to_coast_minutes", "Coast\n(min)", 8,   "center", False),
    ("suitability_score",     "Score",        8,   "center", False),
    ("coast_type",            "Coast Type",  16,   "left",   False),
    ("has_valid_kitchen",     "Kitchen",      8,   "center", False),
    ("has_valid_parking",     "Parking",      8,   "center", False),
    ("has_valid_outdoor",     "Outdoor",      8,   "center", False),
    ("source_family",         "Source",      14,   "left",   False),
    ("source_label",          "Src Label",   18,   "left",   True),
    ("validation_method",     "Val. Method", 20,   "left",   True),
    ("direct_booking_link",   "Link",         8,   "center", False),
    ("timestamp_checked",     "Checked",     18,   "left",   True),
]

BOOLEAN_COLS  = {"has_valid_kitchen", "has_valid_parking", "has_valid_outdoor"}
LINK_COL      = "direct_booking_link"

# ── helpers ──────────────────────────────────────────────────────────────────
def score_fill(value):
    if value is None:
        return None
    if value >= 140:
        return fill(SCORE_HIGH)
    if value >= 120:
        return fill(SCORE_MID)
    return fill(SCORE_LOW)

# ── main ─────────────────────────────────────────────────────────────────────
wb = load_workbook(INPUT)
ws = wb["Verified Results"]

# Build mapping: original_header -> column letter in the sheet
header_to_col = {}
for cell in ws[1]:
    if cell.value:
        header_to_col[cell.value] = cell.column_letter

# ── 1. Style header row ──────────────────────────────────────────────────────
header_font   = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
header_fill   = fill(HEADER_BG)
header_align  = Alignment(horizontal="center", vertical="center",
                           wrap_text=True)

for orig_header, new_header, col_width, h_align, hidden in COL_SPEC:
    col_letter = header_to_col.get(orig_header)
    if not col_letter:
        continue
    cell = ws[f"{col_letter}1"]
    cell.value = new_header
    cell.font  = header_font
    cell.fill  = header_fill
    cell.alignment = header_align
    cell.border = thin_border()
    ws.column_dimensions[col_letter].width  = col_width
    ws.column_dimensions[col_letter].hidden = hidden

ws.row_dimensions[1].height = 36

# ── 2. Style data rows ───────────────────────────────────────────────────────
score_col_letter = header_to_col.get("suitability_score")
link_col_letter  = header_to_col.get(LINK_COL)

link_font    = Font(name="Calibri", size=11, color=LINK_COLOR, underline="single")
default_font = Font(name="Calibri", size=11)
bool_font    = Font(name="Calibri", size=11)

for row_idx in range(2, ws.max_row + 1):
    row_fill  = fill(ROW_ODD) if row_idx % 2 == 1 else fill(ROW_EVEN)
    row_height = 22

    for orig_header, new_header, col_width, h_align, hidden in COL_SPEC:
        col_letter = header_to_col.get(orig_header)
        if not col_letter:
            continue
        cell = ws[f"{col_letter}{row_idx}"]
        raw  = cell.value

        # Boolean columns -> ✓ / ✗
        if orig_header in BOOLEAN_COLS:
            cell.value = "✓" if raw == 1 else ("✗" if raw == 0 else raw)
            cell.font  = Font(name="Calibri", size=12,
                              color=("2D6A4F" if cell.value == "✓" else "C0392B"))
        # Link column -> short clickable text
        elif orig_header == LINK_COL:
            url = str(raw) if raw else ""
            if url.startswith("http"):
                cell.value     = "Book ↗"
                cell.hyperlink = url
                cell.font      = link_font
            else:
                cell.font = default_font
        # Score column -> conditional colour fill override
        elif orig_header == "suitability_score" and raw is not None:
            cell.fill = score_fill(raw) or row_fill
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal=h_align, vertical="center")
            cell.border    = thin_border()
            continue
        else:
            cell.font = default_font

        # Apply common styles (skip score col handled above)
        if orig_header != "suitability_score":
            cell.fill = row_fill
        cell.alignment = Alignment(horizontal=h_align, vertical="center",
                                   wrap_text=(orig_header == "property_name"))
        cell.border    = thin_border()

    ws.row_dimensions[row_idx].height = row_height

# ── 3. Freeze header + auto-filter ──────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ── 4. Style Summary sheet ───────────────────────────────────────────────────
ws2 = wb["Summary"]
ws2["A1"].font  = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
ws2["B1"].font  = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
ws2["A1"].fill  = fill(HEADER_BG)
ws2["B1"].fill  = fill(HEADER_BG)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2["B1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

for r in range(2, ws2.max_row + 1):
    row_fill = fill(ROW_ODD) if r % 2 == 1 else fill(ROW_EVEN)
    for col in ["A", "B"]:
        c = ws2[f"{col}{r}"]
        c.fill = row_fill
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(horizontal=("left" if col == "A" else "center"),
                                vertical="center")
        c.border = thin_border()
    ws2.row_dimensions[r].height = 20

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 12

# ── 5. Save ──────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
