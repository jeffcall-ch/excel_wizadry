from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

DEFAULT_MISSING_REPORT = "KKS_not_in_batch_report.xlsx"
DEFAULT_BATCHED_FILE = (
    "CA100-KVI-50296373_2.0 - BoM General Hangers and support material "
    "LS_erection_split_batched.xlsx"
)
DEFAULT_ORIGINAL_FILE = (
    "CA100-KVI-50296373_2.0 - BoM General Hangers and support material "
    "ORIGINAL DATA .xlsm"
)
DEFAULT_OUTPUT_FILE = (
    "CA100-KVI-50296373_2.0 - BoM General Hangers and support material "
    "LS_erection_split_batched_with_missing_integrated.xlsx"
)

MISSING_SHEET = "Missing_Final"
BATCHED_SHEET = "Detail"
ORIGINAL_SHEET = "Per_Support_Structure"
DEFAULT_SECTION_ORDER = ["UHA", "UMA", "N/A"]
DEFAULT_NA_WORKING_AREAS = {"NO_AREA", "BQ NOT EXIST"}
DEFAULT_NA_SECTION_LABEL = "N/A"

SECTION_ORDER: list[str] = DEFAULT_SECTION_ORDER.copy()
SECTION_ORDER_MAP: dict[str, int] = {
    section: index for index, section in enumerate(SECTION_ORDER)
}
NA_WORKING_AREAS: set[str] = DEFAULT_NA_WORKING_AREAS.copy()
NA_SECTION_LABEL = DEFAULT_NA_SECTION_LABEL


def normalize_kks(value: object) -> str | None:
    if pd.isna(value):
        return None

    text = str(value).strip().upper().replace(" ", "")
    if not text or text == "NAN":
        return None

    if "/SU" in text and not text.startswith("/"):
        text = f"/{text}"

    return text


def normalize_label(value: object) -> str:
    text = str(value).replace("\n", " ").strip().lower()
    return " ".join(text.split())


def parse_csv_upper(value: str) -> list[str]:
    values = [part.strip().upper() for part in str(value).split(",") if part.strip()]
    deduped: list[str] = []
    seen: set[str] = set()
    for item in values:
        if item not in seen:
            seen.add(item)
            deduped.append(item)
    return deduped


def parse_bool(value: str) -> bool:
    normalized = str(value).strip().lower()
    return normalized in {"1", "true", "yes", "y", "on"}


def set_section_order(section_order: list[str]) -> None:
    global SECTION_ORDER
    global SECTION_ORDER_MAP

    SECTION_ORDER = section_order if section_order else DEFAULT_SECTION_ORDER.copy()
    SECTION_ORDER_MAP = {section: index for index, section in enumerate(SECTION_ORDER)}


def set_na_handling(na_working_areas: list[str], na_section_label: str) -> None:
    global NA_WORKING_AREAS
    global NA_SECTION_LABEL

    NA_WORKING_AREAS = set(na_working_areas) if na_working_areas else DEFAULT_NA_WORKING_AREAS.copy()
    NA_SECTION_LABEL = str(na_section_label).strip() or DEFAULT_NA_SECTION_LABEL


def natural_sort_token(value: object) -> str:
    text = "" if pd.isna(value) else str(value).strip().upper()
    return re.sub(r"\d+", lambda m: f"{int(m.group()):06d}", text)


def section_sort_token(value: object) -> str:
    section = "" if pd.isna(value) else str(value).strip().upper()
    rank = SECTION_ORDER_MAP.get(section)
    if rank is not None:
        return f"{rank:03d}_{section}"
    return f"900_{section}"


def derive_building_section(working_area: str) -> str:
    if str(working_area).strip().upper() in NA_WORKING_AREAS:
        return NA_SECTION_LABEL

    match = re.search(r"U[A-Z]{2}", str(working_area).upper())
    return match.group(0) if match else "UNKNOWN"


def find_column(columns: list[str], required_terms: list[str]) -> str | None:
    for col in columns:
        label = normalize_label(col)
        if all(term in label for term in required_terms):
            return col
    return None


def find_header_row(workbook: Path, sheet_name: str) -> int | None:
    wb = load_workbook(workbook, data_only=True)
    ws = wb[sheet_name]

    for row_idx in range(1, 250):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 60)]
        labels = [str(v).strip() if v is not None else "" for v in values]
        if "Support Structure" in labels and "Total Weight [kg]" in labels:
            return row_idx

    return None


def load_missing_candidates(
    missing_report: Path,
    exclude_deleted: bool,
    exclude_bq_not_exist: bool,
    exclude_solid_not_gp: bool,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, int]]:
    missing = pd.read_excel(missing_report, sheet_name=MISSING_SHEET, engine="openpyxl")

    mask_deleted = pd.Series(False, index=missing.index)
    if "Deleted in status" in missing.columns:
        mask_deleted = missing["Deleted in status"].fillna(False).astype(bool)

    mask_bq_not_exist = pd.Series(False, index=missing.index)
    mask_solid_not_gp = pd.Series(False, index=missing.index)

    for col in missing.columns:
        as_text = missing[col].astype(str).str.lower()
        mask_bq_not_exist |= as_text.str.contains("bq not exist", na=False)
        mask_solid_not_gp |= as_text.str.contains(
            "solid handling support not in gp lot", na=False
        )

    exclude_mask = pd.Series(False, index=missing.index)
    if exclude_deleted:
        exclude_mask |= mask_deleted
    if exclude_bq_not_exist:
        exclude_mask |= mask_bq_not_exist
    if exclude_solid_not_gp:
        exclude_mask |= mask_solid_not_gp

    excluded = missing[exclude_mask].copy()
    reason_list: list[str] = []
    for idx in missing.index:
        reasons: list[str] = []
        if exclude_deleted and bool(mask_deleted.loc[idx]):
            reasons.append("Deleted")
        if exclude_bq_not_exist and bool(mask_bq_not_exist.loc[idx]):
            reasons.append("BQ not exist")
        if exclude_solid_not_gp and bool(mask_solid_not_gp.loc[idx]):
            reasons.append("Solid handling not in GP")
        reason_list.append("; ".join(reasons))
    missing["Exclusion reason"] = reason_list
    excluded = missing[exclude_mask].copy()

    included = missing[~exclude_mask].copy()

    included["KKS /SU"] = included["KKS /SU"].map(normalize_kks)
    included = included[included["KKS /SU"].notna()].copy()

    included["Working area"] = included["Working area"].fillna("").astype(str).str.strip()
    included.loc[included["Working area"].isin(["", "nan", "NaN"]), "Working area"] = "NO_AREA"
    included["Building section"] = included["Working area"].map(derive_building_section)

    stats = {
        "missing_total": int(len(missing)),
        "excluded_total": int(exclude_mask.sum()),
        "excluded_deleted": int((mask_deleted & exclude_mask).sum()),
        "excluded_bq_not_exist": int((mask_bq_not_exist & exclude_mask).sum()),
        "excluded_solid_not_gp": int((mask_solid_not_gp & exclude_mask).sum()),
        "included_bq_not_exist_as_na": int((mask_bq_not_exist & ~exclude_mask).sum()),
        "included_deleted": int((mask_deleted & ~exclude_mask).sum()),
        "included_solid_not_gp": int((mask_solid_not_gp & ~exclude_mask).sum()),
        "included_for_integration": int(len(included)),
    }

    return included, excluded, stats


def load_batched_detail(batched_file: Path) -> tuple[pd.DataFrame, dict[str, int], dict[int, float], dict[str, set[int]]]:
    detail = pd.read_excel(batched_file, sheet_name=BATCHED_SHEET, engine="openpyxl")

    required = ["KKS code", "Working area", "KKS Total Weight [kg]", "Building section", "Batch number"]
    missing_cols = [col for col in required if col not in detail.columns]
    if missing_cols:
        raise ValueError(f"Batched detail missing required columns: {missing_cols}")

    detail = detail[required].copy()
    detail = detail[detail["KKS code"].notna() & detail["Batch number"].notna()].copy()
    detail["Batch number"] = pd.to_numeric(detail["Batch number"], errors="coerce").fillna(0).astype(int)
    detail = detail[detail["Batch number"] > 0].copy()

    detail["KKS code"] = detail["KKS code"].map(normalize_kks)
    detail = detail[detail["KKS code"].notna()].copy()

    detail["Working area"] = detail["Working area"].fillna("").astype(str).str.strip()
    detail["Building section"] = detail["Building section"].fillna("").astype(str).str.strip()
    detail.loc[detail["Building section"] == "", "Building section"] = detail["Working area"].map(
        derive_building_section
    )

    detail["KKS Total Weight [kg]"] = pd.to_numeric(
        detail["KKS Total Weight [kg]"], errors="coerce"
    )

    working_area_to_batch: dict[str, int] = {}
    for _, row in (
        detail.drop_duplicates(subset=["Working area", "Batch number"])
        .sort_values(["Working area", "Batch number"])
        .iterrows()
    ):
        wa = str(row["Working area"])
        if wa not in working_area_to_batch:
            working_area_to_batch[wa] = int(row["Batch number"])

    batch_weights = (
        detail.assign(_w=detail["KKS Total Weight [kg]"].fillna(0.0))
        .groupby("Batch number", as_index=True)["_w"]
        .sum()
        .to_dict()
    )

    section_batches: dict[str, set[int]] = {}
    for section, grp in detail.groupby("Building section"):
        section_batches[str(section)] = set(grp["Batch number"].astype(int).tolist())

    return detail, working_area_to_batch, batch_weights, section_batches


def load_weight_lookup(original_file: Path) -> dict[str, float]:
    header_row = find_header_row(original_file, ORIGINAL_SHEET)
    if not header_row:
        return {}

    src = pd.read_excel(
        original_file,
        sheet_name=ORIGINAL_SHEET,
        engine="openpyxl",
        header=header_row - 1,
    )

    if "Support Structure" not in src.columns or "Total Weight [kg]" not in src.columns:
        return {}

    src = src[["Support Structure", "Total Weight [kg]"]].copy()
    src["Support Structure"] = src["Support Structure"].map(normalize_kks)
    src = src[src["Support Structure"].notna()].copy()
    src = src[src["Support Structure"].str.contains("/SU", na=False)].copy()

    src["Total Weight [kg]"] = pd.to_numeric(src["Total Weight [kg]"], errors="coerce")

    lookup = (
        src.groupby("Support Structure", as_index=False)["Total Weight [kg]"]
        .sum()
        .set_index("Support Structure")["Total Weight [kg]"]
        .to_dict()
    )
    return lookup


def pick_batch(
    working_area: str,
    building_section: str,
    wa_to_batch: dict[str, int],
    section_batches: dict[str, set[int]],
    running_batch_weights: dict[int, float],
) -> tuple[int, str]:
    if working_area in wa_to_batch:
        return wa_to_batch[working_area], "Mapped by working area"

    section = str(building_section).strip()
    if section.upper() == NA_SECTION_LABEL.upper():
        batch = min(running_batch_weights, key=lambda b: (running_batch_weights[b], b))
        return int(batch), f"{NA_SECTION_LABEL} fallback to lightest batch"

    section_set = section_batches.get(section, set())
    if section_set:
        batch = min(section_set, key=lambda b: (running_batch_weights.get(b, 0.0), b))
        return int(batch), "Fallback by building section"

    batch = min(running_batch_weights, key=lambda b: (running_batch_weights[b], b))
    return int(batch), "Fallback to overall lightest batch"


def integrate_missing(
    included_missing: pd.DataFrame,
    existing_detail: pd.DataFrame,
    wa_to_batch: dict[str, int],
    batch_weights: dict[int, float],
    section_batches: dict[str, set[int]],
    weight_lookup: dict[str, float],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, int]]:
    integrated_rows: list[dict[str, object]] = []

    running_weights = {int(k): float(v) for k, v in batch_weights.items()}

    for _, row in included_missing.iterrows():
        kks = normalize_kks(row.get("KKS /SU"))
        if not kks:
            continue

        if kks in set(existing_detail["KKS code"].dropna().tolist()):
            continue

        working_area = str(row.get("Working area", "")).strip() or "NO_AREA"
        building_section = str(row.get("Building section", "")).strip() or derive_building_section(
            working_area
        )

        batch_number, assignment_note = pick_batch(
            working_area=working_area,
            building_section=building_section,
            wa_to_batch=wa_to_batch,
            section_batches=section_batches,
            running_batch_weights=running_weights,
        )

        weight = weight_lookup.get(kks)
        weight_found = pd.notna(weight)
        weight_value = float(weight) if weight_found else None

        if weight_found:
            running_weights[batch_number] = running_weights.get(batch_number, 0.0) + weight_value

        integrated_rows.append(
            {
                "KKS code": kks,
                "Working area": working_area,
                "KKS Total Weight [kg]": weight_value,
                "Building section": building_section,
                "Batch number": batch_number,
                "Source": "Integrated from Missing_Final",
                "Weight found in ORIGINAL DATA": bool(weight_found),
                "Assignment note": assignment_note,
                "Status": row.get("Status"),
                "Note": row.get("Note"),
                "System": row.get("System"),
                "AIC": row.get("AIC"),
            }
        )

    integrated = pd.DataFrame(integrated_rows)

    existing_out = existing_detail.copy()
    existing_out["Source"] = "Existing batch"
    existing_out["Weight found in ORIGINAL DATA"] = existing_out["KKS Total Weight [kg]"].notna()
    existing_out["Assignment note"] = "Original assignment"
    for col in ["Status", "Note", "System", "AIC"]:
        existing_out[col] = ""

    updated_detail = pd.concat([existing_out, integrated], ignore_index=True, sort=False)

    updated_detail["_section_sort"] = updated_detail["Building section"].map(section_sort_token)
    updated_detail["_wa_sort"] = updated_detail["Working area"].map(natural_sort_token)
    updated_detail = updated_detail.sort_values(
        ["Batch number", "_section_sort", "_wa_sort", "KKS code"],
        ascending=[True, True, True, True],
    ).drop(columns=["_section_sort", "_wa_sort"]) 
    updated_detail = updated_detail.reset_index(drop=True)

    stats = {
        "integrated_rows": int(len(integrated)),
        "integrated_weight_found": int(integrated["Weight found in ORIGINAL DATA"].sum())
        if not integrated.empty
        else 0,
        "integrated_weight_missing": int((~integrated["Weight found in ORIGINAL DATA"]).sum())
        if not integrated.empty
        else 0,
    }

    return integrated, updated_detail, stats


def build_updated_summary(updated_detail: pd.DataFrame) -> pd.DataFrame:
    temp = updated_detail.copy()
    temp["KKS Total Weight [kg]"] = pd.to_numeric(temp["KKS Total Weight [kg]"], errors="coerce")

    summary = (
        temp.assign(_w=temp["KKS Total Weight [kg]"].fillna(0.0))
        .groupby("Batch number", as_index=False)
        .agg(
            Total_Weight_kg=("_w", "sum"),
            Known_Weight_KKS_Count=("KKS Total Weight [kg]", lambda s: int(s.notna().sum())),
            Missing_Weight_KKS_Count=("KKS Total Weight [kg]", lambda s: int(s.isna().sum())),
            Total_KKS_Count=("KKS code", "nunique"),
            Building_Sections=(
                "Building section",
                lambda s: ", ".join(sorted(set([str(x) for x in s if str(x).strip()]), key=section_sort_token)),
            ),
        )
        .sort_values("Batch number")
        .reset_index(drop=True)
    )

    return summary


def write_output(
    output_file: Path,
    updated_detail: pd.DataFrame,
    updated_summary: pd.DataFrame,
    integrated: pd.DataFrame,
    excluded: pd.DataFrame,
    overview: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Overview", index=False)
        updated_summary.to_excel(writer, sheet_name="Summary_Updated", index=False)
        updated_detail.to_excel(writer, sheet_name="Detail_Updated", index=False)
        integrated.to_excel(writer, sheet_name="Integrated_Additions", index=False)
        excluded.to_excel(writer, sheet_name="Excluded_From_Missing", index=False)


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(
        description=(
            "Integrate in-scope missing KKS into existing batches while excluding "
            "configurable categories and assigning outdoor areas to N/A."
        )
    )
    parser.add_argument(
        "--missing-report",
        type=Path,
        default=script_dir / DEFAULT_MISSING_REPORT,
        help="Path to KKS_not_in_batch_report.xlsx",
    )
    parser.add_argument(
        "--batched-file",
        type=Path,
        default=script_dir / DEFAULT_BATCHED_FILE,
        help="Path to existing batched workbook",
    )
    parser.add_argument(
        "--original-file",
        type=Path,
        default=script_dir / DEFAULT_ORIGINAL_FILE,
        help="Path to ORIGINAL DATA workbook used for weight lookup",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / DEFAULT_OUTPUT_FILE,
        help="Output workbook path",
    )
    parser.add_argument(
        "--section-order",
        default=",".join(DEFAULT_SECTION_ORDER),
        help=(
            "Comma-separated section priority for sorting, for example "
            "UHA,UMA,N/A"
        ),
    )
    parser.add_argument(
        "--na-working-areas",
        default=",".join(sorted(DEFAULT_NA_WORKING_AREAS)),
        help=(
            "Comma-separated working-area aliases that should map to N/A, "
            "for example NO_AREA,BQ NOT EXIST"
        ),
    )
    parser.add_argument(
        "--na-section-label",
        default=DEFAULT_NA_SECTION_LABEL,
        help="Label used for outdoor/unassigned building section, for example N/A",
    )
    parser.add_argument(
        "--exclude-deleted",
        default="true",
        help="Whether to exclude deleted rows from integration (true/false)",
    )
    parser.add_argument(
        "--exclude-bq-not-exist",
        default="false",
        help="Whether to exclude BQ not exist rows from integration (true/false)",
    )
    parser.add_argument(
        "--exclude-solid-not-gp",
        default="true",
        help="Whether to exclude solid-handling-not-in-GP rows (true/false)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    section_order = parse_csv_upper(args.section_order)
    na_working_areas = parse_csv_upper(args.na_working_areas)
    set_section_order(section_order)
    set_na_handling(na_working_areas, args.na_section_label)

    exclude_deleted = parse_bool(args.exclude_deleted)
    exclude_bq_not_exist = parse_bool(args.exclude_bq_not_exist)
    exclude_solid_not_gp = parse_bool(args.exclude_solid_not_gp)

    missing_report = args.missing_report.resolve()
    batched_file = args.batched_file.resolve()
    original_file = args.original_file.resolve()
    output_file = args.output.resolve()

    for file_path in [missing_report, batched_file, original_file]:
        if not file_path.exists():
            raise FileNotFoundError(f"Input file not found: {file_path}")

    included_missing, excluded_missing, stats_missing = load_missing_candidates(
        missing_report=missing_report,
        exclude_deleted=exclude_deleted,
        exclude_bq_not_exist=exclude_bq_not_exist,
        exclude_solid_not_gp=exclude_solid_not_gp,
    )
    existing_detail, wa_to_batch, batch_weights, section_batches = load_batched_detail(batched_file)
    weight_lookup = load_weight_lookup(original_file)

    integrated, updated_detail, stats_integrated = integrate_missing(
        included_missing=included_missing,
        existing_detail=existing_detail,
        wa_to_batch=wa_to_batch,
        batch_weights=batch_weights,
        section_batches=section_batches,
        weight_lookup=weight_lookup,
    )

    updated_summary = build_updated_summary(updated_detail)

    overview_rows = [
        ("Missing_Final rows", stats_missing["missing_total"]),
        ("Excluded total", stats_missing["excluded_total"]),
        ("Excluded: deleted", stats_missing["excluded_deleted"]),
        ("Excluded: BQ not exist", stats_missing["excluded_bq_not_exist"]),
        ("Excluded: Solid handling not in GP", stats_missing["excluded_solid_not_gp"]),
        ("Included: deleted", stats_missing["included_deleted"]),
        ("Included: BQ not exist as N/A", stats_missing["included_bq_not_exist_as_na"]),
        ("Included: Solid handling not in GP", stats_missing["included_solid_not_gp"]),
        ("Included for integration", stats_missing["included_for_integration"]),
        ("Integrated rows added", stats_integrated["integrated_rows"]),
        ("Integrated with weight found", stats_integrated["integrated_weight_found"]),
        ("Integrated with weight missing", stats_integrated["integrated_weight_missing"]),
        ("Final Detail_Updated total rows", len(updated_detail)),
    ]
    overview = pd.DataFrame(overview_rows, columns=["Metric", "Value"])

    write_output(
        output_file=output_file,
        updated_detail=updated_detail,
        updated_summary=updated_summary,
        integrated=integrated,
        excluded=excluded_missing,
        overview=overview,
    )

    print(f"Missing report: {missing_report}")
    print(f"Batched input: {batched_file}")
    print(f"Original data: {original_file}")
    print(f"Output: {output_file}")
    print(overview.to_string(index=False))


if __name__ == "__main__":
    main()
