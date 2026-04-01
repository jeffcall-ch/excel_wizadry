#!/usr/bin/env python3
"""
Compare two Sikla multi-sheet Excel exports and generate a color-coded difference report.

Default behavior:
- Scans a folder for exactly two .xlsx files.
- Applies sheet-specific key and quantity comparison logic.
- Creates an output workbook with summary + per-sheet difference tabs.
- Adds a scope analysis hint based on Support List differences.

Usage examples:
  python sikla_sheet_compare.py --folder .
  python sikla_sheet_compare.py --folder . --output compared.xlsx
  python sikla_sheet_compare.py --file-a "A.xlsx" --file-b "B.xlsx"
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


STATUS_ONLY_A = "Only in BQandSU"
STATUS_ONLY_B = "Only in SU"
STATUS_MISMATCH = "Mismatch"
STATUS_MATCH = "Match"

FILL_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FILL_ONLY_A = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_ONLY_B = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_MISMATCH = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_MATCH = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


@dataclass
class SheetRule:
    name: str
    key_columns: List[str]
    qty_column: Optional[str] = None


@dataclass
class AggregateEntry:
    row_count: int = 0
    qty_sum: float = 0.0
    sample_a: Optional[Dict[str, Any]] = None
    sample_b: Optional[Dict[str, Any]] = None


SHEET_RULES: List[SheetRule] = [
    SheetRule("Total Qty", ["Item Number"], "Qty"),
    SheetRule("Per Support Structure", ["Support Structure", "Item Number", "Cut Length [mm]"], "Qty"),
    SheetRule(
        "Per Pipe Support",
        ["Support Structure", "Pipe Support", "Item Number", "Cut Length [mm]", "Branch", "Pipe", "Nominal Bore"],
        "Qty",
    ),
    SheetRule("Cut List", ["Item Number", "Cut Length [mm]"], "Qty"),
    SheetRule("Total Cut Lengths", ["Item Number"], "Total Cut Length [mm]"),
    SheetRule("Support List", ["Support Structure"], "Total Weight [kg]"),
    SheetRule("Pipe Report", ["Pipe", "Branch", "Pipe Support", "Support Structure", "Nominal Bore"], None),
]


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value).strip()


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return 0.0


def detect_header_row(rows: List[Tuple[Any, ...]]) -> int:
    header_tokens = {
        "#",
        "item number",
        "qty",
        "description",
        "cut length [mm]",
        "support structure",
        "pipe support",
        "pipe",
        "branch",
        "nominal bore",
        "total cut length [mm]",
        "total weight [kg]",
    }

    for i, row in enumerate(rows):
        non_empty = [normalize_value(v) for v in row if normalize_value(v)]
        if len(non_empty) < 3:
            continue

        token_hits = 0
        for token in non_empty:
            if token.lower() in header_tokens:
                token_hits += 1

        if token_hits >= 2:
            return i

    raise ValueError("Could not detect table header row")


def read_sheet_rows(file_path: Path, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        header_idx = detect_header_row(raw_rows)
        header_row = raw_rows[header_idx]

        headers = []
        for i, value in enumerate(header_row):
            if value is None:
                headers.append(f"Col_{i + 1}")
            else:
                headers.append(normalize_value(value))

        data_rows: List[Dict[str, Any]] = []
        for row in raw_rows[header_idx + 1 :]:
            if row is None:
                continue
            row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            if any(normalize_value(v) for v in row_dict.values()):
                data_rows.append(row_dict)

        return headers, data_rows
    finally:
        wb.close()


def make_key(row: Dict[str, Any], key_columns: Iterable[str]) -> Tuple[str, ...]:
    return tuple(normalize_value(row.get(col)) for col in key_columns)


def aggregate_rows(rows: List[Dict[str, Any]], key_columns: List[str], qty_column: Optional[str]) -> Dict[Tuple[str, ...], AggregateEntry]:
    result: Dict[Tuple[str, ...], AggregateEntry] = defaultdict(AggregateEntry)

    for row in rows:
        key = make_key(row, key_columns)
        if all(not part for part in key):
            continue

        entry = result[key]
        entry.row_count += 1
        if qty_column:
            entry.qty_sum += to_number(row.get(qty_column))

    return result


def choose_labels(file_a: Path, file_b: Path) -> Tuple[str, str]:
    a_name = file_a.name.lower()
    b_name = file_b.name.lower()

    a_is_bq = "bqandsu" in a_name
    b_is_bq = "bqandsu" in b_name
    a_is_su = "overall_su" in a_name or a_name.endswith("_su.xlsx")
    b_is_su = "overall_su" in b_name or b_name.endswith("_su.xlsx")

    if a_is_bq and b_is_su:
        return "BQandSU", "SU"
    if b_is_bq and a_is_su:
        return "SU", "BQandSU"
    if a_is_bq and not b_is_bq:
        return "BQandSU", "SU"
    if b_is_bq and not a_is_bq:
        return "SU", "BQandSU"

    return "File A", "File B"


def collect_sheet_differences(
    rule: SheetRule,
    rows_a: List[Dict[str, Any]],
    rows_b: List[Dict[str, Any]],
    label_a: str,
    label_b: str,
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    agg_a = aggregate_rows(rows_a, rule.key_columns, rule.qty_column)
    agg_b = aggregate_rows(rows_b, rule.key_columns, rule.qty_column)

    all_keys = set(agg_a.keys()) | set(agg_b.keys())

    diffs: List[Dict[str, Any]] = []
    stats = {
        "only_a": 0,
        "only_b": 0,
        "mismatch": 0,
        "match": 0,
    }

    for key in sorted(all_keys):
        in_a = key in agg_a
        in_b = key in agg_b
        a_entry = agg_a.get(key, AggregateEntry())
        b_entry = agg_b.get(key, AggregateEntry())

        if in_a and not in_b:
            status = STATUS_ONLY_A.replace("BQandSU", label_a).replace("SU", label_b)
            stats["only_a"] += 1
        elif in_b and not in_a:
            status = STATUS_ONLY_B.replace("BQandSU", label_a).replace("SU", label_b)
            stats["only_b"] += 1
        else:
            qty_delta = a_entry.qty_sum - b_entry.qty_sum
            row_delta = a_entry.row_count - b_entry.row_count
            qty_mismatch = rule.qty_column is not None and abs(qty_delta) > 1e-9
            row_mismatch = row_delta != 0

            if qty_mismatch or row_mismatch:
                status = STATUS_MISMATCH
                stats["mismatch"] += 1
            else:
                status = STATUS_MATCH
                stats["match"] += 1

        if status == STATUS_MATCH:
            continue

        diff_row: Dict[str, Any] = {
            "Status": status,
            f"{label_a} Row Count": a_entry.row_count,
            f"{label_b} Row Count": b_entry.row_count,
            "Row Count Delta": a_entry.row_count - b_entry.row_count,
        }

        for i, col in enumerate(rule.key_columns):
            diff_row[col] = key[i] if i < len(key) else ""

        if rule.qty_column:
            diff_row[f"{label_a} {rule.qty_column}"] = round(a_entry.qty_sum, 6)
            diff_row[f"{label_b} {rule.qty_column}"] = round(b_entry.qty_sum, 6)
            diff_row[f"Delta {rule.qty_column}"] = round(a_entry.qty_sum - b_entry.qty_sum, 6)

        diffs.append(diff_row)

    diffs.sort(
        key=lambda r: (
            0 if "Only in" in str(r.get("Status")) else 1,
            abs(to_number(r.get("Row Count Delta", 0))),
            abs(to_number(r.get("Delta Qty", 0))),
        ),
        reverse=True,
    )

    return diffs, stats


def autosize_worksheet(ws) -> None:
    width_by_col: Dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            width_by_col[cell.column] = max(width_by_col.get(cell.column, 0), len(value))

    for col_idx, width in width_by_col.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 55)


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN


def apply_status_row_color(ws, status_col: int = 1) -> None:
    for row_idx in range(2, ws.max_row + 1):
        status = normalize_value(ws.cell(row=row_idx, column=status_col).value)
        if "Only in" in status and "BQandSU" in status:
            fill = FILL_ONLY_A
        elif "Only in" in status and "SU" in status:
            fill = FILL_ONLY_B
        elif status == STATUS_MISMATCH:
            fill = FILL_MISMATCH
        elif status == STATUS_MATCH:
            fill = FILL_MATCH
        else:
            fill = None

        if fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill


def ensure_two_files(folder: Path, file_a: Optional[Path], file_b: Optional[Path]) -> Tuple[Path, Path]:
    if file_a and file_b:
        return file_a, file_b

    files = sorted(p for p in folder.glob("*.xlsx") if not p.name.startswith("~$"))
    if len(files) != 2:
        raise ValueError(
            f"Expected exactly 2 .xlsx files in {folder}, but found {len(files)}. "
            "Use --file-a and --file-b if needed."
        )

    return files[0], files[1]


def build_summary_sheet(
    wb: Workbook,
    file_a: Path,
    file_b: Path,
    label_a: str,
    label_b: str,
    per_sheet_stats: Dict[str, Dict[str, int]],
    support_scope_hint: Dict[str, Any],
) -> None:
    ws = wb.active
    ws.title = "Summary"

    rows = [
        ["Comparison Report", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [f"{label_a} File", file_a.name],
        [f"{label_b} File", file_b.name],
        [],
        ["Sheet", f"Only in {label_a}", f"Only in {label_b}", "Mismatch", "Match"],
    ]

    for sheet_name in [rule.name for rule in SHEET_RULES] + ["Cut List (Item Aggregate)"]:
        stats = per_sheet_stats.get(sheet_name, {"only_a": 0, "only_b": 0, "mismatch": 0, "match": 0})
        rows.append([sheet_name, stats["only_a"], stats["only_b"], stats["mismatch"], stats["match"]])

    rows.extend(
        [
            [],
            ["Scope Analysis", ""],
            ["Supports only in BQandSU", support_scope_hint.get("only_a_supports", 0)],
            ["Supports only in BQandSU without /SU suffix", support_scope_hint.get("only_a_non_su", 0)],
            [
                "Observation",
                support_scope_hint.get(
                    "message",
                    "No specific scope pattern detected from Support List naming.",
                ),
            ],
        ]
    )

    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = BOLD_FONT

    for cell in ws[6]:
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    ws[ws.max_row - 4][1].font = BOLD_FONT
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:E{6 + len(SHEET_RULES) + 1}"
    autosize_worksheet(ws)


def extract_support_scope_hint(
    support_diffs: List[Dict[str, Any]],
    key_column: str = "Support Structure",
) -> Dict[str, Any]:
    only_a = [r for r in support_diffs if normalize_value(r.get("Status")).startswith("Only in") and "BQandSU" in normalize_value(r.get("Status"))]
    only_a_support_names = [normalize_value(r.get(key_column)) for r in only_a if normalize_value(r.get(key_column))]
    only_a_non_su = [s for s in only_a_support_names if not s.endswith("/SU")]

    if only_a_support_names and len(only_a_support_names) == len(only_a_non_su):
        message = (
            "All support structures that exist only in BQandSU do not use the /SU suffix. "
            "This strongly indicates added BQ scope in the BQandSU export."
        )
    elif only_a_non_su:
        message = (
            "Many support structures only in BQandSU do not use /SU suffix. "
            "Likely additional non-SU scope in the BQandSU export."
        )
    else:
        message = "No strong /SU naming pattern detected for support-only differences."

    return {
        "only_a_supports": len(only_a_support_names),
        "only_a_non_su": len(only_a_non_su),
        "message": message,
    }


def write_diff_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title=title[:31])

    if not rows:
        ws.append(["No differences found"])
        ws["A1"].font = BOLD_FONT
        autosize_worksheet(ws)
        return

    columns = list(rows[0].keys())
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])

    style_header(ws)
    apply_status_row_color(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    autosize_worksheet(ws)


def compare_workbooks(file_a: Path, file_b: Path, output_path: Path) -> None:
    wb_a = load_workbook(file_a, read_only=True, data_only=True)
    wb_b = load_workbook(file_b, read_only=True, data_only=True)
    try:
        sheets_a = set(wb_a.sheetnames)
        sheets_b = set(wb_b.sheetnames)
    finally:
        wb_a.close()
        wb_b.close()

    expected_sheets = [rule.name for rule in SHEET_RULES]
    missing_a = [s for s in expected_sheets if s not in sheets_a]
    missing_b = [s for s in expected_sheets if s not in sheets_b]

    if missing_a or missing_b:
        msg = []
        if missing_a:
            msg.append(f"Missing in file A: {missing_a}")
        if missing_b:
            msg.append(f"Missing in file B: {missing_b}")
        raise ValueError(" | ".join(msg))

    label_a, label_b = choose_labels(file_a, file_b)

    report_wb = Workbook()
    per_sheet_stats: Dict[str, Dict[str, int]] = {}
    support_diffs: List[Dict[str, Any]] = []

    for rule in SHEET_RULES:
        _, rows_a = read_sheet_rows(file_a, rule.name)
        _, rows_b = read_sheet_rows(file_b, rule.name)

        diffs, stats = collect_sheet_differences(rule, rows_a, rows_b, label_a, label_b)
        per_sheet_stats[rule.name] = stats

        if rule.name == "Support List":
            support_diffs = diffs

        sheet_title_map = {
            "Total Qty": "Diff Total Qty",
            "Per Support Structure": "Diff Per Support",
            "Per Pipe Support": "Diff Per Pipe",
            "Cut List": "Diff Cut List",
            "Total Cut Lengths": "Diff Cut Lengths",
            "Support List": "Diff Supports",
            "Pipe Report": "Diff Pipe Report",
        }
        write_diff_sheet(report_wb, sheet_title_map.get(rule.name, f"Diff {rule.name}"), diffs)

        if rule.name == "Cut List":
            cut_item_rule = SheetRule("Cut List Item Aggregate", ["Item Number"], "Qty")
            cut_item_diffs, cut_item_stats = collect_sheet_differences(cut_item_rule, rows_a, rows_b, label_a, label_b)
            per_sheet_stats["Cut List (Item Aggregate)"] = cut_item_stats
            write_diff_sheet(report_wb, "Diff CutList by Item", cut_item_diffs)

    scope_hint = extract_support_scope_hint(support_diffs)
    build_summary_sheet(report_wb, file_a, file_b, label_a, label_b, per_sheet_stats, scope_hint)

    report_wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare two Sikla Excel exports and generate a differences workbook.")
    parser.add_argument("--folder", type=Path, default=Path("."), help="Folder containing the two .xlsx files.")
    parser.add_argument("--file-a", type=Path, default=None, help="Explicit first file path.")
    parser.add_argument("--file-b", type=Path, default=None, help="Explicit second file path.")
    parser.add_argument("--output", type=Path, default=None, help="Output workbook path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    folder = args.folder.resolve()

    file_a, file_b = ensure_two_files(folder, args.file_a, args.file_b)

    if args.output:
        output_path = args.output
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = folder / f"sikla_comparison_{stamp}.xlsx"

    compare_workbooks(file_a, file_b, output_path)

    print(f"Comparison workbook created: {output_path}")


if __name__ == "__main__":
    main()
