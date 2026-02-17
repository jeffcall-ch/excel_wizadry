import pandas as pd
import re
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File paths
input_file = r'C:\Users\szil\Repos\excel_wizadry\00_THROWAWAY\Thameside_stress_calcs_Sikla_comp_selector\extremesupportloads12.02.2026.xlsx'
output_file = r'C:\Users\szil\Repos\excel_wizadry\00_THROWAWAY\Thameside_stress_calcs_Sikla_comp_selector\extremesupportloads12.02.2026_processed_2026-02-17.xlsx'

# Support types we're interested in
target_support_types = [
    "Guided sliding support",
    "Anchor point",
    "Lateral stop all-round, vertical pipe"
]

# Read the sheets
print("Reading Excel file...")
df_extreme = pd.read_excel(input_file, sheet_name='extremesupportloads', header=7)
df_allowable = pd.read_excel(input_file, sheet_name='Allowable_loads')

print(f"Extreme support loads sheet has {len(df_extreme)} rows")
print(f"Allowable loads sheet has {len(df_allowable)} rows")

# Get the actual column names - they might have unnamed columns
# Let's use the column indices to be sure
# Based on the Excel file: A=0, B=1, C=2, D=3, E=4, ... L=11, M=12, N=13, ... R=17, ... T=19

# Let's print the columns to be sure
print("\nColumn names in extreme sheet:")
for i, col in enumerate(df_extreme.columns):
    print(f"  Index {i}: {col}")

# Extract DN number from dimension name string (e.g., "EEDX_DN150" -> 150)
def extract_dn(dimension_name):
    if pd.isna(dimension_name):
        return None
    match = re.search(r'DN(\d+)', str(dimension_name))
    if match:
        return int(match.group(1))
    return None

# Calculate forces based on axial direction
def calculate_forces(qx, qy, qz, axial_direction):
    # Take absolute values first!
    qx = abs(qx)
    qy = abs(qy)
    qz = abs(qz)
    
    if pd.isna(axial_direction):
        axial_direction = ''
    
    axial_dir = str(axial_direction).strip().lower()
    
    if axial_dir == 'x':
        fax = qx
        flat = qy
        fz = qz
    elif axial_dir == 'y':
        fax = qy
        flat = qx
        fz = qz
    elif axial_dir == 'z':
        fax = qz
        flat = max(qx, qy)
        fz = max(qx, qy)
    else:
        # No axial direction specified - treat as no guidance
        fax = 0
        flat = max(qx, qy, qz)
        fz = max(qx, qy, qz)
    
    return fax, flat, fz

# Find matching component from allowable loads
def find_matching_component(dn, fax, flat, fz, support_type, df_allowable, debug=False):
    # Filter allowable loads by DN
    matching_dn = df_allowable[df_allowable['DN'] == dn].copy()
    
    if len(matching_dn) == 0:
        return "NO COMPONENT FOUND - DN not in allowable loads"
    
    # Get the support type column name
    support_type_col = support_type
    
    if support_type_col not in df_allowable.columns:
        return f"NO COMPONENT FOUND - Support type '{support_type}' not in allowable loads"
    
    # Filter rows where all forces are higher than required AND component is not "DO NOT USE"
    valid_components = []
    
    for idx, row in matching_dn.iterrows():
        component_name = row[support_type_col]
        
        # Skip if component is "DO NOT USE" or NaN
        if pd.isna(component_name) or str(component_name).strip().upper() == "DO NOT USE":
            continue
        
        # Get allowable values
        fax_allow = row['Fax [N]']
        flat_allow = row['Flat [N]']
        fz_allow = row['Fz [N]']
        
        # Check EACH force component independently - ALL must qualify
        fax_ok = fax_allow >= fax
        flat_ok = flat_allow >= flat
        fz_ok = fz_allow >= fz
        
        if debug:
            print(f"    Checking {component_name}: Fax:{fax_allow}≥{fax}?{fax_ok}, Flat:{flat_allow}≥{flat}?{flat_ok}, Fz:{fz_allow}≥{fz}?{fz_ok}")
        
        # Only accept if ALL three force components qualify
        if fax_ok and flat_ok and fz_ok:
            # Calculate the total allowable capacity
            total_capacity = fax_allow + flat_allow + fz_allow
            valid_components.append({
                'component': component_name,
                'total_capacity': total_capacity,
                'fax_allowable': fax_allow,
                'flat_allowable': flat_allow,
                'fz_allowable': fz_allow
            })
    
    if len(valid_components) == 0:
        return "NO COMPONENT FOUND"
    
    # Select the component with the lowest total capacity (most efficient)
    best_component = min(valid_components, key=lambda x: x['total_capacity'])
    
    return best_component['component']

# Prepare results list
results = []

# Process each row in extreme support loads
print("\nProcessing rows...")
row_num = 0
for idx, row in df_extreme.iterrows():
    row_num += 1
    
    # Get support type from column B (index 1)
    support_type = row[df_extreme.columns[1]]
    
    # Skip if not in target support types
    if pd.isna(support_type) or support_type not in target_support_types:
        results.append(None)
        continue
    
    # Get dimension name from column E (index 4)
    dimension_name = row[df_extreme.columns[4]]
    dn = extract_dn(dimension_name)
    
    if dn is None:
        results.append("ERROR: Could not extract DN")
        print(f"  Row {row_num}: Could not extract DN from '{dimension_name}'")
        continue
    
    # Get forces from columns L, M, N (indices 11, 12, 13)
    qx = row[df_extreme.columns[11]]
    qy = row[df_extreme.columns[12]]
    qz = row[df_extreme.columns[13]]
    
    # Check if forces are valid numbers
    if pd.isna(qx) or pd.isna(qy) or pd.isna(qz):
        results.append("ERROR: Missing force values")
        print(f"  Row {row_num}: Missing force values")
        continue
    
    # Get axial direction from column R (index 17)
    axial_direction = row[df_extreme.columns[17]]
    
    # Calculate Fax, Flat, Fz (absolute values)
    fax, flat, fz = calculate_forces(qx, qy, qz, axial_direction)
    
    # Find matching component
    component = find_matching_component(dn, fax, flat, fz, support_type, df_allowable)
    
    results.append(component)
    
    print(f"  Row {row_num}: {support_type}, DN{dn}, QX={qx:.1f}, QY={qy:.1f}, QZ={qz:.1f}, Axial={axial_direction} -> Fax={fax:.1f}, Flat={flat:.1f}, Fz={fz:.1f} -> {component}")

# Add results to column T (index 19)
# If column T doesn't exist, we need to add it
if len(df_extreme.columns) < 20:
    # Add columns until we have column T
    while len(df_extreme.columns) < 20:
        df_extreme[f'Unnamed: {len(df_extreme.columns)}'] = None

df_extreme.iloc[:, 19] = results

# Save to Excel
print(f"\nSaving to {output_file}...")
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_extreme.to_excel(writer, sheet_name='extremesupportloads', index=False, header=False, startrow=8)
    df_allowable.to_excel(writer, sheet_name='Allowable_loads', index=False)

# Load the workbook to apply formatting
print("Applying formatting...")
wb = load_workbook(output_file)
ws = wb['extremesupportloads']

# Copy original header rows (rows 1-8)
wb_orig = load_workbook(input_file)
ws_orig = wb_orig['extremesupportloads']

for row_idx in range(1, 9):
    for col_idx in range(1, ws_orig.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).value = ws_orig.cell(row=row_idx, column=col_idx).value

# Apply red fill to "NO COMPONENT FOUND" cells in column T
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

for idx, result in enumerate(results):
    if result and "NO COMPONENT FOUND" in str(result):
        cell = ws.cell(row=idx + 9, column=20)  # Column T is 20, starting from row 9 (after header)
        cell.fill = red_fill

wb.save(output_file)
print(f"\nProcessing complete! Output saved to:\n{output_file}")
