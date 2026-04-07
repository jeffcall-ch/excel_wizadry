"""
Reads Per_Support_Structure (FROM) and Total Qty (TO) sheets.
Matches rows on Item Number, Qty, Description, Cut Length [mm],
Weight [kg], Total Weight [kg], Remarks (FROM C-I == TO B-H).
When a match exists and FROM Coating (col J) == 'Intumescent',
copies 'Intumescent' into TO Coating (col I) and marks the row yellow.
Saves a timestamped copy in the same folder.
"""

import math
import os
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
FILE_PATH = (
    r"C:\Users\szil\Repos\excel_wizadry\00_THROWAWAY\Intumescent copier"
    r"\CA100-KVI-50296373_1.0 - BoM General Hangers and support material.xlsm"
)
FROM_SHEET = "Per_Support_Structure"
TO_SHEET = "Total Qty"

MATCH_COLS = [
    "Item Number",
    "Qty",
    "Description",
    "Cut Length [mm]",
    "Weight [kg]",
    "Total Weight [kg]",
    "Remarks",
]

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# TO sheet column indices (1-based, openpyxl)
TO_COL = {
    "Item Number":      2,   # B
    "Qty":              3,   # C
    "Description":      4,   # D
    "Cut Length [mm]":  5,   # E
    "Weight [kg]":      6,   # F
    "Total Weight [kg]":7,   # G
    "Remarks":          8,   # H
    "Coating":          9,   # I
}
# ---------------------------------------------------------------------------


def normalize(v):
    """Canonicalise a cell value for comparison: NaN→None, whole-floats→int."""
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v):
            return None
        if v == int(v):
            return int(v)
    return v


def make_key(row_dict):
    return tuple(normalize(row_dict[c]) for c in MATCH_COLS)


# ── 1. Read FROM sheet with pandas (uses cached formula results) ─────────────
print("Reading FROM sheet …")
df_from = pd.read_excel(
    FILE_PATH,
    sheet_name=FROM_SHEET,
    engine="openpyxl",
    header=0,
    usecols="B:J",          # B=#  C=Item Number … I=Remarks  J=Coating
)
# Rename '#' column so it doesn't collide; Coating is the last column
df_from.columns = [
    "#", "Item Number", "Qty", "Description",
    "Cut Length [mm]", "Weight [kg]", "Total Weight [kg]",
    "Remarks", "Coating",
]

intumescent_keys: set = set()
for _, row in df_from[df_from["Coating"] == "Intumescent"].iterrows():
    intumescent_keys.add(make_key(row))

print(f"  {len(intumescent_keys)} unique Intumescent keys found in FROM sheet.")

# ── 2. Open workbook with openpyxl for writing ───────────────────────────────
print("Opening workbook with openpyxl …")
wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True)
ws_to = wb[TO_SHEET]

modified = 0
max_col = ws_to.max_column

for row_idx in range(2, ws_to.max_row + 1):
    row_dict = {
        col_name: ws_to.cell(row=row_idx, column=col_num).value
        for col_name, col_num in TO_COL.items()
    }

    # Skip completely empty rows
    if all(row_dict[c] is None for c in MATCH_COLS):
        continue

    key = make_key(row_dict)
    if key not in intumescent_keys:
        continue

    # Write Intumescent into Coating column
    ws_to.cell(row=row_idx, column=TO_COL["Coating"]).value = "Intumescent"

    # Highlight entire row yellow
    for col in range(1, max_col + 1):
        ws_to.cell(row=row_idx, column=col).fill = YELLOW_FILL

    modified += 1
    print(
        f"  Row {row_idx:>5}: {str(row_dict.get('Item Number', '')):>8} "
        f"| {row_dict.get('Description', '')}"
    )

print(f"\nModified {modified} rows in '{TO_SHEET}'.")

# ── 3. Save with timestamp ───────────────────────────────────────────────────
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
folder = os.path.dirname(FILE_PATH)
base = os.path.basename(FILE_PATH)
name, ext = base.rsplit(".", 1)
out_path = os.path.join(folder, f"{name}_{timestamp}.{ext}")

wb.save(out_path)
print(f"Saved: {out_path}")
