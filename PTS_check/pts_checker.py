"""
PTS Activity Matcher Script
Matches activities from project plan with important activities from general file
using fuzzy string matching and adds category markers.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from thefuzz import fuzz
import sys


def find_rightmost_used_column(df):
    """Find the index of the rightmost column with data."""
    for i in range(len(df.columns) - 1, -1, -1):
        if df.iloc[:, i].notna().sum() > 1:  # More than 1 non-null value
            return i
    return len(df.columns) - 1


def match_activity(activity_name, important_activities, threshold=90):
    """
    Match an activity name against a list of important activities using fuzzy matching.
    Returns a list of tuples (activity, score, categories) for matches above threshold.
    """
    if pd.isna(activity_name):
        return []
    
    activity_name_lower = str(activity_name).lower().strip()
    matches = []
    
    for _, row in important_activities.iterrows():
        important_activity = str(row['Activity']).lower().strip()
        
        # Calculate similarity score
        score = fuzz.ratio(activity_name_lower, important_activity)
        
        if score >= threshold:
            # Extract categories where value is 'x' (case insensitive)
            categories = {}
            for col in ['Important for GP', 'Piping', 'Hangers', 'Additional Material', 'Valves']:
                if pd.notna(row[col]) and str(row[col]).lower().strip() == 'x':
                    categories[col] = 'X'
                else:
                    categories[col] = ''
            
            matches.append((row['Activity'], score, categories))
    
    return matches


def format_date_swiss(date_value):
    """
    Convert date to Swiss format dd/mm/yyyy, removing any extra text like 'A'.
    Handles both datetime objects and strings.
    Handles 2-digit years with asterisk (e.g., '26*' means 2026).
    """
    if pd.isna(date_value):
        return ''
    
    date_str = str(date_value).strip()
    
    # Remove trailing ' A' or similar letters (but preserve asterisks for now)
    date_str = date_str.rstrip(' ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    
    try:
        # Try parsing as datetime if it's a timestamp string
        if isinstance(date_value, pd.Timestamp):
            return date_value.strftime('%d/%m/%Y')
        
        # Try parsing common formats
        # Format: dd.mm.yy or dd.mm.yy*
        if '.' in date_str and len(date_str.split('.')) == 3:
            parts = date_str.split('.')
            day, month, year = parts[0], parts[1], parts[2]
            
            # Handle asterisk in year (e.g., '26*')
            year = year.rstrip('*')
            
            # Handle 2-digit year
            if len(year) == 2:
                year = '20' + year if int(year) < 50 else '19' + year
            return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
        
        # Try pandas datetime parsing
        dt = pd.to_datetime(date_value, errors='coerce')
        if pd.notna(dt):
            return dt.strftime('%d/%m/%Y')
    except:
        pass
    
    return date_str


def parse_holiday_periods(df_holidays, year):
    """
    Parse holiday periods from text descriptions to actual date ranges.
    
    Args:
        df_holidays: DataFrame with columns ['From', 'To', 'Length adder weeks']
        year: Year to apply the holiday periods to
    
    Returns:
        List of tuples: (start_date, end_date, adder_weeks, description)
    """
    periods = []
    
    month_map = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'may': 5, 'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    
    for _, row in df_holidays.iterrows():
        from_text = str(row['From']).lower()
        to_text = str(row['To']).lower()
        adder_weeks = int(row['Length adder weeks'])
        
        # Parse "middle of Month" or "end of Month"
        from_month = None
        from_day = 1
        to_month = None
        to_day = 1
        
        for month_name, month_num in month_map.items():
            if month_name in from_text:
                from_month = month_num
                if 'middle' in from_text:
                    from_day = 15
                elif 'end' in from_text:
                    from_day = 28 if month_num == 2 else (30 if month_num in [4, 6, 9, 11] else 31)
                break
        
        for month_name, month_num in month_map.items():
            if month_name in to_text:
                to_month = month_num
                if 'middle' in to_text:
                    to_day = 15
                elif 'end' in to_text:
                    to_day = 28 if month_num == 2 else (30 if month_num in [4, 6, 9, 11] else 31)
                break
        
        if from_month and to_month:
            start_date = datetime(year, from_month, from_day)
            end_date = datetime(year, to_month, to_day)
            description = f"{row['From']} to {row['To']} holiday period"
            periods.append((start_date, end_date, adder_weeks, description))
    
    return periods


def check_holiday_conflict(start_date, end_date, holiday_periods):
    """
    Check if an activity execution period overlaps with a holiday period or within 1 week after.
    
    Args:
        start_date: datetime object - activity start date
        end_date: datetime object - activity end date
        holiday_periods: List of tuples (start_date, end_date, adder_weeks, description)
    
    Returns:
        Tuple (has_conflict, adder_weeks, description) or (False, 0, None)
    """
    for holiday_start, holiday_end, adder_weeks, description in holiday_periods:
        # Extend holiday period by 1 week after
        extended_holiday_end = holiday_end + timedelta(weeks=1)
        
        # Check if activity execution period overlaps with holiday period (including 1 week after)
        # Activity overlaps if: activity_start <= extended_holiday_end AND activity_end >= holiday_start
        if start_date <= extended_holiday_end and end_date >= holiday_start:
            return (True, adder_weeks, description)
    
    return (False, 0, None)


def calculate_requested_dates(df_plan, df_general, df_constraints, matched_activities, df_holidays=None):
    """
    Calculate requested dates by working backwards AND forwards from the mobilisation milestone.
    Detects and reports conflicts when multiple dependencies try to set different dates.
    Ensures all calculated dates fall on weekdays (Monday-Friday).
    Applies holiday adjustments if holiday periods are provided.
    
    Args:
        df_plan: Project plan dataframe
        df_general: General activities dataframe
        df_constraints: Constraints/relationships dataframe
        matched_activities: Dictionary mapping important activity names to their row indices in df_plan
        df_holidays: Optional DataFrame with holiday periods
    
    Returns:
        List of conflict messages (empty if no conflicts)
    """
    
    def ensure_weekday(date):
        """Adjust date to next weekday if it falls on weekend."""
        # 5 = Saturday, 6 = Sunday
        if date.weekday() == 5:  # Saturday
            return date + timedelta(days=2)  # Move to Monday
        elif date.weekday() == 6:  # Sunday
            return date + timedelta(days=1)  # Move to Monday
        return date
    
    # Find the mobilisation milestone in the matched activities
    mobilisation_name = "Lot 5 - Erection General Piping - Contractor On Site Mobilisation"
    
    if mobilisation_name not in matched_activities:
        print(f"Warning: Mobilisation milestone '{mobilisation_name}' not found in matched activities")
        return []
    
    mobilisation_idx = matched_activities[mobilisation_name]
    mobilisation_start = df_plan.at[mobilisation_idx, 'Start']
    
    # Convert to datetime
    try:
        anchor_date = pd.to_datetime(mobilisation_start, format='%d/%m/%Y')
        anchor_date = ensure_weekday(anchor_date)  # Ensure anchor is on weekday
        print(f"\nAnchor point: {mobilisation_name}")
        print(f"  Start date: {anchor_date.strftime('%d/%m/%Y')}")
    except Exception as e:
        print(f"Error parsing mobilisation start date: {e}")
        return []
    
    # Parse holiday periods if provided
    holiday_periods = []
    if df_holidays is not None:
        anchor_year = anchor_date.year
        # Generate holiday periods for anchor year and adjacent years
        for year in [anchor_year - 1, anchor_year, anchor_year + 1]:
            holiday_periods.extend(parse_holiday_periods(df_holidays, year))
        print(f"  Loaded {len(holiday_periods) // 3} holiday period(s) for years {anchor_year-1} to {anchor_year+1}")
    
    # Build a dictionary to store calculated dates and their sources
    # Structure: {activity_name: {'date': datetime, 'sources': [list of calculation sources]}}
    activity_dates = {
        mobilisation_name: {
            'date': anchor_date,
            'sources': ['Anchor milestone (reference point)']
        }
    }
    
    # Track conflicts
    conflicts = []
    
    # Process constraints - work both backwards and forwards
    max_iterations = 100  # Prevent infinite loops
    iteration = 0
    
    while iteration < max_iterations:
        iteration += 1
        progress_made = False
        
        for _, constraint in df_constraints.iterrows():
            from_activity = constraint['From']
            to_activity = constraint['To']
            duration_weeks = constraint['Duration weeks']
            comment = constraint['Comments']
            
            # Skip if duration is not valid
            if pd.isna(duration_weeks):
                continue
            
            # Build comment description
            comment_suffix = f". {comment}" if pd.notna(comment) and str(comment).strip() else ""
            
            # BACKWARD calculation: if "To" date is known but "From" date needs to be calculated/verified
            if to_activity in activity_dates and from_activity in matched_activities:
                to_date = activity_dates[to_activity]['date']
                from_date_initial = to_date - timedelta(weeks=duration_weeks)
                from_date_initial = ensure_weekday(from_date_initial)  # Ensure weekday
                
                # Check if activity execution period (from_date_initial TO to_date) overlaps with holidays
                has_conflict, adder_weeks, holiday_desc = check_holiday_conflict(from_date_initial, to_date, holiday_periods)
                if has_conflict:
                    # For backward calculation, add adder to duration (start earlier)
                    total_weeks = duration_weeks + adder_weeks
                    from_date = to_date - timedelta(weeks=total_weeks)
                    from_date = ensure_weekday(from_date)
                    reason = f"{int(total_weeks)}W before '{to_activity}'{comment_suffix}. Originally {from_date_initial.strftime('%d/%m/%Y')} (based on {int(duration_weeks)}W), shifted by {adder_weeks}W due to {holiday_desc}"
                else:
                    from_date = from_date_initial
                    reason = f"{int(duration_weeks)}W before '{to_activity}'{comment_suffix}"
                
                if from_activity in activity_dates:
                    # Check for conflict
                    existing_date = activity_dates[from_activity]['date']
                    if abs((existing_date - from_date).days) > 0:  # Different dates
                        conflict_msg = (
                            f"CONFLICT: '{from_activity}'\n"
                            f"  Existing: {existing_date.strftime('%d/%m/%Y')} from {activity_dates[from_activity]['sources'][-1]}\n"
                            f"  New calc: {from_date.strftime('%d/%m/%Y')} from {reason}"
                        )
                        conflicts.append(conflict_msg)
                        print(f"⚠️  {conflict_msg}")
                    else:
                        # Same date from different path - just add source
                        activity_dates[from_activity]['sources'].append(reason)
                else:
                    # New calculation
                    activity_dates[from_activity] = {
                        'date': from_date,
                        'sources': [reason]
                    }
                    progress_made = True
            
            # FORWARD calculation: if "From" date is known but "To" date needs to be calculated/verified
            if from_activity in activity_dates and to_activity in matched_activities:
                from_date = activity_dates[from_activity]['date']
                to_date_initial = from_date + timedelta(weeks=duration_weeks)
                to_date_initial = ensure_weekday(to_date_initial)  # Ensure weekday
                
                # Check if activity execution period (from_date TO to_date_initial) overlaps with holidays
                has_conflict, adder_weeks, holiday_desc = check_holiday_conflict(from_date, to_date_initial, holiday_periods)
                if has_conflict:
                    # For forward calculation, add adder to duration (end later)
                    total_weeks = duration_weeks + adder_weeks
                    to_date = from_date + timedelta(weeks=total_weeks)
                    to_date = ensure_weekday(to_date)
                    reason = f"{int(total_weeks)}W after '{from_activity}'{comment_suffix}. Originally {to_date_initial.strftime('%d/%m/%Y')} (based on {int(duration_weeks)}W), shifted by {adder_weeks}W due to {holiday_desc}"
                else:
                    to_date = to_date_initial
                    reason = f"{int(duration_weeks)}W after '{from_activity}'{comment_suffix}"
                
                if to_activity in activity_dates:
                    # Check for conflict
                    existing_date = activity_dates[to_activity]['date']
                    if abs((existing_date - to_date).days) > 0:  # Different dates
                        conflict_msg = (
                            f"CONFLICT: '{to_activity}'\n"
                            f"  Existing: {existing_date.strftime('%d/%m/%Y')} from {activity_dates[to_activity]['sources'][-1]}\n"
                            f"  New calc: {to_date.strftime('%d/%m/%Y')} from {reason}"
                        )
                        conflicts.append(conflict_msg)
                        print(f"⚠️  {conflict_msg}")
                    else:
                        # Same date from different path - just add source
                        activity_dates[to_activity]['sources'].append(reason)
                else:
                    # New calculation
                    activity_dates[to_activity] = {
                        'date': to_date,
                        'sources': [reason]
                    }
                    progress_made = True
        
        # If no progress was made, we're done
        if not progress_made:
            break
    
    print(f"\nCalculated dates for {len(activity_dates)} activities")
    if conflicts:
        print(f"⚠️  Found {len(conflicts)} conflict(s) - see details above")
    
    # Now update the df_plan with calculated dates
    for activity_name, date_info in activity_dates.items():
        if activity_name in matched_activities:
            row_idx = matched_activities[activity_name]
            
            # Set requested date
            df_plan.at[row_idx, 'Requested date'] = date_info['date'].strftime('%d/%m/%Y')
            
            # Set comment (use the first source)
            df_plan.at[row_idx, 'Comment'] = date_info['sources'][0]
    
    return conflicts


def calculate_current_lengths(df_plan, df_constraints, matched_activities):
    """
    Calculate the current length in weeks between dependent activities based on their actual dates in the plan.
    Uses Finish dates for all activities except the anchor (which uses Start date).
    Calculates for the same activity pairs and direction as the requested dates/comments.
    
    Args:
        df_plan: Project plan dataframe
        df_constraints: Constraints/relationships dataframe
        matched_activities: Dictionary mapping important activity names to their row indices in df_plan
    """
    from datetime import timedelta
    
    # Anchor activity uses Start date, all others use Finish date
    anchor_name = "Lot 5 - Erection General Piping - Contractor On Site Mobilisation"
    
    # Dictionary to track current lengths for each activity
    current_lengths = {}
    
    for _, constraint in df_constraints.iterrows():
        from_activity = constraint['From']
        to_activity = constraint['To']
        duration_weeks = constraint['Duration weeks']
        comment = constraint['Comments']
        
        # Skip if duration is not valid or activities not matched
        if pd.isna(duration_weeks) or from_activity not in matched_activities or to_activity not in matched_activities:
            continue
        
        # Get the row indices
        from_idx = matched_activities[from_activity]
        to_idx = matched_activities[to_activity]
        
        # Get the appropriate dates: Finish for most activities, Start for anchor
        # "From" activity: use Finish date
        from_date_value = df_plan.at[from_idx, 'Finish']
        
        # "To" activity: use Start date if it's the anchor, otherwise use Finish date
        if to_activity == anchor_name:
            to_date_value = df_plan.at[to_idx, 'Start']
        else:
            to_date_value = df_plan.at[to_idx, 'Finish']
        
        # Parse dates - skip if either is missing
        if pd.isna(from_date_value) or pd.isna(to_date_value) or from_date_value == '' or to_date_value == '':
            continue
        
        try:
            from_date = pd.to_datetime(from_date_value, format='%d/%m/%Y', dayfirst=True)
            to_date = pd.to_datetime(to_date_value, format='%d/%m/%Y', dayfirst=True)
            
            # Calculate actual duration in weeks
            actual_duration = (to_date - from_date).days / 7
            
            # Build description matching the comment format
            comment_suffix = f". {comment}" if pd.notna(comment) and str(comment).strip() else ""
            
            # BACKWARD relationship: Add to "From" activity (it comes BEFORE "To")
            backward_desc = f"{actual_duration:.1f}W (planned: {int(duration_weeks)}W) before '{to_activity}'{comment_suffix}"
            if from_activity not in current_lengths:
                current_lengths[from_activity] = []
            current_lengths[from_activity].append(backward_desc)
            
            # FORWARD relationship: Add to "To" activity (it comes AFTER "From")
            forward_desc = f"{actual_duration:.1f}W (planned: {int(duration_weeks)}W) after '{from_activity}'{comment_suffix}"
            if to_activity not in current_lengths:
                current_lengths[to_activity] = []
            current_lengths[to_activity].append(forward_desc)
            
        except Exception as e:
            print(f"Warning: Could not parse dates for '{from_activity}' -> '{to_activity}': {e}")
            continue
    
    # Update the dataframe with current lengths
    for activity_name, length_list in current_lengths.items():
        if activity_name in matched_activities:
            row_idx = matched_activities[activity_name]
            # Join multiple lengths with line breaks
            df_plan.at[row_idx, 'Current length'] = '\n'.join(length_list)
    
    print(f"Calculated current lengths for {len(current_lengths)} activities")


def process_pts_file(project_plan_path, general_file_path, similarity_threshold=95):
    """
    Process the PTS project plan file and add category markers for important activities.
    """
    print(f"Loading project plan: {project_plan_path}")
    df_plan = pd.read_excel(project_plan_path)
    
    print(f"Loading general file: {general_file_path}")
    df_general = pd.read_excel(general_file_path, sheet_name=0)
    df_constraints = pd.read_excel(general_file_path, sheet_name=1)  # Load constraints sheet
    
    print(f"\nProject plan shape: {df_plan.shape}")
    print(f"Important activities: {len(df_general)}")
    print(f"Constraints/relationships: {len(df_constraints)}")
    
    # Format date columns to Swiss format
    if 'Start' in df_plan.columns:
        df_plan['Start'] = df_plan['Start'].apply(format_date_swiss)
    if 'Finish' in df_plan.columns:
        df_plan['Finish'] = df_plan['Finish'].apply(format_date_swiss)
    
    # Find rightmost used column
    rightmost_col_idx = find_rightmost_used_column(df_plan)
    print(f"Rightmost used column: {df_plan.columns[rightmost_col_idx]} (index {rightmost_col_idx})")
    
    # Create new columns after the rightmost column
    insert_position = rightmost_col_idx + 1
    
    # Add "Current length", "Requested date" and "Comment" columns
    df_plan.insert(insert_position, 'Current length', '')
    df_plan.insert(insert_position + 1, 'Requested date', '')
    df_plan.insert(insert_position + 2, 'Comment', '')
    df_plan.insert(insert_position + 3, 'Similarity Score %', '')
    
    # Add category columns
    category_columns = ['Important for GP', 'Piping', 'Hangers', 'Additional Material', 'Valves']
    for i, cat_col in enumerate(category_columns):
        df_plan.insert(insert_position + 4 + i, cat_col, '')
    
    # Process each activity - NEW APPROACH: Find best match for each important activity
    print(f"\nMatching activities (threshold: {similarity_threshold}%)...")
    matched_count = 0
    match_details = []
    low_score_rows = []  # Track rows with score < 95%
    matched_activities = {}  # Dictionary to store activity name -> row index mapping
    
    # For each important activity, find the best matching project activity
    for _, important_row in df_general.iterrows():
        important_activity = important_row['Activity']
        best_score = 0
        best_plan_idx = None
        
        # Search through all project plan activities
        for idx, plan_row in df_plan.iterrows():
            if pd.notna(plan_row['Activity Name']):
                activity_name_lower = str(plan_row['Activity Name']).lower().strip()
                important_activity_lower = str(important_activity).lower().strip()
                
                # Calculate similarity score
                score = fuzz.ratio(activity_name_lower, important_activity_lower)
                
                if score >= similarity_threshold and score > best_score:
                    best_score = score
                    best_plan_idx = idx
        
        # If a match was found, mark it
        if best_plan_idx is not None:
            # Store the mapping
            matched_activities[important_activity] = best_plan_idx
            
            # Extract categories where value is 'x' (case insensitive)
            categories = {}
            for col in category_columns:
                if pd.notna(important_row[col]) and str(important_row[col]).lower().strip() == 'x':
                    categories[col] = 'X'
                else:
                    categories[col] = ''
            
            # Fill in category columns
            for cat_col in category_columns:
                df_plan.at[best_plan_idx, cat_col] = categories[cat_col]
            
            # Fill in similarity score
            df_plan.at[best_plan_idx, 'Similarity Score %'] = best_score
            
            # Track low score matches
            if best_score < 95:
                low_score_rows.append(best_plan_idx)
            
            matched_count += 1
            match_details.append({
                'Activity ID': df_plan.at[best_plan_idx, 'Activity ID'],
                'Activity Name': df_plan.at[best_plan_idx, 'Activity Name'],
                'Matched To': important_activity,
                'Score': best_score,
                'Row': best_plan_idx
            })
    
    print(f"\nMatched {matched_count} activities out of {df_plan['Activity Name'].notna().sum()} total activities")
    
    # Load holidays sheet if available
    df_holidays = None
    try:
        df_holidays = pd.read_excel(general_file_path, sheet_name='Holidays')
        print(f"Loaded {len(df_holidays)} holiday period(s)")
    except Exception as e:
        print(f"No Holidays sheet found or error loading: {e}")
    
    # Calculate requested dates based on constraints
    print(f"\nCalculating requested dates based on dependencies...")
    conflicts = calculate_requested_dates(df_plan, df_general, df_constraints, matched_activities, df_holidays)
    
    # Calculate current lengths between activities
    print(f"\nCalculating current durations between activities...")
    calculate_current_lengths(df_plan, df_constraints, matched_activities)
    
    # Print some match details
    if match_details:
        print("\nSample matches:")
        for detail in match_details[:10]:
            print(f"  {detail['Activity ID']}: {detail['Activity Name'][:60]}...")
            print(f"    -> {detail['Matched To'][:60]}... (Score: {detail['Score']}%)")
    
    # Sort by Activity ID for better readability
    df_plan_sorted = df_plan.sort_values('Activity ID').reset_index(drop=True)
    
    # Count important activities
    matched_indices = list(matched_activities.values())
    important_count = len(matched_indices)
    
    print(f"\nExporting all {len(df_plan_sorted)} rows with {important_count} important activities marked")
    
    # Generate output filename
    input_path = Path(project_plan_path)
    current_date = datetime.now().strftime('%Y-%m-%d')
    output_filename = f"{input_path.stem}_process_{current_date}{input_path.suffix}"
    output_path = input_path.parent / output_filename
    
    # Save to Excel with formatting
    print(f"\nSaving output to: {output_path}")
    
    # Use openpyxl for formatting
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.worksheet.filters import FilterColumn, Filters
    
    # First save the full dataframe with pandas
    df_plan_sorted.to_excel(output_path, index=False, engine='openpyxl')
    
    # Add conflicts sheet if there are any conflicts
    if conflicts:
        from openpyxl import load_workbook as load_wb_for_conflicts
        wb_conflicts = load_wb_for_conflicts(output_path)
        ws_conflicts = wb_conflicts.create_sheet(title='Conflicts')
        
        # Add header
        ws_conflicts['A1'] = 'Conflict Description'
        ws_conflicts.column_dimensions['A'].width = 100
        
        # Add conflicts
        for idx, conflict in enumerate(conflicts, start=2):
            ws_conflicts[f'A{idx}'] = conflict
            ws_conflicts[f'A{idx}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        wb_conflicts.save(output_path)
        print(f"Added 'Conflicts' sheet with {len(conflicts)} conflict(s)")
    
    # Load workbook for formatting the main sheet
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Find column indices for category columns, Activity Name, and Similarity Score
    activity_name_col = None
    similarity_score_col = None
    important_for_gp_col = None
    category_col_indices = []
    for idx, col in enumerate(df_plan_sorted.columns, 1):
        if col == 'Activity Name':
            activity_name_col = idx
        if col == 'Similarity Score %':
            similarity_score_col = idx
        if col == 'Important for GP':
            important_for_gp_col = idx
        if col in category_columns:
            category_col_indices.append(idx)
    
    # Add autofilter to top row
    ws.auto_filter.ref = ws.dimensions
    
    # Apply filter to show only important activities (rows with "X" in "Important for GP" column)
    if important_for_gp_col:
        # Filter to show only rows with "X" in Important for GP column
        filter_col = FilterColumn(colId=important_for_gp_col - 1)  # 0-indexed
        filter_col.filters = Filters()
        filter_col.filters.filter.append('X')
        ws.auto_filter.filterColumn.append(filter_col)
    
    # Auto-width all columns first
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap at 100
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set specific widths for columns A and B
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 75
    
    # Apply wrap text and alignment to all cells
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell_idx, cell in enumerate(row, 1):
            # Center align category columns horizontally and vertically
            if cell_idx in category_col_indices:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            else:
                # All other cells: wrap text and center vertically
                cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # Highlight Activity Name cells with score < 95%
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for low_score_idx in low_score_rows:
        # Excel rows start at 1, and we have a header row, so add 2
        excel_row = low_score_idx + 2
        if activity_name_col:
            cell = ws.cell(row=excel_row, column=activity_name_col)
            cell.fill = yellow_fill
    
    # Auto-fit row heights
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].auto_size = True
    
    # Freeze panes at C2 (freezes first row and columns A-B)
    ws.freeze_panes = 'C2'
    
    # Save formatted workbook
    wb.save(output_path)
    
    print(f"\n[OK] Successfully created {output_filename}")
    print(f"  - Total rows exported: {len(df_plan_sorted)}")
    print(f"  - Important activities: {important_count} (filtered by default)")
    print(f"  - Activities with score < 95%: {len(low_score_rows)} (highlighted in yellow)")
    print(f"  - Activities with calculated dates: {df_plan_sorted['Requested date'].notna().sum()}")
    if conflicts:
        print(f"  - Conflicts detected: {len(conflicts)} (see 'Conflicts' sheet)")
    print(f"  - Applied autofilter (showing important activities), column widths, wrap text, and freeze panes at C2")
    
    return output_path, matched_count, match_details


def main():
    # File paths
    project_plan_path = r'C:\Users\szil\Repos\excel_wizadry\PTS_check\PTS_project_specific.xlsx'
    general_file_path = r'C:\Users\szil\Repos\excel_wizadry\PTS_check\PTS_General_input_data.xlsx'
    
    # Check if files exist
    if not Path(project_plan_path).exists():
        print(f"Error: Project plan file not found: {project_plan_path}")
        sys.exit(1)
    
    if not Path(general_file_path).exists():
        print(f"Error: General file not found: {general_file_path}")
        sys.exit(1)
    
    # Process files
    try:
        output_path, matched_count, match_details = process_pts_file(
            project_plan_path, 
            general_file_path, 
            similarity_threshold=90
        )
        print("\n" + "="*80)
        print("Processing complete!")
        print("="*80)
        
    except Exception as e:
        print(f"\nError processing files: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
