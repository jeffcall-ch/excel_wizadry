"""
PTS Activity Matcher Script
Matches activities from project plan with important activities from general file
using fuzzy string matching and adds category markers.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from thefuzz import fuzz
import sys
import html
import re


def find_rightmost_used_column(df):
    """Find the index of the rightmost column with data."""
    for i in range(len(df.columns) - 1, -1, -1):
        if df.iloc[:, i].notna().sum() > 1:  # More than 1 non-null value
            return i
    return len(df.columns) - 1


def match_activity(activity_name, important_activities, threshold=90):
    """
    Match an activity name against a list of important activities using fuzzy matching.
    Returns a list of tuples (activity, score, categories) for matches above threshold.
    """
    if pd.isna(activity_name):
        return []
    
    activity_name_lower = str(activity_name).lower().strip()
    matches = []
    
    for _, row in important_activities.iterrows():
        important_activity = str(row['Activity']).lower().strip()
        
        # Calculate similarity score
        score = fuzz.ratio(activity_name_lower, important_activity)
        
        if score >= threshold:
            # Extract categories where value is 'x' (case insensitive)
            categories = {}
            for col in ['Important for GP', 'Piping', 'Hangers', 'Additional Material', 'Valves']:
                if pd.notna(row[col]) and str(row[col]).lower().strip() == 'x':
                    categories[col] = 'X'
                else:
                    categories[col] = ''
            
            matches.append((row['Activity'], score, categories))
    
    return matches


def format_date_swiss(date_value):
    """
    Convert date to Swiss format dd/mm/yyyy, removing any extra text like 'A'.
    Handles both datetime objects and strings.
    Handles 2-digit years with asterisk (e.g., '26*' means 2026).
    """
    if pd.isna(date_value):
        return ''
    
    date_str = str(date_value).strip()
    
    # Remove trailing ' A' or similar letters (but preserve asterisks for now)
    date_str = date_str.rstrip(' ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    
    try:
        # Try parsing as datetime if it's a timestamp string
        if isinstance(date_value, pd.Timestamp):
            return date_value.strftime('%d/%m/%Y')
        
        # Try parsing common formats
        # Format: dd.mm.yy or dd.mm.yy*
        if '.' in date_str and len(date_str.split('.')) == 3:
            parts = date_str.split('.')
            day, month, year = parts[0], parts[1], parts[2]
            
            # Handle asterisk in year (e.g., '26*')
            year = year.rstrip('*')
            
            # Handle 2-digit year
            if len(year) == 2:
                year = '20' + year if int(year) < 50 else '19' + year
            return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
        
        # Try pandas datetime parsing
        dt = pd.to_datetime(date_value, errors='coerce')
        if pd.notna(dt):
            return dt.strftime('%d/%m/%Y')
    except:
        pass
    
    return date_str


def parse_holiday_periods(df_holidays, year):
    """
    Parse holiday periods from text descriptions to actual date ranges.
    
    Args:
        df_holidays: DataFrame with columns ['From', 'To', 'Length adder weeks']
        year: Year to apply the holiday periods to
    
    Returns:
        List of tuples: (start_date, end_date, adder_weeks, description)
    """
    periods = []
    
    month_map = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'may': 5, 'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    
    for _, row in df_holidays.iterrows():
        from_text = str(row['From']).lower()
        to_text = str(row['To']).lower()
        adder_weeks = int(row['Length adder weeks'])
        
        # Parse "middle of Month" or "end of Month"
        from_month = None
        from_day = 1
        to_month = None
        to_day = 1
        
        for month_name, month_num in month_map.items():
            if month_name in from_text:
                from_month = month_num
                if 'middle' in from_text:
                    from_day = 15
                elif 'end' in from_text:
                    from_day = 28 if month_num == 2 else (30 if month_num in [4, 6, 9, 11] else 31)
                break
        
        for month_name, month_num in month_map.items():
            if month_name in to_text:
                to_month = month_num
                if 'middle' in to_text:
                    to_day = 15
                elif 'end' in to_text:
                    to_day = 28 if month_num == 2 else (30 if month_num in [4, 6, 9, 11] else 31)
                break
        
        if from_month and to_month:
            start_date = datetime(year, from_month, from_day)
            end_date = datetime(year, to_month, to_day)
            description = f"{row['From']} to {row['To']} holiday period"
            periods.append((start_date, end_date, adder_weeks, description))
    
    return periods


def check_holiday_conflict(start_date, end_date, holiday_periods):
    """
    Check if an activity execution period overlaps with a holiday period or within 1 week after.
    
    Args:
        start_date: datetime object - activity start date
        end_date: datetime object - activity end date
        holiday_periods: List of tuples (start_date, end_date, adder_weeks, description)
    
    Returns:
        Tuple (has_conflict, adder_weeks, description) or (False, 0, None)
    """
    for holiday_start, holiday_end, adder_weeks, description in holiday_periods:
        # Extend holiday period by 1 week after
        extended_holiday_end = holiday_end + timedelta(weeks=1)
        
        # Check if activity execution period overlaps with holiday period (including 1 week after)
        # Activity overlaps if: activity_start <= extended_holiday_end AND activity_end >= holiday_start
        if start_date <= extended_holiday_end and end_date >= holiday_start:
            return (True, adder_weeks, description)
    
    return (False, 0, None)


def calculate_requested_dates(df_plan, df_general, df_constraints, matched_activities, df_holidays=None):
    """
    Calculate requested dates by working backwards AND forwards from the mobilisation milestone.
    Detects and reports conflicts when multiple dependencies try to set different dates.
    Ensures all calculated dates fall on weekdays (Monday-Friday).
    Applies holiday adjustments if holiday periods are provided.
    
    Args:
        df_plan: Project plan dataframe
        df_general: General activities dataframe
        df_constraints: Constraints/relationships dataframe
        matched_activities: Dictionary mapping important activity names to their row indices in df_plan
        df_holidays: Optional DataFrame with holiday periods
    
    Returns:
        Tuple of (conflicts list, dependency_map) where dependency_map tracks formula relationships
    """
    
    def ensure_weekday(date):
        """Adjust date to next weekday if it falls on weekend."""
        # 5 = Saturday, 6 = Sunday
        if date.weekday() == 5:  # Saturday
            return date + timedelta(days=2)  # Move to Monday
        elif date.weekday() == 6:  # Sunday
            return date + timedelta(days=1)  # Move to Monday
        return date
    
    # Find the mobilisation milestone in the matched activities
    mobilisation_name = "Lot 5 - Erection General Piping - Contractor On Site Mobilisation"
    
    if mobilisation_name not in matched_activities:
        print(f"Warning: Mobilisation milestone '{mobilisation_name}' not found in matched activities")
        return [], {}
    
    mobilisation_idx = matched_activities[mobilisation_name]
    mobilisation_start = df_plan.at[mobilisation_idx, 'Start']
    
    # Convert to datetime
    try:
        anchor_date = pd.to_datetime(mobilisation_start, format='%d/%m/%Y')
        anchor_date = ensure_weekday(anchor_date)  # Ensure anchor is on weekday
        print(f"\nAnchor point: {mobilisation_name}")
        print(f"  Start date: {anchor_date.strftime('%d/%m/%Y')}")
    except Exception as e:
        print(f"Error parsing mobilisation start date: {e}")
        return [], {}
    
    # Parse holiday periods if provided
    holiday_periods = []
    if df_holidays is not None:
        anchor_year = anchor_date.year
        # Generate holiday periods for anchor year and adjacent years
        for year in [anchor_year - 1, anchor_year, anchor_year + 1]:
            holiday_periods.extend(parse_holiday_periods(df_holidays, year))
        print(f"  Loaded {len(holiday_periods) // 3} holiday period(s) for years {anchor_year-1} to {anchor_year+1}")
    
    # Build a dictionary to store calculated dates and their sources
    # Structure: {activity_name: {'date': datetime, 'sources': [list of calculation sources]}}
    activity_dates = {
        mobilisation_name: {
            'date': anchor_date,
            'sources': ['Anchor milestone (reference point)']
        }
    }
    
    # Track dependency relationships for formula building
    # Structure: {activity_name: {'reference_activity': name, 'weeks': num, 'direction': 'before'/'after'}}
    dependency_map = {}
    dependency_map[mobilisation_name] = {'reference': 'anchor', 'weeks': 0, 'direction': 'anchor'}
    
    # Track conflicts
    conflicts = []
    
    # Process constraints - work both backwards and forwards
    max_iterations = 100  # Prevent infinite loops
    iteration = 0
    
    while iteration < max_iterations:
        iteration += 1
        progress_made = False
        
        for _, constraint in df_constraints.iterrows():
            from_activity = constraint['From']
            to_activity = constraint['To']
            duration_weeks = constraint['Duration weeks']
            comment = constraint['Comments']
            
            # Skip if duration is not valid
            if pd.isna(duration_weeks):
                continue
            
            # Build comment description
            comment_suffix = f". {comment}" if pd.notna(comment) and str(comment).strip() else ""
            
            # BACKWARD calculation: if "To" date is known but "From" date needs to be calculated/verified
            if to_activity in activity_dates and from_activity in matched_activities:
                to_date = activity_dates[to_activity]['date']
                from_date_initial = to_date - timedelta(weeks=duration_weeks)
                from_date_initial = ensure_weekday(from_date_initial)  # Ensure weekday
                
                # Check if activity execution period (from_date_initial TO to_date) overlaps with holidays
                has_conflict, adder_weeks, holiday_desc = check_holiday_conflict(from_date_initial, to_date, holiday_periods)
                if has_conflict:
                    # For backward calculation, add adder to duration (start earlier)
                    total_weeks = duration_weeks + adder_weeks
                    from_date = to_date - timedelta(weeks=total_weeks)
                    from_date = ensure_weekday(from_date)
                    reason = f"{int(total_weeks)}W before '{to_activity}'{comment_suffix}. Originally {from_date_initial.strftime('%d/%m/%Y')} (based on {int(duration_weeks)}W), shifted by {adder_weeks}W due to {holiday_desc}"
                else:
                    from_date = from_date_initial
                    total_weeks = duration_weeks
                    reason = f"{int(duration_weeks)}W before '{to_activity}'{comment_suffix}"
                
                if from_activity in activity_dates:
                    # Check for conflict
                    existing_date = activity_dates[from_activity]['date']
                    if abs((existing_date - from_date).days) > 0:  # Different dates
                        conflict_msg = (
                            f"CONFLICT: '{from_activity}'\n"
                            f"  Existing: {existing_date.strftime('%d/%m/%Y')} from {activity_dates[from_activity]['sources'][-1]}\n"
                            f"  New calc: {from_date.strftime('%d/%m/%Y')} from {reason}"
                        )
                        conflicts.append(conflict_msg)
                        print(f"âš ï¸  {conflict_msg}")
                    else:
                        # Same date from different path - just add source
                        activity_dates[from_activity]['sources'].append(reason)
                else:
                    # New calculation
                    activity_dates[from_activity] = {
                        'date': from_date,
                        'sources': [reason]
                    }
                    # Store dependency for formula building
                    if from_activity not in dependency_map:
                        dependency_map[from_activity] = {
                            'reference': to_activity,
                            'weeks': total_weeks,
                            'direction': 'before'
                        }
                    progress_made = True
            
            # FORWARD calculation: if "From" date is known but "To" date needs to be calculated/verified
            if from_activity in activity_dates and to_activity in matched_activities:
                from_date = activity_dates[from_activity]['date']
                to_date_initial = from_date + timedelta(weeks=duration_weeks)
                to_date_initial = ensure_weekday(to_date_initial)  # Ensure weekday
                
                # Check if activity execution period (from_date TO to_date_initial) overlaps with holidays
                has_conflict, adder_weeks, holiday_desc = check_holiday_conflict(from_date, to_date_initial, holiday_periods)
                if has_conflict:
                    # For forward calculation, add adder to duration (end later)
                    total_weeks = duration_weeks + adder_weeks
                    to_date = from_date + timedelta(weeks=total_weeks)
                    to_date = ensure_weekday(to_date)
                    reason = f"{int(total_weeks)}W after '{from_activity}'{comment_suffix}. Originally {to_date_initial.strftime('%d/%m/%Y')} (based on {int(duration_weeks)}W), shifted by {adder_weeks}W due to {holiday_desc}"
                else:
                    to_date = to_date_initial
                    total_weeks = duration_weeks
                    reason = f"{int(duration_weeks)}W after '{from_activity}'{comment_suffix}"
                
                if to_activity in activity_dates:
                    # Check for conflict
                    existing_date = activity_dates[to_activity]['date']
                    if abs((existing_date - to_date).days) > 0:  # Different dates
                        conflict_msg = (
                            f"CONFLICT: '{to_activity}'\n"
                            f"  Existing: {existing_date.strftime('%d/%m/%Y')} from {activity_dates[to_activity]['sources'][-1]}\n"
                            f"  New calc: {to_date.strftime('%d/%m/%Y')} from {reason}"
                        )
                        conflicts.append(conflict_msg)
                        print(f"âš ï¸  {conflict_msg}")
                    else:
                        # Same date from different path - just add source
                        activity_dates[to_activity]['sources'].append(reason)
                else:
                    # New calculation
                    activity_dates[to_activity] = {
                        'date': to_date,
                        'sources': [reason]
                    }
                    # Store dependency for formula building
                    if to_activity not in dependency_map:
                        dependency_map[to_activity] = {
                            'reference': from_activity,
                            'weeks': total_weeks,
                            'direction': 'after'
                        }
                    progress_made = True
        
        # If no progress was made, we're done
        if not progress_made:
            break
    
    print(f"\nCalculated dates for {len(activity_dates)} activities")
    if conflicts:
        print(f"âš ï¸  Found {len(conflicts)} conflict(s) - see details above")
    
    # Now update the df_plan with calculated dates
    for activity_name, date_info in activity_dates.items():
        if activity_name in matched_activities:
            row_idx = matched_activities[activity_name]
            
            # Set requested date
            df_plan.at[row_idx, 'Requested date'] = date_info['date'].strftime('%d/%m/%Y')
            
            # Set comment (use the first source)
            df_plan.at[row_idx, 'Comment'] = date_info['sources'][0]
    
    return conflicts, dependency_map


def calculate_current_lengths(df_plan, df_constraints, matched_activities):
    """
    Calculate the current length in weeks between dependent activities based on their actual dates in the plan.
    Uses Finish dates for all activities except the anchor (which uses Start date).
    Calculates for the same activity pairs and direction as the requested dates/comments.
    
    Args:
        df_plan: Project plan dataframe
        df_constraints: Constraints/relationships dataframe
        matched_activities: Dictionary mapping important activity names to their row indices in df_plan
    """
    from datetime import timedelta
    
    # Anchor activity uses Start date, all others use Finish date
    anchor_name = "Lot 5 - Erection General Piping - Contractor On Site Mobilisation"
    
    # Dictionary to track current lengths for each activity
    current_lengths = {}
    
    for _, constraint in df_constraints.iterrows():
        from_activity = constraint['From']
        to_activity = constraint['To']
        duration_weeks = constraint['Duration weeks']
        comment = constraint['Comments']
        
        # Skip if duration is not valid or activities not matched
        if pd.isna(duration_weeks) or from_activity not in matched_activities or to_activity not in matched_activities:
            continue
        
        # Get the row indices
        from_idx = matched_activities[from_activity]
        to_idx = matched_activities[to_activity]
        
        # Get the appropriate dates: Finish for most activities, Start for anchor
        # "From" activity: use Finish date
        from_date_value = df_plan.at[from_idx, 'Finish']
        
        # "To" activity: use Start date if it's the anchor, otherwise use Finish date
        if to_activity == anchor_name:
            to_date_value = df_plan.at[to_idx, 'Start']
        else:
            to_date_value = df_plan.at[to_idx, 'Finish']
        
        # Parse dates - skip if either is missing
        if pd.isna(from_date_value) or pd.isna(to_date_value) or from_date_value == '' or to_date_value == '':
            continue
        
        try:
            from_date = pd.to_datetime(from_date_value, format='%d/%m/%Y', dayfirst=True)
            to_date = pd.to_datetime(to_date_value, format='%d/%m/%Y', dayfirst=True)
            
            # Calculate actual duration in weeks
            actual_duration = (to_date - from_date).days / 7
            
            # Build description matching the comment format
            comment_suffix = f". {comment}" if pd.notna(comment) and str(comment).strip() else ""
            
            # BACKWARD relationship: Add to "From" activity (it comes BEFORE "To")
            backward_desc = f"{actual_duration:.1f}W (planned: {int(duration_weeks)}W) before '{to_activity}'{comment_suffix}"
            if from_activity not in current_lengths:
                current_lengths[from_activity] = []
            current_lengths[from_activity].append(backward_desc)
            
            # FORWARD relationship: Add to "To" activity (it comes AFTER "From")
            forward_desc = f"{actual_duration:.1f}W (planned: {int(duration_weeks)}W) after '{from_activity}'{comment_suffix}"
            if to_activity not in current_lengths:
                current_lengths[to_activity] = []
            current_lengths[to_activity].append(forward_desc)
            
        except Exception as e:
            print(f"Warning: Could not parse dates for '{from_activity}' -> '{to_activity}': {e}")
            continue
    
    # Update the dataframe with current lengths
    for activity_name, length_list in current_lengths.items():
        if activity_name in matched_activities:
            row_idx = matched_activities[activity_name]
            # Join multiple lengths with line breaks
            df_plan.at[row_idx, 'Current length'] = '\n'.join(length_list)
    
    print(f"Calculated current lengths for {len(current_lengths)} activities")


def _clean_text(value):
    """Return safe string representation for export."""
    if pd.isna(value):
        return ''
    return str(value).strip()


def _parse_date(value):
    """Parse date values robustly using day-first format."""
    if pd.isna(value) or str(value).strip() == '':
        return None
    dt = pd.to_datetime(value, dayfirst=True, errors='coerce')
    if pd.isna(dt):
        return None
    return dt.to_pydatetime()


def _date_to_text(value):
    """Return date text in dd/mm/yyyy or empty."""
    dt = _parse_date(value)
    if dt is None:
        return ''
    return dt.strftime('%d/%m/%Y')


def validate_generated_html_layout(html_path):
    """
    Read generated HTML and check whether node and edge label text should fit the SVG geometry.
    Returns (is_ok, issues_list).
    """
    content = html_path.read_text(encoding='utf-8')
    issues = []

    # Check edge span vs label width (requires clear space for readable label and arrow).
    edge_spans = [
        float(m.group(7)) - float(m.group(1))
        for m in re.finditer(
            r"<path d='M ([0-9.]+),([0-9.]+) C ([0-9.]+),([0-9.]+) ([0-9.]+),([0-9.]+) ([0-9.]+),([0-9.]+)' stroke='",
            content,
        )
    ]
    label_widths = [
        float(m.group(1))
        for m in re.finditer(r"<rect x='[0-9.\-]+' y='[0-9.\-]+' width='([0-9.]+)' height='24' rx='8' ry='8' class='edgeLabelBg'/>", content)
    ]

    for idx, (span, lw) in enumerate(zip(edge_spans, label_widths), 1):
        if span < (lw + 18):
            issues.append(f"edge-{idx}: span {span:.1f}px < label width {lw:.1f}px + margin")

    # Check REQ label should stay inside each node rect.
    node_blocks = re.findall(
        r"<g class=\"nodeGroup\">(.*?)</g>",
        content,
        flags=re.S,
    )
    for idx, block in enumerate(node_blocks, 1):
        rect = re.search(r"<rect x=\"([0-9.\-]+)\" y=\"([0-9.\-]+)\" width=\"([0-9.]+)\" height=\"([0-9.]+)\"", block)
        req = re.search(r"class=\"nodeMeta\">REQ: .*?</text>", block)
        req_y_match = re.search(r"<text x=\"[0-9.\-]+\" y=\"([0-9.\-]+)\" class=\"nodeMeta\">REQ:", block)
        if rect and req and req_y_match:
            y = float(rect.group(2))
            h = float(rect.group(4))
            req_y = float(req_y_match.group(1))
            if req_y > (y + h - 6):
                issues.append(f"node-{idx}: REQ baseline outside rect")

    return (len(issues) == 0, issues)


def generate_visualization_html(
    df_plan,
    df_general,
    df_constraints,
    matched_activities,
    dependency_map,
    html_output_path,
    df_holidays=None,
    conflicts=None,
    unmatched_activities=None,
    layout_profile='normal',
    _auto_retry=True,
):
    """
    Create an HTML report that visualizes current plan dates, requested dates, and dependencies.
    """
    conflicts = conflicts or []
    unmatched_activities = unmatched_activities or []
    category_columns = ['Important for GP', 'Piping', 'Hangers', 'Additional Material', 'Valves']

    records = []
    for _, general_row in df_general.iterrows():
        important_activity = _clean_text(general_row.get('Activity', ''))
        if not important_activity:
            continue

        if important_activity not in matched_activities:
            records.append({
                'important_activity': important_activity,
                'matched': False,
            })
            continue

        row_idx = matched_activities[important_activity]
        plan_row = df_plan.loc[row_idx]

        record = {
            'important_activity': important_activity,
            'matched': True,
            'activity_id': _clean_text(plan_row.get('Activity ID', '')),
            'activity_name': _clean_text(plan_row.get('Activity Name', '')),
            'start': _date_to_text(plan_row.get('Start', '')),
            'finish': _date_to_text(plan_row.get('Finish', '')),
            'requested': _date_to_text(plan_row.get('Requested date', '')),
            'current_length': _clean_text(plan_row.get('Current length', '')),
            'comment': _clean_text(plan_row.get('Comment', '')),
            'similarity': _clean_text(plan_row.get('Similarity Score %', '')),
            'categories': [col for col in category_columns if _clean_text(plan_row.get(col, '')).upper() == 'X'],
        }
        records.append(record)

    record_by_activity = {
        r['important_activity']: r for r in records if r.get('matched')
    }

    # Date range for timeline axis
    timeline_dates = []
    for r in records:
        if not r.get('matched'):
            continue
        for key in ['start', 'finish', 'requested']:
            dt = _parse_date(r.get(key, ''))
            if dt is not None:
                timeline_dates.append(dt)

    if timeline_dates:
        min_date = min(timeline_dates)
        max_date = max(timeline_dates)
    else:
        min_date = datetime.now()
        max_date = min_date + timedelta(days=1)

    total_days = max((max_date - min_date).days, 1)

    def to_percent(date_text):
        dt = _parse_date(date_text)
        if dt is None:
            return None
        return max(0, min(100, ((dt - min_date).days / total_days) * 100))

    # Build dependency data for table and arrow visualization
    anchor_name = "Lot 5 - Erection General Piping - Contractor On Site Mobilisation"
    holiday_periods = []
    if df_holidays is not None and len(df_holidays) > 0:
        requested_dates = []
        if 'Requested date' in df_plan.columns:
            for idx in matched_activities.values():
                dt = _parse_date(df_plan.at[idx, 'Requested date'])
                if dt is not None:
                    requested_dates.append(dt)
        years = [datetime.now().year]
        if requested_dates:
            years = list(range(min(d.year for d in requested_dates) - 1, max(d.year for d in requested_dates) + 2))
        for y in years:
            holiday_periods.extend(parse_holiday_periods(df_holidays, y))
    holiday_info_available = len(holiday_periods) > 0

    dependency_rows = []
    for _, constraint in df_constraints.iterrows():
        from_activity = _clean_text(constraint.get('From', ''))
        to_activity = _clean_text(constraint.get('To', ''))
        duration_weeks_raw = constraint.get('Duration weeks', '')
        duration_weeks = _clean_text(duration_weeks_raw)
        comment = _clean_text(constraint.get('Comments', ''))
        if from_activity in matched_activities and to_activity in matched_activities:
            from_idx = matched_activities[from_activity]
            to_idx = matched_activities[to_activity]

            from_finish = _parse_date(df_plan.at[from_idx, 'Finish']) if 'Finish' in df_plan.columns else None
            to_date_key = 'Start' if to_activity == anchor_name else 'Finish'
            to_value = df_plan.at[to_idx, to_date_key] if to_date_key in df_plan.columns else None
            to_ref_date = _parse_date(to_value)

            actual_weeks = None
            delta_weeks = None
            try:
                planned_weeks_val = float(duration_weeks_raw)
            except Exception:
                planned_weeks_val = None

            std_weeks_effective = planned_weeks_val
            holiday_adder = 0
            # Use requested dates to estimate holiday-adjusted standard duration.
            from_req = _parse_date(df_plan.at[from_idx, 'Requested date']) if 'Requested date' in df_plan.columns else None
            to_req = _parse_date(df_plan.at[to_idx, 'Requested date']) if 'Requested date' in df_plan.columns else None
            if planned_weeks_val is not None and from_req is not None and to_req is not None and holiday_periods:
                req_start = min(from_req, to_req)
                req_end = max(from_req, to_req)
                has_conflict, adder_weeks, _ = check_holiday_conflict(req_start, req_end, holiday_periods)
                if has_conflict:
                    holiday_adder = int(adder_weeks)
                    std_weeks_effective = planned_weeks_val + holiday_adder

            if from_finish is not None and to_ref_date is not None:
                actual_weeks = (to_ref_date - from_finish).days / 7
                if std_weeks_effective is not None:
                    delta_weeks = actual_weeks - std_weeks_effective

            dependency_rows.append({
                'from_activity': from_activity,
                'to_activity': to_activity,
                'duration_weeks': f"{std_weeks_effective:.1f}" if std_weeks_effective is not None else duration_weeks,
                'duration_weeks_base': f"{planned_weeks_val:.1f}" if planned_weeks_val is not None else duration_weeks,
                'holiday_adder': holiday_adder,
                'comment': comment,
                'from_date': from_finish.strftime('%d/%m/%Y') if from_finish else '',
                'to_date': to_ref_date.strftime('%d/%m/%Y') if to_ref_date else '',
                'actual_weeks': f"{actual_weeks:.1f}" if actual_weeks is not None else '',
                'delta_weeks': f"{delta_weeks:+.1f}" if delta_weeks is not None else '',
                'status_class': (
                    'ok' if delta_weeks is not None and abs(delta_weeks) < 1.0
                    else 'bad' if actual_weeks is not None and std_weeks_effective is not None
                    else 'na'
                ),
            })

    activity_rows_html = []
    for idx, r in enumerate([x for x in records if x.get('matched')], 1):
        start_pct = to_percent(r.get('start', ''))
        finish_pct = to_percent(r.get('finish', ''))
        requested_pct = to_percent(r.get('requested', ''))

        range_left = None
        range_width = None
        if start_pct is not None and finish_pct is not None:
            range_left = min(start_pct, finish_pct)
            range_width = max(abs(finish_pct - start_pct), 1.0)

        cat_text = ', '.join(r['categories']) if r['categories'] else 'None'
        current_length_html = html.escape(r.get('current_length', '')).replace('\n', '<br>')
        comment_html = html.escape(r.get('comment', '')).replace('\n', '<br>')

        bar_html = ''
        if range_left is not None and range_width is not None:
            bar_html += (
                f"<div class='bar current' style='left:{range_left:.2f}%;width:{range_width:.2f}%;' "
                f"title='Current span: {html.escape(r.get('start', ''))} to {html.escape(r.get('finish', ''))}'></div>"
            )
        if requested_pct is not None:
            bar_html += (
                f"<div class='pin requested' style='left:{requested_pct:.2f}%;' "
                f"title='Requested date: {html.escape(r.get('requested', ''))}'></div>"
            )

        activity_rows_html.append(
            f"""
            <tr data-categories="{html.escape(' '.join(r['categories']).lower())}">
              <td>{idx}</td>
              <td class="name">{html.escape(r.get('activity_name', ''))}</td>
              <td>{html.escape(r.get('activity_id', ''))}</td>
              <td>{html.escape(r.get('similarity', ''))}</td>
              <td>{html.escape(r.get('start', ''))}</td>
              <td>{html.escape(r.get('finish', ''))}</td>
              <td>{html.escape(r.get('requested', ''))}</td>
              <td class="pre">{current_length_html}</td>
              <td class="pre">{comment_html}</td>
              <td>{html.escape(cat_text)}</td>
              <td>
                <div class="track">
                  {bar_html}
                </div>
              </td>
            </tr>
            """
        )

    unmatched_html = ''.join(f"<li>{html.escape(x)}</li>" for x in unmatched_activities)
    conflicts_html = ''.join(f"<li>{html.escape(x)}</li>" for x in conflicts)
    dependency_html = ''.join(
        f"<tr><td>{html.escape(x['from_activity'])}</td><td>{html.escape(x['to_activity'])}</td>"
        f"<td>{html.escape(x['duration_weeks'])}</td><td>{html.escape(x['actual_weeks'])}</td>"
        f"<td>{html.escape(x['delta_weeks'])}</td><td>{html.escape(x['comment'])}</td></tr>"
        for x in dependency_rows
    )
    dependency_flow_html = ''.join(
        f"""
        <div class="dep-flow">
          <div class="node from">
            <div class="node-title">{html.escape(x['from_activity'])}</div>
            <div class="node-meta">PTS ref date: {html.escape(x['from_date']) or 'n/a'}</div>
          </div>
          <div class="edge-wrap">
            <div class="edge-line"></div>
            <div class="edge-head"></div>
            <div class="edge-label {x['status_class']}">
              STD: {html.escape(x['duration_weeks']) or 'n/a'}W
              &nbsp;|&nbsp;
              PTS: {html.escape(x['actual_weeks']) or 'n/a'}W
              &nbsp;|&nbsp;
              DELTA: {html.escape(x['delta_weeks']) or 'n/a'}W
            </div>
          </div>
          <div class="node to">
            <div class="node-title">{html.escape(x['to_activity'])}</div>
            <div class="node-meta">PTS ref date: {html.escape(x['to_date']) or 'n/a'}</div>
          </div>
        </div>
        """
        for x in dependency_rows
    )

    # Build left-to-right branch graph (SVG)
    def wrap_text(text, max_chars=36):
        words = str(text).split()
        if not words:
            return ['']
        lines = []
        current = words[0]
        for w in words[1:]:
            if len(current) + 1 + len(w) <= max_chars:
                current += ' ' + w
            else:
                lines.append(current)
                current = w
        lines.append(current)
        return lines

    def as_float(value):
        try:
            return float(value)
        except Exception:
            return None

    graph_nodes = set()
    outgoing = {}
    incoming = {}
    edge_info = {}
    for d in dependency_rows:
        u = d['from_activity']
        v = d['to_activity']
        graph_nodes.add(u)
        graph_nodes.add(v)
        outgoing.setdefault(u, []).append(v)
        incoming.setdefault(v, []).append(u)
        outgoing.setdefault(v, [])
        incoming.setdefault(u, [])
        edge_info[(u, v)] = d

    indegree = {n: len(incoming.get(n, [])) for n in graph_nodes}
    queue = sorted([n for n in graph_nodes if indegree[n] == 0])
    topo = []
    while queue:
        n = queue.pop(0)
        topo.append(n)
        for m in outgoing.get(n, []):
            indegree[m] -= 1
            if indegree[m] == 0:
                queue.append(m)
                queue.sort()

    if len(topo) != len(graph_nodes):
        topo = sorted(graph_nodes)

    levels = {n: 0 for n in graph_nodes}
    for n in topo:
        for m in outgoing.get(n, []):
            levels[m] = max(levels[m], levels[n] + 1)

    profile = str(layout_profile).lower().strip()
    if profile == 'safe':
        box_w = 430
        wrap_chars = 32
        px_per_week = 24
        min_arrow_span = 300
        name_line_h = 22
        meta_line_h = 19
    else:
        box_w = 390
        wrap_chars = 37
        px_per_week = 22
        min_arrow_span = 250
        name_line_h = 22
        meta_line_h = 18

    gap_y = 22
    margin_x, margin_y = 28, 28
    min_level_stride = box_w + 70

    node_dims = {}
    for n in graph_nodes:
        name_lines = wrap_text(n, max_chars=wrap_chars)
        top_pad = 14
        bottom_pad = 12
        box_h = max(top_pad + len(name_lines) * name_line_h + 2 * meta_line_h + bottom_pad, 94)
        node_dims[n] = {
            'w': box_w,
            'h': box_h,
            'name_lines': name_lines,
            'name_line_h': name_line_h,
            'meta_line_h': meta_line_h,
            'top_pad': top_pad,
        }

    level_nodes = {}
    for n, lvl in levels.items():
        level_nodes.setdefault(lvl, []).append(n)

    edge_requirements = {}
    step_values = []
    for d in dependency_rows:
        # Use effective STD duration (already holiday-adjusted when applicable) as spacing driver.
        step = as_float(d.get('duration_weeks', ''))
        if step is None:
            step = as_float(d.get('actual_weeks', ''))
        step = max(step if step is not None else 1.0, 0.5)
        step_values.append(step)
    min_step = min(step_values) if step_values else 1.0

    for (u, v), d in edge_info.items():
        step = as_float(d.get('duration_weeks', ''))
        if step is None:
            step = as_float(d.get('actual_weeks', ''))
        step = max(step if step is not None else 1.0, 0.5)
        holiday_adder = int(d.get('holiday_adder', 0) or 0)
        if holiday_info_available and holiday_adder > 0:
            std_label = f"{d['duration_weeks']}W (base {d['duration_weeks_base']}W, holiday +{holiday_adder}W)"
        else:
            std_label = f"{d['duration_weeks']}W"
        lbl = f"STD {std_label} | PTS {d['actual_weeks']}W | DELTA {d['delta_weeks']}W"
        label_w = max(190, len(lbl) * 7.4)
        # Smallest duration gets minimum readable arrow span; larger durations always add extra length.
        min_readable_span = max(min_arrow_span, label_w + 28)
        proportional_extra = max(0.0, (step - min_step)) * px_per_week
        req_span = min_readable_span + proportional_extra
        edge_requirements[(u, v)] = {
            'label': lbl,
            'label_w': label_w,
            'span': req_span,
        }

    # Solve x positions using edge span constraints, preserving left->right order and proportional duration.
    node_x = {n: margin_x for n in graph_nodes}
    for n in topo:
        node_x[n] = max(node_x[n], margin_x + levels[n] * min_level_stride)
        for m in outgoing.get(n, []):
            req = edge_requirements[(n, m)]['span']
            needed_x = node_x[n] + node_dims[n]['w'] + req
            if needed_x > node_x[m]:
                node_x[m] = needed_x

    node_pos = {}
    for lvl in sorted(level_nodes.keys()):
        nodes_here = sorted(level_nodes[lvl])
        if lvl > 0:
            nodes_here = sorted(
                nodes_here,
                key=lambda name: (
                    sum((node_pos[p][1] + node_dims[p]['h'] / 2) for p in incoming.get(name, []) if p in node_pos) / max(
                        len([p for p in incoming.get(name, []) if p in node_pos]), 1
                    )
                ),
            )
        y = margin_y
        for n in nodes_here:
            x = node_x[n]
            node_pos[n] = [x, y]
            y += node_dims[n]['h'] + gap_y

    max_x = max((node_pos[n][0] + node_dims[n]['w'] for n in node_pos), default=0)
    max_y = max((node_pos[n][1] + node_dims[n]['h'] for n in node_pos), default=0)
    svg_w = int(max_x + margin_x + 40)
    svg_h = int(max_y + margin_y + 20)

    def resolve_label_y(lx, base_ly, label_w, placed):
        """Nudge label up/down to avoid overlap with existing labels."""
        h = 24.0
        offsets = [0, -30, 30, -60, 60, -90, 90, -120, 120]
        for off in offsets:
            ly = base_ly + off
            hit = False
            for (px, py, pw, ph) in placed:
                if abs(lx - px) < ((label_w + pw) / 2 + 8) and abs(ly - py) < ((h + ph) / 2 + 6):
                    hit = True
                    break
            if not hit:
                placed.append((lx, ly, label_w, h))
                return ly
        ly = base_ly + 150
        placed.append((lx, ly, label_w, h))
        return ly

    edge_svg_parts = []
    placed_edge_labels = []
    for (u, v), d in edge_info.items():
        ux, uy = node_pos[u]
        vx, vy = node_pos[v]
        uw = node_dims[u]['w']
        uh = node_dims[u]['h']
        vh = node_dims[v]['h']
        x1, y1 = ux + uw, uy + uh / 2
        x2, y2 = vx, vy + vh / 2
        span = x2 - x1
        c1x = x1 + max(28, span * 0.45)
        c2x = x2 - max(28, span * 0.45)
        path = f"M {x1:.1f},{y1:.1f} C {c1x:.1f},{y1:.1f} {c2x:.1f},{y2:.1f} {x2:.1f},{y2:.1f}"

        status = d['status_class']
        color = '#6b7f95'
        if status == 'ok':
            color = '#1f7a3f'
        elif status == 'bad':
            color = '#b42318'

        lx = (x1 + x2) / 2
        base_ly = (y1 + y2) / 2 - 10
        lbl = edge_requirements[(u, v)]['label']
        label_w = edge_requirements[(u, v)]['label_w']
        ly = resolve_label_y(lx, base_ly, label_w, placed_edge_labels)
        edge_svg_parts.append(
            f"<path d='{path}' stroke='{color}' stroke-width='2.4' fill='none' marker-end='url(#arrow-{status})' opacity='0.95'/>"
            f"<rect x='{(lx - label_w / 2):.1f}' y='{(ly - 17):.1f}' width='{label_w:.1f}' height='24' rx='8' ry='8' class='edgeLabelBg'/>"
            f"<text x='{lx:.1f}' y='{ly:.1f}' class='edgeText {status}' text-anchor='middle'><title>{html.escape(lbl)}</title>{html.escape(lbl)}</text>"
        )

    node_svg_parts = []
    for n in topo:
        if n not in node_pos:
            continue
        x, y = node_pos[n]
        dims = node_dims[n]
        node_record = record_by_activity.get(n, {})
        pts_date = node_record.get('finish', '') or node_record.get('start', '')
        req_date = node_record.get('requested', '')

        tspans = []
        line_y = y + dims['top_pad'] + dims['name_line_h'] - 2
        for line in dims['name_lines']:
            tspans.append(f"<tspan x='{x + 12}' y='{line_y}'><title>{html.escape(n)}</title>{html.escape(line)}</tspan>")
            line_y += dims['name_line_h']

        name_block_end_y = y + dims['top_pad'] + len(dims['name_lines']) * dims['name_line_h']
        pts_y = name_block_end_y + dims['meta_line_h']
        req_y = name_block_end_y + 2 * dims['meta_line_h']

        node_svg_parts.append(
            f"""
            <g class="nodeGroup">
              <title>{html.escape(n)}</title>
              <rect x="{x}" y="{y}" width="{dims['w']}" height="{dims['h']}" rx="10" ry="10" class="nodeRect"></rect>
              <text class="nodeName">{''.join(tspans)}</text>
              <text x="{x + 12}" y="{pts_y}" class="nodeMeta"><title>PTS date: {html.escape(pts_date or 'n/a')}</title>PTS: {html.escape(pts_date or 'n/a')}</text>
              <text x="{x + 12}" y="{req_y}" class="nodeMeta"><title>REQ date: {html.escape(req_date or 'n/a')}</title>REQ: {html.escape(req_date or 'n/a')}</text>
            </g>
            """
        )

    network_svg = f"""
    <svg viewBox="0 0 {svg_w} {svg_h}" width="{svg_w}" height="{svg_h}" role="img" aria-label="Dependency network">
      <defs>
        <marker id="arrow-ok" markerWidth="11" markerHeight="8" refX="9" refY="4" orient="auto">
          <path d="M0,0 L10,4 L0,8 z" fill="#1f7a3f"></path>
        </marker>
        <marker id="arrow-warn" markerWidth="11" markerHeight="8" refX="9" refY="4" orient="auto">
          <path d="M0,0 L10,4 L0,8 z" fill="#b45309"></path>
        </marker>
        <marker id="arrow-bad" markerWidth="11" markerHeight="8" refX="9" refY="4" orient="auto">
          <path d="M0,0 L10,4 L0,8 z" fill="#b42318"></path>
        </marker>
        <marker id="arrow-na" markerWidth="11" markerHeight="8" refX="9" refY="4" orient="auto">
          <path d="M0,0 L10,4 L0,8 z" fill="#6b7f95"></path>
        </marker>
      </defs>
      {''.join(edge_svg_parts)}
      {''.join(node_svg_parts)}
    </svg>
    """

    # REQ-only diagram (same topology, only requested-date flow information)
    req_edge_svg_parts = []
    placed_req_labels = []
    for (u, v), d in edge_info.items():
        ux, uy = node_pos[u]
        vx, vy = node_pos[v]
        uw = node_dims[u]['w']
        uh = node_dims[u]['h']
        vh = node_dims[v]['h']
        x1, y1 = ux + uw, uy + uh / 2
        x2, y2 = vx, vy + vh / 2
        span = x2 - x1
        c1x = x1 + max(28, span * 0.45)
        c2x = x2 - max(28, span * 0.45)
        path = f"M {x1:.1f},{y1:.1f} C {c1x:.1f},{y1:.1f} {c2x:.1f},{y2:.1f} {x2:.1f},{y2:.1f}"

        from_idx = matched_activities.get(u)
        to_idx = matched_activities.get(v)
        req_u = _parse_date(df_plan.at[from_idx, 'Requested date']) if from_idx is not None and 'Requested date' in df_plan.columns else None
        req_v = _parse_date(df_plan.at[to_idx, 'Requested date']) if to_idx is not None and 'Requested date' in df_plan.columns else None
        req_weeks = ''
        if req_u is not None and req_v is not None:
            req_weeks = f"{((req_v - req_u).days / 7):.1f}"

        holiday_adder = int(d.get('holiday_adder', 0) or 0)
        if holiday_info_available and holiday_adder > 0:
            std_holiday_text = f"STD {d['duration_weeks']}W (base {d['duration_weeks_base']}W, holiday +{holiday_adder}W)"
        else:
            std_holiday_text = f"STD {d['duration_weeks']}W"

        req_lbl = f"REQ {req_weeks}W | {std_holiday_text}" if req_weeks != '' else f"REQ n/a | {std_holiday_text}"
        req_label_w = max(130, len(req_lbl) * 7.0)
        lx = (x1 + x2) / 2
        base_ly = (y1 + y2) / 2 - 10
        ly = resolve_label_y(lx, base_ly, req_label_w, placed_req_labels)
        req_edge_svg_parts.append(
            f"<path d='{path}' stroke='#1d4ed8' stroke-width='2.4' fill='none' marker-end='url(#arrow-req)' opacity='0.95'/>"
            f"<rect x='{(lx - req_label_w / 2):.1f}' y='{(ly - 17):.1f}' width='{req_label_w:.1f}' height='24' rx='8' ry='8' class='reqEdgeLabelBg'/>"
            f"<text x='{lx:.1f}' y='{ly:.1f}' class='reqEdgeText' text-anchor='middle'><title>{html.escape(req_lbl)}</title>{html.escape(req_lbl)}</text>"
        )

    req_node_svg_parts = []
    for n in topo:
        if n not in node_pos:
            continue
        x, y = node_pos[n]
        dims = node_dims[n]
        node_record = record_by_activity.get(n, {})
        req_date = node_record.get('requested', '')

        tspans = []
        line_y = y + dims['top_pad'] + dims['name_line_h'] - 2
        for line in dims['name_lines']:
            tspans.append(f"<tspan x='{x + 12}' y='{line_y}'><title>{html.escape(n)}</title>{html.escape(line)}</tspan>")
            line_y += dims['name_line_h']
        req_y = y + dims['h'] - 14

        req_node_svg_parts.append(
            f"""
            <g class="nodeGroup">
              <title>{html.escape(n)}</title>
              <rect x="{x}" y="{y}" width="{dims['w']}" height="{dims['h']}" rx="10" ry="10" class="reqNodeRect"></rect>
              <text class="nodeName">{''.join(tspans)}</text>
              <text x="{x + 12}" y="{req_y}" class="reqNodeMeta"><title>REQ date: {html.escape(req_date or 'n/a')}</title>REQ: {html.escape(req_date or 'n/a')}</text>
            </g>
            """
        )

    req_network_svg = f"""
    <svg viewBox="0 0 {svg_w} {svg_h}" width="{svg_w}" height="{svg_h}" role="img" aria-label="REQ dependency network">
      <defs>
        <marker id="arrow-req" markerWidth="11" markerHeight="8" refX="9" refY="4" orient="auto">
          <path d="M0,0 L10,4 L0,8 z" fill="#1d4ed8"></path>
        </marker>
      </defs>
      {''.join(req_edge_svg_parts)}
      {''.join(req_node_svg_parts)}
    </svg>
    """

    report_title = f"PTS Visualization - {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    html_content = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(report_title)}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, Arial, sans-serif;
      background: #f4f6fb;
      color: #122233;
    }}
    .page {{
      padding: 14px;
    }}
    .title {{
      font-size: 24px;
      font-weight: 700;
      margin: 0 0 8px;
    }}
    .subtitle {{
      margin: 0 0 12px;
      font-size: 15px;
      color: #355067;
    }}
    .legend {{
      margin: 0 0 14px;
      font-size: 15px;
      color: #1f2937;
    }}
    .legend span {{
      display: inline-block;
      margin-right: 16px;
      font-weight: 700;
    }}
    .legend .ok {{ color: #1f7a3f; }}
    .legend .bad {{ color: #b42318; }}
    .legend .req {{ color: #1d4ed8; }}
    .canvas {{
      border: 2px solid #b9c5d5;
      border-radius: 12px;
      background: #fff;
      padding: 12px;
      overflow-x: auto;
      overflow-y: auto;
      width: 100%;
      max-height: calc(100vh - 170px);
    }}
    .nodeRect {{
      fill: #ffffff;
      stroke: #95a6bb;
      stroke-width: 2;
    }}
    .nodeName {{
      font-size: 16px;
      font-weight: 700;
      fill: #132a3a;
    }}
    .nodeMeta {{
      font-size: 13px;
      fill: #355067;
    }}
    .edgeText {{
      font-size: 14px;
      font-weight: 700;
      paint-order: stroke;
      stroke: #fff;
      stroke-width: 4px;
      stroke-linecap: butt;
      stroke-linejoin: round;
    }}
    .edgeLabelBg {{
      fill: #ffffff;
      stroke: #d6dee9;
      stroke-width: 1;
      opacity: 0.94;
    }}
    .edgeText.ok {{ fill: #1f7a3f; }}
    .edgeText.warn {{ fill: #b45309; }}
    .edgeText.bad {{ fill: #b42318; }}
    .edgeText.na {{ fill: #334155; }}
    .divider {{
      height: 0;
      border-top: 4px solid #1f2937;
      border-bottom: 1px solid #c7d2e0;
      margin: 22px 0 14px;
      opacity: 0.85;
    }}
    .reqNodeRect {{
      fill: #ffffff;
      stroke: #6b86b5;
      stroke-width: 2;
    }}
    .reqNodeMeta {{
      font-size: 13px;
      fill: #1e3a8a;
      font-weight: 700;
    }}
    .reqEdgeLabelBg {{
      fill: #eff6ff;
      stroke: #93c5fd;
      stroke-width: 1;
      opacity: 0.96;
    }}
    .reqEdgeText {{
      font-size: 14px;
      font-weight: 700;
      fill: #1d4ed8;
      paint-order: stroke;
      stroke: #fff;
      stroke-width: 4px;
      stroke-linecap: butt;
      stroke-linejoin: round;
    }}
  </style>
</head>
<body>
  <div class="page">
    <h1 class="title">{html.escape(report_title)}</h1>
    <p class="subtitle">Dependency flow from left (early) to right (late). Scroll horizontally to see full chain and branches.</p>
    <p class="legend">
      <span class="ok">Green: PTS duration &gt; STD duration</span>
      <span class="bad">Red: PTS duration &lt;= STD duration</span>
    </p>
    <div class="canvas">
      {network_svg if graph_nodes else "<p>No dependency links could be rendered.</p>"}
    </div>

    <div class="divider"></div>
    <p class="subtitle"><b>REQ Flow Only</b> - this second diagram shows only requested-date flow and REQ durations on arrows.</p>
    <p class="legend">
      <span class="req">Blue arrows: REQ duration between dependency points</span>
    </p>
    <div class="canvas">
      {req_network_svg if graph_nodes else "<p>No REQ dependency links could be rendered.</p>"}
    </div>
  </div>
</body>
</html>
"""

    html_output_path.write_text(html_content, encoding='utf-8')

    is_ok, issues = validate_generated_html_layout(html_output_path)
    if not is_ok and _auto_retry and layout_profile != 'safe':
        print(f"HTML layout check found {len(issues)} visibility issue(s), regenerating with safe layout profile...")
        return generate_visualization_html(
            df_plan=df_plan,
            df_general=df_general,
            df_constraints=df_constraints,
            matched_activities=matched_activities,
            dependency_map=dependency_map,
            html_output_path=html_output_path,
            df_holidays=df_holidays,
            conflicts=conflicts,
            unmatched_activities=unmatched_activities,
            layout_profile='safe',
            _auto_retry=False,
        )
    if not is_ok:
        print(f"Warning: HTML layout check still reports {len(issues)} issue(s)")

    return html_output_path


def process_pts_file(project_plan_path, general_file_path, similarity_threshold=95):
    """
    Process the PTS project plan file and add category markers for important activities.
    """
    print(f"Loading project plan: {project_plan_path}")
    df_plan = pd.read_excel(project_plan_path)
    
    print(f"Loading general file: {general_file_path}")
    df_general = pd.read_excel(general_file_path, sheet_name=0)
    df_constraints = pd.read_excel(general_file_path, sheet_name=1)  # Load constraints sheet
    
    print(f"\nProject plan shape: {df_plan.shape}")
    print(f"Important activities: {len(df_general)}")
    print(f"Constraints/relationships: {len(df_constraints)}")
    
    # Format date columns to Swiss format
    if 'Start' in df_plan.columns:
        df_plan['Start'] = df_plan['Start'].apply(format_date_swiss)
    if 'Finish' in df_plan.columns:
        df_plan['Finish'] = df_plan['Finish'].apply(format_date_swiss)
    
    # Find rightmost used column
    rightmost_col_idx = find_rightmost_used_column(df_plan)
    print(f"Rightmost used column: {df_plan.columns[rightmost_col_idx]} (index {rightmost_col_idx})")
    
    # Create new columns after the rightmost column
    insert_position = rightmost_col_idx + 1
    
    # Add "Current length", "Requested date", "Requested dates editable" and "Comment" columns
    df_plan.insert(insert_position, 'Current length', '')
    df_plan.insert(insert_position + 1, 'Requested date', '')
    df_plan.insert(insert_position + 2, 'Requested dates editable', '')
    df_plan.insert(insert_position + 3, 'Comment', '')
    df_plan.insert(insert_position + 4, 'Similarity Score %', '')
    
    # Add category columns
    category_columns = ['Important for GP', 'Piping', 'Hangers', 'Additional Material', 'Valves']
    for i, cat_col in enumerate(category_columns):
        df_plan.insert(insert_position + 5 + i, cat_col, '')
    
    # Process each activity - NEW APPROACH: Find best match for each important activity
    print(f"\nMatching activities (threshold: {similarity_threshold}%)...")
    matched_count = 0
    match_details = []
    low_score_rows = []  # Track rows with score < 95%
    matched_activities = {}  # Dictionary to store activity name -> row index mapping
    unmatched_activities = []  # Track important activities that couldn't be matched
    
    # For each important activity, find the best matching project activity
    for _, important_row in df_general.iterrows():
        important_activity = important_row['Activity']
        best_score = 0
        best_plan_idx = None
        
        # Search through all project plan activities
        for idx, plan_row in df_plan.iterrows():
            if pd.notna(plan_row['Activity Name']):
                activity_name_lower = str(plan_row['Activity Name']).lower().strip()
                important_activity_lower = str(important_activity).lower().strip()
                
                # Calculate similarity score
                score = fuzz.ratio(activity_name_lower, important_activity_lower)
                
                if score >= similarity_threshold and score > best_score:
                    best_score = score
                    best_plan_idx = idx
        
        # If a match was found, mark it
        if best_plan_idx is not None:
            # Store the mapping
            matched_activities[important_activity] = best_plan_idx
            
            # Extract categories where value is 'x' (case insensitive)
            categories = {}
            for col in category_columns:
                if pd.notna(important_row[col]) and str(important_row[col]).lower().strip() == 'x':
                    categories[col] = 'X'
                else:
                    categories[col] = ''
            
            # Fill in category columns
            for cat_col in category_columns:
                df_plan.at[best_plan_idx, cat_col] = categories[cat_col]
            
            # Fill in similarity score
            df_plan.at[best_plan_idx, 'Similarity Score %'] = best_score
            
            # Track low score matches
            if best_score < 95:
                low_score_rows.append(best_plan_idx)
            
            matched_count += 1
            match_details.append({
                'Activity ID': df_plan.at[best_plan_idx, 'Activity ID'],
                'Activity Name': df_plan.at[best_plan_idx, 'Activity Name'],
                'Matched To': important_activity,
                'Score': best_score,
                'Row': best_plan_idx
            })
        else:
            # No match found - track it
            unmatched_activities.append(important_activity)
    
    print(f"\nMatched {matched_count} activities out of {df_plan['Activity Name'].notna().sum()} total activities")
    
    # Load holidays sheet if available
    df_holidays = None
    try:
        df_holidays = pd.read_excel(general_file_path, sheet_name='Holidays')
        print(f"Loaded {len(df_holidays)} holiday period(s)")
    except Exception as e:
        print(f"No Holidays sheet found or error loading: {e}")
    
    # Calculate requested dates based on constraints
    print(f"\nCalculating requested dates based on dependencies...")
    conflicts, dependency_map = calculate_requested_dates(df_plan, df_general, df_constraints, matched_activities, df_holidays)
    
    # Calculate current lengths between activities
    print(f"\nCalculating current durations between activities...")
    calculate_current_lengths(df_plan, df_constraints, matched_activities)
    
    # Print some match details
    if match_details:
        print("\nSample matches:")
        for detail in match_details[:10]:
            print(f"  {detail['Activity ID']}: {detail['Activity Name'][:60]}...")
            print(f"    -> {detail['Matched To'][:60]}... (Score: {detail['Score']}%)")
    
    # Sort by Activity ID for better readability
    df_plan_sorted = df_plan.sort_values('Activity ID').reset_index(drop=True)
    
    # Count important activities
    matched_indices = list(matched_activities.values())
    important_count = len(matched_indices)
    
    print(f"\nExporting all {len(df_plan_sorted)} rows with {important_count} important activities marked")
    
    # Generate output filename
    input_path = Path(project_plan_path)
    current_date = datetime.now().strftime('%Y-%m-%d')
    output_filename = f"{input_path.stem}_process_{current_date}{input_path.suffix}"
    output_path = input_path.parent / output_filename
    
    # Save to Excel with formatting
    print(f"\nSaving output to: {output_path}")
    
    # Use openpyxl for formatting
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill, numbers
    from openpyxl.worksheet.filters import FilterColumn, Filters
    
    # First save the full dataframe with pandas
    df_plan_sorted.to_excel(output_path, index=False, engine='openpyxl')
    
    # Add conflicts sheet if there are any conflicts
    if conflicts:
        from openpyxl import load_workbook as load_wb_for_conflicts
        wb_conflicts = load_wb_for_conflicts(output_path)
        ws_conflicts = wb_conflicts.create_sheet(title='Conflicts')
        
        # Add header
        ws_conflicts['A1'] = 'Conflict Description'
        ws_conflicts.column_dimensions['A'].width = 100
        
        # Add conflicts
        for idx, conflict in enumerate(conflicts, start=2):
            ws_conflicts[f'A{idx}'] = conflict
            ws_conflicts[f'A{idx}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        wb_conflicts.save(output_path)
        print(f"Added 'Conflicts' sheet with {len(conflicts)} conflict(s)")
    
    # Load workbook for formatting the main sheet
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Find column indices for various columns
    activity_name_col = None
    similarity_score_col = None
    important_for_gp_col = None
    requested_date_col = None
    requested_dates_editable_col = None
    category_col_indices = []
    
    for idx, col in enumerate(df_plan_sorted.columns, 1):
        if col == 'Activity Name':
            activity_name_col = idx
        if col == 'Similarity Score %':
            similarity_score_col = idx
        if col == 'Important for GP':
            important_for_gp_col = idx
        if col == 'Requested date':
            requested_date_col = idx
        if col == 'Requested dates editable':
            requested_dates_editable_col = idx
        if col in category_columns:
            category_col_indices.append(idx)
    
    # Build formulas for "Requested dates editable" column
    print(f"\nBuilding Excel formulas for editable date column...")
    
    # Create a mapping of activity name to Excel row number
    activity_row_map = {}
    for idx, row in df_plan_sorted.iterrows():
        for activity_name, plan_idx in matched_activities.items():
            if df_plan_sorted.at[idx, 'Activity ID'] == df_plan.at[plan_idx, 'Activity ID']:
                # Excel row = pandas index + 2 (1 for 0-indexing, 1 for header)
                activity_row_map[activity_name] = idx + 2
                break
    
    # Find anchor activity name
    anchor_name = "Lot 5 - Erection General Piping - Contractor On Site Mobilisation"
    
    # Set formulas for each activity in "Requested dates editable" column
    formula_count = 0
    for activity_name, excel_row in activity_row_map.items():
        cell = ws.cell(row=excel_row, column=requested_dates_editable_col)
        
        if activity_name == anchor_name:
            # Anchor: reference the "Requested date" column
            ref_cell = get_column_letter(requested_date_col) + str(excel_row)
            cell.value = f"={ref_cell}"
            formula_count += 1
        elif activity_name in dependency_map:
            dep_info = dependency_map[activity_name]
            reference_activity = dep_info['reference']
            weeks = dep_info['weeks']
            direction = dep_info['direction']
            
            if reference_activity in activity_row_map:
                # Reference cell in the "Requested dates editable" column
                ref_row = activity_row_map[reference_activity]
                ref_cell = get_column_letter(requested_dates_editable_col) + str(ref_row)
                
                if direction == 'before':
                    # This activity is BEFORE the reference, so subtract weeks
                    cell.value = f"={ref_cell}-{weeks}*7"
                elif direction == 'after':
                    # This activity is AFTER the reference, so add weeks
                    cell.value = f"={ref_cell}+{weeks}*7"
                
                formula_count += 1
        
        # Apply date format to the cell
        cell.number_format = 'DD/MM/YYYY'
    
    print(f"Added {formula_count} formulas to 'Requested dates editable' column")
    
    # Apply date format to "Requested date" column
    if requested_date_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=requested_date_col)
            cell.number_format = 'DD/MM/YYYY'
    
    # Add autofilter to top row
    ws.auto_filter.ref = ws.dimensions
    
    # Apply filter to show only important activities (rows with "X" in "Important for GP" column)
    if important_for_gp_col:
        # Filter to show only rows with "X" in Important for GP column
        filter_col = FilterColumn(colId=important_for_gp_col - 1)  # 0-indexed
        filter_col.filters = Filters()
        filter_col.filters.filter.append('X')
        ws.auto_filter.filterColumn.append(filter_col)
    
    # Auto-width all columns first
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap at 100
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set specific widths for columns A and B
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 75
    
    # Apply wrap text and alignment to all cells
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell_idx, cell in enumerate(row, 1):
            # Center align category columns horizontally and vertically
            if cell_idx in category_col_indices:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            else:
                # All other cells: wrap text and center vertically
                cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # Highlight Activity Name cells with score < 95%
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for low_score_idx in low_score_rows:
        # Excel rows start at 1, and we have a header row, so add 2
        excel_row = low_score_idx + 2
        if activity_name_col:
            cell = ws.cell(row=excel_row, column=activity_name_col)
            cell.fill = yellow_fill
    
    # Auto-fit row heights
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].auto_size = True
    
    # Freeze panes at C2 (freezes first row and columns A-B)
    ws.freeze_panes = 'C2'
    
    # Save formatted workbook
    wb.save(output_path)

    # Build HTML visualization report
    html_output_path = output_path.with_suffix('.html')
    generate_visualization_html(
        df_plan=df_plan,
        df_general=df_general,
        df_constraints=df_constraints,
        matched_activities=matched_activities,
        dependency_map=dependency_map,
        html_output_path=html_output_path,
        df_holidays=df_holidays,
        conflicts=conflicts,
        unmatched_activities=unmatched_activities,
    )
    
    print(f"\n[OK] Successfully created {output_filename}")
    print(f"  - Total rows exported: {len(df_plan_sorted)}")
    print(f"  - Important activities: {important_count} (filtered by default)")
    print(f"  - Activities with score < 95%: {len(low_score_rows)} (highlighted in yellow)")
    print(f"  - Activities with calculated dates: {df_plan_sorted['Requested date'].notna().sum()}")
    print(f"  - Formulas in 'Requested dates editable': {formula_count}")
    if conflicts:
        print(f"  - Conflicts detected: {len(conflicts)} (see 'Conflicts' sheet)")
    print(f"  - Applied autofilter (showing important activities), column widths, wrap text, and freeze panes at C2")
    print(f"  - HTML visualization: {html_output_path}")
    
    # Report unmatched important activities at the very end
    if unmatched_activities:
        print(f"\n{'='*80}")
        print(f"âš ï¸  WARNING: {len(unmatched_activities)} IMPORTANT ACTIVITY(IES) NOT FOUND IN PROJECT PLAN")
        print(f"{'='*80}")
        print("These activities won't have calculated dates and may break dependency chains:")
        for idx, activity in enumerate(unmatched_activities, 1):
            print(f"  {idx}. {activity}")
        print(f"\nThis will prevent date calculation for activities that depend on these!")
        print(f"{'='*80}")
    
    return output_path, matched_count, match_details


def main():
    # File paths
    project_plan_path = r'C:\Users\szil\Repos\excel_wizadry\PTS_check\PTS_project_specific.xlsx'
    general_file_path = r'C:\Users\szil\Repos\excel_wizadry\PTS_check\PTS_General_input_data.xlsx'
    
    # Check if files exist
    if not Path(project_plan_path).exists():
        print(f"Error: Project plan file not found: {project_plan_path}")
        sys.exit(1)
    
    if not Path(general_file_path).exists():
        print(f"Error: General file not found: {general_file_path}")
        sys.exit(1)
    
    # Process files
    try:
        output_path, matched_count, match_details = process_pts_file(
            project_plan_path, 
            general_file_path, 
            similarity_threshold=90
        )
        print("\n" + "="*80)
        print("Processing complete!")
        print("="*80)
        
    except Exception as e:
        print(f"\nError processing files: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

