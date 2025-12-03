"""
PTS Activity Matcher Script
Matches activities from project plan with important activities from general file
using fuzzy string matching and adds category markers.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from thefuzz import fuzz
import sys


def find_rightmost_used_column(df):
    """Find the index of the rightmost column with data."""
    for i in range(len(df.columns) - 1, -1, -1):
        if df.iloc[:, i].notna().sum() > 1:  # More than 1 non-null value
            return i
    return len(df.columns) - 1


def match_activity(activity_name, important_activities, threshold=90):
    """
    Match an activity name against a list of important activities using fuzzy matching.
    Returns a list of tuples (activity, score, categories) for matches above threshold.
    """
    if pd.isna(activity_name):
        return []
    
    activity_name_lower = str(activity_name).lower().strip()
    matches = []
    
    for _, row in important_activities.iterrows():
        important_activity = str(row['Activity']).lower().strip()
        
        # Calculate similarity score
        score = fuzz.ratio(activity_name_lower, important_activity)
        
        if score >= threshold:
            # Extract categories where value is 'x' (case insensitive)
            categories = {}
            for col in ['Important for GP', 'Piping', 'Hangers', 'Additional Material', 'Valves']:
                if pd.notna(row[col]) and str(row[col]).lower().strip() == 'x':
                    categories[col] = 'X'
                else:
                    categories[col] = ''
            
            matches.append((row['Activity'], score, categories))
    
    return matches


def format_date_swiss(date_value):
    """
    Convert date to Swiss format dd/mm/yyyy, removing any extra text like 'A'.
    Handles both datetime objects and strings.
    Handles 2-digit years with asterisk (e.g., '26*' means 2026).
    """
    if pd.isna(date_value):
        return ''
    
    date_str = str(date_value).strip()
    
    # Remove trailing ' A' or similar letters (but preserve asterisks for now)
    date_str = date_str.rstrip(' ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    
    try:
        # Try parsing as datetime if it's a timestamp string
        if isinstance(date_value, pd.Timestamp):
            return date_value.strftime('%d/%m/%Y')
        
        # Try parsing common formats
        # Format: dd.mm.yy or dd.mm.yy*
        if '.' in date_str and len(date_str.split('.')) == 3:
            parts = date_str.split('.')
            day, month, year = parts[0], parts[1], parts[2]
            
            # Handle asterisk in year (e.g., '26*')
            year = year.rstrip('*')
            
            # Handle 2-digit year
            if len(year) == 2:
                year = '20' + year if int(year) < 50 else '19' + year
            return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
        
        # Try pandas datetime parsing
        dt = pd.to_datetime(date_value, errors='coerce')
        if pd.notna(dt):
            return dt.strftime('%d/%m/%Y')
    except:
        pass
    
    return date_str


def calculate_requested_dates(df_plan, df_general, df_constraints, matched_activities):
    """
    Calculate requested dates by working backwards from the mobilisation milestone.
    
    Args:
        df_plan: Project plan dataframe
        df_general: General activities dataframe
        df_constraints: Constraints/relationships dataframe
        matched_activities: Dictionary mapping important activity names to their row indices in df_plan
    """
    from datetime import timedelta
    
    # Find the mobilisation milestone in the matched activities
    mobilisation_name = "Lot 5 - Erection General Piping - Contractor On Site Mobilisation"
    
    if mobilisation_name not in matched_activities:
        print(f"Warning: Mobilisation milestone '{mobilisation_name}' not found in matched activities")
        return
    
    mobilisation_idx = matched_activities[mobilisation_name]
    mobilisation_start = df_plan.at[mobilisation_idx, 'Start']
    
    # Convert to datetime
    try:
        anchor_date = pd.to_datetime(mobilisation_start, format='%d/%m/%Y')
        print(f"\nAnchor point: {mobilisation_name}")
        print(f"  Start date: {anchor_date.strftime('%d/%m/%Y')}")
    except Exception as e:
        print(f"Error parsing mobilisation start date: {e}")
        return
    
    # Build a dictionary to store calculated dates for each activity
    activity_dates = {mobilisation_name: anchor_date}
    activity_comments = {}
    
    # Process constraints - work backwards from mobilisation
    # We need to process in reverse dependency order
    max_iterations = 100  # Prevent infinite loops
    iteration = 0
    
    while iteration < max_iterations:
        iteration += 1
        progress_made = False
        
        for _, constraint in df_constraints.iterrows():
            from_activity = constraint['From']
            to_activity = constraint['To']
            duration_weeks = constraint['Duration weeks']
            comment = constraint['Comments']
            
            # If we know the "To" date but not the "From" date, calculate it
            if to_activity in activity_dates and from_activity not in activity_dates:
                to_date = activity_dates[to_activity]
                
                # Calculate from_date by subtracting duration
                from_date = to_date - timedelta(weeks=duration_weeks)
                activity_dates[from_activity] = from_date
                
                # Build comment
                if pd.notna(comment) and str(comment).strip():
                    reason = f"{duration_weeks}W before '{to_activity}'. {comment}"
                else:
                    reason = f"{duration_weeks}W before '{to_activity}'"
                
                activity_comments[from_activity] = reason
                progress_made = True
        
        # If no progress was made, we're done
        if not progress_made:
            break
    
    print(f"\nCalculated dates for {len(activity_dates)} activities")
    
    # Now update the df_plan with calculated dates
    for activity_name, calculated_date in activity_dates.items():
        if activity_name in matched_activities:
            row_idx = matched_activities[activity_name]
            
            # Set requested date
            df_plan.at[row_idx, 'Requested date'] = calculated_date.strftime('%d/%m/%Y')
            
            # Set comment
            if activity_name in activity_comments:
                df_plan.at[row_idx, 'Comment'] = activity_comments[activity_name]
            elif activity_name == mobilisation_name:
                df_plan.at[row_idx, 'Comment'] = "Anchor milestone (reference point)"


def process_pts_file(project_plan_path, general_file_path, similarity_threshold=95):
    """
    Process the PTS project plan file and add category markers for important activities.
    """
    print(f"Loading project plan: {project_plan_path}")
    df_plan = pd.read_excel(project_plan_path)
    
    print(f"Loading general file: {general_file_path}")
    df_general = pd.read_excel(general_file_path, sheet_name=0)
    df_constraints = pd.read_excel(general_file_path, sheet_name=1)  # Load constraints sheet
    
    print(f"\nProject plan shape: {df_plan.shape}")
    print(f"Important activities: {len(df_general)}")
    print(f"Constraints/relationships: {len(df_constraints)}")
    
    # Format date columns to Swiss format
    if 'Start' in df_plan.columns:
        df_plan['Start'] = df_plan['Start'].apply(format_date_swiss)
    if 'Finish' in df_plan.columns:
        df_plan['Finish'] = df_plan['Finish'].apply(format_date_swiss)
    
    # Find rightmost used column
    rightmost_col_idx = find_rightmost_used_column(df_plan)
    print(f"Rightmost used column: {df_plan.columns[rightmost_col_idx]} (index {rightmost_col_idx})")
    
    # Create new columns after the rightmost column
    insert_position = rightmost_col_idx + 1
    
    # Add "Requested date" and "Comment" columns
    df_plan.insert(insert_position, 'Requested date', '')
    df_plan.insert(insert_position + 1, 'Comment', '')
    df_plan.insert(insert_position + 2, 'Similarity Score %', '')
    
    # Add category columns
    category_columns = ['Important for GP', 'Piping', 'Hangers', 'Additional Material', 'Valves']
    for i, cat_col in enumerate(category_columns):
        df_plan.insert(insert_position + 3 + i, cat_col, '')
    
    # Process each activity - NEW APPROACH: Find best match for each important activity
    print(f"\nMatching activities (threshold: {similarity_threshold}%)...")
    matched_count = 0
    match_details = []
    low_score_rows = []  # Track rows with score < 95%
    matched_activities = {}  # Dictionary to store activity name -> row index mapping
    
    # For each important activity, find the best matching project activity
    for _, important_row in df_general.iterrows():
        important_activity = important_row['Activity']
        best_score = 0
        best_plan_idx = None
        
        # Search through all project plan activities
        for idx, plan_row in df_plan.iterrows():
            if pd.notna(plan_row['Activity Name']):
                activity_name_lower = str(plan_row['Activity Name']).lower().strip()
                important_activity_lower = str(important_activity).lower().strip()
                
                # Calculate similarity score
                score = fuzz.ratio(activity_name_lower, important_activity_lower)
                
                if score >= similarity_threshold and score > best_score:
                    best_score = score
                    best_plan_idx = idx
        
        # If a match was found, mark it
        if best_plan_idx is not None:
            # Store the mapping
            matched_activities[important_activity] = best_plan_idx
            
            # Extract categories where value is 'x' (case insensitive)
            categories = {}
            for col in category_columns:
                if pd.notna(important_row[col]) and str(important_row[col]).lower().strip() == 'x':
                    categories[col] = 'X'
                else:
                    categories[col] = ''
            
            # Fill in category columns
            for cat_col in category_columns:
                df_plan.at[best_plan_idx, cat_col] = categories[cat_col]
            
            # Fill in similarity score
            df_plan.at[best_plan_idx, 'Similarity Score %'] = best_score
            
            # Track low score matches
            if best_score < 95:
                low_score_rows.append(best_plan_idx)
            
            matched_count += 1
            match_details.append({
                'Activity ID': df_plan.at[best_plan_idx, 'Activity ID'],
                'Activity Name': df_plan.at[best_plan_idx, 'Activity Name'],
                'Matched To': important_activity,
                'Score': best_score,
                'Row': best_plan_idx
            })
    
    print(f"\nMatched {matched_count} activities out of {df_plan['Activity Name'].notna().sum()} total activities")
    
    # Calculate requested dates based on constraints
    print(f"\nCalculating requested dates based on dependencies...")
    calculate_requested_dates(df_plan, df_general, df_constraints, matched_activities)
    
    # Print some match details
    if match_details:
        print("\nSample matches:")
        for detail in match_details[:10]:
            print(f"  {detail['Activity ID']}: {detail['Activity Name'][:60]}...")
            print(f"    -> {detail['Matched To'][:60]}... (Score: {detail['Score']}%)")
    
    # Generate output filename
    input_path = Path(project_plan_path)
    current_date = datetime.now().strftime('%Y-%m-%d')
    output_filename = f"{input_path.stem}_process_{current_date}{input_path.suffix}"
    output_path = input_path.parent / output_filename
    
    # Save to Excel with formatting
    print(f"\nSaving output to: {output_path}")
    
    # Use openpyxl for formatting
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill
    
    # First save with pandas
    df_plan.to_excel(output_path, index=False, engine='openpyxl')
    
    # Load workbook for formatting
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Find column indices for category columns and Activity Name
    activity_name_col = None
    category_col_indices = []
    for idx, col in enumerate(df_plan.columns, 1):
        if col == 'Activity Name':
            activity_name_col = idx
        if col in category_columns:
            category_col_indices.append(idx)
    
    # Add autofilter to top row
    ws.auto_filter.ref = ws.dimensions
    
    # Auto-width all columns first
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap at 100
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set specific widths for columns A and B
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 75
    
    # Apply wrap text and alignment to all cells
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell_idx, cell in enumerate(row, 1):
            # Center align category columns horizontally and vertically
            if cell_idx in category_col_indices:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            else:
                # All other cells: wrap text and center vertically
                cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # Highlight Activity Name cells with score < 95%
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for low_score_idx in low_score_rows:
        # Excel rows start at 1, and we have a header row, so add 2
        excel_row = low_score_idx + 2
        if activity_name_col:
            cell = ws.cell(row=excel_row, column=activity_name_col)
            cell.fill = yellow_fill
    
    # Auto-fit row heights
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].auto_size = True
    
    # Freeze panes at C2 (freezes first row and columns A-B)
    ws.freeze_panes = 'C2'
    
    # Save formatted workbook
    wb.save(output_path)
    
    print(f"\n[OK] Successfully created {output_filename}")
    print(f"  - Total rows: {len(df_plan)}")
    print(f"  - Matched activities: {matched_count}")
    print(f"  - Activities with score < 95%: {len(low_score_rows)} (highlighted in yellow)")
    print(f"  - Applied autofilter, column widths, wrap text, and freeze panes at C2")
    
    return output_path, matched_count, match_details


def main():
    # File paths
    project_plan_path = r'C:\Users\szil\Repos\excel_wizadry\PTS_check\BE020-HZI-50230983_5.0_PTS L3.xlsx'
    general_file_path = r'C:\Users\szil\Repos\excel_wizadry\PTS_check\PTS_General_input_data.xlsx'
    
    # Check if files exist
    if not Path(project_plan_path).exists():
        print(f"Error: Project plan file not found: {project_plan_path}")
        sys.exit(1)
    
    if not Path(general_file_path).exists():
        print(f"Error: General file not found: {general_file_path}")
        sys.exit(1)
    
    # Process files
    try:
        output_path, matched_count, match_details = process_pts_file(
            project_plan_path, 
            general_file_path, 
            similarity_threshold=90
        )
        print("\n" + "="*80)
        print("Processing complete!")
        print("="*80)
        
    except Exception as e:
        print(f"\nError processing files: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
