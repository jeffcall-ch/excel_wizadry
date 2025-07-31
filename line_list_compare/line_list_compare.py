
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(message)s')

def compare_pipeline_lists(
    old_file=None, new_file=None, output_file=None
):
    print()  # Add blank line for readability
    logging.info(f"compare_pipeline_lists called with old_file={old_file}, new_file={new_file}, output_file={output_file}")
    # Default filenames
    if old_file is None:
        old_file = os.path.join(os.path.dirname(__file__), 'pipeline_list_old.xlsx')
    if new_file is None:
        new_file = os.path.join(os.path.dirname(__file__), 'pipeline_list_old_new.xlsx')
    if output_file is None:
        output_file = os.path.splitext(new_file)[0] + '_COMPARE_WITH_PREV_REV.xlsx'

    # Colors
    YELLOW = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')  # changed
    BLUE = PatternFill(start_color='FF00B0F0', end_color='FF00B0F0', fill_type='solid')    # new row
    RED = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')     # deleted

    # Read files

    def read_excel_skip_to_kks(path):
        # Read all rows, find first row where any column is 'KKS', drop above, use as header
        # Always read from the 'Query' sheet
        # Read all rows, find first row where any column is 'KKS', use as header
        raw = pd.read_excel(path, dtype=str, header=None, sheet_name='Query')
        kks_row_idx = None
        for i, row in raw.iterrows():
            if any(str(cell).strip() == 'KKS' for cell in row):
                kks_row_idx = i
                break
        if kks_row_idx is None:
            print(f"ERROR: No 'KKS' header found in any column of {path}")
            import sys
            sys.exit(1)
        df = pd.read_excel(path, dtype=str, header=kks_row_idx, sheet_name='Query')
        return df


    try:
        old_df = read_excel_skip_to_kks(old_file)
    except FileNotFoundError as e:
        print()
        logging.error(f"Failed to read old_file: {e}")
        old_df = pd.DataFrame(columns=['KKS'])
    except Exception as e:
        # For other exceptions like RuntimeError from sys.exit, re-raise
        print()
        logging.error(f"Failed to read old_file: {e}")
        raise
    
    try:
        new_df = read_excel_skip_to_kks(new_file)
    except FileNotFoundError as e:
        print()
        logging.error(f"Failed to read new_file: {e}")
        new_df = pd.DataFrame(columns=['KKS'])
    except Exception as e:
        # For other exceptions like RuntimeError from sys.exit, re-raise
        print()
        logging.error(f"Failed to read new_file: {e}")
        raise


    old_df = old_df.fillna('')
    new_df = new_df.fillna('')

    # Ensure KKS is string
    old_df['KKS'] = old_df['KKS'].astype(str)
    new_df['KKS'] = new_df['KKS'].astype(str)

    # Check uniqueness of KKS in both files
    def check_kks_unique(df, label):
        duplicated = df['KKS'][df['KKS'].duplicated(keep=False)]
        if not duplicated.empty:
            print(f"ERROR: Non-unique KKS values found in {label} file:")
            for kks_val in duplicated.unique():
                # Find all row indices (0-based DataFrame index)
                row_indices = df.index[df['KKS'] == kks_val].tolist()
                # Convert to Excel row numbers (header row is 1, so add 2 for 0-based index)
                excel_rows = [i + 2 for i in row_indices]
                print(f"  KKS value '{kks_val}' found in Excel rows: {excel_rows}")
            import sys
            sys.exit(1)
    check_kks_unique(old_df, 'old')
    check_kks_unique(new_df, 'new')

    # All columns (union), without 'Changes' (we'll add it at the end)
    all_columns = list(dict.fromkeys(list(new_df.columns) + list(old_df.columns)))
    if 'Changes' in all_columns:
        all_columns.remove('Changes')

    # Index by KKS
    old_df = old_df.set_index('KKS', drop=False)
    new_df = new_df.set_index('KKS', drop=False)

    # Prepare output rows, color map, and changes column
    output_rows = []
    color_map = []
    changes_col = []

    # Helper: columns to check for changes/deletes (exclude KKS)
    check_columns = [col for col in all_columns if col != 'KKS']

    # Rows in new (existing or new)
    for kks, new_row in new_df.iterrows():
        if kks not in old_df.index:
            # Entirely new row
            output_rows.append([new_row.get(col, '') for col in all_columns])
            color_map.append([BLUE] * len(all_columns))
            changes_col.append('N')
        else:
            old_row = old_df.loc[kks]
            row_vals = []
            row_colors = []
            has_ch = False
            has_del = False
            for col in all_columns:
                new_val = new_row.get(col, '')
                old_val = old_row.get(col, '')
                if new_val == old_val:
                    row_vals.append(new_val)
                    row_colors.append(None)
                elif old_val == '' and new_val != '':
                    # Was empty, now filled: yellow (change)
                    row_vals.append(new_val)
                    row_colors.append(YELLOW)
                    if col in check_columns:
                        has_ch = True
                elif new_val == '':
                    # Was filled, now empty: red (deleted cell)
                    row_vals.append(new_val)
                    row_colors.append(RED)
                    if col in check_columns:
                        has_del = True
                else:
                    # Changed value: yellow
                    row_vals.append(new_val)
                    row_colors.append(YELLOW)
                    if col in check_columns:
                        has_ch = True
            output_rows.append(row_vals)
            color_map.append(row_colors)
            # Determine Changes column value
            if has_ch and has_del:
                changes_col.append('D,Ch')
            elif has_del:
                changes_col.append('D')
            elif has_ch:
                changes_col.append('Ch')
            else:
                changes_col.append('')

    # Rows deleted (in old, not in new): append at end, all red, Changes = 'D'
    deleted_kks = [kks for kks in old_df.index if kks not in new_df.index]
    for kks in deleted_kks:
        old_row = old_df.loc[kks]
        output_rows.append([old_row.get(col, '') for col in all_columns])
        color_map.append([RED] * len(all_columns))
        changes_col.append('D')

    # Write to Excel with openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparison'
    # Split and freeze panes at C2
    ws.freeze_panes = 'C2'
    ws.sheet_view.splitTop = 1
    ws.sheet_view.splitLeft = 2
    # Write header (add 'Changes' as last column)
    for j, col in enumerate(all_columns, 1):
        ws.cell(row=1, column=j, value=col)
    ws.cell(row=1, column=len(all_columns) + 1, value='Changes')
    # Write data and apply colors
    for i, (row, colors, changes) in enumerate(zip(output_rows, color_map, changes_col), 2):
        for j, (val, fill) in enumerate(zip(row, colors), 1):
            cell = ws.cell(row=i, column=j, value=val)
            if fill is not None:
                cell.fill = fill
        # Write Changes column with color coding
        changes_cell = ws.cell(row=i, column=len(all_columns) + 1, value=changes)
        if changes == 'Ch':
            changes_cell.fill = YELLOW
        elif changes == 'N':
            changes_cell.fill = BLUE
        elif changes == 'D' or changes == 'D,Ch':
            changes_cell.fill = RED

    # Auto-fit column widths (including Changes)
    for j, col in enumerate(all_columns + ['Changes'], 1):
        max_length = len(str(col))
        for i in range(2, len(output_rows) + 2):
            val = ws.cell(row=i, column=j).value
            if val is not None:
                max_length = max(max_length, len(str(val)))
        # Add a little extra space
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = max_length + 2

    try:
        wb.save(output_file)
        print()
        logging.info(f"Output file saved: {output_file}")
    except Exception as e:
        print()
        logging.error(f"Failed to save output file: {e}")


DATA_DIR = os.path.dirname(__file__)
OLD_FILE = os.path.join(DATA_DIR, 'pipeline_list_old.xlsx')
NEW_FILE = os.path.join(DATA_DIR, 'pipeline_list_new.xlsx')
OUTPUT_FILE = os.path.join(DATA_DIR, 'pipeline_list_new_COMPARE_WITH_PREV_REV.xlsx')

if __name__ == "__main__":
    compare_pipeline_lists(old_file=OLD_FILE, new_file=NEW_FILE, output_file=OUTPUT_FILE)