#!/usr/bin/env python3
"""
Excel Line List Comparator - Single File Implementation

A high-performance Python tool for comparing Excel piping line lists with KKS-based identification.
Generates detailed change analysis with visual formatting in a single 'new_color_coded' sheet.

Usage:
    python excel_compare.py compare.xlsx                    # Auto-generated timestamped output
    python excel_compare.py compare.xlsx custom_result.xlsx # Custom output filename

Input file must contain 'old' and 'new' sheets with piping data.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import sys
import locale
import logging
import argparse
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any


# Color constants (matching VBA RGB values)
class Colors:
    BLUE = "0000FF"      # RGB(0, 0, 255) - Blue background
    GREEN = "00FF00"     # RGB(0, 255, 0) - Green background  
    YELLOW = "FFFF00"    # RGB(255, 255, 0) - Yellow background
    RED = "FF0000"       # RGB(255, 0, 0) - Red font color


@dataclass
class ComparisonResult:
    """Stores comparison results with formatting information."""
    data: pd.DataFrame
    is_old_base: bool
    colors: Dict[Tuple[int, str], str] = field(default_factory=dict)
    comments: Dict[Tuple[int, str], str] = field(default_factory=dict)
    markers: Dict[int, str] = field(default_factory=dict)
    duplicate_rows: set = field(default_factory=set)
    added_rows: List[Dict[str, Any]] = field(default_factory=list)


class ExcelTableComparator:
    """
    Pythonic Excel table comparator that produces 'new_color_coded' sheet.
    
    Generates a single comparison sheet with NEW table as base, showing changes from OLD.
    """
    
    def __init__(self, old_file: str = 'old_list.xlsx', 
                 new_file: str = 'new_list.xlsx',
                 output_file: str = None):
        self.old_file = Path(old_file)
        self.new_file = Path(new_file)
        
        if output_file is None:
            # Auto-generate output filename: "filename_YYYYMMDD_HHMMSS_comparison.xlsx"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_file_stem = self.new_file.stem  # filename without extension
            output_filename = f"{new_file_stem}_{timestamp}_comparison.xlsx"
            self.output_file = self.new_file.parent / output_filename
        else:
            self.output_file = Path(output_file)
        
        # Setup logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('comparison.log', mode='w'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
        # Detect system language (matches VBA language detection intent)
        self.is_german = self._detect_german_locale()
        
        # Will be populated during loading
        self.old_workbook: Optional[load_workbook] = None
        self.new_workbook: Optional[load_workbook] = None
        self.old_data: Optional[pd.DataFrame] = None
        self.new_data: Optional[pd.DataFrame] = None
        self.column_names: List[str] = []
        
    def _detect_german_locale(self) -> bool:
        """Detect if system is German (matches VBA LanguageID intent)."""
        try:
            lang_code = locale.getdefaultlocale()[0]
            return lang_code is not None and lang_code.startswith('de')
        except:
            return False
    
    def _get_localized_text(self, english: str, german: str) -> str:
        """Return localized text based on system language."""
        return german if self.is_german else english
    
    def validate_files_and_load_data(self) -> None:
        """Load and validate Excel files."""
        self.logger.info("Validating files and loading data...")
        
        # Load workbooks
        self.old_workbook = load_workbook(self.old_file, data_only=True)
        self.new_workbook = load_workbook(self.new_file, data_only=True)
        
        # Load data from first sheet of each workbook
        old_sheet_name = self.old_workbook.sheetnames[0]
        new_sheet_name = self.new_workbook.sheetnames[0]
        
        self.old_data = pd.read_excel(self.old_file, sheet_name=old_sheet_name)
        self.new_data = pd.read_excel(self.new_file, sheet_name=new_sheet_name)
        
        # Get column names (assuming both tables have same structure)
        self.column_names = list(self.new_data.columns)
        
        self.logger.info(f"Old table: {len(self.old_data)} rows, {len(self.old_data.columns)} columns")
        self.logger.info(f"New table: {len(self.new_data)} rows, {len(self.new_data.columns)} columns")
    
    def create_unique_keys(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Create unique keys for duplicate handling (OPTIMIZED VERSION)."""
        self.logger.info("Creating unique keys for duplicate handling...")
        
        def process_dataframe(df: pd.DataFrame, name: str) -> pd.DataFrame:
            df = df.copy()
            df['_unique_key'] = df['KKS'].astype(str)
            
            # OPTIMIZED: Use pandas value_counts() for O(n) duplicate detection
            kks_counts = df['KKS'].value_counts()
            duplicated_kks = set(kks_counts[kks_counts > 1].index)
            
            # Mark duplicate rows efficiently
            df['_is_duplicate'] = df['KKS'].isin(duplicated_kks)
            
            return df
        
        old_with_keys = process_dataframe(self.old_data, "OLD")
        new_with_keys = process_dataframe(self.new_data, "NEW")
        
        return old_with_keys, new_with_keys
    
    def perform_comparison(self, old_with_keys: pd.DataFrame, new_with_keys: pd.DataFrame) -> ComparisonResult:
        """Perform table comparison (NEW base only)."""
        self.logger.info("Performing table comparison (NEW base only)...")
        
        # NEW base comparison
        new_base_result = self._compare_single_pass(new_with_keys, old_with_keys, is_old_base=False)
        return new_base_result
    
    def _compare_single_pass(self, base_df: pd.DataFrame, compare_df: pd.DataFrame,
                           is_old_base: bool) -> ComparisonResult:
        """
        Perform single comparison pass (OPTIMIZED VERSION).
        """
        result = ComparisonResult(
            data=base_df[self.column_names].copy(),
            is_old_base=is_old_base
        )
        
        # Track duplicate rows for red font marking
        for idx in range(len(base_df)):
            if base_df.iloc[idx]['_is_duplicate']:
                result.duplicate_rows.add(idx)
        
        # PERFORMANCE OPTIMIZATION: Create lookup dictionaries
        # Create a mapping from unique_key to row index for fast lookups
        base_key_to_idx = {row['_unique_key']: idx for idx, row in base_df.iterrows()}
        
        # Set base background color for all cells (VBA: UsedRange.Interior.Color)
        base_color = Colors.BLUE if is_old_base else Colors.GREEN
        for row_idx in range(len(result.data)):
            for col_name in self.column_names:
                result.colors[(row_idx, col_name)] = base_color
        
        # VECTORIZED COMPARISON: Process all rows at once where possible
        compare_keys = compare_df['_unique_key'].tolist()
        
        # Find which keys exist in base (vectorized operation)
        existing_keys = set(base_key_to_idx.keys())
        
        # Separate found and not found keys
        found_comparisons = []
        
        for compare_idx, compare_key in enumerate(compare_keys):
            if compare_key in existing_keys:
                base_idx = base_key_to_idx[compare_key]
                found_comparisons.append((base_idx, compare_idx))
            else:
                # Row not found - add to bottom
                row_data = compare_df.iloc[compare_idx].to_dict()
                result.added_rows.append(row_data)
        
        # Process found rows in batches for better performance
        for base_idx, compare_idx in found_comparisons:
            # Clear base color for this row (VBA: Interior.Pattern = xlNone)
            for col_name in self.column_names:
                if (base_idx, col_name) in result.colors:
                    del result.colors[(base_idx, col_name)]
            
            # Compare individual cells (optimized)
            changes_in_row = []
            
            # VECTORIZED: Get all values for this row at once
            base_row = result.data.iloc[base_idx]
            compare_row = compare_df.iloc[compare_idx]
            
            for col_name in self.column_names:
                base_value = self._get_cell_text_value_fast(base_row[col_name])
                compare_value = self._get_cell_text_value_fast(compare_row[col_name])
                
                if base_value == compare_value:
                    continue  # No change
                
                # Determine change type and apply formatting
                change_type, color = self._analyze_cell_change(
                    base_value, compare_value, is_old_base
                )
                
                result.colors[(base_idx, col_name)] = color
                changes_in_row.append(change_type)
                
                # Add comment if value changed
                if change_type in ["Changed", "Added", "Deleted"]:
                    compare_table_name = "New Table" if is_old_base else "Old Table"
                    result.comments[(base_idx, col_name)] = f"{compare_table_name} value:\n{compare_value}"
            
            # Combine change markers for this row
            if changes_in_row:
                result.markers[base_idx] = self._combine_change_markers(changes_in_row)
        
        return result
    
    def _get_cell_text_value_fast(self, value) -> str:
        """Fast cell value to text conversion (optimized version)."""
        if pd.isna(value):
            return ""
        
        # Match Excel's .Text property behavior (optimized)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))  # Remove .0 from whole numbers
        
        return str(value)
    
    def _get_cell_text_value(self, df: pd.DataFrame, row_idx: int, col_name: str) -> str:
        """Get cell value as text (matches VBA .Text property)."""
        value = df.iloc[row_idx][col_name]
        if pd.isna(value):
            return ""
        
        # Match Excel's .Text property behavior
        if isinstance(value, float) and value.is_integer():
            return str(int(value))  # Remove .0 from whole numbers
        
        return str(value)
    
    def _analyze_cell_change(self, base_value: str, compare_value: str, 
                           is_old_base: bool) -> Tuple[str, str]:
        """
        Analyze cell change and return change type with color (matches VBA change logic).
        """
        # Both empty - no change
        if not base_value and not compare_value:
            return "No Change", ""
        
        # Value added (was empty, now has value)
        if not base_value and compare_value:
            return "Added", Colors.YELLOW
        
        # Value deleted (had value, now empty)
        if base_value and not compare_value:
            return "Deleted", Colors.YELLOW
        
        # Value changed (different non-empty values)
        if base_value != compare_value:
            return "Changed", Colors.YELLOW
        
        return "No Change", ""
    
    def _combine_change_markers(self, changes: List[str]) -> str:
        """Combine multiple change types into a single marker."""
        unique_changes = list(set(changes))
        
        if len(unique_changes) == 1:
            return unique_changes[0]
        
        # Multiple change types - create combined marker
        priority_order = ["Changed", "Added", "Deleted"]
        for change_type in priority_order:
            if change_type in unique_changes:
                return change_type
        
        return "Changed"
    
    def generate_excel_output(self, new_base_result: ComparisonResult) -> None:
        """Generate Excel output file with comparison results."""
        self.logger.info("Generating Excel output...")
        
        # Create new workbook with single sheet
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create new_color_coded sheet
        ws_new = wb.create_sheet("new_color_coded")
        self._write_comparison_sheet(ws_new, new_base_result)
        
        # Save workbook
        wb.save(self.output_file)
        self.logger.info(f"Excel output saved to: {self.output_file}")
    
    def _write_comparison_sheet(self, worksheet, result: ComparisonResult) -> None:
        """Write complete comparison sheet with all elements."""
        # Header row
        self._format_header_row(worksheet, result.is_old_base)
        
        # Data headers and rows
        self._write_data_headers(worksheet)
        self._write_data_rows(worksheet, result)
        
        # Added rows section
        if result.added_rows:
            self._write_added_rows_section(worksheet, result)
        
        # Change marker column
        self._write_change_marker_column(worksheet, result)
        
        # Apply AutoFilter
        self._apply_autofilter(worksheet, result)
    
    def _format_header_row(self, worksheet, is_old_base: bool) -> None:
        """Format the header row with title and legend."""
        # Main title
        base_text = self._get_localized_text("OLD", "ALT") if is_old_base else self._get_localized_text("NEW", "NEU")
        title = f"Comparison, base is the {base_text} table"
        worksheet.cell(1, 2, title)
        
        # Legend
        worksheet.cell(1, 5, self._get_localized_text("Changed", "Geändert"))
        worksheet.cell(1, 6, self._get_localized_text("Added", "Hinzugefügt"))
        worksheet.cell(1, 7, self._get_localized_text("Deleted", "Gelöscht"))
        
        # Apply legend colors
        worksheet.cell(1, 5).fill = PatternFill(start_color=Colors.YELLOW, end_color=Colors.YELLOW, fill_type='solid')
        worksheet.cell(1, 6).fill = PatternFill(start_color=Colors.YELLOW, end_color=Colors.YELLOW, fill_type='solid')
        worksheet.cell(1, 7).fill = PatternFill(start_color=Colors.YELLOW, end_color=Colors.YELLOW, fill_type='solid')
    
    def _write_data_headers(self, worksheet) -> None:
        """Write column headers in row 2."""
        for col_idx, col_name in enumerate(self.column_names, start=1):
            worksheet.cell(2, col_idx, col_name)
    
    def _write_data_rows(self, worksheet, result: ComparisonResult) -> None:
        """Write data rows with colors and comments."""
        for row_idx in range(len(result.data)):
            excel_row = row_idx + 3  # Start from row 3
            
            for col_idx, col_name in enumerate(self.column_names, start=1):
                cell = worksheet.cell(row=excel_row, column=col_idx)
                cell.value = result.data.iloc[row_idx][col_name]
                
                # Apply colors
                if (row_idx, col_name) in result.colors:
                    color = result.colors[(row_idx, col_name)]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                # Apply comments
                if (row_idx, col_name) in result.comments:
                    cell.comment = Comment(result.comments[(row_idx, col_name)], "System")
                    cell.comment.width = 200
                    cell.comment.height = 50
            
            # Mark duplicates with red font (VBA: Font.Color = vbRed) in KKS column
            if row_idx in result.duplicate_rows:
                worksheet.cell(row=excel_row, column=2).font = Font(color=Colors.RED)  # Column B (KKS)
    
    def _write_added_rows_section(self, worksheet, result: ComparisonResult) -> None:
        """Write added rows section at the bottom."""
        start_row = len(result.data) + 5  # Leave some space
        
        # Section header
        header_text = self._get_localized_text("Added Rows", "Hinzugefügte Zeilen")
        worksheet.cell(start_row, 1, header_text)
        
        # Column headers
        for col_idx, col_name in enumerate(self.column_names, start=1):
            worksheet.cell(start_row + 1, col_idx, col_name)
        
        # Added row data
        for idx, row_data in enumerate(result.added_rows):
            excel_row = start_row + 2 + idx
            for col_idx, col_name in enumerate(self.column_names, start=1):
                value = row_data.get(col_name, "")
                worksheet.cell(excel_row, col_idx, value)
    
    def _write_change_marker_column(self, worksheet, result: ComparisonResult) -> None:
        """Write change marker column."""
        marker_col = len(self.column_names) + 2  # After data columns + empty column
        
        # Header
        header_text = self._get_localized_text("Changed", "Geändert")
        worksheet.cell(2, marker_col, header_text)
        
        # Markers for main data
        for row_idx, marker_value in result.markers.items():
            excel_row = row_idx + 3
            worksheet.cell(excel_row, marker_col, marker_value)
        
        # Markers for added rows
        if result.added_rows:
            marker_value = self._get_localized_text("Added", "Hinzugefügt")
            start_row = len(result.data) + 5
            for idx in range(len(result.added_rows)):
                worksheet.cell(row=start_row + idx, column=marker_col, value=marker_value)
    
    def _apply_autofilter(self, worksheet, result: ComparisonResult) -> None:
        """Apply AutoFilter to the data range."""
        marker_col = len(self.column_names) + 2
        last_row = len(result.data) + 3
        if result.added_rows:
            last_row += len(result.added_rows) + 1
        
        last_col_letter = get_column_letter(marker_col)
        worksheet.auto_filter.ref = f"A2:{last_col_letter}{last_row}"
    
    def run(self) -> None:
        """Execute complete comparison workflow."""
        start_time = datetime.now()
        
        try:
            self.logger.info("=" * 60)
            self.logger.info("Excel Table Comparison - Starting")
            self.logger.info("=" * 60)
            
            # Step 1: Validate and load data
            self.validate_files_and_load_data()
            
            # Step 2: Create unique keys for duplicate handling
            old_with_keys, new_with_keys = self.create_unique_keys()
            
            # Step 3: Perform comparison (NEW base only)
            new_base_result = self.perform_comparison(old_with_keys, new_with_keys)
            
            # Step 4: Generate Excel output
            self.generate_excel_output(new_base_result)
            
            duration = (datetime.now() - start_time).total_seconds()
            
            self.logger.info("=" * 60)
            self.logger.info(f"Comparison completed successfully in {duration:.2f} seconds")
            self.logger.info("=" * 60)
            
            print("\n" + "=" * 60)
            print("COMPARISON COMPLETE!")
            print("=" * 60)
            print(f"✓ Output file created: {self.output_file}")
            print(f"✓ Duration: {duration:.2f} seconds")
            print("=" * 60 + "\n")
            
        except Exception as e:
            self.logger.error(f"Error during comparison: {str(e)}", exc_info=True)
            print(f"\n❌ ERROR: {str(e)}\n")


def extract_sheets_and_compare(compare_file, output_file=None):
    """
    Extract old/new sheets from compare.xlsx and run comparison.
    
    Args:
        compare_file: Path to Excel file with 'old' and 'new' sheets
        output_file: Output filename for the new_color_coded sheet (auto-generates if None)
    """
    try:
        print(f"🔍 Processing: {compare_file}")
        print("=" * 50)
        
        # Load the compare.xlsx file
        wb_compare = load_workbook(compare_file)
        
        if 'old' not in wb_compare.sheetnames or 'new' not in wb_compare.sheetnames:
            print("❌ Error: File must contain 'old' and 'new' sheets")
            return False
        
        # Create temporary old.xlsx and new.xlsx files
        print("📋 Extracting sheets...")
        
        # Create old.xlsx
        wb_old = Workbook()
        wb_old.remove(wb_old.active)
        old_sheet = wb_old.create_sheet("Old Table")
        compare_old_sheet = wb_compare["old"]
        
        # Copy data from compare.xlsx old sheet
        for row in compare_old_sheet.iter_rows():
            for cell in row:
                old_sheet[cell.coordinate].value = cell.value
        
        wb_old.save("temp_old.xlsx")
        
        # Create new.xlsx  
        wb_new = Workbook()
        wb_new.remove(wb_new.active)
        new_sheet = wb_new.create_sheet("New Table")
        compare_new_sheet = wb_compare["new"]
        
        # Copy data from compare.xlsx new sheet
        for row in compare_new_sheet.iter_rows():
            for cell in row:
                new_sheet[cell.coordinate].value = cell.value
        
        wb_new.save("temp_new.xlsx")
        
        # Run the comparator
        print("🚀 Running comparison (KKS-based)...")
        
        # If output_file not specified, generate based on original input file
        if output_file is None:
            input_path = Path(compare_file)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = input_path.parent / f"{input_path.stem}_{timestamp}_comparison.xlsx"
        
        comparator = ExcelTableComparator(
            old_file='temp_old.xlsx',
            new_file='temp_new.xlsx', 
            output_file=str(output_file)
        )
        comparator.run()
        
        # Clean up temporary files
        Path("temp_old.xlsx").unlink(missing_ok=True)
        Path("temp_new.xlsx").unlink(missing_ok=True)
        
        print(f"✅ Success! Output saved to: {comparator.output_file}")
        print("📊 Contains 'new_color_coded' sheet with change analysis")
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        # Clean up on error
        Path("temp_old.xlsx").unlink(missing_ok=True)
        Path("temp_new.xlsx").unlink(missing_ok=True)
        return False


def main():
    """Main command-line interface."""
    parser = argparse.ArgumentParser(
        description='Compare Excel tables using KKS as unique identifier',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_compare.py compare.xlsx
  python excel_compare.py compare.xlsx my_result.xlsx
  
The input file must contain 'old' and 'new' sheets.
Output will contain a single 'new_color_coded' sheet with changes highlighted.
        """
    )
    
    parser.add_argument('input_file', 
                       help='Excel file containing old and new sheets')
    parser.add_argument('output_file', nargs='?', 
                       default=None,
                       help='Output Excel file (auto-generates with datetime if not provided)')
    
    args = parser.parse_args()
    
    # Validate input file
    if not Path(args.input_file).exists():
        print(f"❌ Error: Input file '{args.input_file}' not found")
        sys.exit(1)
    
    # Run comparison
    success = extract_sheets_and_compare(args.input_file, args.output_file)
    
    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()