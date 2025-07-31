
import sys
import os
import pytest
import pandas as pd
import shutil
from pathlib import Path

# Ensure parent directory is in sys.path for module import
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from line_list_compare import compare_pipeline_lists

@pytest.fixture(scope="session")
def test_files(tmp_path_factory):
    """Create test files for pipeline comparison."""
    test_dir = tmp_path_factory.mktemp("test_data")
    
    # Create old file with test data
    old_file = test_dir / 'pipeline_list_old.xlsx'
    old_df = pd.DataFrame({
        'KKS': ['A1', 'B1', 'C1', 'D1'],
        'Description': ['Item A', 'Item B', 'Item C', 'To be deleted'],
        'Value': [100, 200, 300, 400]
    })
    
    # Create new file with changes
    new_file = test_dir / 'pipeline_list_new.xlsx'
    new_df = pd.DataFrame({
        'KKS': ['A1', 'B1', 'C1', 'E1'],  # D1 removed, E1 added
        'Description': ['Item A', 'Item B2', 'Item C', 'New item'],  # B changed
        'Value': [100, 250, 300, 500],     # B value changed
        'NewCol': ['', 'New', '', 'New']   # New column added
    })
    
    # Write to Excel files with Query sheet
    with pd.ExcelWriter(old_file, engine='openpyxl') as writer:
        old_df.to_excel(writer, sheet_name='Query', index=False)
    
    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
        new_df.to_excel(writer, sheet_name='Query', index=False)
    
    # Output path
    output_file = test_dir / 'pipeline_list_new_COMPARE_WITH_PREV_REV.xlsx'
    
    return {'old_file': str(old_file), 'new_file': str(new_file), 'output_file': str(output_file)}

def test_compare_pipeline_lists_basic(test_files):
    compare_pipeline_lists(
        old_file=test_files['old_file'],
        new_file=test_files['new_file'],
        output_file=test_files['output_file']
    )
    assert os.path.exists(test_files['output_file'])
    # Check that the output file is a valid Excel file
    df = pd.read_excel(test_files['output_file'])
    assert not df.empty

def test_compare_pipeline_lists_content(test_files):
    compare_pipeline_lists(
        old_file=test_files['old_file'],
        new_file=test_files['new_file'],
        output_file=test_files['output_file']
    )
    df = pd.read_excel(test_files['output_file'])
    # Check that all KKS from new and deleted from old are present
    old_df = pd.read_excel(test_files['old_file'], sheet_name='Query')
    new_df = pd.read_excel(test_files['new_file'], sheet_name='Query')
    all_kks = set(old_df['KKS']).union(set(new_df['KKS']))
    assert set(df['KKS']) == all_kks

def test_compare_pipeline_lists_deleted_rows_at_end(test_files):
    compare_pipeline_lists(
        old_file=test_files['old_file'],
        new_file=test_files['new_file'],
        output_file=test_files['output_file']
    )
    df = pd.read_excel(test_files['output_file'])
    old_df = pd.read_excel(test_files['old_file'], sheet_name='Query')
    new_df = pd.read_excel(test_files['new_file'], sheet_name='Query')
    deleted_kks = set(old_df['KKS']) - set(new_df['KKS'])
    if deleted_kks:
        deleted_rows = df[df['KKS'].isin(deleted_kks)]
        # Deleted rows should be at the end
        assert all(deleted_rows.index >= (len(df) - len(deleted_rows)))

def test_compare_pipeline_lists_colors(test_files):
    # This test checks that openpyxl formatting is applied
    import openpyxl
    compare_pipeline_lists(
        old_file=test_files['old_file'],
        new_file=test_files['new_file'],
        output_file=test_files['output_file']
    )
    wb = openpyxl.load_workbook(test_files['output_file'])
    ws = wb.active
    # Check for at least one yellow, blue, and red fill
    yellow = 'FFFFFF00'
    blue = 'FF00B0F0'
    red = 'FFFF0000'
    fills = [cell.fill.start_color.rgb for row in ws.iter_rows(min_row=2) for cell in row]
    assert any(f == yellow for f in fills)
    assert any(f == blue for f in fills)
    assert any(f == red for f in fills)

def test_freeze_panes_and_column_widths(test_files):
    """Test that freeze panes are set at C2 and column widths are auto-adjusted."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    compare_pipeline_lists(
        old_file=test_files['old_file'],
        new_file=test_files['new_file'],
        output_file=test_files['output_file']
    )
    wb = openpyxl.load_workbook(test_files['output_file'])
    ws = wb.active
    
    # Check freeze panes
    assert ws.freeze_panes == 'C2'
    
    # Check that column widths are set (not default)
    assert any(ws.column_dimensions[get_column_letter(i+1)].width > 0 for i in range(len(list(ws.columns))))

def test_compare_pipeline_lists_edge_empty_files(test_files, tmp_path):
    # Edge case: both files empty
    empty_old = tmp_path / 'empty_old.xlsx'
    empty_new = tmp_path / 'empty_new.xlsx'
    pd.DataFrame({'KKS': []}).to_excel(empty_old, index=False, sheet_name='Query')
    pd.DataFrame({'KKS': []}).to_excel(empty_new, index=False, sheet_name='Query')
    output = tmp_path / 'empty_new_COMPARE_WITH_PREV_REV.xlsx'
    compare_pipeline_lists(str(empty_old), str(empty_new), str(output))
    df = pd.read_excel(output)
    assert len(df) == 0

def test_compare_pipeline_lists_edge_only_new_rows(test_files, tmp_path):
    # Edge case: all rows are new
    new_df = pd.DataFrame({'KKS': ['A', 'B'], 'Col1': [1, 2]})
    old_df = pd.DataFrame({'KKS': [], 'Col1': []})
    old = tmp_path / 'old.xlsx'
    new = tmp_path / 'new.xlsx'
    out = tmp_path / 'new_COMPARE_WITH_PREV_REV.xlsx'
    old_df.to_excel(old, index=False, sheet_name='Query')
    new_df.to_excel(new, index=False, sheet_name='Query')
    compare_pipeline_lists(str(old), str(new), str(out))
    df = pd.read_excel(out)
    assert set(df['KKS'].astype(str)) == {'A', 'B'}

def test_column_value_changes_colored_correctly(test_files):
    """Test that cell value changes are colored correctly."""
    import openpyxl
    
    compare_pipeline_lists(
        old_file=test_files['old_file'],
        new_file=test_files['new_file'],
        output_file=test_files['output_file']
    )
    wb = openpyxl.load_workbook(test_files['output_file'])
    ws = wb.active
    
    # Find rows by KKS value to check color logic
    yellow = 'FFFFFF00'
    blue = 'FF00B0F0'
    red = 'FFFF0000'
    
    # Find B1 row (should have yellow cells for changed values)
    b1_row = None
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if row[0].value == 'B1':  # First column is KKS
            b1_row = i + 2  # Convert to 1-based Excel row
            break
    
    # Check that at least one cell in B1 row has yellow fill (changed value)
    if b1_row:
        b1_cells = list(ws.iter_rows(min_row=b1_row, max_row=b1_row))[0]
        assert any(cell.fill.start_color.rgb == yellow for cell in b1_cells)


def test_changes_column_values_and_colors(test_files):
    """Test that the 'Changes' column has correct values and color coding."""
    import openpyxl

    compare_pipeline_lists(
        old_file=test_files['old_file'],
        new_file=test_files['new_file'],
        output_file=test_files['output_file']
    )
    wb = openpyxl.load_workbook(test_files['output_file'])
    ws = wb.active

    # Find the index of the 'Changes' column
    header = [cell.value for cell in ws[1]]
    changes_idx = header.index('Changes') + 1  # 1-based

    # Color codes
    yellow = 'FFFFFF00'
    blue = 'FF00B0F0'
    red = 'FFFF0000'

    # Build expected values for the test data
    # Test data: old KKS: A1, B1, C1, D1; new KKS: A1, B1, C1, E1
    # - A1: unchanged
    # - B1: changed (Description and Value)
    # - C1: unchanged
    # - E1: new
    # - D1: deleted
    expected = {
        'A1': ('', None),
        'B1': ('Ch', yellow),
        'C1': ('', None),
        'E1': ('N', blue),
        'D1': ('D', red),
    }

    # Map KKS to row index
    kks_col_idx = header.index('KKS') + 1
    kks_to_row = {}
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        kks = ws.cell(row=i, column=kks_col_idx).value
        kks_to_row[kks] = i

    for kks, (exp_val, exp_color) in expected.items():
        row_idx = kks_to_row[kks]
        cell = ws.cell(row=row_idx, column=changes_idx)
        assert cell.value == exp_val, f"KKS {kks}: expected '{exp_val}', got '{cell.value}'"
        if exp_color:
            assert cell.fill.start_color.rgb == exp_color, f"KKS {kks}: expected color {exp_color}, got {cell.fill.start_color.rgb}"
        else:
            # No fill means fill_type is None or start_color is not set
            assert cell.fill is not None and (cell.fill.fill_type is None or cell.fill.start_color.rgb in (None, '00000000', '000000'))
