#!/usr/bin/env python3
"""
AGG Data Comparator — Price Sheet Sikla

Compares AGGREGATED_DATA tables from two SQLite databases (old vs new revision)
and writes a colour-coded Excel workbook highlighting changes, additions, and
deletions.

The output sheet mirrors the AGGREGATED_DATA sheet produced by
price_sheet_db_tool.py:
  - Same column order (Order_Qty / Order_Weight_kg shifted right, Unit_Price /
    Total_Price appended as the order block)
  - Rows grouped by Material_Type with separator rows and per-group SUBTOTAL rows
  - Grand total row
  - Deleted rows section at the bottom, also grouped by Material_Type

Match key: Article_Number + Coating
  (same article may appear multiple times with different Coating; rows within a
  key group are matched positionally, mirroring the source order.)

Colour scheme  (data rows only; separator / subtotal / grand-total rows keep
                their original structural styling)
  GREEN      — new row or cell value added (absent in OLD)
  YELLOW     — value changed (increased, or non-numeric change)
  LIGHT BLUE — numeric value decreased (NEW < OLD)
  RED        — value deleted or entire row deleted (appended at the bottom)

Extra columns appended after the order block:
  Status           — e.g. "New Row", "Changed", "Smaller, Changed", "Deleted"
  Order_Qty_Change — signed integer: NEW Order_Qty minus OLD Order_Qty

Usage:
    python price_sheet_agg_compare.py
    python price_sheet_agg_compare.py --old rev0.sqlite --new rev1.sqlite
    python price_sheet_agg_compare.py --old rev0.sqlite --new rev1.sqlite --out compare.xlsx
"""

import argparse
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Display column order — mirrors _write_agg_sheet in price_sheet_db_tool.py
# ---------------------------------------------------------------------------
# The 15 DB columns in their natural schema order:
_DB_COLS: list[str] = [
    "Material_Type", "Article_Number", "Description", "Qty",
    "Total_Weight_kg", "Remarks", "Coating", "Spare", "Order_Qty",
    "Order_Weight_kg", "Net_Cut_Length_m", "Net_Cut_Weight_kg",
    "Total_Order_Length_m", "Utilisation_pct", "Spare_Calculation_Rule",
]
# Order_Qty and Order_Weight_kg are pulled to the right (order block):
_SHIFT_COLS: frozenset[str] = frozenset(("Order_Qty", "Order_Weight_kg"))
# Two non-DB columns appended at the far right (supplier fills Unit_Price):
_ORD_XTRA: tuple[str, ...] = ("Unit_Price", "Total_Price")
# Quantity columns get a _pcs display suffix:
_PCS_COLS: frozenset[str] = frozenset(("Qty", "Spare", "Order_Qty"))

# Final display sequence (17 columns) — importable by the test suite:
_DISP_SEQ: list[str] = (
    [c for c in _DB_COLS if c not in _SHIFT_COLS]
    + [c for c in _DB_COLS if c in _SHIFT_COLS]
    + list(_ORD_XTRA)
)
_N_DISP = len(_DISP_SEQ)   # 17

# Numeric columns that receive SUBTOTAL formulas on subtotal / grand-total rows:
_NUMERIC_COLS: tuple[str, ...] = (
    "Qty", "Total_Weight_kg", "Spare", "Order_Qty", "Order_Weight_kg", "Total_Price"
)


# ---------------------------------------------------------------------------
# Colour fills
# ---------------------------------------------------------------------------
_FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # new / added
_FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # changed (increased / non-numeric)
_FILL_LBLUE  = PatternFill("solid", fgColor="DDEBF7")   # numeric decreased
_FILL_RED    = PatternFill("solid", fgColor="FFC7CE")   # deleted

_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------
def _find_agg_table(conn: sqlite3.Connection) -> str:
    """Return the first AGGREGATED_DATA_* table found in the database."""
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'AGGREGATED_DATA_%'"
        " ORDER BY name"
    ).fetchall()
    if not rows:
        raise ValueError("No AGGREGATED_DATA_* table found in database.")
    return rows[0][0]


def _load_agg(db_path: Path) -> tuple[list[str], list[tuple]]:
    """Return (column_names, rows) from the AGGREGATED_DATA_* table."""
    conn = sqlite3.connect(db_path)
    try:
        table = _find_agg_table(conn)
        cur   = conn.cursor()
        cols  = [r[1] for r in cur.execute(f'PRAGMA table_info("{table}")')]
        rows  = cur.execute(f'SELECT * FROM "{table}"').fetchall()
        return cols, rows
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Comparison helpers
# ---------------------------------------------------------------------------
def _cell_text(value: Any) -> str:
    """Normalise a cell value to a comparable string."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _match_key(row: tuple, col_names: list[str]) -> tuple:
    """Return (Article_Number, Coating) as the unique match key."""
    art_idx  = col_names.index("Article_Number") if "Article_Number" in col_names else -1
    coat_idx = col_names.index("Coating")         if "Coating"         in col_names else -1
    art  = row[art_idx]  if art_idx  >= 0 and art_idx  < len(row) else None
    coat = row[coat_idx] if coat_idx >= 0 and coat_idx < len(row) else None
    return (art, coat)


# ---------------------------------------------------------------------------
# Excel writer — grouped, subtotalled, styled like the AGG export sheet
# ---------------------------------------------------------------------------
def _write_comparison_sheet(
    wb: Workbook,
    col_names: list[str],           # 15 DB column names
    new_rows: list[tuple],          # rows from the NEW database
    old_lookup: dict[tuple, list[tuple]],   # key → [old DB rows]
) -> None:
    ws = wb.active
    ws.title = "AGG_COMPARISON"

    # ── Map DB column names to display positions ─────────────────────────
    db_pos     = {c: i for i, c in enumerate(col_names)}
    disp_to_db = [db_pos.get(c, -1) for c in _DISP_SEQ]   # -1 = no DB source

    col_idx = {name: i + 1 for i, name in enumerate(_DISP_SEQ)}   # 1-based
    status_col    = _N_DISP + 1   # 18
    qty_chg_col   = _N_DISP + 2   # 19
    n_total_cols  = qty_chg_col

    def cl(name: str) -> str:
        return get_column_letter(col_idx[name])

    def _reorder(db_row: tuple) -> list:
        return [db_row[i] if i >= 0 else None for i in disp_to_db]

    # ── Order block column boundaries ────────────────────────────────────
    ord_ci_start = col_idx["Order_Qty"]     # first order block column
    ord_ci_end   = col_idx["Total_Price"]   # last order block column

    # ── Shared style objects ─────────────────────────────────────────────
    hdr_fill_base = PatternFill("solid", fgColor="FF203864")   # dark navy
    hdr_fill_ord  = PatternFill("solid", fgColor="FF7B3F00")   # dark amber (order block)
    hdr_fill_stat = PatternFill("solid", fgColor="FF2E4E7E")   # dark steel (status cols)
    hdr_font      = Font(bold=True, color="FFFFFFFF")

    sep_fill  = PatternFill("solid", fgColor="FF2E4E7E")   # dark steel blue
    sep_font  = Font(bold=True, color="FFFFFFFF", size=10)
    sub_fill  = PatternFill("solid", fgColor="FFDCE6F1")   # pale blue
    sub_font  = Font(bold=True, italic=True)
    tot_fill  = PatternFill("solid", fgColor="FFFFF2CC")   # yellow
    tot_font  = Font(bold=True)
    unit_fill = PatternFill("solid", fgColor="FFBDD7EE")   # light blue — supplier input
    del_title_fill = PatternFill("solid", fgColor="FFCC0000")  # dark red

    # ── Header row (row 1) ───────────────────────────────────────────────
    headers = [f"{c}_pcs" if c in _PCS_COLS else c for c in _DISP_SEQ]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill_ord if ord_ci_start <= ci <= ord_ci_end else hdr_fill_base
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for ci, label in [(status_col, "Status"), (qty_chg_col, "Order_Qty_Change")]:
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = hdr_fill_stat
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # ── Matching cursor for positional matching within key groups ─────────
    old_cursor: dict[tuple, int] = defaultdict(int)

    mt_db_idx        = db_pos.get("Material_Type", -1)
    order_qty_db_idx = db_pos.get("Order_Qty", -1)

    # ── Group NEW rows by Material_Type ───────────────────────────────────
    groups: dict[str, list[tuple]] = {}
    for row in new_rows:
        mt = (row[mt_db_idx] if mt_db_idx >= 0 else "") or ""
        groups.setdefault(mt, []).append(row)
    groups = dict(sorted(groups.items()))

    current_row = 2

    # ── Helper: fill + border every cell across a structural row ─────────
    def _style_structural(r: int, fill: PatternFill, font: Font) -> None:
        for ci in range(1, n_total_cols + 1):
            cell = ws.cell(r, ci)
            cell.fill   = fill
            cell.font   = font
            cell.border = _BORDER

    # ── Write each material-type group ───────────────────────────────────
    for group_name, group_rows in groups.items():

        # Separator row
        sep_r = current_row
        _style_structural(sep_r, sep_fill, sep_font)
        ws.cell(sep_r, 1).value     = group_name
        ws.cell(sep_r, 1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[sep_r].height = 18
        current_row += 1

        first_data = current_row

        for db_row in group_rows:
            r   = current_row
            key = _match_key(db_row, col_names)
            key_grp    = old_lookup.get(key, [])
            cursor     = old_cursor[key]
            old_db_row = key_grp[cursor] if cursor < len(key_grp) else None
            old_cursor[key] += 1

            disp_vals    = _reorder(db_row)
            change_types: set[str] = set()

            for ci, val in enumerate(disp_vals, start=1):
                col_name = _DISP_SEQ[ci - 1]
                cell = ws.cell(r, ci, value=val)
                cell.border = _BORDER

                # Unit_Price: always light blue (supplier fills this in)
                if col_name == "Unit_Price":
                    cell.fill = unit_fill
                    continue

                # Total_Price: formula only, no comparison fill
                if col_name == "Total_Price":
                    cell.value = (
                        f"=IF(ISNUMBER({cl('Unit_Price')}{r})"
                        f",{cl('Order_Qty')}{r}*{cl('Unit_Price')}{r},\"\")"
                    )
                    continue

                # ── Comparison fill ───────────────────────────────────────
                if old_db_row is None:
                    cell.fill = _FILL_GREEN
                    change_types.add("New Row")
                else:
                    db_i  = disp_to_db[ci - 1]
                    n_txt = _cell_text(val)
                    o_txt = _cell_text(old_db_row[db_i]) if db_i >= 0 else ""
                    if n_txt != o_txt:
                        if n_txt and not o_txt:
                            cell.fill = _FILL_GREEN
                            change_types.add("Added")
                        elif not n_txt and o_txt:
                            cell.fill = _FILL_RED
                            change_types.add("Deleted")
                        else:
                            try:
                                n_num = float(n_txt)
                                o_num = float(o_txt)
                                if n_num < o_num:
                                    cell.fill = _FILL_LBLUE
                                    change_types.add("Smaller")
                                else:
                                    cell.fill = _FILL_YELLOW
                                    change_types.add("Changed")
                            except ValueError:
                                cell.fill = _FILL_YELLOW
                                change_types.add("Changed")

            # Status column
            sc = ws.cell(r, status_col)
            sc.border = _BORDER
            if change_types:
                priority    = ["Deleted", "Smaller", "Changed", "Added", "New Row"]
                sc.value = ", ".join(t for t in priority if t in change_types)
                if "Deleted"  in change_types: sc.fill = _FILL_RED
                elif "Smaller" in change_types: sc.fill = _FILL_LBLUE
                elif "Changed" in change_types or "Added" in change_types: sc.fill = _FILL_YELLOW
                elif "New Row" in change_types: sc.fill = _FILL_GREEN

            # Order_Qty_Change column
            qc = ws.cell(r, qty_chg_col)
            qc.border = _BORDER
            if order_qty_db_idx >= 0:
                try:
                    n_qty = float(db_row[order_qty_db_idx] or 0)
                    o_qty = float(old_db_row[order_qty_db_idx] or 0) if old_db_row is not None else 0.0
                    delta = n_qty - o_qty
                    if delta != 0:
                        qc.value = int(delta) if float(delta).is_integer() else round(delta, 2)
                        qc.fill  = _FILL_GREEN if delta > 0 else _FILL_RED
                except (TypeError, ValueError):
                    pass

            current_row += 1

        last_data = current_row - 1

        # Subtotal row
        sub_r = current_row
        _style_structural(sub_r, sub_fill, sub_font)
        ws.cell(sub_r, col_idx["Material_Type"]).value = group_name
        ws.cell(sub_r, col_idx["Description"]).value   = f"\u21b3 Subtotal \u2014 {group_name}"
        for col_name in _NUMERIC_COLS:
            if col_name in col_idx:
                c = ws.cell(sub_r, col_idx[col_name])
                c.value = f"=SUBTOTAL(9,{cl(col_name)}{first_data}:{cl(col_name)}{last_data})"
        ws.cell(sub_r, 1).alignment = Alignment(
            horizontal="left", vertical="center", indent=2
        )
        ws.row_dimensions[sub_r].height = 16
        current_row += 1

    # ── Number formats ────────────────────────────────────────────────────
    _int_fmts = {col_idx[n] for n in ["Qty", "Spare", "Order_Qty"] if n in col_idx}
    _2dp_fmts = {col_idx[n] for n in [
        "Total_Weight_kg", "Order_Weight_kg", "Net_Cut_Weight_kg",
        "Net_Cut_Length_m", "Total_Order_Length_m", "Unit_Price", "Total_Price",
    ] if n in col_idx}
    _1dp_fmts = {col_idx[n] for n in ["Utilisation_pct"] if n in col_idx}

    for row in ws.iter_rows(min_row=2, max_row=current_row - 1):
        for cell in row:
            ci = cell.column
            if ci in _int_fmts:   cell.number_format = "0"
            elif ci in _2dp_fmts: cell.number_format = "0.00"
            elif ci in _1dp_fmts: cell.number_format = "0.0"

    # ── Grand total row ───────────────────────────────────────────────────
    grand_row = current_row
    _style_structural(grand_row, tot_fill, tot_font)
    ws.cell(grand_row, col_idx["Description"]).value = "GRAND TOTAL"
    for col_name in _NUMERIC_COLS:
        if col_name in col_idx:
            c = ws.cell(grand_row, col_idx[col_name])
            c.value = f"=SUBTOTAL(9,{cl(col_name)}2:{cl(col_name)}{grand_row - 1})"
    ws.cell(grand_row, col_idx["Description"]).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.row_dimensions[grand_row].height = 20
    current_row += 1

    # ── Deleted rows section — grouped by Material_Type ───────────────────
    deleted_by_group: dict[str, list[tuple]] = {}
    for key, rows_list in old_lookup.items():
        consumed = old_cursor.get(key, 0)
        for del_row in rows_list[consumed:]:
            mt = (del_row[mt_db_idx] if mt_db_idx >= 0 else "") or ""
            deleted_by_group.setdefault(mt, []).append(del_row)

    if deleted_by_group:
        current_row += 1   # blank gap

        title_cell      = ws.cell(current_row, 1, value="Deleted Rows")
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = del_title_fill
        current_row += 1

        for group_name in sorted(deleted_by_group.keys()):
            sep_r = current_row
            _style_structural(sep_r, sep_fill, sep_font)
            ws.cell(sep_r, 1).value     = group_name
            ws.cell(sep_r, 1).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            ws.row_dimensions[sep_r].height = 18
            current_row += 1

            for del_row in deleted_by_group[group_name]:
                r         = current_row
                disp_vals = _reorder(del_row)
                for ci, val in enumerate(disp_vals, start=1):
                    cell        = ws.cell(r, ci, value=val)
                    cell.fill   = _FILL_RED
                    cell.border = _BORDER
                sc        = ws.cell(r, status_col, value="Deleted")
                sc.fill   = _FILL_RED
                sc.border = _BORDER
                qc = ws.cell(r, qty_chg_col)
                qc.border = _BORDER
                if order_qty_db_idx >= 0:
                    try:
                        o_qty    = float(del_row[order_qty_db_idx] or 0)
                        qc.value = int(-o_qty) if float(-o_qty).is_integer() else round(-o_qty, 2)
                        qc.fill  = _FILL_RED
                    except (TypeError, ValueError):
                        pass
                current_row += 1

    # ── Medium-border frame around the order block ────────────────────────
    for r in range(1, current_row):
        for ci in range(ord_ci_start, ord_ci_end + 1):
            c   = ws.cell(r, ci)
            lft = _MEDIUM if ci == ord_ci_start else c.border.left
            rgt = _MEDIUM if ci == ord_ci_end   else c.border.right
            c.border = Border(
                left=lft, right=rgt,
                top=c.border.top, bottom=c.border.bottom,
            )

    # ── Auto column widths ────────────────────────────────────────────────
    for col_cells in ws.columns:
        max_len = max(
            (
                len(str(c.value))
                if c.value is not None
                and not (isinstance(c.value, str) and c.value.startswith("="))
                else 0
            )
            for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 5, 60)

    for ci, min_w in [(status_col, 16), (qty_chg_col, 18)]:
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = max(
            ws.column_dimensions[letter].width, min_w
        )

    # ── Hide columns I–M (same as the regular AGG export) ─────────────────
    for hide_col in ("I", "J", "K", "L", "M"):
        ws.column_dimensions[hide_col].hidden = True

    # ── Autofilter covers all columns including Status / Qty_Change ───────
    ws.auto_filter.ref = f"A1:{get_column_letter(n_total_cols)}1"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def run(old_db: Path, new_db: Path, out_xlsx: Path | None = None) -> Path:
    print(f"OLD : {old_db}")
    print(f"NEW : {new_db}")

    old_cols, old_rows = _load_agg(old_db)
    new_cols, new_rows = _load_agg(new_db)

    if old_cols != new_cols:
        print(f"WARNING: column mismatch\n  OLD={old_cols}\n  NEW={new_cols}")

    # Build lookup: key → list of old rows (preserves order for positional matching)
    old_lookup: dict[tuple, list[tuple]] = defaultdict(list)
    for row in old_rows:
        old_lookup[_match_key(row, old_cols)].append(row)

    # Determine output path
    if out_xlsx is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_xlsx  = old_db.parent / f"{timestamp}_AGG_COMPARE.xlsx"

    wb = Workbook()
    wb.calculation.calcMode    = "auto"
    wb.calculation.fullCalcOnLoad = True
    _write_comparison_sheet(wb, new_cols, new_rows, old_lookup)
    wb.save(out_xlsx)

    # ── Summary stats ─────────────────────────────────────────────────────
    old_cursor_s: dict[tuple, int] = defaultdict(int)
    n_new = n_changed = 0
    for new_r in new_rows:
        key  = _match_key(new_r, new_cols)
        grp  = old_lookup.get(key, [])
        idx  = old_cursor_s[key]
        old_cursor_s[key] += 1
        if idx >= len(grp):
            n_new += 1
        else:
            old_r = grp[idx]
            if any(_cell_text(new_r[i]) != _cell_text(old_r[i]) for i in range(len(new_cols))):
                n_changed += 1
    n_deleted = sum(
        max(0, len(v) - old_cursor_s.get(k, 0)) for k, v in old_lookup.items()
    )

    print(f"\nResults : {n_new} new  |  {n_changed} changed  |  {n_deleted} deleted")
    print(f"Output  : {out_xlsx}")
    return out_xlsx


def main() -> None:
    _DEFAULT_OLD = (
        r"C:\Users\szil\Repos\excel_wizadry\Price Sheet Sikla"
        r"\_test_runs\20260520_110103_REV0.0.sqlite"
    )
    _DEFAULT_NEW = (
        r"C:\Users\szil\Repos\excel_wizadry\Price Sheet Sikla"
        r"\_test_runs\20260520_110103_REV1.0.sqlite"
    )

    parser = argparse.ArgumentParser(
        description="Compare AGGREGATED_DATA between two Price Sheet Sikla SQLite databases.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python price_sheet_agg_compare.py
  python price_sheet_agg_compare.py --old rev0.sqlite --new rev1.sqlite
  python price_sheet_agg_compare.py --old rev0.sqlite --new rev1.sqlite --out compare.xlsx
        """,
    )
    parser.add_argument("--old", default=_DEFAULT_OLD, help="Path to OLD SQLite database")
    parser.add_argument("--new", default=_DEFAULT_NEW, help="Path to NEW SQLite database")
    parser.add_argument("--out", default=None,         help="Output xlsx path (auto-named if omitted)")
    args = parser.parse_args()

    run(Path(args.old), Path(args.new), Path(args.out) if args.out else None)


if __name__ == "__main__":
    main()
