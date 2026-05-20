#!/usr/bin/env python3
"""
Validation test suite for price_sheet_agg_compare.py

Each test builds synthetic SQLite databases with known data, runs the
comparator, reloads the Excel output and asserts exact fills, status
labels and Order_Qty_Change values.

Run from the 'Price Sheet Sikla' folder:
    python price_sheet_agg_compare_test.py
"""

import sqlite3
import sys
import tempfile
import traceback
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
import price_sheet_agg_compare as cmp   # module under test

# ---------------------------------------------------------------------------
# DB schema (15 columns) and display column positions
# The display order is defined in price_sheet_agg_compare._DISP_SEQ.
# ---------------------------------------------------------------------------
_DB_COLS = cmp._DB_COLS                          # 15 DB column names
_N_DB_COLS = len(_DB_COLS)                       # 15
_IDX = {c: i for i, c in enumerate(_DB_COLS)}   # DB schema index (0-based)

# Display column sequence and 1-based sheet positions (matches _write_agg_sheet):
_DISP_SEQ    = cmp._DISP_SEQ                     # 17 display columns
_N_DISP_COLS = cmp._N_DISP                       # 17
_DISP_IDX    = {c: i + 1 for i, c in enumerate(_DISP_SEQ)}  # 1-based sheet col

_PCS_COLS    = cmp._PCS_COLS

_STATUS_COL  = _N_DISP_COLS + 1   # 18
_QTY_CHG_COL = _N_DISP_COLS + 2   # 19

# For _make_row the DB column indices still apply:
_COLS = _DB_COLS
_N_COLS = _N_DB_COLS

# Expected fill colours (6-char, no alpha prefix)
_GREEN  = "C6EFCE"
_YELLOW = "FFEB9C"
_LBLUE  = "DDEBF7"
_RED    = "FFC7CE"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_row(
    art: int,
    coating: str | None = None,
    order_qty: int = 10,
    desc: str = "Test Item",
    remarks: str | None = None,
    spare: int = 0,
    mat_type: str = "01 Primary Supports",
) -> tuple:
    """Build a full 15-column row tuple.  Unspecified fields are None / 0."""
    row = [None] * _N_COLS
    row[_IDX["Material_Type"]]          = mat_type
    row[_IDX["Article_Number"]]         = art
    row[_IDX["Description"]]            = desc
    row[_IDX["Qty"]]                    = order_qty
    row[_IDX["Remarks"]]                = remarks
    row[_IDX["Coating"]]                = coating
    row[_IDX["Spare"]]                  = spare
    row[_IDX["Order_Qty"]]              = order_qty
    row[_IDX["Spare_Calculation_Rule"]] = "5% spare"
    return tuple(row)


def _create_db(path: Path, rows: list[tuple]) -> None:
    """Create a minimal AGG SQLite DB with the given rows."""
    def _col_type(name: str) -> str:
        if name in ("Article_Number", "Qty", "Spare", "Order_Qty"):
            return "INTEGER"
        if "kg" in name or name.endswith("_m") or "pct" in name:
            return "REAL"
        return "TEXT"

    conn = sqlite3.connect(path)
    cur  = conn.cursor()
    col_defs = ", ".join(f'"{c}" {_col_type(c)}' for c in _DB_COLS)
    cur.execute(f'CREATE TABLE "AGGREGATED_DATA_REV0" ({col_defs})')
    ph = ", ".join("?" * _N_DB_COLS)
    cur.executemany(f'INSERT INTO "AGGREGATED_DATA_REV0" VALUES ({ph})', rows)
    conn.commit()
    conn.close()


def _run(old_rows: list[tuple], new_rows: list[tuple], tmp: Path):
    """Create DBs, run compare, return loaded workbook (fills preserved)."""
    old_db   = tmp / "old.sqlite"
    new_db   = tmp / "new.sqlite"
    out_xlsx = tmp / "compare.xlsx"
    _create_db(old_db, old_rows)
    _create_db(new_db, new_rows)
    cmp.run(old_db, new_db, out_xlsx)
    return load_workbook(out_xlsx, data_only=True)


def _rgb(cell) -> str | None:
    """Return last 6-char fill colour (upper-case) or None if no solid fill."""
    try:
        fill = cell.fill
        if fill and fill.fill_type == "solid":
            rgb = fill.fgColor.rgb or ""
            # Transparent / unset colours
            if rgb in ("00000000", "FF000000", ""):
                return None
            return rgb[-6:].upper()
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Test runner
# ---------------------------------------------------------------------------
_pass = _fail = 0


def _chk(name: str, ok: bool, detail: str = "") -> None:
    global _pass, _fail
    if ok:
        _pass += 1
        print(f"    PASS  {name}")
    else:
        _fail += 1
        extra = f"  [{detail}]" if detail else ""
        print(f"    FAIL  {name}{extra}")


def _section(title: str) -> None:
    print(f"\n  {'─'*56}")
    print(f"  {title}")
    print(f"  {'─'*56}")


# ---------------------------------------------------------------------------
# Individual tests
# ---------------------------------------------------------------------------

def t01_unchanged_row(tmp: Path) -> None:
    _section("T01  Unchanged row — no fill on data row, Status and Qty_Change empty")
    # Row layout: 1=header, 2=separator, 3=data, 4=subtotal, 5=grand total
    row = _make_row(art=1000, order_qty=10)
    ws  = _run([row], [row], tmp).active
    r   = 3   # first data row

    _unit_price_col  = _DISP_IDX["Unit_Price"]
    _total_price_col = _DISP_IDX["Total_Price"]
    skip = {_unit_price_col, _total_price_col}  # always styled
    any_fill = any(_rgb(ws.cell(r, c)) for c in range(1, _N_DISP_COLS + 3) if c not in skip)
    _chk("no fill on data row cells (excl Unit_Price/Total_Price)",     not any_fill,
         str([_rgb(ws.cell(r, c)) for c in range(1, _N_DISP_COLS + 3) if c not in skip]))
    _chk("Status cell empty",             ws.cell(r, _STATUS_COL).value   is None)
    _chk("Order_Qty_Change cell empty",   ws.cell(r, _QTY_CHG_COL).value is None)


def t02_new_row(tmp: Path) -> None:
    _section("T02  New row (key in NEW only) — all green, Status='New Row', Qty_Change=+qty")
    # Row layout: 1=header, 2=separator, 3=existing(unchanged), 4=brand_new(new)
    existing = _make_row(art=1001, order_qty=10)
    brand_new = _make_row(art=2000, order_qty=5, desc="Brand New")
    ws = _run([existing], [existing, brand_new], tmp).active
    r  = 4   # brand_new is the 2nd data row in the group

    # Check all display data columns (excluding Unit_Price col 16 and Total_Price col 17
    # which are always styled independently)
    _unit_price_col  = _DISP_IDX["Unit_Price"]
    _total_price_col = _DISP_IDX["Total_Price"]
    skip = {_unit_price_col, _total_price_col}
    data_cols = [c for c in range(1, _N_DISP_COLS + 1) if c not in skip]
    all_green = all(_rgb(ws.cell(r, c)) == _GREEN for c in data_cols)
    _chk("all data cells green (excl Unit_Price/Total_Price)", all_green,
         str([_rgb(ws.cell(r, c)) for c in data_cols]))
    _chk("Status = 'New Row'",           ws.cell(r, _STATUS_COL).value == "New Row")
    _chk("Order_Qty_Change = +5",        ws.cell(r, _QTY_CHG_COL).value == 5)
    _chk("Status fill green",            _rgb(ws.cell(r, _STATUS_COL))   == _GREEN)
    _chk("Qty_Change fill green",        _rgb(ws.cell(r, _QTY_CHG_COL)) == _GREEN)


def t03_deleted_row(tmp: Path) -> None:
    _section("T03  Deleted row — red section at bottom, Status='Deleted', Qty_Change=-qty")
    # Deleted section layout: ... | blank | 'Deleted Rows' title | separator | data row
    deleted = _make_row(art=3000, order_qty=3, desc="Vanishing Part")
    kept    = _make_row(art=1001, order_qty=10)
    ws = _run([deleted, kept], [kept], tmp).active

    del_title_row = next(
        (r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Deleted Rows"),
        None,
    )
    _chk("'Deleted Rows' title exists",  del_title_row is not None)
    if del_title_row is None:
        return

    # title + separator row + data row
    data_r = del_title_row + 2
    all_red = all(_rgb(ws.cell(data_r, c)) == _RED for c in range(1, _N_DISP_COLS + 1))
    _chk("all deleted cells red",        all_red)
    _chk("Status = 'Deleted'",           ws.cell(data_r, _STATUS_COL).value  == "Deleted")
    _chk("Order_Qty_Change = -3",        ws.cell(data_r, _QTY_CHG_COL).value == -3)
    _chk("Status fill red",              _rgb(ws.cell(data_r, _STATUS_COL))   == _RED)
    _chk("Qty_Change fill red",          _rgb(ws.cell(data_r, _QTY_CHG_COL)) == _RED)


def t04_order_qty_increased(tmp: Path) -> None:
    _section("T04  Order_Qty increased — yellow cell, Status has 'Changed', Qty_Change=+delta")
    # Row layout: 1=header, 2=separator, 3=data, 4=subtotal, 5=grand total
    old = _make_row(art=4000, order_qty=10)
    new = _make_row(art=4000, order_qty=20)
    ws  = _run([old], [new], tmp).active
    r   = 3
    qc  = _DISP_IDX["Order_Qty"]   # 1-based display column

    _chk("Order_Qty cell yellow",        _rgb(ws.cell(r, qc)) == _YELLOW,
         _rgb(ws.cell(r, qc)))
    _chk("Status contains 'Changed'",   "Changed" in (ws.cell(r, _STATUS_COL).value or ""))
    _chk("Order_Qty_Change = +10",       ws.cell(r, _QTY_CHG_COL).value == 10)
    _chk("Qty_Change fill green",        _rgb(ws.cell(r, _QTY_CHG_COL)) == _GREEN)


def t05_order_qty_decreased(tmp: Path) -> None:
    _section("T05  Order_Qty decreased — light-blue cell, Status has 'Smaller', Qty_Change=-delta")
    old = _make_row(art=5000, order_qty=20)
    new = _make_row(art=5000, order_qty=10)
    ws  = _run([old], [new], tmp).active
    r   = 3
    qc  = _DISP_IDX["Order_Qty"]

    _chk("Order_Qty cell light blue",    _rgb(ws.cell(r, qc)) == _LBLUE,
         _rgb(ws.cell(r, qc)))
    _chk("Status contains 'Smaller'",   "Smaller" in (ws.cell(r, _STATUS_COL).value or ""))
    _chk("Order_Qty_Change = -10",       ws.cell(r, _QTY_CHG_COL).value == -10)
    _chk("Qty_Change fill red",          _rgb(ws.cell(r, _QTY_CHG_COL)) == _RED)


def t06_description_changed(tmp: Path) -> None:
    _section("T06  Description changed (non-numeric) — yellow, Status 'Changed', no Qty_Change")
    old = _make_row(art=6000, order_qty=10, desc="Old Description")
    new = _make_row(art=6000, order_qty=10, desc="New Description")
    ws  = _run([old], [new], tmp).active
    r   = 3
    dc  = _DISP_IDX["Description"]

    _chk("Description cell yellow",      _rgb(ws.cell(r, dc)) == _YELLOW,
         _rgb(ws.cell(r, dc)))
    _chk("Status contains 'Changed'",   "Changed" in (ws.cell(r, _STATUS_COL).value or ""))
    _chk("Order_Qty_Change empty",       ws.cell(r, _QTY_CHG_COL).value is None)
    # Unchanged Order_Qty cell must have no fill
    _chk("Order_Qty cell no fill",       _rgb(ws.cell(r, _DISP_IDX["Order_Qty"])) is None)


def t07_cell_added(tmp: Path) -> None:
    _section("T07  Cell added (None → value) — green fill, Status has 'Added'")
    old = _make_row(art=7000, order_qty=10, remarks=None)
    new = _make_row(art=7000, order_qty=10, remarks="Rush order")
    ws  = _run([old], [new], tmp).active
    r   = 3
    rc  = _DISP_IDX["Remarks"]

    _chk("Remarks cell green",           _rgb(ws.cell(r, rc)) == _GREEN,
         _rgb(ws.cell(r, rc)))
    _chk("Status contains 'Added'",     "Added" in (ws.cell(r, _STATUS_COL).value or ""))


def t08_cell_deleted(tmp: Path) -> None:
    _section("T08  Cell deleted (value → None) — red fill, Status has 'Deleted'")
    old = _make_row(art=8000, order_qty=10, remarks="Was here")
    new = _make_row(art=8000, order_qty=10, remarks=None)
    ws  = _run([old], [new], tmp).active
    r   = 3
    rc  = _DISP_IDX["Remarks"]

    _chk("Remarks cell red",             _rgb(ws.cell(r, rc)) == _RED,
         _rgb(ws.cell(r, rc)))
    _chk("Status contains 'Deleted'",   "Deleted" in (ws.cell(r, _STATUS_COL).value or ""))


def t09_coating_as_key(tmp: Path) -> None:
    _section("T09  Same article + different Coating matched independently")
    # Row layout: 1=header, 2=separator, 3=HCP(changed), 4=C5L(unchanged), 5=subtotal
    old_hcp = _make_row(art=9000, coating="HCP", order_qty=5)
    old_c5  = _make_row(art=9000, coating="C5L", order_qty=3)
    new_hcp = _make_row(art=9000, coating="HCP", order_qty=7)   # changed
    new_c5  = _make_row(art=9000, coating="C5L", order_qty=3)   # unchanged
    ws = _run([old_hcp, old_c5], [new_hcp, new_c5], tmp).active
    qc = _DISP_IDX["Order_Qty"]

    _chk("HCP Order_Qty yellow",         _rgb(ws.cell(3, qc)) == _YELLOW,
         _rgb(ws.cell(3, qc)))
    _chk("HCP Qty_Change = +2",          ws.cell(3, _QTY_CHG_COL).value == 2)
    _chk("C5L Order_Qty no fill",        _rgb(ws.cell(4, qc)) is None,
         _rgb(ws.cell(4, qc)))
    _chk("C5L Status empty",             ws.cell(4, _STATUS_COL).value is None)
    _chk("No Deleted Rows section",
         not any(ws.cell(r, 1).value == "Deleted Rows" for r in range(1, ws.max_row + 1)))


def t10_positional_matching_within_key(tmp: Path) -> None:
    _section("T10  Two OLD rows / one NEW with same key — first matched, second deleted")
    # Row layout: 1=header, 2=separator, 3=new1(matched/unchanged), 4=subtotal ...
    # Deleted section: ... | blank | 'Deleted Rows' | separator | old2
    old1 = _make_row(art=10000, coating=None, order_qty=5, desc="First")
    old2 = _make_row(art=10000, coating=None, order_qty=3, desc="Second")
    new1 = _make_row(art=10000, coating=None, order_qty=5, desc="First")  # identical to old1
    ws   = _run([old1, old2], [new1], tmp).active

    _unit_price_col  = _DISP_IDX["Unit_Price"]
    _total_price_col = _DISP_IDX["Total_Price"]
    skip = {_unit_price_col, _total_price_col}
    any_fill_r3 = any(_rgb(ws.cell(3, c)) for c in range(1, _N_DISP_COLS + 3) if c not in skip)
    _chk("Matched row has no fill (excl Unit_Price/Total_Price)",      not any_fill_r3)
    _chk("Matched row Status empty",     ws.cell(3, _STATUS_COL).value is None)

    del_title_row = next(
        (r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Deleted Rows"),
        None,
    )
    _chk("'Deleted Rows' section exists", del_title_row is not None)
    if del_title_row:
        dr = del_title_row + 2   # title + separator + data
        _chk("Deleted row desc = 'Second'",  ws.cell(dr, _DISP_IDX["Description"]).value == "Second")
        _chk("Deleted Qty_Change = -3",      ws.cell(dr, _QTY_CHG_COL).value == -3)


def t11_column_positions(tmp: Path) -> None:
    _section("T11  Header labels — 17 display cols, Status at 18, Order_Qty_Change at 19")
    row = _make_row(art=11000, order_qty=10)
    ws  = _run([row], [row], tmp).active

    # Expected display headers (with _pcs suffixes for quantity columns)
    expected = [f"{c}_pcs" if c in _PCS_COLS else c for c in _DISP_SEQ]
    headers  = [ws.cell(1, c).value for c in range(1, _N_DISP_COLS + 1)]
    _chk("17 display headers match expected", headers == expected, str(headers))
    _chk(f"Col {_STATUS_COL} = 'Status'",            ws.cell(1, _STATUS_COL).value   == "Status")
    _chk(f"Col {_QTY_CHG_COL} = 'Order_Qty_Change'", ws.cell(1, _QTY_CHG_COL).value == "Order_Qty_Change")


def t12_no_deleted_section_when_absent(tmp: Path) -> None:
    _section("T12  No 'Deleted Rows' section when nothing is deleted")
    row = _make_row(art=12000, order_qty=10)
    ws  = _run([row], [row], tmp).active

    has_section = any(
        ws.cell(r, 1).value == "Deleted Rows" for r in range(1, ws.max_row + 1)
    )
    _chk("No 'Deleted Rows' title cell", not has_section)


def t13_net_qty_change_sum(tmp: Path) -> None:
    _section("T13  Net Order_Qty_Change sum (body + deleted section) = SQL diff")
    old_rows = [
        _make_row(art=100, order_qty=10),
        _make_row(art=200, order_qty=20),
        _make_row(art=300, order_qty=5),   # will be deleted
    ]
    new_rows = [
        _make_row(art=100, order_qty=15),  # +5
        _make_row(art=200, order_qty=18),  # -2
        _make_row(art=400, order_qty=7),   # new +7
        # art=300 absent → deleted → -5
    ]
    expected_net = (15 - 10) + (18 - 20) + 7 + (-5)   # = +5
    ws = _run(old_rows, new_rows, tmp).active

    total = sum(
        v for r in range(1, ws.max_row + 1)
        if isinstance(v := ws.cell(r, _QTY_CHG_COL).value, (int, float))
    )
    _chk(f"Sum of Qty_Change column = {expected_net}", total == expected_net,
         f"got {total}")


def t14_mixed_changes_in_one_row(tmp: Path) -> None:
    _section("T14  Row with mixed changes — Status lists all applicable types")
    # Description changed + Order_Qty decreased
    old = _make_row(art=14000, order_qty=20, desc="Old")
    new = _make_row(art=14000, order_qty=10, desc="New")
    ws  = _run([old], [new], tmp).active
    r   = 3
    status = ws.cell(r, _STATUS_COL).value or ""

    _chk("Status contains 'Smaller'",    "Smaller" in status)
    _chk("Status contains 'Changed'",    "Changed" in status)
    _chk("Description cell yellow",      _rgb(ws.cell(r, _DISP_IDX["Description"])) == _YELLOW)
    _chk("Order_Qty cell light blue",    _rgb(ws.cell(r, _DISP_IDX["Order_Qty"]))   == _LBLUE)


def t15_new_row_order_qty_change_is_full_qty(tmp: Path) -> None:
    _section("T15  New row Order_Qty_Change equals full Order_Qty (no old baseline)")
    # Row layout: 1=header, 2=separator, 3=old(unchanged), 4=new(new row)
    old = _make_row(art=1111, order_qty=50)
    new = _make_row(art=9999, order_qty=33, desc="Entirely New")
    ws  = _run([old], [old, new], tmp).active
    r   = 4   # row 2=separator, row 3=old unchanged, row 4=new

    _chk("Status = 'New Row'",           ws.cell(r, _STATUS_COL).value == "New Row")
    _chk("Order_Qty_Change = 33",        ws.cell(r, _QTY_CHG_COL).value == 33)


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 62)
    print("  price_sheet_agg_compare.py  —  Validation Test Suite")
    print("=" * 62)

    tests = [
        t01_unchanged_row,
        t02_new_row,
        t03_deleted_row,
        t04_order_qty_increased,
        t05_order_qty_decreased,
        t06_description_changed,
        t07_cell_added,
        t08_cell_deleted,
        t09_coating_as_key,
        t10_positional_matching_within_key,
        t11_column_positions,
        t12_no_deleted_section_when_absent,
        t13_net_qty_change_sum,
        t14_mixed_changes_in_one_row,
        t15_new_row_order_qty_change_is_full_qty,
    ]

    for fn in tests:
        with tempfile.TemporaryDirectory() as td:
            try:
                fn(Path(td))
            except Exception as exc:
                _section(f"EXCEPTION in {fn.__name__}")
                traceback.print_exc()

    print(f"\n{'='*62}")
    verdict = "ALL PASS" if _fail == 0 else f"{_fail} FAIL"
    print(f"  {verdict}  —  {_pass} passed  {_fail} failed  ({_pass + _fail} total checks)")
    print(f"{'='*62}")
    sys.exit(0 if _fail == 0 else 1)


if __name__ == "__main__":
    main()
