import argparse
import math
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_HEADER_TITLES = {"item number", "qty", "description", "decscription"}
UNIT_PATTERN = re.compile(r"^\s*([+-]?\d+(?:[\.,]\d+)?)\s*(kg|mm|m)\s*$", re.IGNORECASE)
NUMBER_PATTERN = re.compile(r"^\s*([+-]?\d+(?:[\.,]\d+)?)\s*$")

MATERIAL_TYPE_RULES: list[dict] = [
    {
        "label": "04 Beam Sections",
        "prefixes": ["TP F", "FLS", "MS"],
        "keywords": ["channel", "rail", "profile", "beam section ms", "beam section tp"],
    },
    {
        "label": "02 Brackets Consoles",
        "prefixes": ["WD F", "WBD", "SA F"],
        "keywords": ["bracket", "console", "base", "pivot joint", "connector", "mounting angle", "beam clip", "beam connection", "holder", "eye", "plate", "end support", "adapter", "joining"],
    },
    {
        "label": "01 Primary Supports",
        "prefixes": ["Stabil", "Ratio"],
        "keywords": ["clamp", "hanger", "pipe clamp", "assembly set", "fixed point", "guided", "sliding", "guiding", "assembly", "axial", "threaded rod"],
    },
    {
        "label": "03 Bolts Screws Nuts",
        "prefixes": ["HZM", "SK", "Formlock"],
        "keywords": ["bolt", "screw", "nut", "washer", "resin anchor"],
    },
    {
        "label": "05 Installation Material",
        "prefixes": [],
        "keywords": ["glass fabric", "end cap"],
    },
]


# ---------------------------------------------------------------------------
#  CUT OPTIMIZER — generic constants and FFD helpers
# ---------------------------------------------------------------------------
_KERF_MM: int = 3            # saw-cut allowance per cut
_ORDER_BUFFER: float = 1.05  # 5% workshop safety buffer

# Beam sections: 6000mm physical, no end-loss scrap
_BEAM_PHYSICAL_MM: int = 6000
_BEAM_BAR_MM: int = 6000     # = _BEAM_PHYSICAL_MM (no end-loss)

# Spare percentages by material group — ceil(qty * pct) added on top of net qty.
# Groups 04/09 Beam Sections use FFD-based spare; not listed here.
# Group 01 threaded rods use FFD-based spare; non-rod 01 items use 5% rule below.
# Group 05 glass fabric tape uses _TAPE_SPARE_FACTOR; other 05 items (end caps) use 15% below.
# For 5% groups: spare = 0 when qty < 10; spare = ceil(qty * 0.05) when qty >= 10.
_SPARE_PCT: dict[str, float] = {
    "01 Primary Supports":         0.05,
    "02 Brackets Consoles":        0.05,
    "03 Bolts Screws Nuts":        0.15,
    "05 Installation Material":    0.15,
    "06 Primary Supports C5":      0.05,
    "07 Brackets Consoles C5":     0.05,
    "08 Bolts Screws Nuts C5":     0.15,
    "10 Installation Material C5": 0.15,
}


def _ffd_bars(pieces: list[int], effective_bar_mm: int) -> int:
    """First Fit Decreasing — pieces must be sorted descending. Returns bar count."""
    bins: list[int] = [effective_bar_mm]
    for piece in pieces:
        needed = min(piece + _KERF_MM, effective_bar_mm)
        placed = False
        for i, rem in enumerate(bins):
            if rem >= needed:
                bins[i] -= needed
                placed = True
                break
        if not placed:
            bins.append(effective_bar_mm - needed)
    return len(bins)


def _calc_ordered_bars(num_bars: int, physical_bar_mm: int) -> int:
    """Apply 5% buffer then round up to whole bars."""
    return math.ceil(num_bars * physical_bar_mm * _ORDER_BUFFER / physical_bar_mm)


def _spare_for_pct(qty: int, pct: float) -> int:
    """Compute spare quantity for a given spare percentage.
    For 5% groups: 0 spare when qty < 10; ceil(qty * 0.05) otherwise.
    For all other percentages: ceil(qty * pct) unconditionally.
    """
    if pct == 0.0 or qty == 0:
        return 0
    if pct == 0.05:
        return 0 if qty < 10 else math.ceil(qty * pct)
    return math.ceil(qty * pct)


def _beam_ffd_bars(pieces: list[int]) -> int:
    return _ffd_bars(pieces, _BEAM_BAR_MM)


def _beam_ordered_bars(num_bars: int) -> int:
    return _calc_ordered_bars(num_bars, _BEAM_PHYSICAL_MM)


_ROD_BAR_LEN_RE = re.compile(r'/(\d+)', re.IGNORECASE)


def _rod_bar_length_mm(description: str) -> int | None:
    """Parse nominal bar length from description, e.g. 'M10/2000 HCP' -> 2000."""
    m = _ROD_BAR_LEN_RE.search(description or "")
    if m:
        return int(m.group(1))
    return None


# ---------------------------------------------------------------------------


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def sanitize_column_name(name: str, fallback_index: int) -> str:
    cleaned = re.sub(r"\s+", "_", name.strip())
    cleaned = re.sub(r"[^0-9a-zA-Z_]", "", cleaned)
    if not cleaned:
        cleaned = f"COL_{fallback_index}"
    if cleaned[0].isdigit():
        cleaned = f"C_{cleaned}"
    return cleaned


def unique_column_names(raw_headers: list[Any]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for idx, header in enumerate(raw_headers, start=1):
        base = sanitize_column_name(str(header) if header is not None else "", idx)
        count = seen.get(base, 0)
        seen[base] = count + 1
        final_name = base if count == 0 else f"{base}_{count + 1}"
        names.append(final_name)
    return names


def parse_numeric_text(text: str) -> tuple[bool, Any]:
    m_unit = UNIT_PATTERN.match(text)
    if m_unit:
        num = m_unit.group(1).replace(",", ".")
        if "." in num:
            return True, round(float(num), 2)
        return True, int(num)

    m_num = NUMBER_PATTERN.match(text)
    if m_num:
        num = m_num.group(1).replace(",", ".")
        if "." in num:
            return True, round(float(num), 2)
        return True, int(num)

    return False, text.strip()


def normalize_cell_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return round(value, 2)

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        is_num, parsed = parse_numeric_text(stripped)
        if is_num:
            return parsed
        return stripped

    return str(value)


def sql_quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def infer_sqlite_types(rows: list[list[Any]], col_count: int) -> list[str]:
    inferred: list[str] = []
    for col in range(col_count):
        non_null = [r[col] for r in rows if r[col] is not None]
        if not non_null:
            inferred.append("TEXT")
            continue

        all_int = all(isinstance(v, int) and not isinstance(v, bool) for v in non_null)
        all_num = all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in non_null)
        if all_int:
            inferred.append("INTEGER")
        elif all_num:
            inferred.append("REAL")
        else:
            inferred.append("TEXT")

    return inferred


def classify_material_type(description: str | None) -> str:
    if not description:
        return ""
    desc_lower = description.lower()
    for rule in MATERIAL_TYPE_RULES:
        for prefix in rule["prefixes"]:
            if desc_lower.startswith(prefix.lower()):
                return rule["label"]
    for rule in MATERIAL_TYPE_RULES:
        for keyword in rule["keywords"]:
            if keyword in desc_lower:
                return rule["label"]
    return ""


def _find_header_row_in_df(df: "pd.DataFrame") -> int:
    for i, row in df.iterrows():
        normalized = [normalize_text(v) for v in row]
        if (
            "item number" in normalized
            and "qty" in normalized
            and ("description" in normalized or "decscription" in normalized)
        ):
            return int(i)
    raise ValueError(
        "Could not find header row with Item Number, Qty and Description/Decscription in sheet 'Total Qty'."
    )


def import_total_qty_to_sqlite(
    excel_path: Path, revision: str = "REV0", db_path: Path | None = None,
    sheet_name: str = "Total Qty",
) -> Path:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    raw_df = pd.read_excel(
        excel_path,
        sheet_name=sheet_name,
        header=None,
        engine="calamine",
    )

    header_row_idx = _find_header_row_in_df(raw_df)
    raw_headers: list[Any] = raw_df.iloc[header_row_idx].tolist()
    column_names = unique_column_names(raw_headers)

    # skip the header row itself and any immediately following blank row
    data_df = raw_df.iloc[header_row_idx + 1 :].reset_index(drop=True)
    # drop rows that are entirely NaN/None
    data_df = data_df.dropna(how="all")

    rows: list[list[Any]] = []
    for _, row in data_df.iterrows():
        cleaned = [normalize_cell_value(None if (v is None or (isinstance(v, float) and pd.isna(v))) else v) for v in row]
        if any(v is not None and v != "" for v in cleaned):
            rows.append(cleaned)

    first_data_row = header_row_idx + 2  # 1-based for display
    last_useful_row = first_data_row + len(rows) - 1

    if db_path is None:
        db_path = excel_path.with_suffix(".sqlite")
    raw_table = f"RAW_DATA_{revision}"
    aggregated_table = f"AGGREGATED_DATA_{revision}"

    types = infer_sqlite_types(rows, len(column_names))

    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()

        cur.execute(f"DROP TABLE IF EXISTS {sql_quote_identifier(raw_table)}")

        column_defs = ", ".join(
            f"{sql_quote_identifier(col)} {col_type}"
            for col, col_type in zip(column_names, types, strict=True)
        )
        cur.execute(f"CREATE TABLE {sql_quote_identifier(raw_table)} ({column_defs})")

        placeholders = ", ".join(["?"] * len(column_names))
        insert_sql = f"INSERT INTO {sql_quote_identifier(raw_table)} VALUES ({placeholders})"
        cur.executemany(insert_sql, rows)

        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
            (aggregated_table,),
        )
        aggregated_exists = cur.fetchone() is not None
        if not aggregated_exists:
            cur.execute(
                f"CREATE TABLE {sql_quote_identifier(aggregated_table)} AS "
                f"SELECT * FROM {sql_quote_identifier(raw_table)} WHERE 1=0"
            )

        conn.commit()
    finally:
        conn.close()

    print(f"Header row index: {header_row_idx}")
    print(f"Data rows: {first_data_row} to {last_useful_row}")
    print(f"Imported {len(rows)} rows from 'Total Qty' into {raw_table}")
    if aggregated_exists:
        print(f"Table {aggregated_table} already existed and was kept")
    else:
        print(f"Created empty table {aggregated_table}")
    print(f"Database created: {db_path}")
    return db_path


def populate_aggregated_table(db_path: Path, revision: str = "REV0") -> int:
    if not db_path.exists():
        raise FileNotFoundError(f"DB file not found: {db_path}")

    raw_table = f"RAW_DATA_{revision}"
    agg_table = f"AGGREGATED_DATA_{revision}"

    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (raw_table,))
        if cur.fetchone() is None:
            raise ValueError(f"Table {raw_table} does not exist.")

        # Coating column is optional — not all source files include it.
        _raw_cols = {r[1] for r in cur.execute(f"PRAGMA table_info({sql_quote_identifier(raw_table)})")}
        _coating_expr = "Coating" if "Coating" in _raw_cols else "NULL"
        cur.execute(
            f"SELECT Item_Number, Qty, Description, Cut_Length_mm, Total_Weight_kg, Remarks, {_coating_expr} "
            f"FROM {sql_quote_identifier(raw_table)}"
        )
        raw_rows = cur.fetchall()

        individual: list[dict] = []
        groups: dict[tuple, list[dict]] = {}
        bracket_groups: dict[tuple, list[dict]] = {}
        beam_groups: dict[tuple, dict] = {}
        rod_groups: dict[tuple, dict] = {}
        tape_groups: dict[tuple, dict] = {}
        _bracket_x_re = re.compile(r" x \d{3,4}(.*)")

        for item_number, qty, description, cut_length_mm, total_weight_kg, remarks, coating in raw_rows:
            material_type = classify_material_type(description)
            row_dict = {
                "Article_Number": item_number,
                "Qty": qty,
                "Description": description,
                "Cut_Length_mm": cut_length_mm,
                "Total_Weight_kg": total_weight_kg,
                "Remarks": remarks,
                "Coating": coating,
                "Material_Type": material_type,
            }
            # 04 Beam Sections: clean " x NNN[suffix]" from description, keep cut_length as-is
            if material_type == "04 Beam Sections" and description and " x " in description:
                cleaned_desc = _bracket_x_re.sub(
                    lambda m: (" " + m.group(1).strip()) if m.group(1).strip() else "",
                    description,
                ).strip()
                row_dict["Description"] = cleaned_desc
            # 04 Beam Sections (beam section ms / beam section tp): FFD cut optimizer
            _bdesc = (row_dict["Description"] or "").lower()
            if (
                material_type == "04 Beam Sections"
                and cut_length_mm is not None
                and ("beam section ms" in _bdesc or "beam section tp" in _bdesc)
            ):
                key = (item_number, coating)
                if key not in beam_groups:
                    beam_groups[key] = {
                        "description": row_dict["Description"],
                        "pieces": [],
                        "weight": 0.0,
                    }
                _qty = max(int(qty or 0), 0)
                beam_groups[key]["pieces"].extend([int(cut_length_mm)] * _qty)
                if total_weight_kg is not None:
                    beam_groups[key]["weight"] += total_weight_kg
            # 01 Primary Supports — Threaded Rods: FFD cut optimizer
            elif material_type == "01 Primary Supports" and cut_length_mm is not None and "threaded rod" in (description or "").lower():
                _rod_phys = _rod_bar_length_mm(description)
                if _rod_phys:
                    key = (item_number, coating)
                    if key not in rod_groups:
                        rod_groups[key] = {
                            "description": description,
                            "physical_bar_mm": _rod_phys,
                            "pieces": [],
                            "weight": 0.0,
                        }
                    _qty = max(int(qty or 0), 0)
                    rod_groups[key]["pieces"].extend([int(cut_length_mm)] * _qty)
                    if total_weight_kg is not None:
                        rod_groups[key]["weight"] += total_weight_kg
                else:
                    individual.append(row_dict)
            # 05 Installation Material — Glass Fabric Tape: roll aggregation
            elif material_type == "05 Installation Material" and "glass fabric" in (description or "").lower() and cut_length_mm is not None:
                key = (item_number, coating)
                if key not in tape_groups:
                    tape_groups[key] = {
                        "description": description,
                        "coating": coating,
                        "net_cut_mm": 0,
                        "weight": 0.0,
                    }
                _qty = max(int(qty or 0), 0)
                tape_groups[key]["net_cut_mm"] += int(cut_length_mm) * _qty
                if total_weight_kg is not None:
                    tape_groups[key]["weight"] += total_weight_kg
            elif material_type == "02 Brackets Consoles":
                # Step 1: strip " x NNN[suffix]" from description if present
                cleaned_desc = description or ""
                if description and " x " in description:
                    cleaned_desc = _bracket_x_re.sub(
                        lambda m: (" " + m.group(1).strip()) if m.group(1).strip() else "",
                        description,
                    ).strip()
                    row_dict["Description"] = cleaned_desc
                # Step 2: angled bracket → keep cut_length, stay as individual
                if description and "angled" in description.lower():
                    individual.append(row_dict)
                else:
                    # clear cut_length, aggregate on (article_number, cleaned_desc, coating)
                    row_dict["Cut_Length_mm"] = None
                    bracket_groups.setdefault((item_number, coating, cleaned_desc), []).append(row_dict)
            elif cut_length_mm is not None:
                individual.append(row_dict)
            else:
                groups.setdefault((item_number, coating), []).append(row_dict)

        aggregated: list[dict] = []
        for (item_number, coating), group in groups.items():
            total_qty = sum(r["Qty"] or 0 for r in group)
            non_null_w = [r["Total_Weight_kg"] for r in group if r["Total_Weight_kg"] is not None]
            total_weight = round(sum(non_null_w), 2) if non_null_w else None
            description = group[0]["Description"]
            seen: set[str] = set()
            parts: list[str] = []
            for r in group:
                rem = r["Remarks"]
                if rem and rem not in seen:
                    seen.add(rem)
                    parts.append(rem)
            agg_remarks = " | ".join(parts) if parts else None
            mat_type = group[0]["Material_Type"]
            if mat_type == "04 Beam Sections" and "channel" in (description or "").lower():
                _sp = 0.05
            else:
                _sp = _SPARE_PCT.get(mat_type, 0.0)
            spare_n = _spare_for_pct(total_qty, _sp)
            order_qty_v = total_qty + spare_n
            order_weight_v = round(total_weight * order_qty_v / total_qty, 2) if total_weight and total_qty else total_weight
            _pct = int(round(_sp * 100))
            if not _sp:
                spare_calc_rule = "No spare"
            elif _sp == 0.05 and total_qty < 10:
                spare_calc_rule = f"5% spare: qty {total_qty} < 10 → 0 spare"
            else:
                spare_calc_rule = f"{_pct}% spare: ceil({total_qty} \u00d7 {_sp}) = {spare_n}"
            aggregated.append({
                "Article_Number": item_number,
                "Qty": total_qty,
                "Description": description,
                "Cut_Length_mm": None,
                "Total_Weight_kg": total_weight,
                "Remarks": agg_remarks,
                "Coating": coating,
                "Material_Type": mat_type,
                "Spare": spare_n if _sp else None,
                "Order_Qty": order_qty_v,
                "Order_Weight_kg": order_weight_v,
                "Spare_Calculation_Rule": spare_calc_rule,
            })

        for (item_number, coating, cleaned_desc), group in bracket_groups.items():
            total_qty = sum(r["Qty"] or 0 for r in group)
            non_null_w = [r["Total_Weight_kg"] for r in group if r["Total_Weight_kg"] is not None]
            total_weight = round(sum(non_null_w), 2) if non_null_w else None
            _sp2 = _SPARE_PCT.get("02 Brackets Consoles", 0.0)
            spare_n2 = _spare_for_pct(total_qty, _sp2)
            order_qty_v2 = total_qty + spare_n2
            order_weight_v2 = round(total_weight * order_qty_v2 / total_qty, 2) if total_weight and total_qty else total_weight
            _pct2 = int(round(_sp2 * 100))
            if not _sp2:
                spare_calc_rule2 = "No spare"
            elif _sp2 == 0.05 and total_qty < 10:
                spare_calc_rule2 = f"5% spare: qty {total_qty} < 10 → 0 spare"
            else:
                spare_calc_rule2 = f"{_pct2}% spare: ceil({total_qty} \u00d7 {_sp2}) = {spare_n2}"
            aggregated.append({
                "Article_Number": item_number,
                "Qty": total_qty,
                "Description": cleaned_desc,
                "Cut_Length_mm": None,
                "Total_Weight_kg": total_weight,
                "Remarks": None,
                "Coating": coating,
                "Material_Type": "02 Brackets Consoles",
                "Spare": spare_n2 if _sp2 else None,
                "Order_Qty": order_qty_v2,
                "Order_Weight_kg": order_weight_v2,
                "Spare_Calculation_Rule": spare_calc_rule2,
            })

        for (item_number, coating), entry in beam_groups.items():
            pieces = sorted(entry["pieces"], reverse=True)
            if not pieces:
                continue
            num_bars = _beam_ffd_bars(pieces)
            ordered = _beam_ordered_bars(num_bars)
            net_cut_mm = sum(pieces)
            naive_qty = math.ceil(net_cut_mm / _BEAM_PHYSICAL_MM)
            net_cut_m = round(net_cut_mm / 1000, 3)
            total_order_m = ordered * 6
            net_cut_weight = round(entry["weight"], 2) if entry["weight"] else None
            if net_cut_weight is not None and net_cut_m > 0:
                scaled_weight = round((net_cut_weight / net_cut_m) * total_order_m, 2)
            else:
                scaled_weight = None
            aggregated.append({
                "Article_Number": item_number,
                "Qty": naive_qty,
                "Description": entry["description"],
                "Cut_Length_mm": None,
                "Total_Weight_kg": net_cut_weight,
                "Remarks": f"{_BEAM_PHYSICAL_MM // 1000}m bar",
                "Coating": coating,
                "Material_Type": "04 Beam Sections",
                "Spare": ordered - naive_qty,
                "Order_Qty": ordered,
                "Order_Weight_kg": scaled_weight,
                "Net_Cut_Length_m": net_cut_m,
                "Net_Cut_Weight_kg": net_cut_weight,
                "Total_Order_Length_m": total_order_m,
                "Utilisation_pct": round(net_cut_m / total_order_m * 100, 1) if total_order_m else None,
                "Spare_Calculation_Rule": (
                    f"FFD: min {naive_qty} bars (ceil({net_cut_m}m / 6m)); "
                    f"FFD\u2192{num_bars} bars; \u00d71.05\u2192{ordered} ordered; spare={ordered - naive_qty}"
                ),
            })

        for (item_number, coating), entry in rod_groups.items():
            phys_mm = entry["physical_bar_mm"]
            eff_mm = phys_mm  # no end-loss scrap
            pieces = sorted(entry["pieces"], reverse=True)
            if not pieces:
                continue
            num_bars = _ffd_bars(pieces, eff_mm)
            ordered = _calc_ordered_bars(num_bars, phys_mm)
            net_cut_mm = sum(pieces)
            naive_qty = math.ceil(net_cut_mm / phys_mm)
            net_cut_m = round(net_cut_mm / 1000, 3)
            total_order_m = round(ordered * phys_mm / 1000, 3)
            net_cut_weight = round(entry["weight"], 2) if entry["weight"] else None
            if net_cut_weight is not None and net_cut_m > 0:
                scaled_weight = round((net_cut_weight / net_cut_m) * total_order_m, 2)
            else:
                scaled_weight = None
            aggregated.append({
                "Article_Number": item_number,
                "Qty": naive_qty,
                "Description": entry["description"],
                "Cut_Length_mm": None,
                "Total_Weight_kg": net_cut_weight,
                "Remarks": f"{phys_mm / 1000:g}m rod",
                "Coating": coating,
                "Material_Type": "01 Primary Supports",
                "Spare": ordered - naive_qty,
                "Order_Qty": ordered,
                "Order_Weight_kg": scaled_weight,
                "Net_Cut_Length_m": net_cut_m,
                "Net_Cut_Weight_kg": net_cut_weight,
                "Total_Order_Length_m": total_order_m,
                "Utilisation_pct": round(net_cut_m / total_order_m * 100, 1) if total_order_m else None,
                "Spare_Calculation_Rule": (
                    f"FFD: min {naive_qty} bars (ceil({net_cut_m}m / {phys_mm / 1000:.1f}m)); "
                    f"FFD\u2192{num_bars} bars; \u00d71.05\u2192{ordered} ordered; spare={ordered - naive_qty}"
                ),
            })

        _TAPE_ROLL_M = 10
        _TAPE_SPARE_FACTOR = 1.30
        _TAPE_DENSITY_KG_M3 = 2600.0  # 2.6 g/cm³ — glass fabric tape material density

        def _tape_weight_per_m(description: str | None) -> float | None:
            """Parse 'GSK WW x TT' from description and return kg/m using material density."""
            import re as _re
            if not description:
                return None
            m = _re.search(r"(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)", description)
            if not m:
                return None
            width_mm = float(m.group(1))
            thickness_mm = float(m.group(2))
            # volume per metre of tape (m³): width × thickness × 1 m length
            vol_per_m = (width_mm / 1000) * (thickness_mm / 1000) * 1.0
            return vol_per_m * _TAPE_DENSITY_KG_M3

        for (item_number, coating), entry in tape_groups.items():
            net_cut_mm = entry["net_cut_mm"]
            if not net_cut_mm:
                continue
            net_cut_m = round(net_cut_mm / 1000, 3)
            needed_m = net_cut_m * _TAPE_SPARE_FACTOR
            ordered_rolls = math.ceil(needed_m / _TAPE_ROLL_M)
            naive_rolls = math.ceil(net_cut_m / _TAPE_ROLL_M)
            total_order_m = round(ordered_rolls * _TAPE_ROLL_M, 3)
            wpm = _tape_weight_per_m(entry["description"])
            if wpm is not None:
                net_cut_weight = round(wpm * net_cut_m, 3)
                naive_rolls_weight = round(wpm * naive_rolls * _TAPE_ROLL_M, 3)
                order_weight = round(wpm * total_order_m, 3)
            else:
                net_cut_weight = naive_rolls_weight = order_weight = None
            aggregated.append({
                "Article_Number": item_number,
                "Qty": naive_rolls,
                "Description": entry["description"],
                "Cut_Length_mm": None,
                "Total_Weight_kg": naive_rolls_weight,
                "Remarks": f"{_TAPE_ROLL_M}m roll",
                "Coating": entry["coating"],
                "Material_Type": "05 Installation Material",
                "Spare": ordered_rolls - naive_rolls,
                "Order_Qty": ordered_rolls,
                "Order_Weight_kg": order_weight,
                "Net_Cut_Length_m": net_cut_m,
                "Net_Cut_Weight_kg": net_cut_weight,
                "Total_Order_Length_m": total_order_m,
                "Utilisation_pct": round(net_cut_m / total_order_m * 100, 1) if total_order_m else None,
                "Spare_Calculation_Rule": (
                    f"50% factor: ceil({net_cut_m}m \u00d7 1.50 / {_TAPE_ROLL_M}m) = {ordered_rolls} rolls; "
                    f"naive {naive_rolls}; spare={ordered_rolls - naive_rolls}"
                ),
            })

        _C5_MAP = {
            "01 Primary Supports":    "06 Primary Supports C5",
            "02 Brackets Consoles":   "07 Brackets Consoles C5",
            "03 Bolts Screws Nuts":   "08 Bolts Screws Nuts C5",
            "04 Beam Sections":       "09 Beam Sections C5",
            "05 Installation Material": "10 Installation Material C5",
        }

        output_rows: list[tuple] = []
        for row_dict in aggregated + individual:
            material_type = row_dict.get("Material_Type") or classify_material_type(row_dict["Description"])
            coating = row_dict["Coating"]
            if coating and "C5" in coating:
                material_type = _C5_MAP.get(material_type, material_type)
            spare = row_dict.get("Spare")
            order_qty = row_dict.get("Order_Qty")
            order_weight = row_dict.get("Order_Weight_kg")
            spare_calc_rule = row_dict.get("Spare_Calculation_Rule")
            if order_qty is None:  # individual item — compute spare from _SPARE_PCT
                if material_type in ("04 Beam Sections", "09 Beam Sections C5") and "channel" in (row_dict["Description"] or "").lower():
                    _sp = 0.05
                else:
                    _sp = _SPARE_PCT.get(material_type, 0.0)
                _base_qty = row_dict["Qty"] or 0
                _spare_n = _spare_for_pct(_base_qty, _sp)
                spare = _spare_n if _sp else None
                order_qty = _base_qty + _spare_n
                _w = row_dict["Total_Weight_kg"]
                order_weight = round(_w * order_qty / _base_qty, 2) if _w and _base_qty else _w
                _pct = int(round(_sp * 100))
                if not _sp:
                    spare_calc_rule = "No spare"
                elif _sp == 0.05 and _base_qty < 10:
                    spare_calc_rule = f"5% spare: qty {_base_qty} < 10 → 0 spare"
                else:
                    spare_calc_rule = (
                        f"{_pct}% spare: ceil({_base_qty} \u00d7 {_sp}) = {_spare_n}"
                    )
            # Fold Cut_Length_mm into Remarks; the column is dropped from AGG schema.
            _cut = row_dict["Cut_Length_mm"]
            _base_remark = row_dict["Remarks"]
            if _cut is not None:
                _cut_label = f"cut: {int(_cut)} mm"
                _remarks = f"{_base_remark} / {_cut_label}" if _base_remark else _cut_label
            else:
                _remarks = _base_remark
            output_rows.append((
                material_type,               # 0
                row_dict["Article_Number"],  # 1
                row_dict["Description"],     # 2
                row_dict["Qty"],             # 3
                row_dict["Total_Weight_kg"], # 4
                _remarks,                    # 5  (was index 6)
                coating,                     # 6  (was index 7)
                spare,                       # 7  (was index 8)
                order_qty,                   # 8  (was index 9)
                order_weight,                # 9  (was index 10)
                row_dict.get("Net_Cut_Length_m"),    # 10
                row_dict.get("Net_Cut_Weight_kg"),   # 11
                row_dict.get("Total_Order_Length_m"),# 12
                row_dict.get("Utilisation_pct"),     # 13
                spare_calc_rule,                     # 14
            ))

        # Final aggregation: group by (Article_Number, Coating, Remarks), sum Qty + weights
        _final: dict[tuple, list] = {}
        for row in output_rows:
            # key: article_number(1), coating(6), remarks(5)
            key = (row[1], row[6], row[5])
            if key not in _final:
                _final[key] = list(row)
            else:
                existing = _final[key]
                existing[3] = (existing[3] or 0) + (row[3] or 0)   # Qty
                if row[4] is not None:
                    existing[4] = round((existing[4] or 0) + row[4], 2)   # Total_Weight_kg
                if row[7] is not None:
                    existing[7] = (existing[7] or 0) + (row[7] or 0)      # Spare
                existing[8] = (existing[8] or 0) + (row[8] or 0)          # Order_Qty
                if row[9] is not None:
                    existing[9] = round((existing[9] or 0) + row[9], 2)   # Order_Weight_kg
                if row[11] is not None:
                    existing[11] = round((existing[11] or 0) + row[11], 2) # Net_Cut_Weight_kg
        output_rows = [tuple(v) for v in _final.values()]

        output_rows.sort(key=lambda r: (r[0] or "", (r[2] or "").lower()))

        cur.execute(f"DROP TABLE IF EXISTS {sql_quote_identifier(agg_table)}")
        cur.execute(
            f"CREATE TABLE {sql_quote_identifier(agg_table)} ("
            f'"Material_Type" TEXT, '
            f'"Article_Number" INTEGER, '
            f'"Description" TEXT, '
            f'"Qty" INTEGER, '
            f'"Total_Weight_kg" REAL, '
            f'"Remarks" TEXT, '
            f'"Coating" TEXT, '
            f'"Spare" INTEGER, '
            f'"Order_Qty" INTEGER, '
            f'"Order_Weight_kg" REAL, '
            f'"Net_Cut_Length_m" REAL, '
            f'"Net_Cut_Weight_kg" REAL, '
            f'"Total_Order_Length_m" REAL, '
            f'"Utilisation_pct" REAL, '
            f'"Spare_Calculation_Rule" TEXT'
            f")"
        )
        cur.executemany(
            f"INSERT INTO {sql_quote_identifier(agg_table)} VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            output_rows,
        )
        conn.commit()

        print(
            f"Populated {agg_table}: {len(output_rows)} rows "
            f"({len(aggregated)} aggregated [{len(bracket_groups)} bracket groups], {len(individual)} individual with cut length)"
        )
        return len(output_rows)
    finally:
        conn.close()


def _run_validation_checks(
    conn: sqlite3.Connection,
    raw_table: str,
    agg_table: str,
    has_raw: bool,
) -> list[dict]:
    """Run all validation checks. Returns list of result dicts:
    {id, category, name, status, count, detail, rows}.
    """
    rt = sql_quote_identifier(raw_table)
    at = sql_quote_identifier(agg_table)
    cur = conn.cursor()
    results: list[dict] = []

    # Coating column is optional in the raw table — substitute NULL when absent.
    _raw_cols = {r[1] for r in cur.execute(f"PRAGMA table_info({rt})")}
    _r_coating = "r.Coating" if "Coating" in _raw_cols else "NULL"

    def _chk(check_id: str, cat: str, name: str, rows: list, level: str = "FAIL") -> None:
        results.append({
            "id": check_id, "category": cat, "name": name,
            "status": level if rows else "PASS",
            "count": len(rows), "detail": "", "rows": rows,
        })

    def _stat(check_id: str, cat: str, name: str, status: str, detail: str) -> None:
        results.append({
            "id": check_id, "category": cat, "name": name,
            "status": status, "count": None, "detail": detail, "rows": [],
        })

    # ── 1. RAW DATA INTEGRITY ─────────────────────────────────────────────
    if has_raw:
        _chk("V01", "1 Raw integrity", "No NULL Article Numbers",
             cur.execute(f"SELECT Item_Number, Description FROM {rt} WHERE Item_Number IS NULL").fetchall())

        _chk("V02", "1 Raw integrity", "No NULL/blank Descriptions",
             cur.execute(
                 f"SELECT Item_Number, Description FROM {rt} "
                 f"WHERE Description IS NULL OR trim(Description) = ''"
             ).fetchall())

        _chk("V03", "1 Raw integrity", "All raw Qty values > 0",
             cur.execute(
                 f"SELECT Item_Number, Description, Qty FROM {rt} "
                 f"WHERE Qty IS NULL OR Qty <= 0"
             ).fetchall())

        _chk("V04", "1 Raw integrity", "No NULL Total_Weight_kg in raw",
             cur.execute(
                 f"SELECT Item_Number, Description, Qty, Total_Weight_kg FROM {rt} "
                 f"WHERE Total_Weight_kg IS NULL"
             ).fetchall(), level="WARN")

        # For cut-to-length items Weight_kg is linear density (kg/m), not per-piece weight,
        # so this check is only valid for standard (no cut length) rows.
        # Level=WARN because source data rounds Weight_kg to 2 dp (e.g. 0.128 → 0.13).
        _chk("V05", "1 Raw integrity",
             "Standard items: Unit_Weight_kg x Qty = Total_Weight_kg (tolerance +-0.02 kg)",
             cur.execute(
                 f"SELECT Item_Number, Description, Qty, Weight_kg, Total_Weight_kg FROM {rt} "
                 f"WHERE Cut_Length_mm IS NULL AND Weight_kg IS NOT NULL AND Qty IS NOT NULL "
                 f"  AND ABS(Weight_kg * Qty - Total_Weight_kg) > 0.02"
             ).fetchall(), level="WARN")

        _chk("V06", "1 Raw integrity", "Cut_Length_mm > 0 when present",
             cur.execute(
                 f"SELECT Item_Number, Description, Cut_Length_mm FROM {rt} "
                 f"WHERE Cut_Length_mm IS NOT NULL AND Cut_Length_mm <= 0"
             ).fetchall())

    # ── 2. CLASSIFICATION COMPLETENESS ───────────────────────────────────
    if has_raw:
        raw_arts = {r[0] for r in cur.execute(f"SELECT DISTINCT Item_Number FROM {rt}")}
        agg_arts = {r[0] for r in cur.execute(f"SELECT DISTINCT Article_Number FROM {at}")}

        _chk("V07", "2 Classification", "All raw articles present in aggregated output",
             [(art,) for art in sorted(raw_arts - agg_arts)])

        _chk("V08", "2 Classification", "No unexpected articles in aggregated output",
             [(art,) for art in sorted(agg_arts - raw_arts)])

    # ── 3. AGGREGATION FIDELITY ───────────────────────────────────────────
    if has_raw:
        # Standard no-cut items: AGG Qty == SUM(raw Qty) per Article+Coating
        _chk("V09", "3 Aggregation fidelity",
             "No-cut AGG Qty = SUM(raw Qty) per Article+Coating",
             cur.execute(f"""
                 SELECT a.Article_Number, a.Coating,
                        a.Qty AS agg_qty, SUM(r.Qty) AS raw_qty,
                        a.Qty - SUM(r.Qty) AS diff
                 FROM {at} a
                 JOIN {rt} r ON r.Item_Number = a.Article_Number
                     AND {_r_coating} IS a.Coating
                 WHERE a.Net_Cut_Length_m IS NULL
                   AND (a.Remarks IS NULL OR a.Remarks NOT LIKE '%cut: % mm%')
                 GROUP BY a.Article_Number, a.Coating
                 HAVING ABS(a.Qty - SUM(r.Qty)) > 0.01
             """).fetchall())

        # FFD/tape: Net_Cut_Length_m * 1000 == SUM(raw Qty * Cut_Length_mm)
        _chk("V10", "3 Aggregation fidelity",
             "FFD/tape Net_Cut_Length_m matches SUM(raw Qty x Cut_Length_mm)",
             cur.execute(f"""
                 SELECT a.Article_Number, a.Coating,
                        ROUND(a.Net_Cut_Length_m * 1000) AS agg_net_mm,
                        ROUND(SUM(r.Qty * r.Cut_Length_mm)) AS raw_net_mm,
                        ROUND(a.Net_Cut_Length_m * 1000
                              - SUM(r.Qty * r.Cut_Length_mm)) AS diff_mm
                 FROM {at} a
                 JOIN {rt} r ON r.Item_Number = a.Article_Number
                     AND {_r_coating} IS a.Coating
                     AND r.Cut_Length_mm IS NOT NULL
                 WHERE a.Net_Cut_Length_m IS NOT NULL
                 GROUP BY a.Article_Number, a.Coating
                 HAVING ABS(a.Net_Cut_Length_m * 1000
                            - SUM(r.Qty * r.Cut_Length_mm)) > 1
             """).fetchall())

        # V11: weight conservation for standard (non-FFD/tape) items only.
        # Tape weights are NULL (source uses dummy 0.1 kg/piece), FFD weights
        # are covered by V17. Exclude FFD/tape rows from both sides.
        raw_ffd_articles = cur.execute(
            f"SELECT DISTINCT Article_Number FROM {at} WHERE Net_Cut_Length_m IS NOT NULL"
        ).fetchall()
        ffd_placeholders = ",".join("?" * len(raw_ffd_articles)) if raw_ffd_articles else "NULL"
        ffd_ids = [r[0] for r in raw_ffd_articles]
        raw_w = cur.execute(
            f"SELECT ROUND(SUM(Total_Weight_kg), 2) FROM {rt} "
            f"WHERE Item_Number NOT IN ({ffd_placeholders})",
            ffd_ids,
        ).fetchone()[0] or 0.0
        agg_w = cur.execute(
            f"SELECT ROUND(SUM(Total_Weight_kg), 2) FROM {at} "
            f"WHERE Net_Cut_Length_m IS NULL"
        ).fetchone()[0] or 0.0
        diff_w = round(abs(raw_w - agg_w), 2)
        _stat("V11", "3 Aggregation fidelity",
              "Grand total net weight: raw total = AGG total (threshold < 1 kg)",
              "PASS" if diff_w < 1.0 else "FAIL",
              f"Raw: {raw_w} kg  |  AGG net: {agg_w} kg  |  Diff: {diff_w} kg")

    # ── 4. SPARE / ORDER LOGIC ────────────────────────────────────────────
    _chk("V12", "4 Spare/Order logic", "Order_Qty >= Qty for all rows",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description, Qty, Order_Qty "
             f"FROM {at} WHERE Order_Qty < Qty"
         ).fetchall())

    _chk("V13", "4 Spare/Order logic",
         "Spare = Order_Qty - Qty (when Spare is not NULL)",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Qty, Spare, Order_Qty "
             f"FROM {at} WHERE Spare IS NOT NULL "
             f"  AND ABS(Spare - (Order_Qty - Qty)) > 0.01"
         ).fetchall())

    _chk("V14", "4 Spare/Order logic",
         "Material types with spare rule have Spare set",
         cur.execute(f"""
             SELECT Material_Type, Article_Number, Description, Qty, Spare
             FROM {at}
             WHERE Material_Type IN (
                 '01 Primary Supports',    '02 Brackets Consoles',
                 '03 Bolts Screws Nuts',   '05 Installation Material',
                 '06 Primary Supports C5', '07 Brackets Consoles C5',
                 '08 Bolts Screws Nuts C5','10 Installation Material C5'
             ) AND Spare IS NULL
         """).fetchall(), level="WARN")

    _chk("V15", "4 Spare/Order logic", "FFD bar utilisation <= 100%",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description, Utilisation_pct "
             f"FROM {at} WHERE Utilisation_pct IS NOT NULL AND Utilisation_pct > 100"
         ).fetchall())

    _chk("V16", "4 Spare/Order logic",
         "Total_Order_Length_m >= Net_Cut_Length_m for FFD/tape items",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description, "
             f"       Net_Cut_Length_m, Total_Order_Length_m "
             f"FROM {at} WHERE Net_Cut_Length_m IS NOT NULL "
             f"  AND Total_Order_Length_m < Net_Cut_Length_m - 0.01"
         ).fetchall())

    _chk("V17", "4 Spare/Order logic",
         "FFD/tape Order_Weight_kg >= Net_Cut_Weight_kg",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description, "
             f"       Net_Cut_Weight_kg, Order_Weight_kg "
             f"FROM {at} WHERE Net_Cut_Length_m IS NOT NULL "
             f"  AND Net_Cut_Weight_kg IS NOT NULL AND Order_Weight_kg IS NOT NULL "
             f"  AND Order_Weight_kg < Net_Cut_Weight_kg - 0.01"
         ).fetchall())

    # ── 5. WEIGHT SANITY ─────────────────────────────────────────────────
    _chk("V18", "5 Weight sanity", "No negative Total_Weight_kg in AGG",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description, Total_Weight_kg "
             f"FROM {at} WHERE Total_Weight_kg IS NOT NULL AND Total_Weight_kg < 0"
         ).fetchall())

    _chk("V19", "5 Weight sanity", "No negative Order_Weight_kg in AGG",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description, Order_Weight_kg "
             f"FROM {at} WHERE Order_Weight_kg IS NOT NULL AND Order_Weight_kg < 0"
         ).fetchall())

    _chk("V20", "5 Weight sanity",
         "Pct-spare unit weight preserved: Order_Weight/Order_Qty ~ Total_Weight/Qty (+-0.05)",
         cur.execute(f"""
             SELECT Material_Type, Article_Number, Description,
                    Qty, Total_Weight_kg, Order_Qty, Order_Weight_kg,
                    ROUND(Total_Weight_kg * 1.0 / Qty, 4)       AS net_unit_w,
                    ROUND(Order_Weight_kg * 1.0 / Order_Qty, 4) AS order_unit_w
             FROM {at}
             WHERE Net_Cut_Length_m IS NULL AND Qty > 0 AND Order_Qty > 0
               AND Total_Weight_kg IS NOT NULL AND Order_Weight_kg IS NOT NULL
               AND ABS(Total_Weight_kg * 1.0 / Qty
                       - Order_Weight_kg * 1.0 / Order_Qty) > 0.05
         """).fetchall())

    _chk("V21", "5 Weight sanity", "No NULL Total_Weight_kg in aggregated output",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description "
             f"FROM {at} WHERE Total_Weight_kg IS NULL"
         ).fetchall(), level="WARN")

    # ── 6. AGG DATA QUALITY ───────────────────────────────────────────────
    _chk("V22", "6 AGG quality",
         "No duplicate (Article, Coating, Remarks) keys",
         cur.execute(f"""
             SELECT Article_Number, Coating, Remarks,
                    COUNT(*) AS cnt
             FROM {at}
             GROUP BY Article_Number, Coating, Remarks
             HAVING COUNT(*) > 1
         """).fetchall())

    _chk("V23", "6 AGG quality", "No zero-Qty rows in aggregated output",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description, Qty "
             f"FROM {at} WHERE Qty = 0"
         ).fetchall(), level="WARN")

    _chk("V24", "6 AGG quality", "All rows have Spare_Calculation_Rule",
         cur.execute(
             f"SELECT Material_Type, Article_Number, Description "
             f"FROM {at} WHERE Spare_Calculation_Rule IS NULL "
             f"   OR Spare_Calculation_Rule = ''"
         ).fetchall())

    _chk("V25", "6 AGG quality",
         "04/09 Beam Sections non-FFD non-channel items have spare assigned",
         cur.execute(f"""
             SELECT Material_Type, Article_Number, Description, Qty,
                    COALESCE(Spare_Calculation_Rule, 'NULL') AS calc_rule
             FROM {at}
             WHERE Material_Type IN ('04 Beam Sections', '09 Beam Sections C5')
               AND Net_Cut_Length_m IS NULL
               AND lower(Description) NOT LIKE '%channel%'
               AND (Spare IS NULL OR Spare_Calculation_Rule = 'No spare')
         """).fetchall(), level="WARN")

    # ── 7. SUMMARY ────────────────────────────────────────────────────────
    raw_count = cur.execute(f"SELECT COUNT(*) FROM {rt}").fetchone()[0] if has_raw else "N/A"
    agg_count = cur.execute(f"SELECT COUNT(*) FROM {at}").fetchone()[0]
    _stat("V26", "7 Summary", "Row counts", "INFO",
          f"Raw rows: {raw_count}  |  AGG rows: {agg_count}")

    net_qty   = cur.execute(f"SELECT SUM(Qty)       FROM {at}").fetchone()[0]
    order_qty = cur.execute(f"SELECT SUM(Order_Qty) FROM {at}").fetchone()[0]
    spare_qty = cur.execute(f"SELECT SUM(COALESCE(Spare, 0)) FROM {at}").fetchone()[0]
    _stat("V27", "7 Summary", "Grand totals: Qty / Spare / Order_Qty", "INFO",
          f"Net Qty: {net_qty}  |  Spare: {spare_qty}  |  Order Qty: {order_qty}")

    net_w   = cur.execute(f"SELECT ROUND(SUM(Total_Weight_kg),   1) FROM {at}").fetchone()[0]
    order_w = cur.execute(f"SELECT ROUND(SUM(Order_Weight_kg), 1) FROM {at}").fetchone()[0]
    _stat("V28", "7 Summary", "Grand totals: Net Weight / Order Weight", "INFO",
          f"Net weight: {net_w} kg  |  Order weight: {order_w} kg")

    mt_rows = cur.execute(f"""
        SELECT Material_Type, COUNT(*) AS rows,
               SUM(Qty) AS net_qty, SUM(COALESCE(Spare, 0)) AS spare,
               SUM(Order_Qty) AS order_qty,
               ROUND(SUM(Total_Weight_kg), 1) AS net_w_kg,
               ROUND(SUM(Order_Weight_kg), 1) AS order_w_kg
        FROM {at}
        GROUP BY Material_Type ORDER BY Material_Type
    """).fetchall()
    results.append({
        "id": "V29", "category": "7 Summary", "name": "Per-material-type breakdown",
        "status": "INFO", "count": None, "detail": "", "rows": mt_rows,
        "_table_headers": [
            "Material_Type", "AGG Rows", "Net Qty", "Spare",
            "Order Qty", "Net Weight (kg)", "Order Weight (kg)",
        ],
    })

    return results


def _write_validation_sheet(wb: Workbook, checks: list[dict]) -> None:
    """Append a VALIDATION sheet to wb."""
    ws = wb.create_sheet("VALIDATION")

    _FILLS = {
        "PASS": PatternFill("solid", fgColor="FFC6EFCE"),
        "FAIL": PatternFill("solid", fgColor="FFFFC7CE"),
        "WARN": PatternFill("solid", fgColor="FFFFEB9C"),
        "INFO": PatternFill("solid", fgColor="FFDDEBF7"),
    }
    _BOLD  = Font(bold=True)
    _HEAD  = Font(bold=True, size=11)
    _TITLE = Font(bold=True, size=13)

    def _section(title: str) -> None:
        ws.append([])
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = _HEAD

    # ── TITLE ─────────────────────────────────────────────────────────────
    fails  = sum(1 for c in checks if c["status"] == "FAIL")
    warns  = sum(1 for c in checks if c["status"] == "WARN")
    passes = sum(1 for c in checks if c["status"] == "PASS")
    overall = "FAIL" if fails else ("WARN" if warns else "PASS")
    ws.append([
        f"VALIDATION REPORT  {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Overall: {overall}   {fails} FAIL  {warns} WARN  {passes} PASS",
    ])
    ws.cell(row=1, column=1).font = _TITLE
    ws.cell(row=1, column=2).font = Font(
        bold=True, size=12,
        color="FFCC0000" if overall == "FAIL" else ("FF7A6000" if overall == "WARN" else "FF375623"),
    )

    # ── CHECK SUMMARY TABLE ───────────────────────────────────────────────
    _section("CHECK SUMMARY")
    ws.append(["ID", "Category", "Check", "Status", "Count", "Detail"])
    for cell in ws[ws.max_row]:
        cell.font = _BOLD

    summary_checks = [c for c in checks if "_table_headers" not in c]
    for chk in summary_checks:
        ws.append([
            chk["id"], chk["category"], chk["name"], chk["status"],
            chk["count"] if chk["count"] is not None else "",
            chk.get("detail", ""),
        ])
        fill = _FILLS.get(chk["status"])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # ── MATERIAL TYPE BREAKDOWN ───────────────────────────────────────────
    for chk in checks:
        if "_table_headers" in chk:
            _section("MATERIAL TYPE BREAKDOWN")
            ws.append(chk["_table_headers"])
            for cell in ws[ws.max_row]:
                cell.font = _BOLD
            for row in chk["rows"]:
                ws.append(list(row))

    # ── DETAIL — FAILING / WARNING CHECKS ────────────────────────────────
    _section("DETAIL — FAILING AND WARNING CHECKS")
    any_detail = False
    for chk in summary_checks:
        if chk["status"] in ("FAIL", "WARN") and chk.get("rows"):
            any_detail = True
            tag = "[FAIL]" if chk["status"] == "FAIL" else "[WARN]"
            ws.append([f"{tag} {chk['id']} -- {chk['name']}  ({chk['count']} row(s))"])
            hcell = ws.cell(row=ws.max_row, column=1)
            hcell.font = Font(
                bold=True,
                color="FFCC0000" if chk["status"] == "FAIL" else "FF7A6000",
            )
            hcell.fill = _FILLS[chk["status"]]
            for row_data in chk["rows"][:30]:
                ws.append(list(row_data))
            if len(chk["rows"]) > 30:
                ws.append([f"  ... and {len(chk['rows']) - 30} additional row(s) not shown"])
            ws.append([])

    if not any_detail:
        ws.append(["All checks passed or warning checks have no offending rows."])

    # ── COLUMN WIDTHS ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        max_len = max(
            len(str(c.value)) if c.value is not None else 0 for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 80)


def _write_agg_sheet(ws, columns: list[str], rows: list[tuple]) -> None:
    """Write aggregated sheet grouped by Material_Type with separator, per-group
    subtotal rows, and a grand-total row.  All totals use SUBTOTAL(9,...) which
    automatically ignores other SUBTOTAL rows, so autofilter keeps everything
    consistent without double-counting."""

    # ── Column display order ──────────────────────────────────────────────
    # Order_Qty and Order_Weight_kg shift to the far right so that the four
    # order-related columns (Order_Qty, Order_Weight_kg, Unit_Price, Total_Price)
    # sit together as a visually distinct block.
    _DB_COLS   = list(columns)                          # 15 DB columns (schema order)
    _SHIFT     = ("Order_Qty", "Order_Weight_kg")       # pulled to far right
    _base_cols = [c for c in _DB_COLS if c not in _SHIFT]
    _ORD_XTRA  = ("Unit_Price", "Total_Price")          # appended, no DB source

    # Internal-name sequence: drives col_idx and formula column letters
    _int_seq    = list(_base_cols) + list(_SHIFT) + list(_ORD_XTRA)

    # Mapping: display-position i → original DB tuple index (-1 = no DB col)
    _db_pos     = {c: i for i, c in enumerate(_DB_COLS)}
    _disp_to_db = [_db_pos.get(c, -1) for c in _int_seq]

    def _reorder(row_data) -> list:
        return [row_data[i] if i >= 0 else None for i in _disp_to_db]

    # Quantity columns get a _pcs suffix in the displayed header row
    _PCS_COLS = {"Qty", "Spare", "Order_Qty"}
    headers   = [f"{c}_pcs" if c in _PCS_COLS else c for c in _int_seq]

    # col_idx: internal_name → 1-based sheet column (used for formula building)
    col_idx = {name: i + 1 for i, name in enumerate(_int_seq)}
    n_cols  = len(headers)

    ws.append(headers)

    # ── Shared style objects ──────────────────────────────────────────────
    _thin   = Side(style="thin")
    _medium = Side(style="medium")

    # Order column block boundaries
    _ord_ci_start = col_idx["Order_Qty"]    # first column of the order block
    _ord_ci_end   = col_idx["Total_Price"]  # last column of the order block

    # ── Header row styling ────────────────────────────────────────────────
    hdr_fill_base = PatternFill("solid", fgColor="FF203864")  # dark navy
    hdr_fill_ord  = PatternFill("solid", fgColor="FF7B3F00")  # dark amber (order block)
    hdr_font      = Font(bold=True, color="FFFFFFFF")
    for ci, cell in enumerate(ws[1], start=1):
        cell.fill = hdr_fill_ord if _ord_ci_start <= ci <= _ord_ci_end else hdr_fill_base
        cell.font = hdr_font

    def cl(name: str) -> str:
        return get_column_letter(col_idx[name])

    NUMERIC_COLS = ["Qty", "Total_Weight_kg", "Spare",
                    "Order_Qty", "Order_Weight_kg", "Total_Price"]

    # ── Row styles ────────────────────────────────────────────────────────
    sep_fill  = PatternFill("solid", fgColor="FF2E4E7E")  # dark steel blue
    sep_font  = Font(bold=True, color="FFFFFFFF", size=10)
    sub_fill  = PatternFill("solid", fgColor="FFDCE6F1")  # pale blue
    sub_font  = Font(bold=True, italic=True)
    tot_fill  = PatternFill("solid", fgColor="FFFFF2CC")  # yellow
    tot_font  = Font(bold=True)
    unit_fill = PatternFill("solid", fgColor="FFBDD7EE")  # light blue — supplier input

    # ── Group rows by Material_Type (use original DB index for grouping) ──
    mt_db_idx = _db_pos["Material_Type"]
    groups: dict[str, list] = {}
    for row_data in rows:
        mt = row_data[mt_db_idx] or ""
        groups.setdefault(mt, []).append(row_data)
    groups = dict(sorted(groups.items()))

    current_row = 2

    for group_name, group_rows in groups.items():
        # ── Separator / group header row ──────────────────────────────────
        sep_r = current_row
        ws.cell(sep_r, 1).value = group_name
        for ci in range(1, n_cols + 1):
            cell = ws.cell(sep_r, ci)
            cell.fill = sep_fill
            cell.font = sep_font
        ws.cell(sep_r, 1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[sep_r].height = 18
        current_row += 1

        # ── Data rows ─────────────────────────────────────────────────────
        first_data = current_row
        # DB index for the two shifted columns (needed to pull static values)
        _oqty_db = _db_pos["Order_Qty"]
        _owt_db  = _db_pos["Order_Weight_kg"]
        for row_data in group_rows:
            r = current_row
            ws.append(_reorder(row_data))
            # Order_Qty and Order_Weight_kg are pre-computed by the Python
            # aggregation pipeline and stored in the DB.  Write them as static
            # numbers so cross-sheet SUMPRODUCT in PRICE_SHEET can read them
            # immediately without depending on Excel's formula-evaluation order.
            # (openpyxl writes formula cells with no cached value; if PRICE_SHEET
            # is evaluated before AGG, formula cells appear as 0.)
            oqty_val = row_data[_oqty_db]
            owt_val  = row_data[_owt_db]
            ws.cell(r, col_idx["Order_Qty"]).value        = oqty_val
            ws.cell(r, col_idx["Order_Weight_kg"]).value  = owt_val
            # Total_Price stays as a formula — it depends on Unit_Price which
            # the supplier fills in later, so there is no DB value to write.
            ws.cell(r, col_idx["Total_Price"]).value = (
                f"=IF(ISNUMBER({cl('Unit_Price')}{r})"
                f",{cl('Order_Qty')}{r}*{cl('Unit_Price')}{r},\"\")"
            )
            # Light blue fill signals that the supplier should fill in Unit_Price
            ws.cell(r, col_idx["Unit_Price"]).fill = unit_fill
            current_row += 1
        last_data = current_row - 1

        # ── Group subtotal row ────────────────────────────────────────────
        # Keep col A (Material_Type) = group_name so the subtotal row is part
        # of the same filter value as the data rows — no extra entry in dropdown.
        sub_r = current_row
        ws.cell(sub_r, col_idx["Material_Type"]).value = group_name
        ws.cell(sub_r, col_idx["Description"]).value = f"↳ Subtotal — {group_name}"
        for col_name in NUMERIC_COLS:
            letter = cl(col_name)
            c = ws.cell(sub_r, col_idx[col_name])
            c.value = f"=SUBTOTAL(9,{letter}{first_data}:{letter}{last_data})"
            c.number_format = "0.00"
        for ci in range(1, n_cols + 1):
            cell = ws.cell(sub_r, ci)
            cell.fill = sub_fill
            cell.font = sub_font
        ws.cell(sub_r, 1).alignment = Alignment(
            horizontal="left", vertical="center", indent=2
        )
        ws.row_dimensions[sub_r].height = 16
        current_row += 1

    # ── Number formats (column-based so formula cells are covered too) ────
    FMT_INT  = "0"
    FMT_2DP  = "0.00"
    FMT_1DP  = "0.0"
    _int_fmts = {col_idx[n] for n in ["Qty", "Spare", "Order_Qty"] if n in col_idx}
    _2dp_fmts = {col_idx[n] for n in [
        "Total_Weight_kg", "Order_Weight_kg", "Net_Cut_Weight_kg",
        "Net_Cut_Length_m", "Total_Order_Length_m", "Unit_Price", "Total_Price",
    ] if n in col_idx}
    _1dp_fmts = {col_idx[n] for n in ["Utilisation_pct"] if n in col_idx}

    for row in ws.iter_rows(min_row=2, max_row=current_row - 1):
        for cell in row:
            ci = cell.column
            if ci in _int_fmts:
                cell.number_format = FMT_INT
            elif ci in _2dp_fmts:
                cell.number_format = FMT_2DP
            elif ci in _1dp_fmts:
                cell.number_format = FMT_1DP

    # ── Grand total row ───────────────────────────────────────────────────
    # SUBTOTAL(9,...) over the entire body automatically skips nested SUBTOTALs.
    grand_row = current_row
    # Leave col A blank so this row does not appear in the Material_Type filter.
    ws.cell(grand_row, col_idx["Description"]).value = "GRAND TOTAL"
    for col_name in NUMERIC_COLS:
        letter = cl(col_name)
        c = ws.cell(grand_row, col_idx[col_name])
        c.value = f"=SUBTOTAL(9,{letter}2:{letter}{grand_row - 1})"
        c.number_format = "0.00"
    for ci in range(1, n_cols + 1):
        cell = ws.cell(grand_row, ci)
        cell.fill = tot_fill
        cell.font = tot_font
    ws.cell(grand_row, col_idx["Description"]).alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.row_dimensions[grand_row].height = 20

    # ── Thin borders on every cell ────────────────────────────────────────
    _all_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for r in range(1, grand_row + 1):
        for ci in range(1, n_cols + 1):
            ws.cell(r, ci).border = _all_border

    # ── Medium-border frame around the 4 order columns on every row ───────
    # A vertical "wall" on the left of Order_Qty and right of Total_Price
    # makes the order block visually distinct across the whole sheet.
    for r in range(1, grand_row + 1):
        for ci in range(_ord_ci_start, _ord_ci_end + 1):
            c   = ws.cell(r, ci)
            lft = _medium if ci == _ord_ci_start else c.border.left
            rgt = _medium if ci == _ord_ci_end   else c.border.right
            c.border = Border(left=lft, right=rgt,
                              top=c.border.top, bottom=c.border.bottom)

    # Freeze header row; autofilter over header only (Excel extends it downward)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    # ── Center-align specified columns ───────────────────────────────────
    _center_cols = {"B", "D", "E", "G", "H", "I", "J", "K", "L", "N", "O", "P", "Q"}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if get_column_letter(cell.column) in _center_cols:
                a = cell.alignment
                cell.alignment = Alignment(
                    horizontal="center", vertical=a.vertical,
                    wrap_text=a.wrap_text, indent=a.indent,
                )

    # ── Row 1 (header): height 20, vertical center ────────────────────────
    ws.row_dimensions[1].height = 20
    for cell in ws[1]:
        a = cell.alignment
        cell.alignment = Alignment(
            horizontal=a.horizontal, vertical="center",
            wrap_text=a.wrap_text, indent=a.indent,
        )

    # ── Auto column widths (last step before hiding) ──────────────────────
    # Skip formula cells — their value is the formula string, not the rendered
    # display value, which would inflate the width measurement significantly.
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value))
             if c.value is not None and not (isinstance(c.value, str) and c.value.startswith("="))
             else 0)
            for c in col_cells
        )
        # +5 extra chars: +2 padding + ~3 to account for the autofilter dropdown button
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 5, 60)

    # ── Hide columns I to M ───────────────────────────────────────────────
    for _hide_col in ("I", "J", "K", "L", "M"):
        ws.column_dimensions[_hide_col].hidden = True

    # ── Note at the bottom ────────────────────────────────────────────────
    def _add_note(row: int, text: str) -> None:
        cell = ws.cell(row, 1, value=text)
        cell.font = Font(italic=True, color="FF808080")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=ws.max_column,
        )

    _add_note(ws.max_row + 2,
        "Note: The Supplier shall verify all C5 material article numbers and related details, "
        "and update them as required to ensure full compliance with the C5 corrosion category."
    )
    _add_note(ws.max_row + 1,
        "Spare rules applied:  "
        "Primary Supports \u2014 5%;  "
        "Beam Sections, Threaded Rods \u2014 5%;  "
        "Bolts / Screws / Nuts, Installation Material \u2014 15%;  "
        "Glass Fabric Tape \u2014 30%."
    )


def _write_price_sheet(
    wb: Workbook,
    agg_ws_title: str,
    agg_all_headers: list[str],
    revision: str,
    has_coating: bool = True,
) -> None:
    """Write a formatted PRICE_SHEET summary tab with one SUMIF row per material group."""
    ws = wb.create_sheet(title=f"PRICE_SHEET_{revision}")

    # AGG sheet reference (quote if name contains spaces)
    agg_ref = f"'{agg_ws_title}'" if " " in agg_ws_title else agg_ws_title

    def _acol(name: str) -> str:
        return get_column_letter(agg_all_headers.index(name) + 1)

    mt_c = _acol("Material_Type")
    an_c = _acol("Article_Number")   # used to exclude separator/subtotal rows
    tw_c = _acol("Total_Weight_kg")
    ow_c = _acol("Order_Weight_kg")
    tp_c = _acol("Total_Price")

    # SUMIFS helper: sums col_letter only for real data rows.
    # Separator & subtotal rows have Article_Number = "" → "<>" excludes them.
    # SUMIFS natively skips non-numeric cells, so no IF(ISNUMBER) wrapper needed.
    # Bounded range 2-2000 avoids full-column overhead.
    def _sum_data(crit: str, col_letter: str) -> str:
        return (
            f"=IFERROR(SUMIFS("
            f"{agg_ref}!${col_letter}$2:${col_letter}$2000,"
            f"{agg_ref}!${mt_c}$2:${mt_c}$2000,{crit},"
            f"{agg_ref}!${an_c}$2:${an_c}$2000,\"<>\"),0)"
        )

    # SUMIFS on Total_Price column: sums only rows where Material_Type matches
    # and Article_Number is non-empty (excludes separator/subtotal rows).
    # SUMIFS ignores non-numeric cells (e.g. "" from unpopulated Unit_Price rows)
    # without erroring, so no IF(ISNUMBER) wrapper needed.
    def _sum_total_price(crit: str) -> str:
        return (
            f"=IFERROR(SUMIFS("
            f"{agg_ref}!${tp_c}$2:${tp_c}$2000,"
            f"{agg_ref}!${mt_c}$2:${mt_c}$2000,{crit},"
            f"{agg_ref}!${an_c}$2:${an_c}$2000,\"<>\"),0)"
        )

    # ── Shared styles ──────────────────────────────────────────────────────
    _thin    = Side(style="thin")
    _medium  = Side(style="medium")
    _box     = Border(top=_thin,   bottom=_thin,   left=_thin,   right=_thin)
    _box_hd  = Border(top=_medium, bottom=_medium, left=_thin,   right=_thin)

    _ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    _rgt = Alignment(horizontal="right",  vertical="center")

    _f_title  = Font(bold=True, size=13)
    _f_italic = Font(italic=True, size=10)
    _f_bold   = Font(bold=True)
    _f_hdr2   = Font(bold=True, size=10, color="FFFFFFFF")
    _f_tot    = Font(bold=True, size=11)

    _fill_hdr1   = PatternFill("solid", fgColor="FFD9E1F2")  # light blue
    _fill_hdr2   = PatternFill("solid", fgColor="FF1F3864")  # dark navy
    _fill_blue   = PatternFill("solid", fgColor="FFBDD7EE")  # input cell
    _fill_yellow = PatternFill("solid", fgColor="FFFFFF00")  # order weight
    _fill_total  = PatternFill("solid", fgColor="FFFFF2CC")  # grand total
    _fill_alt    = [
        PatternFill("solid", fgColor="FFF2F2F2"),  # light grey
        PatternFill("solid", fgColor="FFFFFFFF"),  # white
    ]
    _fill_red    = PatternFill("solid", fgColor="FFFF4444")  # red - reminder to verify/update

    NUM = "#,##0.00"

    # ── Column widths ──────────────────────────────────────────────────────
    for col, w in zip("ABCDEF", [9, 54, 16, 16, 16, 20]):
        ws.column_dimensions[col].width = w

    # ── Rows 1-2: Title block ──────────────────────────────────────────────
    ws.append(["Price sheet of Hangers and Supports Material \u2013 General Piping", None, None, None, None, None])
    ws.merge_cells("A1:F1")
    ws["A1"].font = _f_title
    ws["A1"].alignment = _lft
    ws.row_dimensions[1].height = 24

    # ── Row 2: subtitle / reference documents ─────────────────────────────
    _subtitle = (
        'Based on \u201cTSD General Piping Hangers and Supports Material\u201d'
        ' and BOM \u2013 Hangers and Supports Material \u2013 General Piping \u2013 Overall'
    )
    ws.append([_subtitle, None, None, None, None, None])
    ws.merge_cells("A2:F2")
    ws["A2"].font = _f_italic
    ws["A2"].alignment = _lft
    # ── Row 3: blank ───────────────────────────────────────────────────────
    ws.append([None])

    # ── Row 4: Supplier / Currency ─────────────────────────────────────────
    ws.append(["\u25ba  Fields highlighted in light blue are to be filled in by the supplier.", None, None, None, "Currency:", None])
    ws["A4"].font = Font(italic=True, size=10)
    ws["E4"].font = _f_bold
    ws["F4"].value = "EUR"
    ws["F4"].fill  = _fill_blue
    ws["F4"].alignment = _ctr

    # ── Row 5: blank ───────────────────────────────────────────────────────
    ws.append([None])

    # ── Row 6: Supplier name (red = reminder to verify before sending) ─────
    ws.append(["Supplier:", None, None, None, None, None])
    ws["A6"].font = _f_bold
    ws.merge_cells("B6:D6")
    ws["B6"].value = "SIKLA UK LTD"
    ws["B6"].fill = _fill_red

    # ── Row 7: blank ───────────────────────────────────────────────────────
    ws.append([None])

    # ── Row 8: Column header top band ─────────────────────────────────────
    ws.append([None, None, "Material as per BoM Rev0.0", "Spare", "Material for Order", None])
    ws.merge_cells("E8:F8")
    ws.row_dimensions[8].height = 48
    for col in ["C", "D", "E", "F"]:
        c = ws[f"{col}8"]
        c.fill = _fill_hdr1
        c.font = Font(bold=True, size=9, color="FF1F3864")
        c.alignment = _ctr
        c.border = _box
    ws["C8"].fill = _fill_red  # red = reminder to update revision when re-generating

    # ── Row 9: Column header bottom band (dark navy) ───────────────────────
    ws.append(["#", "DESCRIPTION", "WEIGHT\n[kg]", "WEIGHT\n[kg]", "WEIGHT\n[kg]", "Total Price\n[EUR]"])
    ws.row_dimensions[9].height = 34
    for col in "ABCDEF":
        c = ws[f"{col}9"]
        c.fill = _fill_hdr2
        c.font = _f_hdr2
        c.alignment = _ctr
        c.border = _box_hd

    # ── Group rows 11-20 ───────────────────────────────────────────────────
    GROUPS = [
        ("01 Primary Supports",         "Primary Supports"),
        ("02 Brackets Consoles",        "Brackets / Consoles  (Secondary Supports)"),
        ("03 Bolts Screws Nuts",        "Bolts, Screws and Nuts"),
        ("04 Beam Sections",            "Beam Sections"),
        ("05 Installation Material",    "Installation Material"),
    ]
    if has_coating:
        GROUPS += [
            ("06 Primary Supports C5",      "Primary Supports \u2014 Corrosion Category C5"),
            ("07 Brackets Consoles C5",     "Brackets / Consoles \u2014 Corrosion Category C5"),
            ("08 Bolts Screws Nuts C5",     "Bolts, Screws and Nuts \u2014 Corrosion Category C5"),
            ("09 Beam Sections C5",         "Beam Sections \u2014 Corrosion Category C5"),
            ("10 Installation Material C5", "Installation Material \u2014 Corrosion Category C5"),
        ]

    FIRST_ROW = 10
    for idx, (mt_key, label) in enumerate(GROUPS):
        r    = FIRST_ROW + idx
        fill = _fill_alt[idx % 2]
        crit = f'"{mt_key}"'

        bom_w = _sum_data(crit, tw_c)
        ord_w = _sum_data(crit, ow_c)
        spr_w = f"=E{r}-C{r}"
        tot_p = _sum_total_price(crit)

        ws.append([f"{idx + 1:02d}", label, bom_w, spr_w, ord_w, tot_p])
        ws.row_dimensions[r].height = 18

        aligns = [_ctr, _lft, _rgt, _rgt, _rgt, _rgt]
        for col_letter, align in zip("ABCDEF", aligns):
            c = ws[f"{col_letter}{r}"]
            c.fill = fill
            c.alignment = align
            c.border = _box
        for col_letter in "CDEF":
            ws[f"{col_letter}{r}"].number_format = NUM

        ws[f"E{r}"].font = _f_bold

    # ── Grand Total row ────────────────────────────────────────────────────
    LAST_GRP  = FIRST_ROW + len(GROUPS) - 1
    GT_ROW    = LAST_GRP + 1

    ws.append([
        None,
        "GRAND TOTAL PRICE FOR MATERIAL SUPPLY",
        f"=SUM(C{FIRST_ROW}:C{LAST_GRP})",
        f"=E{GT_ROW}-C{GT_ROW}",
        f"=SUM(E{FIRST_ROW}:E{LAST_GRP})",
        f"=SUM(F{FIRST_ROW}:F{LAST_GRP})",
    ])
    ws.row_dimensions[GT_ROW].height = 22
    for col_letter, align in zip("ABCDEF", [_ctr, _lft, _rgt, _rgt, _rgt, _rgt]):
        c = ws[f"{col_letter}{GT_ROW}"]
        c.fill = _fill_total
        c.font = _f_tot
        c.alignment = align
        c.border = _box_hd
    for col_letter in "CDEF":
        ws[f"{col_letter}{GT_ROW}"].number_format = NUM

    # ── Blank row ─────────────────────────────────────────────────────────
    ws.append([None])

    # ── Extra service rows (manual input, price filled by hand) ───────────
    for svc_label in [
        "PROJECT DISCOUNT 5%",
        "PRE-ASSEMBLY OF SECONDARY SUPPORTS IN SHOP",
        "TRANSPORTATION SERVICES TO CONSTRUCTION SITE",
    ]:
        er = ws.max_row + 1
        ws.append([None, svc_label, None, None, None, None])
        ws.merge_cells(f"B{er}:E{er}")
        ws[f"B{er}"].font = _f_bold
        ws[f"B{er}"].alignment = _lft
        ws[f"F{er}"].fill = _fill_blue
        ws[f"F{er}"].number_format = NUM
        ws[f"F{er}"].alignment = _rgt
        ws.row_dimensions[er].height = 18
        for col_letter in "ABCDEF":
            ws[f"{col_letter}{er}"].border = _box

    ws.freeze_panes = f"A{FIRST_ROW}"


def export_aggregated_to_excel(db_path: Path, out_xlsx: Path | None = None, revision: str = "REV0") -> Path:
    if not db_path.exists():
        raise FileNotFoundError(f"DB file not found: {db_path}")

    table = f"AGGREGATED_DATA_{revision}"
    if out_xlsx is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_xlsx = db_path.with_name(f"{db_path.stem}_{table}_{timestamp}.xlsx")

    raw_table = f"RAW_DATA_{revision}"

    validation_checks: list[dict] = []
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
            (table,),
        )
        if cur.fetchone() is None:
            raise ValueError(f"Table {table} does not exist in {db_path}")

        cur.execute(f"SELECT * FROM {sql_quote_identifier(table)}")
        agg_rows = cur.fetchall()
        cur.execute(f"PRAGMA table_info({sql_quote_identifier(table)})")
        agg_columns = [r[1] for r in cur.fetchall()]

        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
            (raw_table,),
        )
        has_raw = cur.fetchone() is not None
        if has_raw:
            cur.execute(f"SELECT * FROM {sql_quote_identifier(raw_table)}")
            raw_rows = cur.fetchall()
            cur.execute(f"PRAGMA table_info({sql_quote_identifier(raw_table)})")
            raw_columns = [r[1] for r in cur.fetchall()]
            has_coating = "Coating" in raw_columns
        else:
            has_coating = True  # no raw table → assume coating may be present

        validation_checks = _run_validation_checks(conn, raw_table, table, has_raw)
    finally:
        conn.close()

    def _write_sheet(ws, columns, rows):
        ws.append(columns)
        for row in rows:
            ws.append(list(row))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"
        # Freeze top row
        ws.freeze_panes = "A2"
        # AutoFilter on header row
        ws.auto_filter.ref = ws.dimensions
        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

    wb = Workbook()
    # Force Excel to fully recalculate all formulas on open.
    # openpyxl writes formula cells with no cached value (default 0), so without
    # this flag Excel may display stale 0s for cells that depend on other formula cells.
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    # Create AGG first via the active sheet, then create PRICE_SHEET and move it to front.
    if has_raw:
        ws_raw = wb.active
        ws_raw.title = raw_table
        _write_sheet(ws_raw, raw_columns, raw_rows)
        ws_agg = wb.create_sheet(title=table)
    else:
        ws_agg = wb.active
        ws_agg.title = table

    _write_agg_sheet(ws_agg, agg_columns, agg_rows)
    # Build the same column order that _write_agg_sheet uses so SUMIF formulas
    # in PRICE_SHEET reference the correct AGG columns.
    _ps_shift = ("Order_Qty", "Order_Weight_kg")
    _ps_base  = [c for c in agg_columns if c not in _ps_shift]
    _ps_cols  = _ps_base + list(_ps_shift) + ["Unit_Price", "Total_Price"]
    _write_price_sheet(wb, ws_agg.title, _ps_cols, revision, has_coating=has_coating)
    # Move PRICE_SHEET to front so it opens first
    wb.move_sheet(f"PRICE_SHEET_{revision}", offset=-len(wb.sheetnames) + 1)

    _write_validation_sheet(wb, validation_checks)
    wb.save(out_xlsx)
    n_fail = sum(1 for c in validation_checks if c["status"] == "FAIL")
    n_warn = sum(1 for c in validation_checks if c["status"] == "WARN")
    n_pass = sum(1 for c in validation_checks if c["status"] == "PASS")
    print(f"Exported {len(agg_rows)} aggregated + {len(raw_rows) if has_raw else 0} raw rows to: {out_xlsx}")
    print(f"Validation: {n_fail} FAIL  {n_warn} WARN  {n_pass} PASS — see VALIDATION sheet")
    return out_xlsx


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import Total Qty sheet to SQLite and export aggregated table to Excel.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    import_parser = subparsers.add_parser("import", help="Import 'Total Qty' from Excel into SQLite.")
    import_parser.add_argument("--excel", required=True, help="Path to source Excel file (.xlsm/.xlsx)")
    import_parser.add_argument("--sheet", default="Total Qty", help="Sheet name to import (default: 'Total Qty')")
    import_parser.add_argument("--revision", default="REV0", help="Revision tag, e.g. REV0, REV1")

    export_parser = subparsers.add_parser("export", help="Export AGGREGATED table from SQLite to Excel.")
    export_parser.add_argument("--db", required=True, help="Path to SQLite database")
    export_parser.add_argument("--revision", default="REV0", help="Revision tag, e.g. REV0, REV1")
    export_parser.add_argument("--out", help="Output xlsx path (optional)")

    populate_parser = subparsers.add_parser("populate", help="Populate AGGREGATED table from RAW and export to Excel.")
    populate_parser.add_argument("--db", required=True, help="Path to SQLite database")
    populate_parser.add_argument("--revision", default="REV0", help="Revision tag, e.g. REV0, REV1")
    populate_parser.add_argument("--out", help="Output xlsx path (optional)")

    test_parser = subparsers.add_parser("test-run", help="Import + populate + export into _test_runs/ with timestamp.")
    test_parser.add_argument("--excel", required=True, help="Path to source Excel file (.xlsm/.xlsx)")
    test_parser.add_argument("--sheet", default="Total Qty", help="Sheet name to import (default: 'Total Qty')")
    test_parser.add_argument("--revision", default="REV0", help="Revision tag, e.g. REV0, REV1")

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.command == "import":
        import_total_qty_to_sqlite(Path(args.excel), revision=args.revision, sheet_name=args.sheet)
        return

    if args.command == "export":
        out = Path(args.out) if args.out else None
        export_aggregated_to_excel(Path(args.db), out_xlsx=out, revision=args.revision)
        return

    if args.command == "populate":
        populate_aggregated_table(Path(args.db), revision=args.revision)
        out = Path(args.out) if args.out else None
        export_aggregated_to_excel(Path(args.db), out_xlsx=out, revision=args.revision)
        return

    if args.command == "test-run":
        excel_path = Path(args.excel)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        test_dir = excel_path.parent / "_test_runs"
        test_dir.mkdir(exist_ok=True)
        db_path = test_dir / f"{timestamp}.sqlite"
        out_xlsx = test_dir / f"{timestamp}_AGGREGATED_DATA_{args.revision}.xlsx"
        print(f"--- Test run: {timestamp} ---")
        import_total_qty_to_sqlite(excel_path, revision=args.revision, db_path=db_path, sheet_name=args.sheet)
        populate_aggregated_table(db_path, revision=args.revision)
        export_aggregated_to_excel(db_path, out_xlsx=out_xlsx, revision=args.revision)
        return

    raise ValueError(f"Unsupported command: {args.command}")


if __name__ == "__main__":
    main()