"""
BoM List vs Shipping List Comparison Script
Compares baseline BoM list with shipping list and generates a color-coded Excel report.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# File paths
BASELINE_FILE = r"C:\Users\szil\Repos\excel_wizadry\LMS_Shipping_list_vs_BoM_list\BoM_List_AA030-HZI-50211861_6.0_BoM List.xlsx"
SHIPPING_FILE = r"C:\Users\szil\Repos\excel_wizadry\LMS_Shipping_list_vs_BoM_list\Shipping_list_281000199_SIK_2024-550000204.xlsx"
OUTPUT_FILE = r"C:\Users\szil\Repos\excel_wizadry\LMS_Shipping_list_vs_BoM_list\Comparison_Report.xlsx"

# Color definitions
COLOR_MATCH = "90EE90"      # Light green
COLOR_MISMATCH = "FFFF99"   # Light yellow
COLOR_MISSING = "FFB6C1"    # Light red/pink
COLOR_EXTRA = "ADD8E6"      # Light blue

def normalize_string(value):
    """Normalize strings for case-insensitive comparison"""
    if pd.isna(value):
        return ""
    return str(value).strip().upper()

def load_baseline(file_path):
    """Load and process baseline BoM list"""
    print(f"Loading baseline file: {file_path}")
    # Read from correct sheet with header on row 7 (0-indexed)
    df = pd.read_excel(file_path, sheet_name='BoM List upload template', header=7)
    
    # Identify the columns we need (note: first column has leading space)
    baseline = df[['* BoM Item No.', ' * Qty', '* UoM      (Unit of Measure)', '* BoM Item Description']].copy()
    baseline.columns = ['BoM_Item_No', 'Baseline_Qty', 'Baseline_UoM', 'BoM_Item_Description']
    
    # Normalize for comparison
    baseline['BoM_Item_No_normalized'] = baseline['BoM_Item_No'].apply(normalize_string)
    baseline['Baseline_UoM_normalized'] = baseline['Baseline_UoM'].apply(normalize_string)
    
    # Remove rows with empty article numbers
    baseline = baseline[baseline['BoM_Item_No_normalized'] != ''].copy()
    
    print(f"Loaded {len(baseline)} items from baseline")
    return baseline

def load_shipping(file_path):
    """Load and process shipping list"""
    print(f"Loading shipping file: {file_path}")
    # Read from correct sheet with header on row 11 (0-indexed)
    df = pd.read_excel(file_path, sheet_name='Template Shipping List', header=11)
    
    # Identify the columns we need
    shipping = df[['BoM Item No *', 'Quantity*', 'Unit of Measure * ']].copy()
    shipping.columns = ['BoM_Item_No', 'Shipping_Qty', 'Shipping_UoM']
    
    # Normalize for comparison
    shipping['BoM_Item_No_normalized'] = shipping['BoM_Item_No'].apply(normalize_string)
    shipping['Shipping_UoM_normalized'] = shipping['Shipping_UoM'].apply(normalize_string)
    
    # Remove rows with empty article numbers
    shipping = shipping[shipping['BoM_Item_No_normalized'] != ''].copy()
    
    # Aggregate quantities by article number and unit
    shipping_agg = shipping.groupby(['BoM_Item_No_normalized', 'Shipping_UoM_normalized']).agg({
        'BoM_Item_No': 'first',
        'Shipping_Qty': 'sum',
        'Shipping_UoM': 'first'
    }).reset_index()
    
    print(f"Loaded {len(shipping)} rows from shipping list")
    print(f"Aggregated to {len(shipping_agg)} unique items")
    return shipping_agg

def compare_lists(baseline, shipping):
    """Compare baseline and shipping lists"""
    print("\nComparing lists...")
    
    results = []
    
    # Process each baseline item
    for idx, base_row in baseline.iterrows():
        item_no = base_row['BoM_Item_No_normalized']
        base_qty = base_row['Baseline_Qty']
        base_uom = base_row['Baseline_UoM_normalized']
        
        # Find matching items in shipping list
        matches = shipping[
            (shipping['BoM_Item_No_normalized'] == item_no) &
            (shipping['Shipping_UoM_normalized'] == base_uom)
        ]
        
        if len(matches) == 0:
            # Item not found in shipping list
            status = 'MISSING'
            ship_qty = None
            ship_uom = None
            qty_diff = None
        elif len(matches) == 1:
            ship_qty = matches.iloc[0]['Shipping_Qty']
            ship_uom = matches.iloc[0]['Shipping_UoM']
            
            # Check if quantities match
            if base_qty == ship_qty:
                status = 'MATCH'
                qty_diff = 0
            else:
                status = 'MISMATCH'
                qty_diff = ship_qty - base_qty
        else:
            # This shouldn't happen after aggregation, but handle it
            ship_qty = matches['Shipping_Qty'].sum()
            ship_uom = matches.iloc[0]['Shipping_UoM']
            
            if base_qty == ship_qty:
                status = 'MATCH'
                qty_diff = 0
            else:
                status = 'MISMATCH'
                qty_diff = ship_qty - base_qty
        
        results.append({
            '* BoM Item No.': base_row['BoM_Item_No'],
            '* BoM Item Description': base_row['BoM_Item_Description'],
            '* Qty (Baseline)': base_qty,
            '* UoM      (Unit of Measure) (Baseline)': base_row['Baseline_UoM'],
            'Quantity (Shipping)': ship_qty,
            'Unit of Measure (Shipping)': ship_uom,
            'Quantity Difference': qty_diff,
            'Status': status
        })
    
    # Find extra items in shipping list (not in baseline)
    baseline_items = set(baseline['BoM_Item_No_normalized'])
    extra_items = []
    
    for idx, ship_row in shipping.iterrows():
        item_no = ship_row['BoM_Item_No_normalized']
        if item_no not in baseline_items:
            extra_items.append({
                '* BoM Item No.': ship_row['BoM_Item_No'],
                '* Qty (Baseline)': None,
                '* UoM      (Unit of Measure) (Baseline)': None,
                'Quantity (Shipping)': ship_row['Shipping_Qty'],
                'Unit of Measure (Shipping)': ship_row['Shipping_UoM'],
                'Quantity Difference': None,
                'Status': 'EXTRA (Not in baseline)'
            })
    
    # Create DataFrames
    df_results = pd.DataFrame(results)
    df_extra = pd.DataFrame(extra_items) if extra_items else pd.DataFrame()
    
    # Count statistics
    match_count = len(df_results[df_results['Status'] == 'MATCH'])
    mismatch_count = len(df_results[df_results['Status'] == 'MISMATCH'])
    missing_count = len(df_results[df_results['Status'] == 'MISSING'])
    extra_count = len(df_extra)
    
    print(f"\nComparison Results:")
    print(f"  Matches: {match_count}")
    print(f"  Mismatches: {mismatch_count}")
    print(f"  Missing in shipping: {missing_count}")
    print(f"  Extra in shipping: {extra_count}")
    
    return df_results, df_extra

def apply_color_coding(output_file, df_results, df_extra):
    """Apply color coding to the output Excel file"""
    print(f"\nApplying color coding to {output_file}...")
    
    wb = load_workbook(output_file)
    ws_baseline = wb['Baseline Comparison']
    
    # Define fills
    fill_match = PatternFill(start_color=COLOR_MATCH, end_color=COLOR_MATCH, fill_type="solid")
    fill_mismatch = PatternFill(start_color=COLOR_MISMATCH, end_color=COLOR_MISMATCH, fill_type="solid")
    fill_missing = PatternFill(start_color=COLOR_MISSING, end_color=COLOR_MISSING, fill_type="solid")
    fill_extra = PatternFill(start_color=COLOR_EXTRA, end_color=COLOR_EXTRA, fill_type="solid")
    
    # Apply colors to baseline comparison sheet (starting from row 2, row 1 is header)
    for idx, row in enumerate(ws_baseline.iter_rows(min_row=2, max_row=len(df_results) + 1), start=0):
        if idx < len(df_results):
            status = df_results.iloc[idx]['Status']
            
            if status == 'MATCH':
                fill = fill_match
            elif status == 'MISMATCH':
                fill = fill_mismatch
            elif status == 'MISSING':
                fill = fill_missing
            else:
                fill = None
            
            if fill:
                for cell in row:
                    cell.fill = fill
    
    # Apply colors to extra items sheet if it exists
    if 'Extra Items (Shipping Only)' in wb.sheetnames and len(df_extra) > 0:
        ws_extra = wb['Extra Items (Shipping Only)']
        for row in ws_extra.iter_rows(min_row=2, max_row=len(df_extra) + 1):
            for cell in row:
                cell.fill = fill_extra
    
    # Apply colors to the Color Legend sheet
    if 'Color Legend' in wb.sheetnames:
        ws_legend = wb['Color Legend']
        # Color code the legend rows
        fills = [fill_match, fill_mismatch, fill_missing, fill_extra]
        for idx, row in enumerate(ws_legend.iter_rows(min_row=2, max_row=5), start=0):
            if idx < len(fills):
                for cell in row:
                    cell.fill = fills[idx]
    
    wb.save(output_file)
    print("Color coding applied successfully")

def main():
    """Main function to run the comparison"""
    print("="*60)
    print("BoM List vs Shipping List Comparison")
    print("="*60)
    
    # Check if files exist
    if not os.path.exists(BASELINE_FILE):
        print(f"ERROR: Baseline file not found: {BASELINE_FILE}")
        return
    
    if not os.path.exists(SHIPPING_FILE):
        print(f"ERROR: Shipping file not found: {SHIPPING_FILE}")
        return
    
    # Load data
    baseline = load_baseline(BASELINE_FILE)
    shipping = load_shipping(SHIPPING_FILE)
    
    # Compare
    df_results, df_extra = compare_lists(baseline, shipping)
    
    # Save to Excel with multiple sheets
    print(f"\nSaving results to {OUTPUT_FILE}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Baseline Comparison', index=False)
        
        if len(df_extra) > 0:
            df_extra.to_excel(writer, sheet_name='Extra Items (Shipping Only)', index=False)
        
        # Create a summary sheet
        summary_data = {
            'Metric': ['Total Baseline Items', 'Matches', 'Mismatches (Qty/UoM differs)', 
                       'Missing in Shipping', 'Extra in Shipping (Not in baseline)'],
            'Count': [
                len(df_results),
                len(df_results[df_results['Status'] == 'MATCH']),
                len(df_results[df_results['Status'] == 'MISMATCH']),
                len(df_results[df_results['Status'] == 'MISSING']),
                len(df_extra)
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Create a color legend sheet
        legend_data = {
            'Status': ['MATCH', 'MISMATCH', 'MISSING', 'EXTRA'],
            'Description': [
                'Perfect match - Article number, quantity, and unit all match',
                'Mismatch - Article exists in both lists but quantity or unit differs',
                'Missing - Article in baseline but not found in shipping list',
                'Extra - Article in shipping list but not in baseline'
            ],
            'Color': ['Light Green', 'Light Yellow', 'Light Red/Pink', 'Light Blue'],
            'RGB': [COLOR_MATCH, COLOR_MISMATCH, COLOR_MISSING, COLOR_EXTRA]
        }
        df_legend = pd.DataFrame(legend_data)
        df_legend.to_excel(writer, sheet_name='Color Legend', index=False)
    
    # Apply color coding
    apply_color_coding(OUTPUT_FILE, df_results, df_extra)
    
    print("\n" + "="*60)
    print("Comparison completed successfully!")
    print(f"Output file: {OUTPUT_FILE}")
    print("="*60)
    print("\nColor Legend:")
    print("  GREEN  - Match (quantities and units match)")
    print("  YELLOW - Mismatch (item exists but quantity differs)")
    print("  RED    - Missing (item not found in shipping list)")
    print("  BLUE   - Extra (item in shipping list but not in baseline)")
    print("="*60)

if __name__ == "__main__":
    main()
