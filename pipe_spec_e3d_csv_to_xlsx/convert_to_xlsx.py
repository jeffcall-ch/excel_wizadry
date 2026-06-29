"""
Convert E3D pipe specification CSV file to XLSX format.

This script processes a semicolon-separated CSV file where:
- First column is "SPRE" (pipe specification reference)
- Remaining columns are in alternating name;value pairs
- Output is an Excel (.xlsx) file with proper column structure
"""

import pandas as pd
import sys
import math
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


MM_VALUE_PATTERN = re.compile(r"^\s*([+-]?\d+(?:[.,]\d+)?)\s*mm\s*$", re.IGNORECASE)
LEADING_QUOTE_PATTERN = re.compile(r"^[\"']+")


def _parse_numeric_token(token):
    """Convert numeric text to int/float while handling comma decimal separators."""
    normalized = token.strip()
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")

    value = float(normalized)
    if value.is_integer():
        return int(value)
    return value


def parse_e3d_csv(file_path):
    """
    Parse E3D CSV file with alternating column name/value pairs.
    
    Args:
        file_path: Path to the input CSV file
        
    Returns:
        pandas.DataFrame with parsed data
    """
    data_rows = []
    
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
                
            # Split by semicolon
            parts = [p.strip() for p in line.split(';')]
            
            # First part is SPRE
            spre = parts[0]
            
            # Create row dictionary starting with SPRE
            row = {'SPRE': spre}
            
            # Process remaining parts as alternating column_name, column_value
            # Start from index 1 (after SPRE)
            for i in range(1, len(parts) - 1, 2):
                if i + 1 < len(parts):
                    col_name = parts[i].strip()
                    col_value = parts[i + 1].strip()
                    
                    # Only add if column name is not empty
                    if col_name:
                        row[col_name] = col_value
            
            data_rows.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(data_rows)
    
    # Clean P1 CONN and P2 CONN columns by removing commas
    if 'P1 CONN' in df.columns:
        df['P1 CONN'] = df['P1 CONN'].str.replace(',', '', regex=False)
    if 'P2 CONN' in df.columns:
        df['P2 CONN'] = df['P2 CONN'].str.replace(',', '', regex=False)
    
    return df


def normalize_mm_columns(df):
    """Move mm unit from cell values to header and convert cells to numeric values."""
    rename_map = {}
    mm_columns = []

    for col in df.columns:
        series = df[col].dropna().astype(str)
        if series.empty:
            continue

        mm_mask = series.str.match(MM_VALUE_PATTERN)
        if not mm_mask.any():
            continue

        mm_columns.append(col)

        def _convert_mm_cell(value):
            if pd.isna(value) or isinstance(value, (int, float)):
                return value

            text = str(value).strip()
            match = MM_VALUE_PATTERN.match(text)
            if not match:
                return value

            return _parse_numeric_token(match.group(1))

        df[col] = df[col].apply(_convert_mm_cell)

        if 'mm' not in col.lower():
            new_name = f"{col} [mm]"
            if new_name not in df.columns:
                rename_map[col] = new_name

    if rename_map:
        df = df.rename(columns=rename_map)

    return df, mm_columns


def clean_first_column(df):
    """Strip leading quote characters from first column values."""
    if df.empty:
        return 0

    first_col = df.columns[0]
    original = df[first_col].copy()
    df[first_col] = df[first_col].apply(
        lambda value: LEADING_QUOTE_PATTERN.sub('', value) if isinstance(value, str) else value
    )
    cleaned_count = (original.fillna('<NA>') != df[first_col].fillna('<NA>')).sum()
    return int(cleaned_count)


def sort_dataframe(df):
    """Sort by A-group, then B, then C (numeric when possible), then full A text."""
    if df.empty:
        return df, []

    sort_columns = []
    helper_columns = []

    col_a = df.columns[0]
    # Use the first path segment (e.g. AHFXR from /AHFXR/...) as A sort group.
    a_group_col = '__sort_a_group__'
    df[a_group_col] = df[col_a].apply(
        lambda value: str(value).strip('/').split('/')[0] if pd.notna(value) and str(value).strip('/') else str(value)
    )
    sort_columns.append(a_group_col)
    helper_columns.append(a_group_col)

    if len(df.columns) > 1:
        col_b = df.columns[1]
        sort_columns.append(col_b)

    if len(df.columns) > 2:
        col_c = df.columns[2]
        c_numeric_col = '__sort_c_numeric__'
        df[c_numeric_col] = pd.to_numeric(df[col_c], errors='coerce')
        if df[c_numeric_col].notna().any():
            sort_columns.append(c_numeric_col)
            helper_columns.append(c_numeric_col)
        else:
            sort_columns.append(col_c)

    # Keep deterministic ordering within same A/B/C groups.
    sort_columns.append(col_a)

    df = df.sort_values(by=sort_columns, ascending=True, na_position='last', kind='mergesort')
    df = df.reset_index(drop=True)
    if helper_columns:
        df = df.drop(columns=helper_columns)

    return df, sort_columns


def apply_excel_formatting(output_file):
    """Apply filter, freeze panes, wrapping, vertical centering and row-height sizing."""
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = 'B2'

    alignment = Alignment(wrap_text=True, vertical='center')

    # Cache column widths used to estimate wrapped line count for row height.
    column_widths = {}
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        width = worksheet.column_dimensions[col_letter].width
        column_widths[col_idx] = float(width) if width else 8.43

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        max_line_count = 1

        for cell in row:
            cell.alignment = alignment
            if cell.value is None:
                continue

            text = str(cell.value)
            for line in text.splitlines() or [text]:
                col_width = max(column_widths.get(cell.column, 8.43), 1)
                estimated_lines = max(1, math.ceil(len(line) / col_width))
                max_line_count = max(max_line_count, estimated_lines)

        worksheet.row_dimensions[row[0].row].height = max(15, max_line_count * 15)

    workbook.save(output_file)


def main():
    """Main function to convert CSV to XLSX."""
    
    # Hard-coded paths for easy execution
    if len(sys.argv) < 2:
        # Use hard-coded default paths
        input_file = Path(__file__).parent / "ADI_PIPE_SPECS_ALL.csv" 
        output_file = input_file.with_suffix('.xlsx')
    else:
        # Get input file path from command line argument
        input_file = Path(sys.argv[1])
        
        # Determine output file path
        if len(sys.argv) >= 3:
            output_file = Path(sys.argv[2])
        else:
            # Use same name as input but with .xlsx extension
            output_file = input_file.with_suffix('.xlsx')
    
    # Check if file exists
    if not input_file.exists():
        print(f"Error: Input file '{input_file}' not found!")
        sys.exit(1)
    
    print(f"Reading input file: {input_file}")
    
    # Parse the CSV file
    df = parse_e3d_csv(input_file)

    # Normalize units in mm columns and keep values numeric for filtering.
    df, mm_columns = normalize_mm_columns(df)
    cleaned_first_col_count = clean_first_column(df)
    df, sort_columns = sort_dataframe(df)
    
    print(f"Parsed {len(df)} rows with {len(df.columns)} columns")
    print(f"Columns: {', '.join(df.columns.tolist())}")
    if mm_columns:
        print(f"Normalized mm columns: {', '.join(mm_columns)}")
    print(f"Stripped leading quotes in first column: {cleaned_first_col_count} rows")
    if sort_columns:
        pretty_sort = [
            'Column A group'
            if col == '__sort_a_group__'
            else 'Column C numeric'
            if col == '__sort_c_numeric__'
            else col
            for col in sort_columns
        ]
        print(f"Sorted by columns: {', '.join(pretty_sort)}")
    
    # Write to Excel
    print(f"\nWriting to Excel file: {output_file}")
    df.to_excel(output_file, index=False, engine='openpyxl')
    apply_excel_formatting(output_file)
    
    print(f"\n✓ Successfully converted to {output_file}")
    print(f"  Total rows: {len(df)}")
    print(f"  Total columns: {len(df.columns)}")
    print("  Applied: filter, freeze pane B2, wrap text, vertical center, row-height sizing")


if __name__ == "__main__":
    main()
