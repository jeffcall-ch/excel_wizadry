#!/usr/bin/env python3
"""
Excel Line List Comparator - Single File Implementation

A high-performance Python tool for comparing Excel piping line lists with Interface No.-based identification.
Generates detailed change analysis with visual formatting in a single 'new_color_coded' sheet.

Usage:
    python excel_compare.py compare.xlsx                    # Auto-generated timestamped output
    python excel_compare.py compare.xlsx custom_result.xlsx # Custom output filename

Input file must contain 'old' and 'new' sheets with piping data.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from copy import copy
from collections import defaultdict
import sys
import locale
import logging
import argparse
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any


# Color constants (matching VBA RGB values)
class Colors:
    BLUE = "0000FF"      # RGB(0, 0, 255) - Blue background
    GREEN = "00FF00"     # RGB(0, 255, 0) - Green background  
    YELLOW = "FFFF00"    # RGB(255, 255, 0) - Yellow background
    RED = "FFC0C0"       # RGB(255, 192, 192) - Light red/pink for deleted items
    LIGHT_BLUE = "C0C0FF"  # RGB(192, 192, 255) - Light blue/cyan (&HC0C0FF from VBA)
    DARK_RED = "FF0000"  # RGB(255, 0, 0) - Dark red for duplicate font color


@dataclass
class ComparisonResult:
    """Stores comparison results with formatting information."""
    data: pd.DataFrame
    is_old_base: bool
    colors: Dict[Tuple[int, str], str] = field(default_factory=dict)
    comments: Dict[Tuple[int, str], str] = field(default_factory=dict)
    markers: Dict[int, str] = field(default_factory=dict)
    duplicate_rows: set = field(default_factory=set)
    added_rows: List[Dict[str, Any]] = field(default_factory=list)
    deleted_rows: List[Dict[str, Any]] = field(default_factory=list)


class ExcelTableComparator:
    """
    Pythonic Excel table comparator that produces 'new_color_coded' sheet.
    
    Generates a single comparison sheet with NEW table as base, showing changes from OLD.
    """
    
    def __init__(self, old_file: str = 'old_list.xlsx', 
                 new_file: str = 'new_list.xlsx',
                 output_file: str = None):
        self.old_file = Path(old_file)
        self.new_file = Path(new_file)
        
        if output_file is None:
            # Auto-generate output filename: "filename_YYYYMMDD_HHMMSS_comparison.xlsx"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_file_stem = self.new_file.stem  # filename without extension
            output_filename = f"{new_file_stem}_{timestamp}_comparison.xlsx"
            self.output_file = self.new_file.parent / output_filename
        else:
            self.output_file = Path(output_file)
        
        # Setup logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('comparison.log', mode='w'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
        # Detect system language (matches VBA language detection intent)
        self.is_german = self._detect_german_locale()
        
        # Will be populated during loading
        self.old_workbook: Optional[load_workbook] = None
        self.new_workbook: Optional[load_workbook] = None
        self.old_data: Optional[pd.DataFrame] = None
        self.new_data: Optional[pd.DataFrame] = None
        self.column_names: List[str] = []
        
    def _detect_german_locale(self) -> bool:
        """Detect if system is German (matches VBA LanguageID intent)."""
        try:
            lang_code = locale.getdefaultlocale()[0]
            return lang_code is not None and lang_code.startswith('de')
        except:
            return False
    
    def _get_localized_text(self, english: str, german: str) -> str:
        """Return localized text based on system language."""
        return german if self.is_german else english
    
    def validate_files_and_load_data(self) -> None:
        """Load and validate Excel files."""
        self.logger.info("Validating files and loading data...")
        
        # Load workbooks
        self.old_workbook = load_workbook(self.old_file, data_only=True)
        self.new_workbook = load_workbook(self.new_file, data_only=True)
        
        # Load data from first sheet of each workbook
        old_sheet_name = self.old_workbook.sheetnames[0]
        new_sheet_name = self.new_workbook.sheetnames[0]
        
        self.old_data = pd.read_excel(self.old_file, sheet_name=old_sheet_name)
        self.new_data = pd.read_excel(self.new_file, sheet_name=new_sheet_name)
        
        # Remove "Changed" column and any "Unnamed" columns (from previous comparisons)
        self.old_data = self._remove_comparison_columns(self.old_data)
        self.new_data = self._remove_comparison_columns(self.new_data)
        
        # Remove "Added Rows" and "Deleted Rows" sections at the bottom
        self.old_data = self._remove_added_deleted_rows_sections(self.old_data)
        self.new_data = self._remove_added_deleted_rows_sections(self.new_data)
        
        # Get column names (assuming both tables have same structure)
        self.column_names = list(self.new_data.columns)
        
        self.logger.info(f"Old table: {len(self.old_data)} rows, {len(self.old_data.columns)} columns")
        self.logger.info(f"New table: {len(self.new_data)} rows, {len(self.new_data.columns)} columns")
    
    def _remove_comparison_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Remove the 'Changed' column and any 'Unnamed' columns that are artifacts 
        from previous comparisons.
        """
        columns_to_drop = []
        
        # Check for "Changed" or "Geändert" column
        for col in df.columns:
            col_str = str(col)
            if col_str in ['Changed', 'Geändert'] or col_str.startswith('Unnamed:'):
                columns_to_drop.append(col)
        
        if columns_to_drop:
            self.logger.info(f"Removing comparison columns: {columns_to_drop}")
            df = df.drop(columns=columns_to_drop)
        
        return df
    
    def _remove_added_deleted_rows_sections(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Remove rows from the bottom by finding the first row with an empty Interface No.
        and dropping that row and everything below it.
        """
        if len(df) == 0:
            return df
        
        # Check if 'Interface No.' column exists
        if 'Interface No.' not in df.columns:
            self.logger.warning("'Interface No.' column not found - skipping cleanup")
            return df
        
        # Find the first row index where 'Interface No.' is empty
        first_empty_idx = None
        for idx in range(len(df)):
            interface_no = df.iloc[idx]['Interface No.']
            if pd.isna(interface_no) or str(interface_no).strip() == '':
                first_empty_idx = idx
                break
        
        # If we found an empty row, drop it and everything below
        if first_empty_idx is not None:
            rows_dropped = len(df) - first_empty_idx
            self.logger.info(f"Found first empty Interface No. at row {first_empty_idx}, dropping {rows_dropped} rows")
            df = df.iloc[:first_empty_idx].copy()
        
        return df
    
    def _remove_legend_row_if_present(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Check if the first row is a color coding legend row and remove it.
        Legend row is detected by checking if the first row has values like 'Changed', 'Added', 'Deleted'
        or their German equivalents 'Geändert', 'Hinzugefügt', 'Gelöscht'.
        """
        if len(df) == 0:
            return df
        
        # Get first row values as strings
        first_row_values = [str(val).strip() for val in df.iloc[0].values if pd.notna(val)]
        
        # Check if any of the legend keywords are in the first row
        legend_keywords = ['Changed', 'Added', 'Deleted', 'Geändert', 'Hinzugefügt', 'Gelöscht']
        
        # If any legend keyword is found in the first row, it's a legend row
        is_legend_row = any(keyword in first_row_values for keyword in legend_keywords)
        
        if is_legend_row:
            self.logger.info("Legend row detected in first row - removing it")
            # Drop the first row and reset index
            df = df.iloc[1:].reset_index(drop=True)
        
        return df
    
    def _capture_formatting_from_old_sheet(self):
        """Capture formatting from the 'old' sheet to apply to output."""
        self.logger.info("Capturing formatting from 'old' sheet...")
        
        old_sheet_name = self.old_workbook.sheetnames[0]
        old_ws = self.old_workbook[old_sheet_name]
        
        formatting = {
            'column_widths': {},
            'row_heights': {},
            'cell_formats': {},
            'freeze_panes': old_ws.freeze_panes,
            'auto_filter': old_ws.auto_filter.ref if old_ws.auto_filter else None,
        }
        
        # Capture column widths
        for col_idx, col_name in enumerate(self.column_names, start=1):
            col_letter = get_column_letter(col_idx)
            if old_ws.column_dimensions[col_letter].width:
                formatting['column_widths'][col_idx] = old_ws.column_dimensions[col_letter].width
        
        # Capture row heights (for header rows)
        for row_idx in range(1, min(10, old_ws.max_row + 1)):  # First 10 rows
            if old_ws.row_dimensions[row_idx].height:
                formatting['row_heights'][row_idx] = old_ws.row_dimensions[row_idx].height
        
        # Capture cell formats from header row and first data row
        for row_idx in [1, 2, 3]:  # Header rows and first data row
            for col_idx in range(1, len(self.column_names) + 1):
                cell = old_ws.cell(row=row_idx, column=col_idx)
                formatting['cell_formats'][(row_idx, col_idx)] = {
                    'font': copy(cell.font),
                    'alignment': copy(cell.alignment),
                    'border': copy(cell.border),
                    'number_format': cell.number_format,
                    'protection': copy(cell.protection),
                }
        
        return formatting
    
    def create_unique_keys(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Create unique keys for duplicate handling (OPTIMIZED VERSION)."""
        self.logger.info("Creating unique keys for duplicate handling...")
        
        def process_dataframe(df: pd.DataFrame, name: str) -> pd.DataFrame:
            df = df.copy()
            df['_unique_key'] = df['Interface No.'].astype(str)
            
            # OPTIMIZED: Use pandas value_counts() for O(n) duplicate detection
            interface_counts = df['Interface No.'].value_counts()
            duplicated_interfaces = set(interface_counts[interface_counts > 1].index)
            
            # Mark duplicate rows efficiently
            df['_is_duplicate'] = df['Interface No.'].isin(duplicated_interfaces)
            
            return df
        
        old_with_keys = process_dataframe(self.old_data, "OLD")
        new_with_keys = process_dataframe(self.new_data, "NEW")
        
        return old_with_keys, new_with_keys
    
    def perform_comparison(self, old_with_keys: pd.DataFrame, new_with_keys: pd.DataFrame) -> ComparisonResult:
        """Perform table comparison (NEW base only)."""
        self.logger.info("Performing table comparison (NEW base only)...")
        
        # NEW base comparison
        new_base_result = self._compare_single_pass(new_with_keys, old_with_keys, is_old_base=False)
        
        # Post-processing: Check for rows where all cells are green (completely new rows)
        self._post_process_all_green_rows(new_base_result)
        
        return new_base_result
    
    def _compare_single_pass(self, base_df: pd.DataFrame, compare_df: pd.DataFrame,
                           is_old_base: bool) -> ComparisonResult:
        """
        Perform single comparison pass (OPTIMIZED VERSION).
        """
        result = ComparisonResult(
            data=base_df[self.column_names].copy(),
            is_old_base=is_old_base
        )
        
        # Track duplicate rows for red font marking
        for idx in range(len(base_df)):
            if base_df.iloc[idx]['_is_duplicate']:
                result.duplicate_rows.add(idx)
        
        # PERFORMANCE OPTIMIZATION: Create lookup dictionaries
        # Create a mapping from unique_key to ALL row indices (to handle duplicates)
        base_key_to_indices = defaultdict(list)
        for idx, row in base_df.iterrows():
            base_key_to_indices[row['_unique_key']].append(idx)
        
        # Set base background color for all cells (VBA: UsedRange.Interior.Color)
        base_color = Colors.LIGHT_BLUE if is_old_base else Colors.GREEN
        for row_idx in range(len(result.data)):
            for col_name in self.column_names:
                result.colors[(row_idx, col_name)] = base_color
        
        # VECTORIZED COMPARISON: Process all rows at once where possible
        compare_keys = compare_df['_unique_key'].tolist()
        
        # Find which keys exist in base (vectorized operation)
        existing_keys = set(base_key_to_indices.keys())
        
        # Separate found and not found keys
        found_comparisons = []
        used_compare_indices = set()  # Track which compare indices have been matched
        
        for compare_idx, compare_key in enumerate(compare_keys):
            if compare_key in existing_keys:
                # Get ALL indices for this key in base table
                base_indices = base_key_to_indices[compare_key]
                
                # Match 1-to-1 in order: find first unused base index for this key
                for base_idx in base_indices:
                    # Check if this base_idx has already been matched to another compare row
                    if not any(base_idx == b_idx for b_idx, _ in found_comparisons):
                        found_comparisons.append((base_idx, compare_idx))
                        used_compare_indices.add(compare_idx)
                        break  # Only match once per compare row
            else:
                # Row not found in base table by key
                # Only include columns that are in self.column_names (exclude metadata columns)
                row_data = {col: compare_df.iloc[compare_idx][col] for col in self.column_names if col in compare_df.columns}
                if is_old_base:
                    # OLD is base, NEW row not found = Added to NEW
                    result.added_rows.append(row_data)
                else:
                    # NEW is base, OLD row not found = Deleted from NEW
                    result.deleted_rows.append(row_data)
        
        # Process found rows in batches for better performance
        for base_idx, compare_idx in found_comparisons:
            # Clear base color for this row (VBA: Interior.Pattern = xlNone)
            for col_name in self.column_names:
                if (base_idx, col_name) in result.colors:
                    del result.colors[(base_idx, col_name)]
            
            # Compare individual cells (optimized)
            changes_in_row = []
            
            # VECTORIZED: Get all values for this row at once
            base_row = result.data.iloc[base_idx]
            compare_row = compare_df.iloc[compare_idx]
            
            for col_name in self.column_names:
                base_value = self._get_cell_text_value_fast(base_row[col_name])
                compare_value = self._get_cell_text_value_fast(compare_row[col_name])
                
                if base_value == compare_value:
                    continue  # No change
                
                # Determine change type and apply formatting
                change_type, color = self._analyze_cell_change(
                    base_value, compare_value, is_old_base
                )
                
                result.colors[(base_idx, col_name)] = color
                changes_in_row.append(change_type)
                
                # Add comment if value changed
                if change_type in ["Changed", "Added", "Deleted"]:
                    compare_table_name = "New Table" if is_old_base else "Old Table"
                    # Show the old value in the comment
                    if compare_value:
                        result.comments[(base_idx, col_name)] = f"{compare_table_name} value:\n{compare_value}"
                    else:
                        result.comments[(base_idx, col_name)] = f"{compare_table_name} value:\n(empty)"
            
            # Combine change markers for this row
            if changes_in_row:
                result.markers[base_idx] = self._combine_change_markers(changes_in_row)
        
        return result
    
    def _post_process_all_green_rows(self, result: ComparisonResult) -> None:
        """
        Post-processing step: If all cells in a row are green (completely new row),
        then mark the entire row as 'Added' in the change marker column.
        """
        self.logger.info("Post-processing: Checking for rows with all green cells...")
        
        for row_idx in range(len(result.data)):
            # Check if all cells in this row have green background
            all_cells_green = True
            cells_with_colors = 0
            
            for col_name in self.column_names:
                color_key = (row_idx, col_name)
                if color_key in result.colors:
                    cells_with_colors += 1
                    if result.colors[color_key] != Colors.GREEN:
                        all_cells_green = False
                        break
                else:
                    # If a cell has no color, it's not green
                    all_cells_green = False
                    break
            
            # If all cells in the row are green, mark the row as "Added"
            if all_cells_green and cells_with_colors == len(self.column_names):
                added_text = self._get_localized_text("Added", "Hinzugefügt")
                result.markers[row_idx] = added_text
                self.logger.debug(f"Row {row_idx}: All cells green -> marked as '{added_text}'")
    
    def _get_cell_text_value_fast(self, value) -> str:
        """Fast cell value to text conversion (optimized version)."""
        if pd.isna(value):
            return ""
        
        # Match Excel's .Text property behavior (optimized)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))  # Remove .0 from whole numbers
        
        return str(value)
    
    def _get_cell_text_value(self, df: pd.DataFrame, row_idx: int, col_name: str) -> str:
        """Get cell value as text (matches VBA .Text property)."""
        value = df.iloc[row_idx][col_name]
        if pd.isna(value):
            return ""
        
        # Match Excel's .Text property behavior
        if isinstance(value, float) and value.is_integer():
            return str(int(value))  # Remove .0 from whole numbers
        
        return str(value)
    
    def _analyze_cell_change(self, base_value: str, compare_value: str, 
                           is_old_base: bool) -> Tuple[str, str]:
        """
        Analyze cell change and return change type with color (matches VBA change logic).
        VBA logic depends on which table is the base!
        """
        # Both empty - no change
        if not base_value and not compare_value:
            return "No Change", ""
        
        # Base cell is empty AND compare cell has value
        if not base_value and compare_value:
            if is_old_base:  # OLD is base, NEW has value = Added to NEW
                return "Added", Colors.GREEN  # "N"
            else:  # NEW is base, OLD has value = Deleted from NEW  
                return "Deleted", Colors.RED  # "D"
        
        # Base cell has value AND compare cell is empty
        if base_value and not compare_value:
            if is_old_base:  # OLD is base, NEW empty = Deleted in NEW
                return "Deleted", Colors.RED  # "D"
            else:  # NEW is base, OLD empty = Added in NEW
                return "Added", Colors.GREEN  # "N"
        
        # Both have different non-empty values = Changed
        if base_value != compare_value:
            return "Changed", Colors.YELLOW  # "Ch"
        
        return "No Change", ""
    
    def _combine_change_markers(self, changes: List[str]) -> str:
        """Combine multiple change types into descriptive text."""
        unique_changes = set(changes)
        unique_changes.discard("No Change")
        
        if not unique_changes:
            return ""
        
        # Convert to descriptive text
        descriptions = []
        if "Deleted" in unique_changes:
            descriptions.append("Deleted")
        if "Changed" in unique_changes:
            descriptions.append("Changed") 
        if "Added" in unique_changes:
            descriptions.append("Added")
            
        return ", ".join(descriptions)
    
    def generate_excel_output(self, new_base_result: ComparisonResult) -> None:
        """Generate Excel output file with comparison results."""
        self.logger.info("Generating Excel output...")
        
        # Capture formatting from old sheet
        formatting = self._capture_formatting_from_old_sheet()
        
        # Create new workbook with single sheet
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create new_color_coded sheet
        ws_new = wb.create_sheet("new_color_coded")
        self._write_comparison_sheet(ws_new, new_base_result, formatting)
        
        # Save workbook
        wb.save(self.output_file)
        self.logger.info(f"Excel output saved to: {self.output_file}")
    
    def _write_comparison_sheet(self, worksheet, result: ComparisonResult, formatting: dict) -> None:
        """Write complete comparison sheet with all elements."""
        # Apply column widths from old sheet
        self._apply_column_formatting(worksheet, formatting)
        
        # Header row
        self._format_header_row(worksheet, result.is_old_base)
        
        # Data headers and rows
        self._write_data_headers(worksheet, formatting)
        self._write_data_rows(worksheet, result, formatting)
        
        # Added rows section
        if result.added_rows:
            self._write_added_rows_section(worksheet, result)
        
        # Deleted rows section
        if result.deleted_rows:
            self._write_deleted_rows_section(worksheet, result)
        
        # Change marker column
        self._write_change_marker_column(worksheet, result)
        
        # Apply AutoFilter
        self._apply_autofilter(worksheet, result)
        
        # Enable auto-height for row 1 (legend row)
        worksheet.row_dimensions[1].height = None
        
        # Enable auto-height for all data rows
        self._enable_auto_height_for_data_rows(worksheet, result)
        
        # Apply freeze panes at C3 as the very last formatting action
        worksheet.freeze_panes = 'C3'
    
    def _enable_auto_height_for_data_rows(self, worksheet, result: ComparisonResult) -> None:
        """Enable auto-height for all data rows based on wrap text."""
        # Calculate the range of data rows
        start_row = 3
        end_row = start_row + len(result.data) - 1
        
        # Check each cell for wrap text and enable auto height if needed
        for row_num in range(start_row, end_row + 1):
            for col_idx in range(1, len(self.column_names) + 1):
                cell = worksheet.cell(row=row_num, column=col_idx)
                if cell.alignment and cell.alignment.wrap_text:
                    # Enable auto height for this row
                    worksheet.row_dimensions[row_num].height = None
                    break  # Only need to set once per row
    
    def _apply_column_formatting(self, worksheet, formatting: dict) -> None:
        """Apply column widths and auto-height for rows from old sheet."""
        # Apply column widths
        for col_idx, width in formatting.get('column_widths', {}).items():
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width
        
        # Apply row heights for header rows
        for row_idx, height in formatting.get('row_heights', {}).items():
            worksheet.row_dimensions[row_idx].height = height
    
    def _format_header_row(self, worksheet, is_old_base: bool) -> None:
        """Format the header row with title and legend."""
        # Clean header row with only the three change type labels
        worksheet.cell(1, 4, self._get_localized_text("Changed", "Geändert"))
        worksheet.cell(1, 5, self._get_localized_text("Added", "Hinzugefügt"))  
        worksheet.cell(1, 6, self._get_localized_text("Deleted", "Gelöscht"))
        
        # Apply background colors to match the change types
        worksheet.cell(1, 4).fill = PatternFill(start_color=Colors.YELLOW, end_color=Colors.YELLOW, fill_type='solid')
        worksheet.cell(1, 5).fill = PatternFill(start_color=Colors.GREEN, end_color=Colors.GREEN, fill_type='solid')
        worksheet.cell(1, 6).fill = PatternFill(start_color=Colors.RED, end_color=Colors.RED, fill_type='solid')
    
    def _write_data_headers(self, worksheet, formatting: dict) -> None:
        """Write column headers in row 2 with formatting from old sheet."""
        for col_idx, col_name in enumerate(self.column_names, start=1):
            cell = worksheet.cell(2, col_idx, col_name)
            
            # Apply formatting from old sheet header row (row 2)
            if (2, col_idx) in formatting.get('cell_formats', {}):
                fmt = formatting['cell_formats'][(2, col_idx)]
                cell.font = copy(fmt['font'])
                cell.alignment = copy(fmt['alignment'])
                cell.border = copy(fmt['border'])
                cell.number_format = fmt['number_format']
            
            # Enable wrap text for header row
            existing_alignment = cell.alignment if cell.alignment else Alignment()
            cell.alignment = Alignment(
                horizontal=existing_alignment.horizontal,
                vertical=existing_alignment.vertical,
                text_rotation=existing_alignment.text_rotation,
                indent=existing_alignment.indent,
                shrink_to_fit=existing_alignment.shrink_to_fit,
                wrap_text=True
            )
            
            # Make all column headers bold
            existing_font = cell.font
            cell.font = Font(
                name=existing_font.name,
                size=existing_font.size,
                bold=True,
                italic=existing_font.italic,
                color=existing_font.color
            )
            
            # Set specific column widths
            col_letter = get_column_letter(col_idx)
            if col_name == "PIPEN - Name":
                worksheet.column_dimensions[col_letter].width = 20
            elif col_name == "PIPEN - comment":
                worksheet.column_dimensions[col_letter].width = 30
            elif col_name == "PIPEN - Date":
                worksheet.column_dimensions[col_letter].width = 20
    
    def _write_data_rows(self, worksheet, result: ComparisonResult, formatting: dict) -> None:
        """Write data rows with colors, comments, and formatting from old sheet."""
        for row_idx in range(len(result.data)):
            excel_row = row_idx + 3  # Start from row 3
            
            for col_idx, col_name in enumerate(self.column_names, start=1):
                cell = worksheet.cell(row=excel_row, column=col_idx)
                cell.value = result.data.iloc[row_idx][col_name]
                
                # Apply formatting from old sheet data row (row 3 as template)
                if (3, col_idx) in formatting.get('cell_formats', {}):
                    fmt = formatting['cell_formats'][(3, col_idx)]
                    cell.font = copy(fmt['font'])
                    cell.alignment = copy(fmt['alignment'])
                    cell.border = copy(fmt['border'])
                    cell.number_format = fmt['number_format']
                
                # Enable wrap text for PIPEN - comment column
                if col_name == "PIPEN - comment":
                    existing_alignment = cell.alignment if cell.alignment else Alignment()
                    cell.alignment = Alignment(
                        horizontal=existing_alignment.horizontal,
                        vertical=existing_alignment.vertical,
                        text_rotation=existing_alignment.text_rotation,
                        indent=existing_alignment.indent,
                        shrink_to_fit=existing_alignment.shrink_to_fit,
                        wrap_text=True
                    )
                
                # Apply colors (this may override some formatting like fill)
                if (row_idx, col_name) in result.colors:
                    color = result.colors[(row_idx, col_name)]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                # Apply comments
                if (row_idx, col_name) in result.comments:
                    cell.comment = Comment(result.comments[(row_idx, col_name)], "System")
                    cell.comment.width = 200
                    cell.comment.height = 50
            
            # Mark duplicates with dark red font in Interface No. column
            # Preserve existing font attributes but change color to dark red
            if row_idx in result.duplicate_rows:
                dup_cell = worksheet.cell(row=excel_row, column=2)
                existing_font = dup_cell.font
                dup_cell.font = Font(
                    name=existing_font.name,
                    size=existing_font.size,
                    bold=existing_font.bold,
                    italic=existing_font.italic,
                    color=Colors.DARK_RED
                )
    
    def _write_added_rows_section(self, worksheet, result: ComparisonResult) -> None:
        """Write added rows section at the bottom."""
        # Filter out completely empty rows (all values are None, NaN, or empty string)
        non_empty_added_rows = []
        for row_data in result.added_rows:
            # Check if at least one value is non-empty
            has_content = any(
                pd.notna(row_data.get(col)) and str(row_data.get(col)).strip() != ''
                for col in self.column_names
            )
            if has_content:
                non_empty_added_rows.append(row_data)
        
        # If no non-empty added rows, skip this section entirely
        if not non_empty_added_rows:
            return
        
        start_row = len(result.data) + 5  # Leave some space
        
        # Section header
        header_text = self._get_localized_text("Added Rows", "Hinzugefügte Zeilen")
        worksheet.cell(start_row, 1, header_text)
        
        # Column headers
        for col_idx, col_name in enumerate(self.column_names, start=1):
            worksheet.cell(start_row + 1, col_idx, col_name)
        
        # Added row data with GREEN background (only non-empty rows)
        for idx, row_data in enumerate(non_empty_added_rows):
            excel_row = start_row + 2 + idx
            for col_idx, col_name in enumerate(self.column_names, start=1):
                cell = worksheet.cell(excel_row, col_idx)
                value = row_data.get(col_name, "")
                cell.value = value
                # Apply GREEN background to added rows
                cell.fill = PatternFill(start_color=Colors.GREEN, end_color=Colors.GREEN, fill_type='solid')
    
    def _write_deleted_rows_section(self, worksheet, result: ComparisonResult) -> None:
        """Write deleted rows section at the bottom."""
        # Filter out completely empty rows (all values are None, NaN, or empty string)
        non_empty_deleted_rows = []
        for row_data in result.deleted_rows:
            # Check if at least one value is non-empty
            has_content = any(
                pd.notna(row_data.get(col)) and str(row_data.get(col)).strip() != ''
                for col in self.column_names
            )
            if has_content:
                non_empty_deleted_rows.append(row_data)
        
        # If no non-empty deleted rows, skip this section entirely
        if not non_empty_deleted_rows:
            return
        
        # Calculate start row based on whether added rows exist
        start_row = len(result.data) + 5
        if result.added_rows:
            start_row += len(result.added_rows) + 4  # Added rows + header + spacing
        
        # Section header
        header_text = self._get_localized_text("Deleted Rows", "Gelöschte Zeilen")
        worksheet.cell(start_row, 1, header_text)
        
        # Column headers
        for col_idx, col_name in enumerate(self.column_names, start=1):
            worksheet.cell(start_row + 1, col_idx, col_name)
        
        # Deleted row data with RED background (only non-empty rows)
        for idx, row_data in enumerate(non_empty_deleted_rows):
            excel_row = start_row + 2 + idx
            for col_idx, col_name in enumerate(self.column_names, start=1):
                cell = worksheet.cell(excel_row, col_idx)
                value = row_data.get(col_name, "")
                cell.value = value
                # Apply RED background to deleted rows
                cell.fill = PatternFill(start_color=Colors.RED, end_color=Colors.RED, fill_type='solid')
    
    def _write_change_marker_column(self, worksheet, result: ComparisonResult) -> None:
        """Write change marker column."""
        marker_col = len(self.column_names) + 2  # After data columns + empty column
        
        # Set column width to 16 for the "Changed" column
        marker_col_letter = get_column_letter(marker_col)
        worksheet.column_dimensions[marker_col_letter].width = 16
        
        # Header
        header_text = self._get_localized_text("Changed", "Geändert")
        worksheet.cell(2, marker_col, header_text)
        
        # Markers for main data
        for row_idx, marker_value in result.markers.items():
            excel_row = row_idx + 3
            worksheet.cell(excel_row, marker_col, marker_value)
        
        # Markers for added rows
        if result.added_rows:
            marker_value = self._get_localized_text("Added", "Hinzugefügt")
            start_row = len(result.data) + 7  # +7 for header row
            for idx in range(len(result.added_rows)):
                worksheet.cell(row=start_row + idx, column=marker_col, value=marker_value)
        
        # Markers for deleted rows
        if result.deleted_rows:
            marker_value = self._get_localized_text("Deleted", "Gelöscht")
            start_row = len(result.data) + 7
            if result.added_rows:
                start_row += len(result.added_rows) + 4  # Added rows + header + spacing
            for idx in range(len(result.deleted_rows)):
                worksheet.cell(row=start_row + idx, column=marker_col, value=marker_value)
    
    def _apply_autofilter(self, worksheet, result: ComparisonResult) -> None:
        """Apply AutoFilter to the data range."""
        marker_col = len(self.column_names) + 2
        last_row = len(result.data) + 3
        if result.added_rows:
            last_row += len(result.added_rows) + 3
        if result.deleted_rows:
            last_row += len(result.deleted_rows) + 3
        
        last_col_letter = get_column_letter(marker_col)
        worksheet.auto_filter.ref = f"A2:{last_col_letter}{last_row}"
    
    def run(self) -> None:
        """Execute complete comparison workflow."""
        start_time = datetime.now()
        
        try:
            self.logger.info("=" * 60)
            self.logger.info("Excel Table Comparison - Starting")
            self.logger.info("=" * 60)
            
            # Step 1: Validate and load data
            self.validate_files_and_load_data()
            
            # Step 2: Create unique keys for duplicate handling
            old_with_keys, new_with_keys = self.create_unique_keys()
            
            # Step 3: Perform comparison (NEW base only)
            new_base_result = self.perform_comparison(old_with_keys, new_with_keys)
            
            # Step 4: Generate Excel output
            self.generate_excel_output(new_base_result)
            
            duration = (datetime.now() - start_time).total_seconds()
            
            self.logger.info("=" * 60)
            self.logger.info(f"Comparison completed successfully in {duration:.2f} seconds")
            self.logger.info("=" * 60)
            
            print("\n" + "=" * 60)
            print("COMPARISON COMPLETE!")
            print("=" * 60)
            print(f"✓ Output file created: {self.output_file}")
            print(f"✓ Duration: {duration:.2f} seconds")
            print("=" * 60 + "\n")
            
        except Exception as e:
            self.logger.error(f"Error during comparison: {str(e)}", exc_info=True)
            print(f"\n❌ ERROR: {str(e)}\n")


def extract_sheets_and_compare(compare_file, output_file=None):
    """
    Extract old/new sheets from compare.xlsx and run comparison.
    The RIGHTMOST sheet is treated as "new" and the sheet immediately to its left is "old".
    
    Args:
        compare_file: Path to Excel file with multiple sheets
        output_file: Output filename for the new_color_coded sheet (auto-generates if None)
    """
    try:
        print(f"🔍 Processing: {compare_file}")
        print("=" * 50)
        
        # Load the compare.xlsx file
        wb_compare = load_workbook(compare_file)
        
        # Get sheet names
        all_sheets = wb_compare.sheetnames
        
        if len(all_sheets) < 2:
            print("❌ Error: File must contain at least 2 sheets")
            return False
        
        # Rightmost sheet is "new", the one to its left is "old"
        new_sheet_name = all_sheets[-1]  # Last sheet (rightmost)
        old_sheet_name = all_sheets[-2]  # Second to last (left of rightmost)
        
        print(f"📊 OLD sheet: '{old_sheet_name}'")
        print(f"📊 NEW sheet: '{new_sheet_name}'")
        
        # Create temporary old.xlsx and new.xlsx files
        print("📋 Extracting sheets...")
        
        # Helper function to check if a row is a legend row
        def is_legend_row(row):
            legend_keywords = ['Changed', 'Added', 'Deleted', 'Geändert', 'Hinzugefügt', 'Gelöscht']
            row_values = [str(cell.value).strip() for cell in row if cell.value is not None]
            return any(keyword in row_values for keyword in legend_keywords)
        
        # Create old.xlsx with formatting
        wb_old = Workbook()
        wb_old.remove(wb_old.active)
        old_sheet = wb_old.create_sheet("Old Table")
        compare_old_sheet = wb_compare[old_sheet_name]
        
        # Check if first row is legend and skip it
        rows_to_copy = list(compare_old_sheet.iter_rows())
        start_row = 0
        if rows_to_copy and is_legend_row(rows_to_copy[0]):
            print(f"  ⚠️  Legend row detected in '{old_sheet_name}' - skipping first row")
            start_row = 1
        
        # Copy data and formatting from compare.xlsx old sheet (skipping legend if present)
        for row_idx, row in enumerate(rows_to_copy[start_row:], start=1):
            for cell in row:
                # Map to new position (row_idx instead of original row number)
                new_cell = old_sheet.cell(row=row_idx, column=cell.column)
                new_cell.value = cell.value
                
                # Copy cell formatting
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        # Copy column widths
        for col_letter in compare_old_sheet.column_dimensions:
            if compare_old_sheet.column_dimensions[col_letter].width:
                old_sheet.column_dimensions[col_letter].width = compare_old_sheet.column_dimensions[col_letter].width
        
        # Copy row heights
        for row_num in compare_old_sheet.row_dimensions:
            if compare_old_sheet.row_dimensions[row_num].height:
                old_sheet.row_dimensions[row_num].height = compare_old_sheet.row_dimensions[row_num].height
        
        # Copy freeze panes
        if compare_old_sheet.freeze_panes:
            old_sheet.freeze_panes = compare_old_sheet.freeze_panes
        
        wb_old.save("temp_old.xlsx")
        
        # Create new.xlsx (formatting not needed as we use old sheet formatting)
        wb_new = Workbook()
        wb_new.remove(wb_new.active)
        new_sheet = wb_new.create_sheet("New Table")
        compare_new_sheet = wb_compare[new_sheet_name]
        
        # Check if first row is legend and skip it
        rows_to_copy_new = list(compare_new_sheet.iter_rows())
        start_row_new = 0
        if rows_to_copy_new and is_legend_row(rows_to_copy_new[0]):
            print(f"  ⚠️  Legend row detected in '{new_sheet_name}' - skipping first row")
            start_row_new = 1
        
        # Copy only data from compare.xlsx new sheet (formatting comes from old)
        for row_idx, row in enumerate(rows_to_copy_new[start_row_new:], start=1):
            for cell in row:
                new_cell = new_sheet.cell(row=row_idx, column=cell.column)
                new_cell.value = cell.value
        
        wb_new.save("temp_new.xlsx")
        
        # Run the comparator
        print("🚀 Running comparison (Interface No.-based)...")
        
        comparator = ExcelTableComparator(
            old_file='temp_old.xlsx',
            new_file='temp_new.xlsx', 
            output_file=None  # Will be written back to input file
        )
        
        # Perform comparison without saving to file yet
        comparator.validate_files_and_load_data()
        old_with_keys, new_with_keys = comparator.create_unique_keys()
        new_base_result = comparator.perform_comparison(old_with_keys, new_with_keys)
        
        # Capture formatting from old sheet
        formatting = comparator._capture_formatting_from_old_sheet()
        
        # Generate sheet name based on today's date: YYMMDD_compared
        today = datetime.now()
        base_sheet_name = today.strftime("%y%m%d") + "_compared"
        
        # Find unique sheet name by appending _1, _2, etc. if needed
        sheet_name = base_sheet_name
        counter = 1
        while sheet_name in wb_compare.sheetnames:
            sheet_name = f"{base_sheet_name}_{counter}"
            counter += 1
        
        print(f"📝 Creating new sheet: '{sheet_name}'")
        
        # Create the new sheet at the rightmost position
        ws_result = wb_compare.create_sheet(sheet_name)
        
        # Write comparison results to the new sheet
        comparator._write_comparison_sheet(ws_result, new_base_result, formatting)
        
        # Deselect all sheets first
        for sheet in wb_compare.worksheets:
            sheet.sheet_view.tabSelected = False
        
        # Make only the new sheet active and selected
        wb_compare.active = wb_compare.sheetnames.index(sheet_name)
        ws_result.sheet_view.tabSelected = True
        
        # Save back to the input file
        wb_compare.save(compare_file)
        
        # Clean up temporary files
        Path("temp_old.xlsx").unlink(missing_ok=True)
        Path("temp_new.xlsx").unlink(missing_ok=True)
        
        print(f"✅ Success! New sheet '{sheet_name}' added to: {compare_file}")
        print("📊 Contains comparison with change analysis")
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        # Clean up on error
        Path("temp_old.xlsx").unlink(missing_ok=True)
        Path("temp_new.xlsx").unlink(missing_ok=True)
        return False


def main():
    """Main command-line interface."""
    # Default input file path
    default_input_file = r"C:\Users\szil\Repos\excel_wizadry\Interface_List_Compare\compare.xlsx"
    
    parser = argparse.ArgumentParser(
        description='Compare Excel tables using Interface No. as unique identifier',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_compare.py                              # Uses default compare.xlsx
  python excel_compare.py compare.xlsx
  python excel_compare.py compare.xlsx my_result.xlsx
  
The input file must contain 'old' and 'new' sheets.
Output will contain a single 'new_color_coded' sheet with changes highlighted.
        """
    )
    
    parser.add_argument('input_file', nargs='?',
                       default=default_input_file,
                       help=f'Excel file containing old and new sheets (default: {default_input_file})')
    parser.add_argument('output_file', nargs='?', 
                       default=None,
                       help='Output Excel file (auto-generates with datetime if not provided)')
    
    args = parser.parse_args()
    
    # Validate input file
    if not Path(args.input_file).exists():
        print(f"❌ Error: Input file '{args.input_file}' not found")
        sys.exit(1)
    
    # Run comparison
    success = extract_sheets_and_compare(args.input_file, args.output_file)
    
    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()